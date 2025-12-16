from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from database import get_db, log_islem
from datetime import datetime, date
from werkzeug.utils import secure_filename
from email_service import yeni_ise_giris_bildirimi, isten_cikis_bildirimi
import os
import sqlite3
import hashlib
import pandas as pd
import io
from functools import wraps
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

app = Flask(__name__)
app.secret_key = 'ozanerenkayan1988'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = 3600

# Dosya yÃ¼kleme ayarlarÄ±
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'gif'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

def allowed_file(filename):
    """Dosya uzantÄ±sÄ± kontrol"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_file_size_mb(size_bytes):
    """Dosya boyutunu MB olarak dÃ¶ndÃ¼r"""
    return round(size_bytes / (1024 * 1024), 2)

# Login gerekli decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Bu sayfayÄ± gÃ¶rÃ¼ntÃ¼lemek iÃ§in giriÅŸ yapmalÄ±sÄ±nÄ±z.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin gerekli decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Bu sayfayÄ± gÃ¶rÃ¼ntÃ¼lemek iÃ§in giriÅŸ yapmalÄ±sÄ±nÄ±z.', 'warning')
            return redirect(url_for('login'))
        
        conn = get_db()
        user = conn.execute('SELECT role FROM users WHERE id = %s', (session['user_id'],)).fetchone()
        conn.close()
        
        if not user or user['role'] != 'admin':
            flash('Bu sayfaya eriÅŸim yetkiniz yok.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function
    
# Manager veya Admin gerekli decorator
def manager_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Bu sayfayÄ± gÃ¶rÃ¼ntÃ¼lemek iÃ§in giriÅŸ yapmalÄ±sÄ±nÄ±z.', 'warning')
            return redirect(url_for('login'))
        
        conn = get_db()
        user = conn.execute('SELECT role FROM users WHERE id = %s', (session['user_id'],)).fetchone()
        conn.close()
        
        if not user or user['role'] not in ['admin', 'manager']:
            flash('Bu sayfaya eriÅŸim yetkiniz yok.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# Yetki kontrol fonksiyonu
def has_permission(action):
    """
    KullanÄ±cÄ±nÄ±n belirli bir iÅŸlem iÃ§in yetkisi var mÄ±?
    action: 'view', 'add', 'edit', 'delete'
    """
    if 'user_id' not in session:
        return False
    
    role = session.get('role', 'user')
    
    permissions = {
        'admin': ['view', 'add', 'edit', 'delete', 'manage_users'],
        'manager': ['view', 'add', 'edit', 'delete'],
        'user': ['view', 'add']
    }
    
    return action in permissions.get(role, [])  

# Template'lerde kullanÄ±lacak fonksiyonlar
@app.context_processor
def utility_processor():
    """Template'lerde kullanÄ±lacak yardÄ±mcÄ± fonksiyonlar"""
    def check_permission(action):
        return has_permission(action)
    return dict(has_permission=check_permission)    
    
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login sayfasÄ±"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        conn = get_db()
        user = conn.execute('''
            SELECT * FROM users 
            WHERE username = %s AND password_hash = %s AND aktif = TRUE
        ''', (username, password_hash)).fetchone()
        
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['full_name'] = user['full_name']
            session['role'] = user['role']
            
            # Son giriÅŸ tarihini gÃ¼ncelle
            conn.execute('UPDATE users SET son_giris_tarihi = %s WHERE id = %s', 
                        (datetime.now(), user['id']))
            conn.commit()
            conn.close()
            
            log_islem('GÄ°RÄ°Åž', 'users', user['id'], f'{username} sisteme giriÅŸ yaptÄ±')
            flash(f'HoÅŸ geldiniz, {user["full_name"]}!', 'success')
            return redirect(url_for('index'))
        else:
            conn.close()
            flash('KullanÄ±cÄ± adÄ± veya ÅŸifre hatalÄ±!', 'danger')
    
    return render_template('login.html')
   


@app.route('/logout')
def logout():
    """Logout"""
    username = session.get('username', 'Bilinmeyen')
    user_id = session.get('user_id')
    
    if user_id:
        log_islem('Ã‡IKIÅž', 'users', user_id, f'{username} sistemden Ã§Ä±kÄ±ÅŸ yaptÄ±')
    
    session.clear()
    flash('BaÅŸarÄ±yla Ã§Ä±kÄ±ÅŸ yaptÄ±nÄ±z.', 'info')
    return redirect(url_for('login'))    

# app.py iÃ§indeki index() fonksiyonunu bu kodla deÄŸiÅŸtir

@app.route('/')
@login_required
def index():
    """Ana sayfa - Dashboard"""
    conn = get_db()
    
    # Proje bazÄ±nda Ã¶zet
    projeler = conn.execute('''
        SELECT 
            p.id,
            p.proje_adi,
            COUNT(DISTINCT hk.id) as toplam_kadro,
            SUM(hk.hedef_kisi_sayisi) as toplam_hedef,
            SUM(hk.dolu_kisi_sayisi) as toplam_dolu,
            COUNT(DISTINCT a.id) as toplam_aday,
            COUNT(DISTINCT c.id) as toplam_calisan
        FROM projeler p
        LEFT JOIN hedef_kadrolar hk ON p.id = hk.proje_id
        LEFT JOIN adaylar a ON hk.id = a.kadro_id
        LEFT JOIN calisanlar c ON hk.id = c.kadro_id AND c.aktif = TRUE
        WHERE p.aktif = TRUE
        GROUP BY p.id, p.proje_adi
    ''').fetchall()
    
    # Kadro bazÄ±nda detay
    kadrolar = conn.execute('''
        SELECT 
            hk.id,
            hk.proje_id,
            hk.mudurluk_id,
            hk.direktorluk_id,
            hk.il_id,
            hk.ilce_id,
            hk.pozisyon_adi,
            hk.magaza_adi,
            hk.hedef_kisi_sayisi,
            hk.dolu_kisi_sayisi,
            hk.durum,
            hk.olusturma_tarihi,
            p.proje_adi,
            m.mudurluk_adi,
            d.direktorluk_adi,
            i.il_adi,
            ic.ilce_adi,
            COUNT(DISTINCT a.id) as aday_sayisi,
            COUNT(DISTINCT CASE WHEN c.aktif = TRUE THEN c.id END) as calisan_sayisi
        FROM hedef_kadrolar hk
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN adaylar a ON hk.id = a.kadro_id
        LEFT JOIN calisanlar c ON hk.id = c.kadro_id
        GROUP BY hk.id, hk.proje_id, hk.mudurluk_id, hk.direktorluk_id, hk.il_id, hk.ilce_id,
                 hk.pozisyon_adi, hk.magaza_adi, hk.hedef_kisi_sayisi, hk.dolu_kisi_sayisi, hk.durum, hk.olusturma_tarihi,
                 p.proje_adi, m.mudurluk_adi, d.direktorluk_adi, i.il_adi, ic.ilce_adi
        ORDER BY p.proje_adi, i.il_adi, ic.ilce_adi
    ''').fetchall()
    
    # Son iÅŸlemler
    son_islemler = conn.execute('''
        SELECT * FROM surec_loglari 
        ORDER BY islem_tarihi DESC 
        LIMIT 10
    ''').fetchall()
    
    # === GRAFÄ°K VERÄ°LERÄ° ===
    
    # 1. Ä°l bazÄ±nda Ã§alÄ±ÅŸan daÄŸÄ±lÄ±mÄ± (Top 10)
    il_dagilim = conn.execute('''
        SELECT 
            i.il_adi,
            COUNT(DISTINCT c.id) as calisan_sayisi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN iller i ON hk.il_id = i.id
        WHERE c.aktif = TRUE
        GROUP BY i.il_adi
        ORDER BY calisan_sayisi DESC
        LIMIT 10
    ''').fetchall()
    
    # 2. AylÄ±k iÅŸe giriÅŸ trendi (Son 6 ay)
    aylik_giris = conn.execute('''
        SELECT 
            TO_CHAR(ise_baslama_tarihi, 'YYYY-MM') as ay,
            COUNT(*) as sayi
        FROM calisanlar
        WHERE ise_baslama_tarihi >= CURRENT_DATE - INTERVAL '6 months'
        AND aktif = TRUE
        GROUP BY TO_CHAR(ise_baslama_tarihi, 'YYYY-MM')
        ORDER BY ay
    ''').fetchall()
    
    # 3. AraÃ§lÄ±/AraÃ§sÄ±z daÄŸÄ±lÄ±m
    arac_dagilim = conn.execute('''
        SELECT 
            hk.aracli_durum,
            COUNT(DISTINCT c.id) as calisan_sayisi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        WHERE c.aktif = TRUE
        GROUP BY hk.aracli_durum
    ''').fetchall()
    
    # 4. MÃ¼dÃ¼rlÃ¼k bazÄ±nda daÄŸÄ±lÄ±m (varsa)
    mudurluk_dagilim = conn.execute('''
        SELECT 
            COALESCE(m.mudurluk_adi, 'TanÄ±msÄ±z') as mudurluk,
            COUNT(DISTINCT c.id) as calisan_sayisi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        WHERE c.aktif = TRUE
        GROUP BY mudurluk
        ORDER BY calisan_sayisi DESC
        LIMIT 8
    ''').fetchall()
    
    # 5. HARÄ°TA Ä°Ã‡Ä°N: TÃœM Ä°LLER VERÄ°SÄ° - YENÄ°
    harita_verisi = conn.execute('''
        SELECT 
            i.il_adi,
            COUNT(DISTINCT hk.id) as kadro_sayisi,
            COUNT(DISTINCT c.id) as calisan_sayisi,
            COUNT(DISTINCT a.id) as aday_sayisi
        FROM iller i
        LEFT JOIN hedef_kadrolar hk ON i.id = hk.il_id
        LEFT JOIN calisanlar c ON hk.id = c.kadro_id AND c.aktif = TRUE
        LEFT JOIN adaylar a ON hk.id = a.kadro_id AND a.durum = 'Aday'
        GROUP BY i.id, i.il_adi
        ORDER BY i.il_adi
    ''').fetchall()
    
    # 6. Genel istatistikler
    stats = {
        'toplam_proje': conn.execute('SELECT COUNT(*) as cnt FROM projeler WHERE aktif = TRUE').fetchone()['cnt'],
        'toplam_kadro': conn.execute('SELECT COUNT(*) as cnt FROM hedef_kadrolar').fetchone()['cnt'],
        'toplam_aday': conn.execute('SELECT COUNT(*) as cnt FROM adaylar').fetchone()['cnt'],
        'toplam_calisan': conn.execute('SELECT COUNT(*) as cnt FROM calisanlar WHERE aktif = TRUE').fetchone()['cnt'],
        'hedef_toplam': conn.execute('SELECT SUM(hedef_kisi_sayisi) as total FROM hedef_kadrolar').fetchone()['total'] or 0,
        'dolu_toplam': conn.execute('SELECT SUM(dolu_kisi_sayisi) as total FROM hedef_kadrolar').fetchone()['total'] or 0,
        'acik_kadro': conn.execute("SELECT COUNT(*) as cnt FROM hedef_kadrolar WHERE durum = 'AÃ§Ä±k'").fetchone()['cnt'],
        'dolu_kadro': conn.execute("SELECT COUNT(*) as cnt FROM hedef_kadrolar WHERE durum = 'Dolu'").fetchone()['cnt']
    }
    
    conn.close()
    
    return render_template('index.html', 
                         projeler=projeler, 
                         kadrolar=kadrolar,
                         son_islemler=son_islemler,
                         il_dagilim=il_dagilim,
                         aylik_giris=aylik_giris,
                         arac_dagilim=arac_dagilim,
                         mudurluk_dagilim=mudurluk_dagilim,
                         harita_verisi=harita_verisi,  # YENÄ°
                         stats=stats)

@app.route('/projeler')
@login_required
def projeler():
    """Proje listesi"""
    conn = get_db()
    projeler = conn.execute('''
        SELECT p.*, m.musteri_adi, m.logo_yolu
        FROM projeler p
        LEFT JOIN musteriler m ON p.musteri_id = m.id
        ORDER BY p.olusturma_tarihi DESC
    ''').fetchall()
    conn.close()
    return render_template('projeler.html', projeler=projeler)

@app.route('/proje/ekle', methods=['GET', 'POST'])
@manager_required
def proje_ekle():
    """Yeni proje ekle"""
    if request.method == 'POST':
        proje_adi = request.form['proje_adi']
        aciklama = request.form.get('aciklama', '')
        musteri_id = request.form.get('musteri_id') or None
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO projeler (proje_adi, aciklama, musteri_id) VALUES (%s, %s, %s)', 
                      (proje_adi, aciklama, musteri_id))
        proje_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'projeler', proje_id, f'{proje_adi} projesi oluÅŸturuldu')
        flash('Proje baÅŸarÄ±yla eklendi!', 'success')
        return redirect(url_for('projeler'))
    
    # MÃ¼ÅŸteri listesi
    conn = get_db()
    musteriler = conn.execute('SELECT * FROM musteriler WHERE aktif = TRUE ORDER BY musteri_adi').fetchall()
    conn.close()
    
    return render_template('proje_form.html', musteriler=musteriler)

@app.route('/kadrolar')
@login_required
def kadrolar():
    """Kadro listesi"""
    conn = get_db()
    kadrolar = conn.execute('''
        SELECT hk.id,
               hk.proje_id,
               hk.mudurluk_id,
               hk.direktorluk_id,
               hk.il_id,
               hk.ilce_id,
               p.proje_adi,
               hk.pozisyon_adi,
               hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               hk.magaza_adi,
               hk.aracli_durum,
               hk.hedef_kisi_sayisi,
               hk.dolu_kisi_sayisi,
               hk.durum,
               hk.olusturma_tarihi,
               COUNT(DISTINCT a.id) as aday_sayisi,
               COUNT(DISTINCT CASE WHEN c.aktif = TRUE THEN c.id END) as calisan_sayisi
        FROM hedef_kadrolar hk
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN adaylar a ON hk.id = a.kadro_id
        LEFT JOIN calisanlar c ON hk.id = c.kadro_id
        GROUP BY hk.id, hk.proje_id, hk.mudurluk_id, hk.direktorluk_id, hk.il_id, hk.ilce_id,
                 p.proje_adi, hk.pozisyon_adi, hk.calisma_sekli, m.mudurluk_adi, d.direktorluk_adi,
                 i.il_adi, ic.ilce_adi, hk.magaza_adi, hk.aracli_durum, hk.hedef_kisi_sayisi,
                 hk.dolu_kisi_sayisi, hk.durum, hk.olusturma_tarihi
        ORDER BY p.proje_adi, hk.olusturma_tarihi DESC
    ''').fetchall()
    conn.close()
    return render_template('kadrolar.html', kadrolar=kadrolar)

@app.route('/kadro/ekle', methods=['GET', 'POST'])
@login_required
def kadro_ekle():
    """Yeni kadro ekle"""
    if request.method == 'POST':
        # Form verilerini al
        proje_id = request.form['proje_id']
        pozisyon_adi = request.form['pozisyon_adi']
        calisma_sekli = request.form['calisma_sekli']  # YENÄ°
        mudurluk_id = request.form.get('mudurluk_id') or None
        direktorluk_id = request.form.get('direktorluk_id') or None
        il_id = request.form['il_id']
        ilce_id = request.form.get('ilce_id') or None
        magaza_adi = request.form.get('magaza_adi', '')
        hedef_kisi_sayisi = request.form['hedef_kisi_sayisi']
        aracli_durum = request.form['aracli_durum']
        
        conn = get_db()
        cursor = conn.cursor()
        
        # Kadro ekle
        cursor.execute('''
    INSERT INTO hedef_kadrolar 
    (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id, 
     magaza_adi, hedef_kisi_sayisi, dolu_kisi_sayisi, aracli_durum, calisma_sekli)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
''', (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id, 
      magaza_adi, hedef_kisi_sayisi, 0, aracli_durum, calisma_sekli))
        
        kadro_id = cursor.lastrowid
        
        # Durum gÃ¼ncelle (baÅŸlangÄ±Ã§ta her zaman AÃ§Ä±k)
        cursor.execute('UPDATE hedef_kadrolar SET durum = %s WHERE id = %s', ('AÃ§Ä±k', kadro_id))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'hedef_kadrolar', kadro_id, f'{pozisyon_adi} kadrosu eklendi')
        flash('Kadro baÅŸarÄ±yla eklendi!', 'success')
        return redirect(url_for('kadrolar'))
    
    # GET request - form iÃ§in gerekli verileri Ã§ek
    conn = get_db()
    
    projeler = conn.execute('''
        SELECT * FROM projeler 
        WHERE aktif = TRUE 
        ORDER BY proje_adi
    ''').fetchall()
    
    mudurluker = conn.execute('''
        SELECT * FROM mudurluker 
        ORDER BY mudurluk_adi
    ''').fetchall()
    
    direktorlukler = conn.execute('''
        SELECT * FROM direktorlukler 
        ORDER BY direktorluk_adi
    ''').fetchall()
    
    iller = conn.execute('''
        SELECT * FROM iller 
        ORDER BY il_adi
    ''').fetchall()
    
    # Ã‡alÄ±ÅŸma ÅŸekilleri - YENÄ°
    calisma_sekilleri = conn.execute('''
        SELECT calisma_sekli FROM calisma_sekilleri 
        WHERE aktif = TRUE 
        ORDER BY calisma_sekli
    ''').fetchall()
    
    conn.close()
    
    return render_template('kadro_form.html', 
                         projeler=projeler, 
                         mudurluker=mudurluker,
                         direktorlukler=direktorlukler,
                         iller=iller,
                         calisma_sekilleri=calisma_sekilleri)

@app.route('/adaylar')
@login_required
def adaylar():
    """Aday listesi"""
    conn = get_db()
    adaylar_raw = conn.execute('''
        SELECT a.*, 
               hk.pozisyon_adi, hk.magaza_adi, hk.aracli_durum,
               hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi,
               k.kaynak_adi
        FROM adaylar a
        LEFT JOIN hedef_kadrolar hk ON a.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN kaynaklar k ON a.kaynak_id = k.id
        ORDER BY a.basvuru_tarihi DESC
    ''').fetchall()
    
    # YaÅŸ hesaplama
    adaylar = []
    for aday in adaylar_raw:
        aday_dict = dict(aday)
        
        # YaÅŸ hesapla
        if aday_dict.get('dogum_tarihi'):
            try:
                dogum_tarihi = aday_dict['dogum_tarihi']
                
                # PostgreSQL date objesi veya string - ikisini de destekle
                if isinstance(dogum_tarihi, str):
                    dogum = datetime.strptime(dogum_tarihi, '%Y-%m-%d')
                elif hasattr(dogum_tarihi, 'year'):  # date objesi
                    dogum = datetime(dogum_tarihi.year, dogum_tarihi.month, dogum_tarihi.day)
                else:
                    raise ValueError("Bilinmeyen tarih formati")
                
                bugun = datetime.now()
                
                yil = bugun.year - dogum.year
                ay = bugun.month - dogum.month
                
                if ay < 0:
                    yil -= 1
                    ay += 12
                
                if bugun.day < dogum.day:
                    ay -= 1
                    if ay < 0:
                        yil -= 1
                        ay += 12
                
                aday_dict['yas'] = f"{yil}.{ay}"
                aday_dict['yas_yil'] = yil
            except Exception as e:
                print(f"Yas hesaplama hatasi (aday): {e}")
                aday_dict['yas'] = None
                aday_dict['yas_yil'] = None
        else:
            aday_dict['yas'] = None
            aday_dict['yas_yil'] = None
        
        adaylar.append(aday_dict)
    
    conn.close()
    return render_template('adaylar.html', adaylar=adaylar)

@app.route('/aday/ekle', methods=['GET', 'POST'])
@login_required
def aday_ekle():
    """Yeni aday ekle"""
    if request.method == 'POST':
        # Form verilerini al
        kadro_id = request.form['kadro_id']
        ad_soyad = request.form['ad_soyad']
        telefon = request.form.get('telefon', '')
        email = request.form.get('email', '')
        tc_kimlik = request.form.get('tc_kimlik', '')
        dogum_tarihi = request.form.get('dogum_tarihi', None)  # â† BURASI VAR MI?
        notlar = request.form.get('notlar', '')
        
        # KAYNAK BÄ°LGÄ°SÄ°
        kaynak_id = request.form.get('kaynak_id') or None
        kaynak_diger = request.form.get('kaynak_diger', '').strip()
        
        # EÄŸer "DiÄŸer" seÃ§ilmiÅŸse kaynak_id'yi NULL yap
        if kaynak_id and int(kaynak_id) == -1:
            kaynak_id = None
        
        # DEBUG: Form verilerini kontrol et
        print(f"DEBUG - Dogum Tarihi: [{dogum_tarihi}]")  # â† DEBUG EKLENDÄ°
        
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO adaylar (kadro_id, ad_soyad, telefon, email, tc_kimlik, dogum_tarihi, notlar, kaynak_id, kaynak_diger)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (kadro_id, ad_soyad, telefon or None, email or None, tc_kimlik or None, 
              dogum_tarihi or None, notlar or None, kaynak_id, kaynak_diger if kaynak_diger else None))
        
        aday_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'adaylar', aday_id, f'{ad_soyad} adayÄ± sisteme eklendi')
        flash('Aday baÅŸarÄ±yla eklendi!', 'success')
        return redirect(url_for('adaylar'))
    
    # GET request iÃ§in
    conn = get_db()
    kadrolar = conn.execute('''
        SELECT hk.*, 
               p.proje_adi,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi
        FROM hedef_kadrolar hk
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        WHERE hk.durum = 'AÃ§Ä±k'
        ORDER BY p.proje_adi, hk.pozisyon_adi
    ''').fetchall()
    
    # Kaynaklar listesi
    kaynaklar = conn.execute('''
        SELECT * FROM kaynaklar 
        WHERE aktif = TRUE 
        ORDER BY kaynak_adi
    ''').fetchall()
    
    conn.close()
    return render_template('aday_form.html', kadrolar=kadrolar, kaynaklar=kaynaklar)

@app.route('/aday/<int:aday_id>/calisana-donustur', methods=['POST'])
@manager_required
def aday_calisana_donustur(aday_id):
    """AdayÄ± Ã§alÄ±ÅŸana dÃ¶nÃ¼ÅŸtÃ¼r"""
    ise_baslama_tarihi = request.form.get('ise_baslama_tarihi', datetime.now().strftime('%Y-%m-%d'))
    
    conn = get_db()
    cursor = conn.cursor()
    
    # Aday ve kadro bilgilerini tek sorguda al
    aday_detay_raw = conn.execute('''
    SELECT a.*, 
           hk.pozisyon_adi, 
           hk.magaza_adi, 
           hk.proje_id,
           hk.aracli_durum,
           hk.calisma_sekli,
           m.mudurluk_adi,
           d.direktorluk_adi,
           i.il_adi,
           ic.ilce_adi,
           p.proje_adi
    FROM adaylar a
    LEFT JOIN hedef_kadrolar hk ON a.kadro_id = hk.id
    LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
    LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
    LEFT JOIN iller i ON hk.il_id = i.id
    LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
    LEFT JOIN projeler p ON hk.proje_id = p.id
    WHERE a.id = %s
    ''', (aday_id,)).fetchone()
    
    if not aday_detay_raw:
        conn.close()
        flash('Aday bulunamadÄ±!', 'danger')
        return redirect(url_for('adaylar'))
    
    # Dict'e Ã§evir
    aday_detay = dict(aday_detay_raw)
    
    # Ã‡alÄ±ÅŸan olarak ekle
    cursor.execute('''
        INSERT INTO calisanlar (aday_id, kadro_id, ad_soyad, telefon, email, tc_kimlik, 
                               dogum_tarihi, ise_baslama_tarihi, aracli_durum)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (aday_id, aday_detay['kadro_id'], aday_detay['ad_soyad'], aday_detay['telefon'], 
          aday_detay['email'], aday_detay['tc_kimlik'], aday_detay.get('dogum_tarihi'),
          ise_baslama_tarihi, aday_detay['aracli_durum']))
    
    # Aday durumunu gÃ¼ncelle
    cursor.execute('UPDATE adaylar SET durum = %s, ise_baslama_tarihi = %s WHERE id = %s',
                   ('Ã‡alÄ±ÅŸan', ise_baslama_tarihi, aday_id))
    
    # Kadro dolu sayÄ±sÄ±nÄ± artÄ±r
    cursor.execute('''
        UPDATE hedef_kadrolar 
        SET dolu_kisi_sayisi = dolu_kisi_sayisi + 1
        WHERE id = %s
    ''', (aday_detay['kadro_id'],))
    
    # Kadro durumunu kontrol et ve gÃ¼ncelle
    cursor.execute('''
        UPDATE hedef_kadrolar 
        SET durum = CASE 
            WHEN dolu_kisi_sayisi >= hedef_kisi_sayisi THEN 'Dolu'
            ELSE 'AÃ§Ä±k'
        END
        WHERE id = %s
    ''', (aday_detay['kadro_id'],))
    
    conn.commit()
    
    # Email bildirimi gÃ¶nder
    print("ðŸ”¥ðŸ”¥ðŸ”¥ EMAIL KODU Ã‡ALIÅžTI! ðŸ”¥ðŸ”¥ðŸ”¥")
    try:
        aday_bilgileri = {
            'id': aday_id,
            'ad_soyad': aday_detay['ad_soyad'],
            'tc_kimlik': aday_detay['tc_kimlik'],
            'dogum_tarihi': aday_detay.get('dogum_tarihi'),
            'telefon': aday_detay['telefon'] or '-',
            'email': aday_detay['email'] or '-',
            'ise_baslama_tarihi': ise_baslama_tarihi,
            'proje_adi': aday_detay['proje_adi'],
            'pozisyon_adi': aday_detay['pozisyon_adi'],
            'mudurluk_adi': aday_detay['mudurluk_adi'] or '-',  # â† YENÄ°
            'direktorluk_adi': aday_detay['direktorluk_adi'] or '-',  # â† YENÄ°
            'il': aday_detay['il_adi'],
            'ilce': aday_detay['ilce_adi'] or '-',
            'magaza_adi': aday_detay['magaza_adi'] or '-',
            'aracli_durum': aday_detay['aracli_durum'],
            'calisma_sekli': aday_detay['calisma_sekli'],
        }  
        
        print(f"Aday bilgileri: {aday_bilgileri}")
        
        email_basarili, email_mesaj = yeni_ise_giris_bildirimi(aday_bilgileri)
        
        print(f"Email sonucu: {email_basarili}, {email_mesaj}")
        
        if email_basarili:
            flash(email_mesaj, 'info')
    except Exception as e:
        print(f"âŒ Email gÃ¶nderimi hatasÄ±: {e}")
        import traceback
        traceback.print_exc()
        flash(f'UyarÄ±: Email bildirimi gÃ¶nderilemedi', 'warning')
    
    conn.close()
    
    log_islem('DÃ–NÃœÅžÃœM', 'calisanlar', aday_id, 
             f'{aday_detay["ad_soyad"]} adaydan Ã§alÄ±ÅŸana dÃ¶nÃ¼ÅŸtÃ¼rÃ¼ldÃ¼', 
             session.get('username'))
    flash('Aday baÅŸarÄ±yla Ã§alÄ±ÅŸana dÃ¶nÃ¼ÅŸtÃ¼rÃ¼ldÃ¼!', 'success')
    return redirect(url_for('calisanlar'))

@app.route('/calisanlar')
@login_required
def calisanlar():
    """Ã‡alÄ±ÅŸan listesi"""
    conn = get_db()
    calisanlar_raw = conn.execute('''
        SELECT c.*, 
               hk.pozisyon_adi, hk.magaza_adi, hk.aracli_durum,
               hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        WHERE c.aktif = TRUE
        ORDER BY c.ise_baslama_tarihi DESC
    ''').fetchall()
    
    # YaÅŸ hesaplama
    calisanlar = []
    for calisan in calisanlar_raw:
        calisan_dict = dict(calisan)
        
        # YaÅŸ hesapla
        if calisan_dict.get('dogum_tarihi'):
            try:
                dogum_tarihi = calisan_dict['dogum_tarihi']
                
                # PostgreSQL date objesi veya string - ikisini de destekle
                if isinstance(dogum_tarihi, str):
                    dogum = datetime.strptime(dogum_tarihi, '%Y-%m-%d')
                elif hasattr(dogum_tarihi, 'year'):  # date objesi
                    dogum = datetime(dogum_tarihi.year, dogum_tarihi.month, dogum_tarihi.day)
                else:
                    raise ValueError("Bilinmeyen tarih formati")
                
                bugun = datetime.now()
                
                yil = bugun.year - dogum.year
                ay = bugun.month - dogum.month
                
                if ay < 0:
                    yil -= 1
                    ay += 12
                
                if bugun.day < dogum.day:
                    ay -= 1
                    if ay < 0:
                        yil -= 1
                        ay += 12
                
                calisan_dict['yas'] = f"{yil}.{ay}"
                calisan_dict['yas_yil'] = yil
            except Exception as e:
                print(f"Yas hesaplama hatasi (calisan): {e}")
                calisan_dict['yas'] = None
                calisan_dict['yas_yil'] = None
        else:
            calisan_dict['yas'] = None
            calisan_dict['yas_yil'] = None
        
        calisanlar.append(calisan_dict)
    
    conn.close()
    return render_template('calisanlar.html', calisanlar=calisanlar)
    
@app.route('/calisanlar/excel-export')
@login_required
def calisanlar_excel_export():
    """Ã‡alÄ±ÅŸanlarÄ± Excel'e aktar"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    conn = get_db()
    calisanlar = conn.execute('''
        SELECT c.*, 
               hk.pozisyon_adi, hk.magaza_adi, hk.aracli_durum, hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        WHERE c.aktif = TRUE
        ORDER BY c.ise_baslama_tarihi DESC
    ''').fetchall()
    conn.close()
    
    # Excel oluÅŸtur
    wb = Workbook()
    ws = wb.active
    ws.title = "Ã‡alÄ±ÅŸanlar"
    
    # BaÅŸlÄ±k satÄ±rÄ±
    headers = ['ID', 'Ad Soyad', 'Telefon', 'Email', 'TC Kimlik', 
               'DoÄŸum Tarihi', 'YaÅŸ',
               'Proje', 'Pozisyon', 'MÃ¼dÃ¼rlÃ¼k', 'DirektÃ¶rlÃ¼k',
               'Ä°l', 'Ä°lÃ§e', 'MaÄŸaza', 'AraÃ§ Durumu', 'Ã‡alÄ±ÅŸma Åžekli',
               'Ä°ÅŸe BaÅŸlama', 'Durum']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Veri satÄ±rlarÄ±
    for row_idx, calisan in enumerate(calisanlar, 2):
        # YaÅŸ hesapla
        yas_str = '-'
        if calisan['dogum_tarihi']:
            try:
                dogum_tarihi = calisan['dogum_tarihi']
                
                if isinstance(dogum_tarihi, str):
                    dogum = datetime.strptime(dogum_tarihi, '%Y-%m-%d')
                elif hasattr(dogum_tarihi, 'year'):
                    dogum = datetime(dogum_tarihi.year, dogum_tarihi.month, dogum_tarihi.day)
                else:
                    raise ValueError("Bilinmeyen tarih formati")
                
                bugun = datetime.now()
                yil = bugun.year - dogum.year
                ay = bugun.month - dogum.month
                if ay < 0:
                    yil -= 1
                    ay += 12
                if bugun.day < dogum.day:
                    ay -= 1
                    if ay < 0:
                        yil -= 1
                        ay += 12
                yas_str = f"{yil}.{ay}"
            except:
                pass
        
        ws.cell(row=row_idx, column=1, value=calisan['id'])
        ws.cell(row=row_idx, column=2, value=calisan['ad_soyad'])
        ws.cell(row=row_idx, column=3, value=calisan['telefon'] or '')
        ws.cell(row=row_idx, column=4, value=calisan['email'] or '')
        ws.cell(row=row_idx, column=5, value=calisan['tc_kimlik'] or '')
        ws.cell(row=row_idx, column=6, value=calisan['dogum_tarihi'] or '')
        ws.cell(row=row_idx, column=7, value=yas_str)
        ws.cell(row=row_idx, column=8, value=calisan['proje_adi'] or '')
        ws.cell(row=row_idx, column=9, value=calisan['pozisyon_adi'] or '')
        ws.cell(row=row_idx, column=10, value=calisan['mudurluk_adi'] or '')
        ws.cell(row=row_idx, column=11, value=calisan['direktorluk_adi'] or '')
        ws.cell(row=row_idx, column=12, value=calisan['il_adi'] or '')
        ws.cell(row=row_idx, column=13, value=calisan['ilce_adi'] or '')
        ws.cell(row=row_idx, column=14, value=calisan['magaza_adi'] or '')
        ws.cell(row=row_idx, column=15, value=calisan['aracli_durum'] or '')
        ws.cell(row=row_idx, column=16, value=calisan['calisma_sekli'] or '')
        ws.cell(row=row_idx, column=17, value=calisan['ise_baslama_tarihi'])
        ws.cell(row=row_idx, column=18, value='Aktif' if calisan['aktif'] else 'Pasif')
    
    # SÃ¼tun geniÅŸliklerini ayarla
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # DosyayÄ± belleÄŸe kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'calisanlar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )    
    
@app.route('/calisan/<int:calisan_id>/cikis', methods=['GET'])
@login_required
def calisan_cikis(calisan_id):
    """Ä°ÅŸten Ã§Ä±kÄ±ÅŸ formu"""
    conn = get_db()
    
    # Ã‡alÄ±ÅŸan bilgilerini getir
    calisan = conn.execute('''
        SELECT c.*, 
               hk.pozisyon_adi,
               p.proje_adi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        WHERE c.id = %s AND c.aktif = TRUE
    ''', (calisan_id,)).fetchone()
    
    if not calisan:
        flash('Ã‡alÄ±ÅŸan bulunamadÄ± veya zaten pasif durumda!', 'error')
        return redirect(url_for('calisanlar'))
    
    # Ã‡Ä±kÄ±ÅŸ nedenleri
    cikis_nedenleri = conn.execute('''
        SELECT * FROM cikis_nedenleri 
        WHERE aktif = TRUE 
        ORDER BY neden
    ''').fetchall()
    
    conn.close()
    
    # BugÃ¼nÃ¼n tarihi
    today = date.today().isoformat()
    now = datetime.now()
    
    return render_template('calisan_cikis.html', 
                         calisan=calisan,
                         cikis_nedenleri=cikis_nedenleri,
                         today=today,
                         now=now)


@app.route('/calisan/<int:calisan_id>/cikis/kaydet', methods=['POST'])
@login_required
def calisan_cikis_kaydet(calisan_id):
    """Ä°ÅŸten Ã§Ä±kÄ±ÅŸÄ± kaydet"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Ã‡alÄ±ÅŸan kontrolÃ¼
    calisan = conn.execute('SELECT * FROM calisanlar WHERE id = %s AND aktif = TRUE', 
                          (calisan_id,)).fetchone()
    
    if not calisan:
        flash('Ã‡alÄ±ÅŸan bulunamadÄ±!', 'error')
        return redirect(url_for('calisanlar'))
    
    # Form verilerini al
    cikis_tarihi = request.form['cikis_tarihi']
    cikis_nedeni = request.form['cikis_nedeni']
    liste_durumu = request.form['liste_durumu']
    tekrar_ise_alinabilir = 1 if request.form.get('tekrar_ise_alinabilir') else 0
    
    # Checklist
    zimmet_teslim = 1 if request.form.get('zimmet_teslim') else 0
    kiyafet_teslim = 1 if request.form.get('kiyafet_teslim') else 0
    anahtar_teslim = 1 if request.form.get('anahtar_teslim') else 0
    kimlik_teslim = 1 if request.form.get('kimlik_teslim') else 0
    
    # Tazminat
    ihbar_tazminat_durumu = request.form.get('ihbar_tazminat_durumu', '')
    kidem_tazminat_durumu = request.form.get('kidem_tazminat_durumu', '')
    
    # Notlar
    yonetici_notu = request.form.get('yonetici_notu', '')
    ik_notu = request.form.get('ik_notu', '')
    genel_degerlendirme = request.form.get('genel_degerlendirme', '')
    
    try:
        # 1. Ã‡Ä±kÄ±ÅŸ kaydÄ± oluÅŸtur
        cursor.execute('''
            INSERT INTO cikis_kayitlari 
            (calisan_id, cikis_tarihi, cikis_nedeni, liste_durumu, tekrar_ise_alinabilir,
             zimmet_teslim, kiyafet_teslim, anahtar_teslim, kimlik_teslim,
             ihbar_tazminat_durumu, kidem_tazminat_durumu,
             yonetici_notu, ik_notu, genel_degerlendirme, islem_yapan)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (calisan_id, cikis_tarihi, cikis_nedeni, liste_durumu, tekrar_ise_alinabilir,
              zimmet_teslim, kiyafet_teslim, anahtar_teslim, kimlik_teslim,
              ihbar_tazminat_durumu, kidem_tazminat_durumu,
              yonetici_notu, ik_notu, genel_degerlendirme,
              session.get('username')))
        
        cikis_kayit_id = cursor.lastrowid
        
        # 2. Ã‡alÄ±ÅŸanÄ± pasif yap ve Ã§Ä±kÄ±ÅŸ bilgilerini gÃ¼ncelle
        cursor.execute('''
            UPDATE calisanlar 
            SET aktif = FALSE,
                cikis_tarihi = %s,
                cikis_nedeni = %s,
                liste_durumu = %s
            WHERE id = %s
        ''', (cikis_tarihi, cikis_nedeni, liste_durumu, calisan_id))
        
        # 3. YENI: Aday durumunu gÃ¼ncelle
        aday_id = calisan['aday_id']
        if aday_id:
            cursor.execute('''
            UPDATE adaylar 
            SET durum = 'Pasif'
               WHERE id = %s
    ''', (aday_id,))
        
        # 4. Kadro doluluk sayÄ±sÄ±nÄ± gÃ¼ncelle
        kadro_id = calisan['kadro_id']
        cursor.execute('''
            UPDATE hedef_kadrolar 
            SET dolu_kisi_sayisi = dolu_kisi_sayisi - 1
            WHERE id = %s
        ''', (kadro_id,))
        
        # 5. Kadro durumunu gÃ¼ncelle (Dolu â†’ AÃ§Ä±k)
        cursor.execute('''
            UPDATE hedef_kadrolar 
            SET durum = CASE 
                WHEN dolu_kisi_sayisi < hedef_kisi_sayisi THEN 'AÃ§Ä±k'
                ELSE 'Dolu'
            END
            WHERE id = %s
        ''', (kadro_id,))
        
        conn.commit()
        
        # Log kaydet
        log_islem('Ã‡IKIÅž', 'calisanlar', calisan_id, 
                 f'{calisan["ad_soyad"]} - {cikis_nedeni} nedeniyle iÅŸten Ã§Ä±kÄ±ÅŸ')
        # ============================================
        # YENÄ°: E-POSTA BÄ°LDÄ°RÄ°MÄ°
        # ============================================
        try:
            # Ã‡alÄ±ÅŸan detay bilgilerini getir
            calisan_detay = conn.execute('''
    SELECT c.*,
           hk.pozisyon_adi,
           hk.magaza_adi,
           hk.aracli_durum,
           m.mudurluk_adi,
           d.direktorluk_adi,
           p.proje_adi,
           i.il_adi,
           ic.ilce_adi
    FROM calisanlar c
    LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
    LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
    LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
    LEFT JOIN projeler p ON hk.proje_id = p.id
    LEFT JOIN iller i ON hk.il_id = i.id
    LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
    WHERE c.id = %s
''', (calisan_id,)).fetchone()
            
            cikis_bilgileri = {
                'id': cikis_kayit_id,
                'ad_soyad': calisan_detay['ad_soyad'],
                'telefon': calisan_detay['telefon'] or '-',
                'email': calisan_detay['email'] or '-',
                'tc_kimlik': calisan_detay['tc_kimlik'] or '-',
                'proje_adi': calisan_detay['proje_adi'],
                'pozisyon_adi': calisan_detay['pozisyon_adi'],
                'mudurluk_adi': calisan_detay['mudurluk_adi'] or '-',  # â† YENÄ°
                'direktorluk_adi': calisan_detay['direktorluk_adi'] or '-',  # â† YENÄ°
                'il': calisan_detay['il_adi'],
                'ilce': calisan_detay['ilce_adi'] or '-',
                'magaza_adi': calisan_detay['magaza_adi'] or '-',
                'aracli_durum': calisan_detay['aracli_durum'] or 'AraÃ§sÄ±z',
                'ise_baslama_tarihi': calisan_detay['ise_baslama_tarihi'],
                'cikis_tarihi': cikis_tarihi,
                'cikis_nedeni': cikis_nedeni,
                'liste_durumu': liste_durumu,
                'tekrar_ise_alinabilir': tekrar_ise_alinabilir,
                'zimmet_teslim': zimmet_teslim,
                'kiyafet_teslim': kiyafet_teslim,
                'anahtar_teslim': anahtar_teslim,
                'kimlik_teslim': kimlik_teslim,
                'ihbar_tazminat_durumu': ihbar_tazminat_durumu,
                'kidem_tazminat_durumu': kidem_tazminat_durumu,
                'yonetici_notu': yonetici_notu,
                'ik_notu': ik_notu,
                'genel_degerlendirme': genel_degerlendirme,
                'islem_yapan': session.get('username')
            }
            
            email_basarili, email_mesaj = isten_cikis_bildirimi(cikis_bilgileri)
            if email_basarili:
                flash(email_mesaj, 'info')
        except Exception as e:
            print(f"Email gÃ¶nderimi hatasÄ±: {e}")
            flash(f'UyarÄ±: Email bildirimi gÃ¶nderilemedi', 'warning')
        
        return redirect(url_for('cikis_kayitlari'))
        
    except Exception as e:
        conn.rollback()
        flash(f'Hata oluÅŸtu: {str(e)}', 'error')
        return redirect(url_for('calisan_cikis', calisan_id=calisan_id))
    finally:
        conn.close()


@app.route('/cikis-kayitlari')
@login_required
def cikis_kayitlari():
    """Ã‡Ä±kÄ±ÅŸ kayÄ±tlarÄ± listesi"""
    conn = get_db()
    
    kayitlar = conn.execute('''
        SELECT ck.*,
               c.ad_soyad,
               c.telefon,
               hk.pozisyon_adi,
               p.proje_adi,
               i.il_adi
        FROM cikis_kayitlari ck
        LEFT JOIN calisanlar c ON ck.calisan_id = c.id
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN iller i ON hk.il_id = i.id
        ORDER BY ck.cikis_tarihi DESC, ck.olusturma_tarihi DESC
    ''').fetchall()
    
    # Ä°statistikler
    stats = {
        'toplam_cikis': len(kayitlar),
        'beyaz_liste': len([k for k in kayitlar if k['liste_durumu'] == 'Beyaz']),
        'gri_liste': len([k for k in kayitlar if k['liste_durumu'] == 'Gri']),
        'kara_liste': len([k for k in kayitlar if k['liste_durumu'] == 'Kara'])
    }
    
    conn.close()
    
    return render_template('cikis_kayitlari.html', kayitlar=kayitlar, stats=stats)


@app.route('/cikis-kaydi/<int:kayit_id>')
@login_required
def cikis_detay(kayit_id):
    conn = get_db()
    
    kayit = conn.execute('''
        SELECT ck.*,
               c.ad_soyad,
               c.telefon,
               c.email,
               c.ise_baslama_tarihi,
               hk.pozisyon_adi,
               hk.magaza_adi,
               p.proje_adi,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi
        FROM cikis_kayitlari ck
        LEFT JOIN calisanlar c ON ck.calisan_id = c.id
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        WHERE ck.id = %s
    ''', (kayit_id,)).fetchone()
    
    if not kayit:
        flash('KayÄ±t bulunamadÄ±!', 'error')
        return redirect(url_for('cikis_kayitlari'))
    
    conn.close()
    
    # YENÄ°: Return ekle!
    return render_template('cikis_detay.html', kayit=kayit)    
    
@app.route('/adaylar/excel-export')
@login_required
def adaylar_excel_export():
    """AdaylarÄ± Excel'e aktar"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    conn = get_db()
    adaylar = conn.execute('''
        SELECT a.*, 
               hk.pozisyon_adi, hk.magaza_adi, hk.aracli_durum, hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi,
               k.kaynak_adi
        FROM adaylar a
        LEFT JOIN hedef_kadrolar hk ON a.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN kaynaklar k ON a.kaynak_id = k.id
        ORDER BY a.basvuru_tarihi DESC
    ''').fetchall()
    conn.close()
    
    # Excel oluÅŸtur
    wb = Workbook()
    ws = wb.active
    ws.title = "Adaylar"
    
    # BaÅŸlÄ±k satÄ±rÄ±
    headers = ['ID', 'Ad Soyad', 'Telefon', 'Email', 'TC Kimlik', 
               'DoÄŸum Tarihi', 'YaÅŸ', 
               'Proje', 'Pozisyon', 'MÃ¼dÃ¼rlÃ¼k', 'DirektÃ¶rlÃ¼k', 
               'Ä°l', 'Ä°lÃ§e', 'MaÄŸaza', 'AraÃ§ Durumu', 'Ã‡alÄ±ÅŸma Åžekli',
               'Kaynak', 'BaÅŸvuru Tarihi', 'Durum']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Veri satÄ±rlarÄ±
    for row_idx, aday in enumerate(adaylar, 2):
        # YaÅŸ hesapla
        yas_str = '-'
        if aday['dogum_tarihi']:
            try:
                dogum_tarihi = aday['dogum_tarihi']
                
                if isinstance(dogum_tarihi, str):
                    dogum = datetime.strptime(dogum_tarihi, '%Y-%m-%d')
                elif hasattr(dogum_tarihi, 'year'):
                    dogum = datetime(dogum_tarihi.year, dogum_tarihi.month, dogum_tarihi.day)
                else:
                    raise ValueError("Bilinmeyen tarih formati")
                
                bugun = datetime.now()
                yil = bugun.year - dogum.year
                ay = bugun.month - dogum.month
                if ay < 0:
                    yil -= 1
                    ay += 12
                if bugun.day < dogum.day:
                    ay -= 1
                    if ay < 0:
                        yil -= 1
                        ay += 12
                yas_str = f"{yil}.{ay}"
            except:
                pass
        
        ws.cell(row=row_idx, column=1, value=aday['id'])
        ws.cell(row=row_idx, column=2, value=aday['ad_soyad'])
        ws.cell(row=row_idx, column=3, value=aday['telefon'] or '')
        ws.cell(row=row_idx, column=4, value=aday['email'] or '')
        ws.cell(row=row_idx, column=5, value=aday['tc_kimlik'] or '')
        ws.cell(row=row_idx, column=6, value=aday['dogum_tarihi'] or '')
        ws.cell(row=row_idx, column=7, value=yas_str)
        ws.cell(row=row_idx, column=8, value=aday['proje_adi'] or '')
        ws.cell(row=row_idx, column=9, value=aday['pozisyon_adi'] or '')
        ws.cell(row=row_idx, column=10, value=aday['mudurluk_adi'] or '')
        ws.cell(row=row_idx, column=11, value=aday['direktorluk_adi'] or '')
        ws.cell(row=row_idx, column=12, value=aday['il_adi'] or '')
        ws.cell(row=row_idx, column=13, value=aday['ilce_adi'] or '')
        ws.cell(row=row_idx, column=14, value=aday['magaza_adi'] or '')
        ws.cell(row=row_idx, column=15, value=aday['aracli_durum'] or '')
        ws.cell(row=row_idx, column=16, value=aday['calisma_sekli'] or '')
        ws.cell(row=row_idx, column=17, value=aday['kaynak_adi'] or aday['kaynak_diger'] or '')
        ws.cell(row=row_idx, column=18, value=aday['basvuru_tarihi'])
        ws.cell(row=row_idx, column=19, value=aday['durum'])
    
    # SÃ¼tun geniÅŸliklerini ayarla
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    # DosyayÄ± belleÄŸe kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'adaylar_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    
# Ã‡ALIÅžAN DÃœZENLE ROUTE'U
# app.py'ye eklenecek (Ä°ÅŸten Ã§Ä±kÄ±ÅŸ route'larÄ±ndan sonra)

@app.route('/calisan/<int:calisan_id>/duzenle', methods=['GET', 'POST'])
@login_required
def calisanlar_duzenle(calisan_id):
    """Ã‡alÄ±ÅŸan dÃ¼zenle"""
    conn = get_db()
    
    if request.method == 'GET':
        # Ã‡alÄ±ÅŸan bilgilerini getir
        calisan = conn.execute('''
            SELECT c.*, hk.proje_id
            FROM calisanlar c
            LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
            WHERE c.id = %s
        ''', (calisan_id,)).fetchone()
        
        if not calisan:
            flash('Ã‡alÄ±ÅŸan bulunamadÄ±!', 'error')
            return redirect(url_for('calisanlar'))
        
        # Dropdown'lar iÃ§in veriler
        projeler = conn.execute('SELECT * FROM projeler ORDER BY proje_adi').fetchall()
        kadrolar = conn.execute('''
            SELECT hk.*, p.proje_adi, i.il_adi, ic.ilce_adi
            FROM hedef_kadrolar hk
            LEFT JOIN projeler p ON hk.proje_id = p.id
            LEFT JOIN iller i ON hk.il_id = i.id
            LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
            ORDER BY p.proje_adi, hk.pozisyon_adi
        ''').fetchall()
        calisma_sekilleri = conn.execute('SELECT * FROM calisma_sekilleri WHERE aktif = TRUE ORDER BY calisma_sekli').fetchall()
        
        conn.close()
        
        return render_template('calisan_duzenle.html', 
                             calisan=calisan,
                             projeler=projeler,
                             kadrolar=kadrolar,
                             calisma_sekilleri=calisma_sekilleri)
    
    else:  # POST
        # Form verilerini al
        ad_soyad = request.form['ad_soyad']
        telefon = request.form.get('telefon', '')
        email = request.form.get('email', '')
        tc_kimlik = request.form.get('tc_kimlik', '')
        dogum_tarihi = request.form.get('dogum_tarihi', None)
        adres = request.form.get('adres', '')
        ise_baslama_tarihi = request.form['ise_baslama_tarihi']
        kadro_id = request.form['kadro_id']
        aracli_durum = request.form.get('aracli_durum', 'AraÃ§sÄ±z')
        calisma_sekli_id = request.form.get('calisma_sekli_id', None)
        aktif = TRUE if request.form.get('aktif') else 0
        
        try:
            cursor = conn.cursor()
            
            # Eski kadro bilgisi
            eski_calisan = conn.execute('SELECT kadro_id FROM calisanlar WHERE id = %s', 
                                       (calisan_id,)).fetchone()
            eski_kadro_id = eski_calisan['kadro_id']
            
            # Ã‡alÄ±ÅŸanÄ± gÃ¼ncelle
            cursor.execute('''
                UPDATE calisanlar 
                SET ad_soyad = %s,
                    telefon = %s,
                    email = %s,
                    tc_kimlik = %s,
                    dogum_tarihi = %s,
                    adres = %s,
                    ise_baslama_tarihi = %s,
                    kadro_id = %s,
                    aracli_durum = %s,
                    calisma_sekli_id = %s,
                    aktif = %s
                WHERE id = %s
            ''', (ad_soyad, telefon, email, tc_kimlik, dogum_tarihi, adres,
                  ise_baslama_tarihi, kadro_id, aracli_durum, calisma_sekli_id,
                  aktif, calisan_id))
            
            # Kadro deÄŸiÅŸtiyse doluluk sayÄ±larÄ±nÄ± gÃ¼ncelle
            if eski_kadro_id != int(kadro_id):
                # Eski kadrodan Ã§Ä±kar
                cursor.execute('''
                    UPDATE hedef_kadrolar 
                    SET dolu_kisi_sayisi = dolu_kisi_sayisi - 1
                    WHERE id = %s
                ''', (eski_kadro_id,))
                
                # Yeni kadroya ekle
                cursor.execute('''
                    UPDATE hedef_kadrolar 
                    SET dolu_kisi_sayisi = dolu_kisi_sayisi + 1
                    WHERE id = %s
                ''', (kadro_id,))
                
                # Her iki kadronun durumunu gÃ¼ncelle
                cursor.execute('''
                    UPDATE hedef_kadrolar 
                    SET durum = CASE 
                        WHEN dolu_kisi_sayisi >= hedef_kisi_sayisi THEN 'Dolu'
                        ELSE 'AÃ§Ä±k'
                    END
                    WHERE id IN (%s, %s)
                ''', (eski_kadro_id, kadro_id))
            
            conn.commit()
            
            # Log kaydet
            log_islem('GÃœNCELLEME', 'calisanlar', calisan_id, 
                     f'{ad_soyad} bilgileri gÃ¼ncellendi')
            
            flash('Ã‡alÄ±ÅŸan baÅŸarÄ±yla gÃ¼ncellendi!', 'success')
            return redirect(url_for('calisan_detay', calisan_id=calisan_id))
            
        except Exception as e:
            conn.rollback()
            flash(f'Hata oluÅŸtu: {str(e)}', 'error')
            return redirect(url_for('calisanlar_duzenle', calisan_id=calisan_id))
        finally:
            conn.close()    
    
@app.route('/calisan/<int:calisan_id>/detay')
@login_required
def calisan_detay(calisan_id):
    """Ã‡alÄ±ÅŸan detay sayfasÄ±"""
    conn = get_db()
    
    # Ã‡alÄ±ÅŸan bilgileri
    calisan_raw = conn.execute('''
        SELECT c.*, 
               hk.pozisyon_adi, hk.magaza_adi, hk.aracli_durum, hk.calisma_sekli,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi
        FROM calisanlar c
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        WHERE c.id = %s
    ''', (calisan_id,)).fetchone()
    
    if not calisan_raw:
        conn.close()
        flash('Ã‡alÄ±ÅŸan bulunamadÄ±!', 'danger')
        return redirect(url_for('calisanlar'))
    
    # Dict'e Ã§evir
    calisan = dict(calisan_raw)
    
    # YAÅž HESAPLA
    if calisan.get('dogum_tarihi'):
        try:
            dogum_tarihi = calisan['dogum_tarihi']
            
            if isinstance(dogum_tarihi, str):
                dogum = datetime.strptime(dogum_tarihi, '%Y-%m-%d')
            elif hasattr(dogum_tarihi, 'year'):
                dogum = datetime(dogum_tarihi.year, dogum_tarihi.month, dogum_tarihi.day)
            else:
                raise ValueError("Bilinmeyen tarih formati")
            
            bugun = datetime.now()
            
            yil = bugun.year - dogum.year
            ay = bugun.month - dogum.month
            
            if ay < 0:
                yil -= 1
                ay += 12
            
            if bugun.day < dogum.day:
                ay -= 1
                if ay < 0:
                    yil -= 1
                    ay += 12
            
            calisan['yas'] = f"{yil}.{ay}"
            calisan['yas_yil'] = yil
        except Exception as e:
            print(f"Yas hesaplama hatasi: {e}")
            calisan['yas'] = None
            calisan['yas_yil'] = None
    else:
        calisan['yas'] = None
        calisan['yas_yil'] = None
    
    # Dosyalar (aday_id Ã¼zerinden de ara - aday dÃ¶neminden kalan dosyalar iÃ§in)
    dosyalar = conn.execute('''
        SELECT d.*, u.username as yukleyen
        FROM dosyalar d
        LEFT JOIN users u ON d.yukleyen_user_id = u.id
        WHERE (d.ilgili_tablo = 'calisanlar' AND d.ilgili_id = %s)
           OR (d.ilgili_tablo = 'adaylar' AND d.ilgili_id = %s)
        ORDER BY d.yukleme_tarihi DESC
    ''', (calisan_id, calisan['aday_id'])).fetchall()
    
    conn.close()
    
    return render_template('calisan_detay.html', calisan=calisan, dosyalar=dosyalar)
    
@app.route('/calisan/<int:calisan_id>/dosya-yukle', methods=['POST'])
@login_required
def calisan_dosya_yukle(calisan_id):
    """Ã‡alÄ±ÅŸana dosya yÃ¼kle"""
    if 'dosya' not in request.files:
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('calisan_detay', calisan_id=calisan_id))
    
    file = request.files['dosya']
    dosya_tipi = request.form.get('dosya_tipi', 'diger')
    aciklama = request.form.get('aciklama', '')
    
    if file.filename == '':
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('calisan_detay', calisan_id=calisan_id))
    
    if not allowed_file(file.filename):
        flash('GeÃ§ersiz dosya formatÄ±! Ä°zin verilen: PDF, DOC, DOCX, JPG, PNG', 'danger')
        return redirect(url_for('calisan_detay', calisan_id=calisan_id))
    
    # GÃ¼venli dosya adÄ± oluÅŸtur
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    yeni_dosya_adi = f"{calisan_id}_{dosya_tipi}_{timestamp}_{filename}"
    
    # KlasÃ¶r yolu
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'calisanlar')
    os.makedirs(upload_path, exist_ok=True)
    
    # DosyayÄ± kaydet
    dosya_yolu = os.path.join(upload_path, yeni_dosya_adi)
    file.save(dosya_yolu)
    
    # Dosya boyutu
    dosya_boyutu = os.path.getsize(dosya_yolu)
    
    # VeritabanÄ±na kaydet
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO dosyalar 
        (ilgili_tablo, ilgili_id, dosya_tipi, dosya_adi, dosya_yolu, dosya_boyutu, yukleyen_user_id, aciklama)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', ('calisanlar', calisan_id, dosya_tipi, filename, dosya_yolu, dosya_boyutu, session['user_id'], aciklama))
    conn.commit()
    conn.close()
    
    log_islem('DOSYA_YUKLEME', 'dosyalar', calisan_id, 
             f'{dosya_tipi} dosyasÄ± yÃ¼klendi: {filename}', session.get('username'))
    
    flash(f'Dosya baÅŸarÄ±yla yÃ¼klendi! ({get_file_size_mb(dosya_boyutu)} MB)', 'success')
    return redirect(url_for('calisan_detay', calisan_id=calisan_id))    

@app.route('/loglar')
@login_required
def loglar():
    """SÃ¼reÃ§ loglarÄ±"""
    conn = get_db()
    loglar = conn.execute('''
        SELECT * FROM surec_loglari 
        ORDER BY islem_tarihi DESC 
        LIMIT 100
    ''').fetchall()
    conn.close()
    return render_template('loglar.html', loglar=loglar)

@app.route('/api/stats')
@login_required
def api_stats():
    """Dashboard iÃ§in istatistikler"""
    conn = get_db()
    
    stats = {
        'toplam_proje': conn.execute('SELECT COUNT(*) as cnt FROM projeler WHERE aktif = TRUE').fetchone()['cnt'],
        'toplam_kadro': conn.execute('SELECT COUNT(*) as cnt FROM hedef_kadrolar').fetchone()['cnt'],
        'toplam_aday': conn.execute('SELECT COUNT(*) as cnt FROM adaylar').fetchone()['cnt'],
        'toplam_calisan': conn.execute('SELECT COUNT(*) as cnt FROM calisanlar WHERE aktif = TRUE').fetchone()['cnt'],
        'hedef_toplam': conn.execute('SELECT SUM(hedef_kisi_sayisi) as total FROM hedef_kadrolar').fetchone()['total'] or 0,
        'dolu_toplam': conn.execute('SELECT SUM(dolu_kisi_sayisi) as total FROM hedef_kadrolar').fetchone()['total'] or 0
    }
    
    conn.close()
    return jsonify(stats)
@app.route('/users')
@admin_required
def users():
    """KullanÄ±cÄ± listesi (Sadece Admin)"""
    conn = get_db()
    users = conn.execute('''
        SELECT id, username, email, full_name, role, aktif, 
               olusturma_tarihi, son_giris_tarihi
        FROM users 
        ORDER BY olusturma_tarihi DESC
    ''').fetchall()
    conn.close()
    return render_template('users.html', users=users)

@app.route('/user/ekle', methods=['GET', 'POST'])
@admin_required
def user_ekle():
    """Yeni kullanÄ±cÄ± ekle (Sadece Admin)"""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        email = request.form.get('email', '')
        full_name = request.form['full_name']
        role = request.form['role']
        
        # Åžifreyi hashle
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO users (username, password_hash, email, full_name, role)
                VALUES (%s, %s, %s, %s, %s)
            ''', (username, password_hash, email, full_name, role))
            user_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            log_islem('EKLEME', 'users', user_id, 
                     f'{username} kullanÄ±cÄ±sÄ± oluÅŸturuldu (Rol: {role})', 
                     session.get('username', 'Sistem'))
            flash(f'KullanÄ±cÄ± "{username}" baÅŸarÄ±yla eklendi!', 'success')
            return redirect(url_for('users'))
        except sqlite3.IntegrityError:
            conn.close()
            flash('Bu kullanÄ±cÄ± adÄ± zaten kullanÄ±lÄ±yor!', 'danger')
    
    return render_template('user_form.html')

@app.route('/user/<int:user_id>/duzenle', methods=['GET', 'POST'])
@admin_required
def user_duzenle(user_id):
    """KullanÄ±cÄ± dÃ¼zenle (Sadece Admin)"""
    conn = get_db()
    
    if request.method == 'POST':
        email = request.form.get('email', '')
        full_name = request.form['full_name']
        role = request.form['role']
        aktif = TRUE if request.form.get('aktif') == 'on' else 0
        
        # Åžifre deÄŸiÅŸtirilmek isteniyorsa
        new_password = request.form.get('new_password', '')
        if new_password:
            password_hash = hashlib.sha256(new_password.encode()).hexdigest()
            conn.execute('''
                UPDATE users 
                SET email = %s, full_name = %s, role = %s, aktif = %s, password_hash = %s
                WHERE id = %s
            ''', (email, full_name, role, aktif, password_hash, user_id))
        else:
            conn.execute('''
                UPDATE users 
                SET email = %s, full_name = %s, role = %s, aktif = %s
                WHERE id = %s
            ''', (email, full_name, role, aktif, user_id))
        
        conn.commit()
        conn.close()
        
        log_islem('GÃœNCELLEME', 'users', user_id, 
                 f'KullanÄ±cÄ± bilgileri gÃ¼ncellendi', 
                 session.get('username', 'Sistem'))
        flash('KullanÄ±cÄ± baÅŸarÄ±yla gÃ¼ncellendi!', 'success')
        return redirect(url_for('users'))
    
    user = conn.execute('SELECT * FROM users WHERE id = %s', (user_id,)).fetchone()
    conn.close()
    
    if not user:
        flash('KullanÄ±cÄ± bulunamadÄ±!', 'danger')
        return redirect(url_for('users'))
    
    return render_template('user_edit.html', user=user)

@app.route('/user/<int:user_id>/sil', methods=['POST'])
@admin_required
def user_sil(user_id):
    """KullanÄ±cÄ± sil (Sadece Admin)"""
    # Kendi hesabÄ±nÄ± silemez
    if user_id == session.get('user_id'):
        flash('Kendi hesabÄ±nÄ±zÄ± silemezsiniz!', 'danger')
        return redirect(url_for('users'))
    
    conn = get_db()
    user = conn.execute('SELECT username FROM users WHERE id = %s', (user_id,)).fetchone()
    
    if user:
        conn.execute('DELETE FROM users WHERE id = %s', (user_id,))
        conn.commit()
        log_islem('SÄ°LME', 'users', user_id, 
                 f'{user["username"]} kullanÄ±cÄ±sÄ± silindi', 
                 session.get('username', 'Sistem'))
        flash(f'KullanÄ±cÄ± "{user["username"]}" baÅŸarÄ±yla silindi!', 'success')
    else:
        flash('KullanÄ±cÄ± bulunamadÄ±!', 'danger')
    
    conn.close()
    return redirect(url_for('users'))

@app.route('/profil')
@login_required
def profil():
    """KullanÄ±cÄ± profili"""
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = %s', (session['user_id'],)).fetchone()
    conn.close()
    return render_template('profil.html', user=user)

@app.route('/profil/sifre-degistir', methods=['POST'])
@login_required
def sifre_degistir():
    """Åžifre deÄŸiÅŸtir"""
    eski_sifre = request.form['eski_sifre']
    yeni_sifre = request.form['yeni_sifre']
    yeni_sifre_tekrar = request.form['yeni_sifre_tekrar']
    
    if yeni_sifre != yeni_sifre_tekrar:
        flash('Yeni ÅŸifreler eÅŸleÅŸmiyor!', 'danger')
        return redirect(url_for('profil'))
    
    # Eski ÅŸifre kontrolÃ¼
    eski_sifre_hash = hashlib.sha256(eski_sifre.encode()).hexdigest()
    
    conn = get_db()
    user = conn.execute('SELECT * FROM users WHERE id = %s AND password_hash = %s', 
                       (session['user_id'], eski_sifre_hash)).fetchone()
    
    if not user:
        conn.close()
        flash('Eski ÅŸifre hatalÄ±!', 'danger')
        return redirect(url_for('profil'))
    
    # Yeni ÅŸifreyi kaydet
    yeni_sifre_hash = hashlib.sha256(yeni_sifre.encode()).hexdigest()
    conn.execute('UPDATE users SET password_hash = %s WHERE id = %s', 
                (yeni_sifre_hash, session['user_id']))
    conn.commit()
    conn.close()
    
    log_islem('GÃœNCELLEME', 'users', session['user_id'], 
             'Åžifre deÄŸiÅŸtirildi', session.get('username', 'Sistem'))
    flash('Åžifreniz baÅŸarÄ±yla deÄŸiÅŸtirildi!', 'success')
    return redirect(url_for('profil')) 

@app.route('/yetkisiz')
@login_required
def yetkisiz():
    """Yetki yok sayfasÄ±"""
    return render_template('yetkisiz.html')
    
@app.route('/aday/<int:aday_id>/dosya-yukle', methods=['POST'])
@login_required
def aday_dosya_yukle(aday_id):
    """Adaya dosya yÃ¼kle"""
    if 'dosya' not in request.files:
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('aday_detay', aday_id=aday_id))
    
    file = request.files['dosya']
    dosya_tipi = request.form.get('dosya_tipi', 'diger')
    aciklama = request.form.get('aciklama', '')
    
    if file.filename == '':
        flash('Dosya seÃ§ilmedi!', 'danger')
        return redirect(url_for('aday_detay', aday_id=aday_id))
    
    if not allowed_file(file.filename):
        flash('GeÃ§ersiz dosya formatÄ±! Ä°zin verilen: PDF, DOC, DOCX, JPG, PNG', 'danger')
        return redirect(url_for('aday_detay', aday_id=aday_id))
    
    # GÃ¼venli dosya adÄ± oluÅŸtur
    filename = secure_filename(file.filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    yeni_dosya_adi = f"{aday_id}_{dosya_tipi}_{timestamp}_{filename}"
    
    # KlasÃ¶r yolu
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'adaylar')
    os.makedirs(upload_path, exist_ok=True)
    
    # DosyayÄ± kaydet
    dosya_yolu = os.path.join(upload_path, yeni_dosya_adi)
    file.save(dosya_yolu)
    
    # Dosya boyutu
    dosya_boyutu = os.path.getsize(dosya_yolu)
    
    # VeritabanÄ±na kaydet
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO dosyalar 
        (ilgili_tablo, ilgili_id, dosya_tipi, dosya_adi, dosya_yolu, dosya_boyutu, yukleyen_user_id, aciklama)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', ('adaylar', aday_id, dosya_tipi, filename, dosya_yolu, dosya_boyutu, session['user_id'], aciklama))
    conn.commit()
    conn.close()
    
    log_islem('DOSYA_YUKLEME', 'dosyalar', aday_id, 
             f'{dosya_tipi} dosyasÄ± yÃ¼klendi: {filename}', session.get('username'))
    
    flash(f'Dosya baÅŸarÄ±yla yÃ¼klendi! ({get_file_size_mb(dosya_boyutu)} MB)', 'success')
    return redirect(url_for('aday_detay', aday_id=aday_id))

@app.route('/dosya/<int:dosya_id>/indir')
@login_required
def dosya_indir(dosya_id):
    """Dosya indir"""
    conn = get_db()
    dosya = conn.execute('SELECT * FROM dosyalar WHERE id = %s', (dosya_id,)).fetchone()
    conn.close()
    
    if not dosya:
        flash('Dosya bulunamadÄ±!', 'danger')
        return redirect(url_for('index'))
    
    try:
        from flask import send_file
        return send_file(dosya['dosya_yolu'], 
                        as_attachment=True, 
                        download_name=dosya['dosya_adi'])
    except Exception as e:
        flash(f'Dosya indirilemedi: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('index'))

@app.route('/dosya/<int:dosya_id>/sil', methods=['POST'])
@login_required
def dosya_sil(dosya_id):
    """Dosya sil"""
    conn = get_db()
    dosya = conn.execute('SELECT * FROM dosyalar WHERE id = %s', (dosya_id,)).fetchone()
    
    if not dosya:
        conn.close()
        flash('Dosya bulunamadÄ±!', 'danger')
        return redirect(request.referrer or url_for('index'))
    
    # Yetki kontrolÃ¼ (sadece admin veya yÃ¼kleyen silebilir)
    if session.get('role') != 'admin' and dosya['yukleyen_user_id'] != session['user_id']:
        conn.close()
        flash('Bu dosyayÄ± silme yetkiniz yok!', 'danger')
        return redirect(request.referrer or url_for('index'))
    
    # Fiziksel dosyayÄ± sil
    try:
        if os.path.exists(dosya['dosya_yolu']):
            os.remove(dosya['dosya_yolu'])
    except Exception as e:
        print(f"Dosya silinemedi: {e}")
    
    # VeritabanÄ±ndan sil
    conn.execute('DELETE FROM dosyalar WHERE id = %s', (dosya_id,))
    conn.commit()
    conn.close()
    
    log_islem('DOSYA_SILME', 'dosyalar', dosya_id, 
             f'Dosya silindi: {dosya["dosya_adi"]}', session.get('username'))
    
    flash('Dosya baÅŸarÄ±yla silindi!', 'success')
    return redirect(request.referrer or url_for('index'))

@app.route('/aday/<int:aday_id>/detay')
@login_required
def aday_detay(aday_id):
    """Aday detay sayfasÄ±"""
    conn = get_db()
    
    # Aday bilgileri
    aday = conn.execute('''
        SELECT a.*, 
               hk.pozisyon_adi, hk.magaza_adi,
               i.il_adi,
               ic.ilce_adi,
               p.proje_adi
        FROM adaylar a
        LEFT JOIN hedef_kadrolar hk ON a.kadro_id = hk.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        WHERE a.id = %s
    ''', (aday_id,)).fetchone()
    
    if not aday:
        conn.close()
        flash('Aday bulunamadÄ±!', 'danger')
        return redirect(url_for('adaylar'))
    
    # Dosyalar
    dosyalar = conn.execute('''
        SELECT d.*, u.username as yukleyen
        FROM dosyalar d
        LEFT JOIN users u ON d.yukleyen_user_id = u.id
        WHERE d.ilgili_tablo = 'adaylar' AND d.ilgili_id = %s
        ORDER BY d.yukleme_tarihi DESC
    ''', (aday_id,)).fetchall()
    
    conn.close()
    
    return render_template('aday_detay.html', aday=aday, dosyalar=dosyalar)


@app.route('/email-ayarlari')
@admin_required
def email_ayarlari():
    """Email ayarlarÄ± sayfasÄ± (Sadece Admin)"""
    conn = get_db()
    
    # Email ayarlarÄ±
    ayarlar = conn.execute('''
        SELECT * FROM email_ayarlari 
        ORDER BY 
            CASE ayar_adi
                WHEN 'smtp_sunucu' THEN 1
                WHEN 'smtp_port' THEN 2
                WHEN 'smtp_kullanici' THEN 3
                WHEN 'smtp_sifre' THEN 4
                WHEN 'gonderici_adi' THEN 5
                ELSE 6
            END
    ''').fetchall()
    
    # Email ÅŸablonlarÄ±
    sablonlar = conn.execute('''
        SELECT * FROM email_sablonlari 
        ORDER BY olusturma_tarihi DESC
    ''').fetchall()
    
    # Email loglarÄ± (son 50)
    loglar = conn.execute('''
        SELECT * FROM email_loglari 
        ORDER BY gonderim_tarihi DESC 
        LIMIT 50
    ''').fetchall()
    
    # Ä°statistikler
    stats = {
        'toplam_gonderim': conn.execute('SELECT COUNT(*) as cnt FROM email_loglari').fetchone()['cnt'],
        'basarili': conn.execute("SELECT COUNT(*) as cnt FROM email_loglari WHERE gonderim_durumu = 'gonderildi'").fetchone()['cnt'],
        'basarisiz': conn.execute("SELECT COUNT(*) as cnt FROM email_loglari WHERE gonderim_durumu = 'hata'").fetchone()['cnt'],
    }
    
    conn.close()
    
    return render_template('email_ayarlari.html', 
                         ayarlar=ayarlar, 
                         sablonlar=sablonlar, 
                         loglar=loglar,
                         stats=stats)

@app.route('/email-ayarlari/guncelle', methods=['POST'])
@admin_required
def email_ayarlari_guncelle():
    """Email ayarlarÄ±nÄ± gÃ¼ncelle"""
    conn = get_db()
    cursor = conn.cursor()
    
    # Form'dan gelen tÃ¼m ayarlarÄ± gÃ¼ncelle
    for key, value in request.form.items():
        if key.startswith('ayar_'):
            ayar_adi = key.replace('ayar_', '')
            cursor.execute('''
                UPDATE email_ayarlari 
                SET ayar_degeri = %s 
                WHERE ayar_adi = %s
            ''', (value, ayar_adi))
    
    conn.commit()
    conn.close()
    
    log_islem('GÃœNCELLEME', 'email_ayarlari', 0, 
             'Email ayarlarÄ± gÃ¼ncellendi', session.get('username'))
    
    flash('Email ayarlarÄ± baÅŸarÄ±yla gÃ¼ncellendi!', 'success')
    return redirect(url_for('email_ayarlari'))

@app.route('/email-sablonu/<int:sablon_id>/duzenle', methods=['GET', 'POST'])
@admin_required
def email_sablon_duzenle(sablon_id):
    """Email ÅŸablonunu dÃ¼zenle"""
    conn = get_db()
    
    if request.method == 'POST':
        sablon_konusu = request.form['sablon_konusu']
        sablon_icerik = request.form['sablon_icerik']
        aktif = TRUE if request.form.get('aktif') == 'on' else 0
        
        conn.execute('''
            UPDATE email_sablonlari 
            SET sablon_konusu = %s, sablon_icerik = %s, aktif = %s
            WHERE id = %s
        ''', (sablon_konusu, sablon_icerik, aktif, sablon_id))
        conn.commit()
        conn.close()
        
        log_islem('GÃœNCELLEME', 'email_sablonlari', sablon_id, 
                 'Email ÅŸablonu gÃ¼ncellendi', session.get('username'))
        
        flash('Email ÅŸablonu baÅŸarÄ±yla gÃ¼ncellendi!', 'success')
        return redirect(url_for('email_ayarlari'))
    
    sablon = conn.execute('SELECT * FROM email_sablonlari WHERE id = %s', (sablon_id,)).fetchone()
    conn.close()
    
    if not sablon:
        flash('Åžablon bulunamadÄ±!', 'danger')
        return redirect(url_for('email_ayarlari'))
    
    return render_template('email_sablon_duzenle.html', sablon=sablon)

@app.route('/email-test-gonder', methods=['POST'])
@admin_required
def email_test_gonder():
    """Test email gÃ¶nder"""
    test_email = request.form.get('test_email')
    
    if not test_email:
        flash('Test email adresi girilmedi!', 'danger')
        return redirect(url_for('email_ayarlari'))
    
    try:
        from email_service import email_gonder
        
        test_icerik = """
        <h2>Test Email</h2>
        <p>Bu bir test emailidir. Email ayarlarÄ±nÄ±z doÄŸru Ã§alÄ±ÅŸÄ±yor!</p>
        <p><strong>GÃ¶nderim ZamanÄ±:</strong> {}</p>
        <hr>
        <p style="color: #666; font-size: 12px;">
            Team Guerilla - Ä°K YÃ¶netim Sistemi<br>
            Test Email
        </p>
        """.format(datetime.now().strftime('%d.%m.%Y %H:%M:%S'))
        
        basarili, mesaj = email_gonder(
            test_email,
            'Test Email - Ä°K Portal',
            test_icerik,
            'test_email'
        )
        
        if basarili:
            flash(f'Test email baÅŸarÄ±yla gÃ¶nderildi: {test_email}', 'success')
        else:
            flash(f'Test email gÃ¶nderilemedi: {mesaj}', 'danger')
            
    except Exception as e:
        flash(f'Hata: {str(e)}', 'danger')
    
    return redirect(url_for('email_ayarlari'))

@app.route('/email-loglari/temizle', methods=['POST'])
@admin_required
def email_loglari_temizle():
    """Email loglarÄ±nÄ± temizle"""
    conn = get_db()
    
    # 30 gÃ¼nden eski loglarÄ± sil
    conn.execute("""
        DELETE FROM email_loglari 
        WHERE gonderim_tarihi < datetime('now', '-30 days')
    """)
    
    silinen = conn.total_changes
    conn.commit()
    conn.close()
    
    log_islem('SÄ°LME', 'email_loglari', 0, 
             f'{silinen} eski email logu temizlendi', session.get('username'))
    
    flash(f'{silinen} eski email logu temizlendi!', 'success')
    return redirect(url_for('email_ayarlari'))    
    
# ============================================
# TANIMLAR YÃ–NETÄ°MÄ° (MÃ¼dÃ¼rlÃ¼k, DirektÃ¶rlÃ¼k)
# ============================================

# tanimlar() route'unu app.py'de gÃ¼ncelle:

@app.route('/tanimlar')
@admin_required
def tanimlar():
    """Sistem tanÄ±mlarÄ± - MÃ¼dÃ¼rlÃ¼k, DirektÃ¶rlÃ¼k, Ã‡alÄ±ÅŸma Åžekilleri, Kaynaklar"""
    conn = get_db()
    
    # Ä°statistikler
    stats = {
        'mudurluk_sayisi': conn.execute('SELECT COUNT(*) as cnt FROM mudurluker WHERE aktif = TRUE').fetchone()['cnt'],
        'direktorluk_sayisi': conn.execute('SELECT COUNT(*) as cnt FROM direktorlukler WHERE aktif = TRUE').fetchone()['cnt'],
        'calisma_sekli_sayisi': conn.execute('SELECT COUNT(*) as cnt FROM calisma_sekilleri WHERE aktif = TRUE').fetchone()['cnt'],
        'kaynak_sayisi': conn.execute('SELECT COUNT(*) as cnt FROM kaynaklar WHERE aktif = TRUE').fetchone()['cnt']
    }
    
    # MÃ¼dÃ¼rlÃ¼kler
    mudurluker = conn.execute('''
        SELECT * FROM mudurluker 
        ORDER BY mudurluk_adi
    ''').fetchall()
    
    # DirektÃ¶rlÃ¼kler
    direktorlukler = conn.execute('''
        SELECT * FROM direktorlukler 
        ORDER BY direktorluk_adi
    ''').fetchall()
    
    # Ã‡alÄ±ÅŸma Åžekilleri - YENÄ°
    calisma_sekilleri = conn.execute('''
        SELECT cs.*,
               COUNT(hk.id) as kadro_sayisi
        FROM calisma_sekilleri cs
        LEFT JOIN hedef_kadrolar hk ON cs.calisma_sekli = hk.calisma_sekli
        GROUP BY cs.id
        ORDER BY cs.calisma_sekli
    ''').fetchall()
    
    # Kaynaklar - YENÄ°
    kaynaklar = conn.execute('''
        SELECT k.*,
               COUNT(a.id) as aday_sayisi
        FROM kaynaklar k
        LEFT JOIN adaylar a ON k.id = a.kaynak_id
        GROUP BY k.id
        ORDER BY k.kaynak_adi
    ''').fetchall()
    
    conn.close()
    
    return render_template('tanimlar.html', 
                         stats=stats,
                         mudurluker=mudurluker, 
                         direktorlukler=direktorlukler,
                         calisma_sekilleri=calisma_sekilleri,
                         kaynaklar=kaynaklar)

@app.route('/mudurluk/ekle', methods=['POST'])
@admin_required
def mudurluk_ekle():
    """Yeni mÃ¼dÃ¼rlÃ¼k ekle"""
    mudurluk_adi = request.form['mudurluk_adi']
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO mudurluker (mudurluk_adi) VALUES (%s)', (mudurluk_adi,))
        mudurluk_id = cursor.lastrowid
        conn.commit()
        log_islem('EKLEME', 'mudurluker', mudurluk_id, f'{mudurluk_adi} mÃ¼dÃ¼rlÃ¼ÄŸÃ¼ eklendi')
        flash(f'MÃ¼dÃ¼rlÃ¼k "{mudurluk_adi}" baÅŸarÄ±yla eklendi!', 'success')
    except sqlite3.IntegrityError:
        flash('Bu mÃ¼dÃ¼rlÃ¼k zaten mevcut!', 'danger')
    
    conn.close()
    return redirect(url_for('tanimlar'))

@app.route('/direktorluk/ekle', methods=['POST'])
@admin_required
def direktorluk_ekle():
    """Yeni direktÃ¶rlÃ¼k ekle"""
    direktorluk_adi = request.form['direktorluk_adi']
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO direktorlukler (direktorluk_adi) VALUES (%s)', (direktorluk_adi,))
        direktorluk_id = cursor.lastrowid
        conn.commit()
        log_islem('EKLEME', 'direktorlukler', direktorluk_id, f'{direktorluk_adi} direktÃ¶rlÃ¼ÄŸÃ¼ eklendi')
        flash(f'DirektÃ¶rlÃ¼k "{direktorluk_adi}" baÅŸarÄ±yla eklendi!', 'success')
    except sqlite3.IntegrityError:
        flash('Bu direktÃ¶rlÃ¼k zaten mevcut!', 'danger')
    
    conn.close()
    return redirect(url_for('tanimlar'))

@app.route('/mudurluk/<int:mudurluk_id>/sil', methods=['POST'])
@admin_required
def mudurluk_sil(mudurluk_id):
    """MÃ¼dÃ¼rlÃ¼k sil"""
    conn = get_db()
    mudurluk = conn.execute('SELECT mudurluk_adi FROM mudurluker WHERE id = %s', (mudurluk_id,)).fetchone()
    
    if mudurluk:
        conn.execute('DELETE FROM mudurluker WHERE id = %s', (mudurluk_id,))
        conn.commit()
        log_islem('SÄ°LME', 'mudurluker', mudurluk_id, f'{mudurluk["mudurluk_adi"]} mÃ¼dÃ¼rlÃ¼ÄŸÃ¼ silindi')
        flash('MÃ¼dÃ¼rlÃ¼k baÅŸarÄ±yla silindi!', 'success')
    
    conn.close()
    return redirect(url_for('tanimlar'))

@app.route('/direktorluk/<int:direktorluk_id>/sil', methods=['POST'])
@admin_required
def direktorluk_sil(direktorluk_id):
    """DirektÃ¶rlÃ¼k sil"""
    conn = get_db()
    direktorluk = conn.execute('SELECT direktorluk_adi FROM direktorlukler WHERE id = %s', (direktorluk_id,)).fetchone()
    
    if direktorluk:
        conn.execute('DELETE FROM direktorlukler WHERE id = %s', (direktorluk_id,))
        conn.commit()
        log_islem('SÄ°LME', 'direktorlukler', direktorluk_id, f'{direktorluk["direktorluk_adi"]} direktÃ¶rlÃ¼ÄŸÃ¼ silindi')
        flash('DirektÃ¶rlÃ¼k baÅŸarÄ±yla silindi!', 'success')
    
    conn.close()
    return redirect(url_for('tanimlar'))

@app.route('/api/ilceler/<int:il_id>')
@login_required
def api_ilceler(il_id):
    """Ä°le gÃ¶re ilÃ§eleri getir (AJAX iÃ§in)"""
    conn = get_db()
    ilceler = conn.execute('''
        SELECT id, ilce_adi 
        FROM ilceler 
        WHERE il_id = %s AND aktif = TRUE 
        ORDER BY ilce_adi
    ''', (il_id,)).fetchall()
    conn.close()
    
    return jsonify([{'id': ilce['id'], 'ilce_adi': ilce['ilce_adi']} for ilce in ilceler]) 

# ============================================
# MÃœÅžTERÄ° YÃ–NETÄ°MÄ°
# ============================================

@app.route('/musteriler')
@admin_required
def musteriler():
    """MÃ¼ÅŸteri listesi (Sadece Admin)"""
    conn = get_db()
    
    stats = {
        'toplam_musteri': conn.execute('SELECT COUNT(*) as cnt FROM musteriler WHERE aktif = TRUE').fetchone()['cnt'],
        'toplam_proje': conn.execute('SELECT COUNT(*) as cnt FROM projeler WHERE musteri_id IS NOT NULL').fetchone()['cnt']
    }
    
    musteriler = conn.execute('''
        SELECT m.*,
               COUNT(DISTINCT p.id) as proje_sayisi
        FROM musteriler m
        LEFT JOIN projeler p ON m.id = p.musteri_id
        GROUP BY m.id
        ORDER BY m.musteri_adi
    ''').fetchall()
    
    conn.close()
    return render_template('musteriler.html', musteriler=musteriler, stats=stats)

@app.route('/musteri/ekle', methods=['GET', 'POST'])
@admin_required
def musteri_ekle():
    """Yeni mÃ¼ÅŸteri ekle"""
    if request.method == 'POST':
        musteri_adi = request.form['musteri_adi']
        sektor = request.form.get('sektor', '')
        yetkili_kisi = request.form.get('yetkili_kisi', '')
        telefon = request.form.get('telefon', '')
        email = request.form.get('email', '')
        adres = request.form.get('adres', '')
        
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO musteriler (musteri_adi, sektor, yetkili_kisi, telefon, email, adres)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (musteri_adi, sektor, yetkili_kisi, telefon, email, adres))
            musteri_id = cursor.lastrowid
            
            # Logo yÃ¼kleme
            if 'logo' in request.files:
                file = request.files['logo']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    yeni_dosya_adi = f"musteri_{musteri_id}_{timestamp}_{filename}"
                    
                    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'musteriler')
                    os.makedirs(upload_path, exist_ok=True)
                    
                    logo_yolu = os.path.join(upload_path, yeni_dosya_adi)
                    file.save(logo_yolu)
                    
                    cursor.execute('UPDATE musteriler SET logo_yolu = %s WHERE id = %s', 
                                 (logo_yolu, musteri_id))
            
            conn.commit()
            log_islem('EKLEME', 'musteriler', musteri_id, f'{musteri_adi} mÃ¼ÅŸterisi eklendi')
            flash(f'MÃ¼ÅŸteri "{musteri_adi}" baÅŸarÄ±yla eklendi!', 'success')
            conn.close()
            return redirect(url_for('musteriler'))
            
        except sqlite3.IntegrityError:
            conn.close()
            flash('Bu mÃ¼ÅŸteri adÄ± zaten kullanÄ±lÄ±yor!', 'danger')
    
    return render_template('musteri_form.html')

@app.route('/musteri/<int:musteri_id>/duzenle', methods=['GET', 'POST'])
@admin_required
def musteri_duzenle(musteri_id):
    """MÃ¼ÅŸteri dÃ¼zenle"""
    conn = get_db()
    
    if request.method == 'POST':
        musteri_adi = request.form['musteri_adi']
        sektor = request.form.get('sektor', '')
        yetkili_kisi = request.form.get('yetkili_kisi', '')
        telefon = request.form.get('telefon', '')
        email = request.form.get('email', '')
        adres = request.form.get('adres', '')
        aktif = TRUE if request.form.get('aktif') == 'on' else 0
        
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE musteriler 
            SET musteri_adi = %s, sektor = %s, yetkili_kisi = %s, 
                telefon = %s, email = %s, adres = %s, aktif = %s
            WHERE id = %s
        ''', (musteri_adi, sektor, yetkili_kisi, telefon, email, adres, aktif, musteri_id))
        
        # Logo gÃ¼ncelleme
        if 'logo' in request.files:
            file = request.files['logo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                yeni_dosya_adi = f"musteri_{musteri_id}_{timestamp}_{filename}"
                
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'musteriler')
                os.makedirs(upload_path, exist_ok=True)
                
                logo_yolu = os.path.join(upload_path, yeni_dosya_adi)
                file.save(logo_yolu)
                
                cursor.execute('UPDATE musteriler SET logo_yolu = %s WHERE id = %s', 
                             (logo_yolu, musteri_id))
        
        conn.commit()
        log_islem('GÃœNCELLEME', 'musteriler', musteri_id, f'{musteri_adi} mÃ¼ÅŸterisi gÃ¼ncellendi')
        flash('MÃ¼ÅŸteri baÅŸarÄ±yla gÃ¼ncellendi!', 'success')
        conn.close()
        return redirect(url_for('musteriler'))
    
    musteri = conn.execute('SELECT * FROM musteriler WHERE id = %s', (musteri_id,)).fetchone()
    conn.close()
    
    if not musteri:
        flash('MÃ¼ÅŸteri bulunamadÄ±!', 'danger')
        return redirect(url_for('musteriler'))
    
    return render_template('musteri_edit.html', musteri=musteri)

@app.route('/musteri/<int:musteri_id>/sil', methods=['POST'])
@admin_required
def musteri_sil(musteri_id):
    """MÃ¼ÅŸteri sil"""
    conn = get_db()
    musteri = conn.execute('SELECT musteri_adi FROM musteriler WHERE id = %s', (musteri_id,)).fetchone()
    
    if musteri:
        conn.execute('DELETE FROM musteriler WHERE id = %s', (musteri_id,))
        conn.commit()
        log_islem('SÄ°LME', 'musteriler', musteri_id, f'{musteri["musteri_adi"]} mÃ¼ÅŸterisi silindi')
        flash('MÃ¼ÅŸteri baÅŸarÄ±yla silindi!', 'success')
    
    conn.close()
    return redirect(url_for('musteriler'))

@app.route('/musteri/logo/<int:musteri_id>')
def musteri_logo(musteri_id):
    """MÃ¼ÅŸteri logosunu gÃ¶ster"""
    conn = get_db()
    musteri = conn.execute('SELECT logo_yolu FROM musteriler WHERE id = %s', (musteri_id,)).fetchone()
    conn.close()
    
    if musteri and musteri['logo_yolu'] and os.path.exists(musteri['logo_yolu']):
        from flask import send_file
        return send_file(musteri['logo_yolu'])
    else:
        # VarsayÄ±lan logo
        from flask import send_file
        default_logo_path = os.path.join('static', 'images', 'default-company.png')
        if os.path.exists(default_logo_path):
            return send_file(default_logo_path)
        else:
            return '', 404    
            
@app.route('/adaylar/import')
@manager_required
def adaylar_import():
    """Adaylar toplu yÃ¼kleme sayfasÄ±"""
    return render_template('adaylar_import.html')

@app.route('/adaylar/import/sablon-indir')
@manager_required
def adaylar_import_sablon():
    """Adaylar import ÅŸablonu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Adaylar"
    
    # BaÅŸlÄ±k satÄ±rÄ±
    headers = [
        'Ad Soyad*', 'Telefon', 'Email', 'TC Kimlik', 
        'Proje AdÄ±*', 'Pozisyon AdÄ±*', 'MÃ¼dÃ¼rlÃ¼k', 'DirektÃ¶rlÃ¼k',
        'Ä°l*', 'Ä°lÃ§e', 'MaÄŸaza AdÄ±', 'AraÃ§ Durumu*', 
        'BaÅŸvuru Tarihi', 'Notlar'
    ]
    
    # Stil tanÄ±mlamalarÄ±
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # BaÅŸlÄ±klarÄ± ekle
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # AÃ§Ä±klama satÄ±rÄ± (2. satÄ±r)
    aciklamalar = [
        'Zorunlu', 'Opsiyonel', 'Opsiyonel', 'Opsiyonel',
        'Sistemde kayÄ±tlÄ± proje adÄ±', 'Kadro pozisyon adÄ±', 'Opsiyonel', 'Opsiyonel',
        'Sistemde kayÄ±tlÄ± il adÄ±', 'Opsiyonel', 'Opsiyonel', 'AraÃ§lÄ± veya AraÃ§sÄ±z',
        'YYYY-MM-DD formatÄ±nda', 'Opsiyonel notlar'
    ]
    
    for col, aciklama in enumerate(aciklamalar, 1):
        cell = ws.cell(row=2, column=col, value=aciklama)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Ã–rnek veri satÄ±rlarÄ± (3-5. satÄ±rlar)
    ornekler = [
        ['Ahmet YÄ±lmaz', '05551234567', 'ahmet@example.com', '12345678901', 
         'Migros Projesi', 'SatÄ±ÅŸ DanÄ±ÅŸmanÄ±', 'Perakende MÃ¼dÃ¼rlÃ¼ÄŸÃ¼', 'Ä°stanbul DirektÃ¶rlÃ¼ÄŸÃ¼',
         'Ä°stanbul', 'KadÄ±kÃ¶y', 'Migros AcÄ±badem', 'AraÃ§sÄ±z', 
         '2025-01-15', 'Ä°yi bir aday'],
        ['AyÅŸe Demir', '05559876543', 'ayse@example.com', '98765432109',
         'Migros Projesi', 'Reyon GÃ¶revlisi', '', '',
         'Ankara', 'Ã‡ankaya', '', 'AraÃ§lÄ±',
         '2025-01-16', ''],
        ['Mehmet Kaya', '', '', '',
         'CarrefourSA', 'Kasa GÃ¶revlisi', '', '',
         'Ä°zmir', 'Konak', 'CarrefourSA Alsancak', 'AraÃ§sÄ±z',
         '', 'Acil deÄŸerlendirilmeli']
    ]
    
    for row_idx, ornek in enumerate(ornekler, 3):
        for col_idx, deger in enumerate(ornek, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=deger)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
    
    # SÃ¼tun geniÅŸliklerini ayarla
    column_widths = [20, 15, 25, 15, 20, 20, 20, 25, 15, 15, 20, 15, 15, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # SatÄ±r yÃ¼ksekliklerini ayarla
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    
    # DosyayÄ± belleÄŸe kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'adaylar_import_sablonu_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/adaylar/import/yukle', methods=['POST'])
@manager_required
def adaylar_import_yukle():
    """Adaylar Excel dosyasÄ±nÄ± yÃ¼kle ve iÅŸle"""
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'Dosya seÃ§ilmedi!'}), 400
    
    file = request.files['excel_file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Dosya seÃ§ilmedi!'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'GeÃ§ersiz dosya formatÄ±! Sadece Excel dosyalarÄ± (.xlsx, .xls) kabul edilir.'}), 400
    
    try:
        # Excel dosyasÄ±nÄ± oku
        df = pd.read_excel(file, sheet_name=0)
        
        # BaÅŸlÄ±k satÄ±rÄ±ndan sonraki aÃ§Ä±klama satÄ±rÄ±nÄ± atla (eÄŸer varsa)
        if len(df) > 0 and 'Zorunlu' in str(df.iloc[0].values):
            df = df.iloc[1:].reset_index(drop=True)
        
        # BoÅŸ satÄ±rlarÄ± temizle
        df = df.dropna(how='all')
        
        conn = get_db()
        cursor = conn.cursor()
        
        basarili_sayisi = 0
        hatalar = []
        
        for index, row in df.iterrows():
            satir_no = index + 3  # Excel'de baÅŸlÄ±k + aÃ§Ä±klama + 1-indexed
            
            try:
                # Zorunlu alanlarÄ± kontrol et
                ad_soyad = str(row.get('Ad Soyad*', '')).strip()
                proje_adi = str(row.get('Proje AdÄ±*', '')).strip()
                pozisyon_adi = str(row.get('Pozisyon AdÄ±*', '')).strip()
                il_adi = str(row.get('Ä°l*', '')).strip()
                aracli_durum = str(row.get('AraÃ§ Durumu*', '')).strip()
                
                if not ad_soyad or ad_soyad == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Ad Soyad zorunludur")
                    continue
                
                if not proje_adi or proje_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Proje AdÄ± zorunludur")
                    continue
                
                if not pozisyon_adi or pozisyon_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Pozisyon AdÄ± zorunludur")
                    continue
                
                if not il_adi or il_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Ä°l zorunludur")
                    continue
                
                if not aracli_durum or aracli_durum == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: AraÃ§ Durumu zorunludur")
                    continue
                
                # AraÃ§ durumu kontrolÃ¼
                if aracli_durum not in ['AraÃ§lÄ±', 'AraÃ§sÄ±z']:
                    hatalar.append(f"SatÄ±r {satir_no}: AraÃ§ Durumu 'AraÃ§lÄ±' veya 'AraÃ§sÄ±z' olmalÄ±dÄ±r")
                    continue
                
                # Proje kontrolÃ¼
                proje = cursor.execute('SELECT id FROM projeler WHERE proje_adi = %s AND aktif = TRUE', 
                                      (proje_adi,)).fetchone()
                if not proje:
                    hatalar.append(f"SatÄ±r {satir_no}: '{proje_adi}' projesi bulunamadÄ±")
                    continue
                proje_id = proje['id']
                
                # Ä°l kontrolÃ¼
                il = cursor.execute('SELECT id FROM iller WHERE il_adi = %s AND aktif = TRUE', 
                                   (il_adi,)).fetchone()
                if not il:
                    hatalar.append(f"SatÄ±r {satir_no}: '{il_adi}' ili bulunamadÄ±")
                    continue
                il_id = il['id']
                
                # Ä°lÃ§e kontrolÃ¼ (opsiyonel)
                ilce_adi = str(row.get('Ä°lÃ§e', '')).strip()
                ilce_id = None
                if ilce_adi and ilce_adi != 'nan':
                    ilce = cursor.execute('SELECT id FROM ilceler WHERE ilce_adi = %s AND il_id = %s AND aktif = TRUE',
                                        (ilce_adi, il_id)).fetchone()
                    if ilce:
                        ilce_id = ilce['id']
                
                # MÃ¼dÃ¼rlÃ¼k kontrolÃ¼ (opsiyonel)
                mudurluk_adi = str(row.get('MÃ¼dÃ¼rlÃ¼k', '')).strip()
                mudurluk_id = None
                if mudurluk_adi and mudurluk_adi != 'nan':
                    mudurluk = cursor.execute('SELECT id FROM mudurluker WHERE mudurluk_adi = %s AND aktif = TRUE',
                                             (mudurluk_adi,)).fetchone()
                    if mudurluk:
                        mudurluk_id = mudurluk['id']
                
                # DirektÃ¶rlÃ¼k kontrolÃ¼ (opsiyonel)
                direktorluk_adi = str(row.get('DirektÃ¶rlÃ¼k', '')).strip()
                direktorluk_id = None
                if direktorluk_adi and direktorluk_adi != 'nan':
                    direktorluk = cursor.execute('SELECT id FROM direktorlukler WHERE direktorluk_adi = %s AND aktif = TRUE',
                                                (direktorluk_adi,)).fetchone()
                    if direktorluk:
                        direktorluk_id = direktorluk['id']
                
                # Kadro ara veya oluÅŸtur
                magaza_adi = str(row.get('MaÄŸaza AdÄ±', '')).strip()
                if magaza_adi == 'nan':
                    magaza_adi = ''
                
                kadro = cursor.execute('''
                    SELECT id FROM hedef_kadrolar 
                    WHERE proje_id = %s AND pozisyon_adi = %s 
                    AND il_id = %s AND COALESCE(ilce_id, 0) = COALESCE(%s, 0)
                    AND COALESCE(magaza_adi, '') = %s
                    AND aracli_durum = %s
                ''', (proje_id, pozisyon_adi, il_id, ilce_id, magaza_adi, aracli_durum)).fetchone()
                
                if not kadro:
                    # Kadro yoksa oluÅŸtur
                    cursor.execute('''
                        INSERT INTO hedef_kadrolar 
                        (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id, 
                         magaza_adi, aracli_durum, hedef_kisi_sayisi)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1)
                    ''', (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id, 
                          magaza_adi, aracli_durum))
                    kadro_id = cursor.lastrowid
                else:
                    kadro_id = kadro['id']
                
                # Opsiyonel alanlar
                telefon = str(row.get('Telefon', '')).strip()
                if telefon == 'nan':
                    telefon = ''
                
                email = str(row.get('Email', '')).strip()
                if email == 'nan':
                    email = ''
                
                tc_kimlik = str(row.get('TC Kimlik', '')).strip()
                if tc_kimlik == 'nan':
                    tc_kimlik = ''
                
                notlar = str(row.get('Notlar', '')).strip()
                if notlar == 'nan':
                    notlar = ''
                
                # BaÅŸvuru tarihi
                basvuru_tarihi = None
                basvuru_tarihi_raw = row.get('BaÅŸvuru Tarihi', '')
                if pd.notna(basvuru_tarihi_raw):
                    try:
                        if isinstance(basvuru_tarihi_raw, str):
                            basvuru_tarihi = datetime.strptime(basvuru_tarihi_raw, '%Y-%m-%d').strftime('%Y-%m-%d')
                        else:
                            basvuru_tarihi = basvuru_tarihi_raw.strftime('%Y-%m-%d')
                    except:
                        pass
                
                # AdayÄ± ekle
                cursor.execute('''
                    INSERT INTO adaylar (kadro_id, ad_soyad, telefon, email, tc_kimlik, basvuru_tarihi, notlar)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                ''', (kadro_id, ad_soyad, telefon or None, email or None, tc_kimlik or None, 
                      basvuru_tarihi, notlar or None))
                
                basarili_sayisi += 1
                
            except Exception as e:
                hatalar.append(f"SatÄ±r {satir_no}: {str(e)}")
                continue
        
        conn.commit()
        conn.close()
        
        # Log kaydÄ±
        log_islem('TOPLU_EKLEME', 'adaylar', 0, 
                 f'{basarili_sayisi} aday Excel ile yÃ¼klendi', 
                 session.get('username'))
        
        # SonuÃ§ mesajÄ±
        if basarili_sayisi > 0 and len(hatalar) == 0:
            return jsonify({
                'success': True,
                'message': f'âœ… {basarili_sayisi} aday baÅŸarÄ±yla eklendi!',
                'basarili_sayisi': basarili_sayisi,
                'hata_sayisi': 0
            })
        elif basarili_sayisi > 0 and len(hatalar) > 0:
            return jsonify({
                'success': True,
                'message': f'âš ï¸ {basarili_sayisi} aday eklendi, {len(hatalar)} satÄ±rda hata oluÅŸtu',
                'basarili_sayisi': basarili_sayisi,
                'hata_sayisi': len(hatalar),
                'hatalar': hatalar[:10]  # Ä°lk 10 hatayÄ± gÃ¶ster
            })
        else:
            return jsonify({
                'success': False,
                'message': f'âŒ HiÃ§bir aday eklenemedi. {len(hatalar)} hata bulundu.',
                'basarili_sayisi': 0,
                'hata_sayisi': len(hatalar),
                'hatalar': hatalar[:10]
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Dosya iÅŸlenirken hata oluÅŸtu: {str(e)}'
        }), 500

# ============================================
# Ã‡ALIÅžANLAR IMPORT
# ============================================

@app.route('/calisanlar/import')
@manager_required
def calisanlar_import():
    """Ã‡alÄ±ÅŸanlar toplu yÃ¼kleme sayfasÄ±"""
    return render_template('calisanlar_import.html')

@app.route('/calisanlar/import/sablon-indir')
@manager_required
def calisanlar_import_sablon():
    """Ã‡alÄ±ÅŸanlar import ÅŸablonu indir"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ã‡alÄ±ÅŸanlar"
    
    # BaÅŸlÄ±k satÄ±rÄ±
    headers = [
        'Ad Soyad*', 'Telefon', 'Email', 'TC Kimlik*', 
        'Proje AdÄ±*', 'Pozisyon AdÄ±*', 'MÃ¼dÃ¼rlÃ¼k', 'DirektÃ¶rlÃ¼k',
        'Ä°l*', 'Ä°lÃ§e', 'MaÄŸaza AdÄ±', 'AraÃ§ Durumu*', 
        'Ä°ÅŸe BaÅŸlama Tarihi*'
    ]
    
    # Stil tanÄ±mlamalarÄ±
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # BaÅŸlÄ±klarÄ± ekle
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # AÃ§Ä±klama satÄ±rÄ±
    aciklamalar = [
        'Zorunlu', 'Opsiyonel', 'Opsiyonel', 'Zorunlu (11 haneli)',
        'Sistemde kayÄ±tlÄ± proje adÄ±', 'Kadro pozisyon adÄ±', 'Opsiyonel', 'Opsiyonel',
        'Sistemde kayÄ±tlÄ± il adÄ±', 'Opsiyonel', 'Opsiyonel', 'AraÃ§lÄ± veya AraÃ§sÄ±z',
        'YYYY-MM-DD formatÄ±nda zorunlu'
    ]
    
    for col, aciklama in enumerate(aciklamalar, 1):
        cell = ws.cell(row=2, column=col, value=aciklama)
        cell.font = Font(italic=True, size=9, color="666666")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Ã–rnek veri satÄ±rlarÄ±
    ornekler = [
        ['Fatma Åžahin', '05551112233', 'fatma@example.com', '11122233344',
         'Migros Projesi', 'SatÄ±ÅŸ DanÄ±ÅŸmanÄ±', 'Perakende MÃ¼dÃ¼rlÃ¼ÄŸÃ¼', 'Ä°stanbul DirektÃ¶rlÃ¼ÄŸÃ¼',
         'Ä°stanbul', 'KadÄ±kÃ¶y', 'Migros AcÄ±badem', 'AraÃ§sÄ±z',
         '2025-01-10'],
        ['Ali YÄ±lmaz', '05559998877', 'ali@example.com', '55566677788',
         'CarrefourSA', 'Reyon GÃ¶revlisi', '', '',
         'Ankara', '', 'CarrefourSA KÄ±zÄ±lay', 'AraÃ§lÄ±',
         '2025-01-05']
    ]
    
    for row_idx, ornek in enumerate(ornekler, 3):
        for col_idx, deger in enumerate(ornek, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=deger)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
    
    # SÃ¼tun geniÅŸliklerini ayarla
    column_widths = [20, 15, 25, 15, 20, 20, 20, 25, 15, 15, 20, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # SatÄ±r yÃ¼ksekliklerini ayarla
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25
    
    # DosyayÄ± belleÄŸe kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'calisanlar_import_sablonu_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/calisanlar/import/yukle', methods=['POST'])
@manager_required
def calisanlar_import_yukle():
    """Ã‡alÄ±ÅŸanlar Excel dosyasÄ±nÄ± yÃ¼kle ve iÅŸle"""
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'Dosya seÃ§ilmedi!'}), 400
    
    file = request.files['excel_file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'Dosya seÃ§ilmedi!'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'GeÃ§ersiz dosya formatÄ±! Sadece Excel dosyalarÄ± (.xlsx, .xls) kabul edilir.'}), 400
    
    try:
        # Excel dosyasÄ±nÄ± oku
        df = pd.read_excel(file, sheet_name=0)
        
        # BaÅŸlÄ±k satÄ±rÄ±ndan sonraki aÃ§Ä±klama satÄ±rÄ±nÄ± atla
        if len(df) > 0 and 'Zorunlu' in str(df.iloc[0].values):
            df = df.iloc[1:].reset_index(drop=True)
        
        # BoÅŸ satÄ±rlarÄ± temizle
        df = df.dropna(how='all')
        
        conn = get_db()
        cursor = conn.cursor()
        
        basarili_sayisi = 0
        hatalar = []
        
        for index, row in df.iterrows():
            satir_no = index + 3
            
            try:
                # Zorunlu alanlarÄ± kontrol et
                ad_soyad = str(row.get('Ad Soyad*', '')).strip()
                tc_kimlik = str(row.get('TC Kimlik*', '')).strip()
                proje_adi = str(row.get('Proje AdÄ±*', '')).strip()
                pozisyon_adi = str(row.get('Pozisyon AdÄ±*', '')).strip()
                il_adi = str(row.get('Ä°l*', '')).strip()
                aracli_durum = str(row.get('AraÃ§ Durumu*', '')).strip()
                
                if not ad_soyad or ad_soyad == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Ad Soyad zorunludur")
                    continue
                
                if not tc_kimlik or tc_kimlik == 'nan' or len(tc_kimlik) != 11:
                    hatalar.append(f"SatÄ±r {satir_no}: TC Kimlik zorunludur ve 11 haneli olmalÄ±dÄ±r")
                    continue
                
                if not proje_adi or proje_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Proje AdÄ± zorunludur")
                    continue
                
                if not pozisyon_adi or pozisyon_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Pozisyon AdÄ± zorunludur")
                    continue
                
                if not il_adi or il_adi == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: Ä°l zorunludur")
                    continue
                
                if not aracli_durum or aracli_durum == 'nan':
                    hatalar.append(f"SatÄ±r {satir_no}: AraÃ§ Durumu zorunludur")
                    continue
                
                if aracli_durum not in ['AraÃ§lÄ±', 'AraÃ§sÄ±z']:
                    hatalar.append(f"SatÄ±r {satir_no}: AraÃ§ Durumu 'AraÃ§lÄ±' veya 'AraÃ§sÄ±z' olmalÄ±dÄ±r")
                    continue
                
                # Ä°ÅŸe baÅŸlama tarihi kontrolÃ¼
                ise_baslama_tarihi = None
                ise_baslama_tarihi_raw = row.get('Ä°ÅŸe BaÅŸlama Tarihi*', '')
                if pd.isna(ise_baslama_tarihi_raw):
                    hatalar.append(f"SatÄ±r {satir_no}: Ä°ÅŸe BaÅŸlama Tarihi zorunludur")
                    continue
                
                try:
                    if isinstance(ise_baslama_tarihi_raw, str):
                        ise_baslama_tarihi = datetime.strptime(ise_baslama_tarihi_raw, '%Y-%m-%d').strftime('%Y-%m-%d')
                    else:
                        ise_baslama_tarihi = ise_baslama_tarihi_raw.strftime('%Y-%m-%d')
                except:
                    hatalar.append(f"SatÄ±r {satir_no}: Ä°ÅŸe BaÅŸlama Tarihi formatÄ± hatalÄ± (YYYY-MM-DD olmalÄ±)")
                    continue
                
                # Proje kontrolÃ¼
                proje = cursor.execute('SELECT id FROM projeler WHERE proje_adi = %s AND aktif = TRUE',
                                      (proje_adi,)).fetchone()
                if not proje:
                    hatalar.append(f"SatÄ±r {satir_no}: '{proje_adi}' projesi bulunamadÄ±")
                    continue
                proje_id = proje['id']
                
                # Ä°l kontrolÃ¼
                il = cursor.execute('SELECT id FROM iller WHERE il_adi = %s AND aktif = TRUE',
                                   (il_adi,)).fetchone()
                if not il:
                    hatalar.append(f"SatÄ±r {satir_no}: '{il_adi}' ili bulunamadÄ±")
                    continue
                il_id = il['id']
                
                # Ä°lÃ§e, MÃ¼dÃ¼rlÃ¼k, DirektÃ¶rlÃ¼k kontrolÃ¼ (opsiyonel)
                ilce_adi = str(row.get('Ä°lÃ§e', '')).strip()
                ilce_id = None
                if ilce_adi and ilce_adi != 'nan':
                    ilce = cursor.execute('SELECT id FROM ilceler WHERE ilce_adi = %s AND il_id = %s AND aktif = TRUE',
                                        (ilce_adi, il_id)).fetchone()
                    if ilce:
                        ilce_id = ilce['id']
                
                mudurluk_adi = str(row.get('MÃ¼dÃ¼rlÃ¼k', '')).strip()
                mudurluk_id = None
                if mudurluk_adi and mudurluk_adi != 'nan':
                    mudurluk = cursor.execute('SELECT id FROM mudurluker WHERE mudurluk_adi = %s AND aktif = TRUE',
                                             (mudurluk_adi,)).fetchone()
                    if mudurluk:
                        mudurluk_id = mudurluk['id']
                
                direktorluk_adi = str(row.get('DirektÃ¶rlÃ¼k', '')).strip()
                direktorluk_id = None
                if direktorluk_adi and direktorluk_adi != 'nan':
                    direktorluk = cursor.execute('SELECT id FROM direktorlukler WHERE direktorluk_adi = %s AND aktif = TRUE',
                                                (direktorluk_adi,)).fetchone()
                    if direktorluk:
                        direktorluk_id = direktorluk['id']
                
                # Kadro ara veya oluÅŸtur
                magaza_adi = str(row.get('MaÄŸaza AdÄ±', '')).strip()
                if magaza_adi == 'nan':
                    magaza_adi = ''
                
                kadro = cursor.execute('''
                    SELECT id FROM hedef_kadrolar
                    WHERE proje_id = %s AND pozisyon_adi = %s
                    AND il_id = %s AND COALESCE(ilce_id, 0) = COALESCE(%s, 0)
                    AND COALESCE(magaza_adi, '') = %s
                    AND aracli_durum = %s
                ''', (proje_id, pozisyon_adi, il_id, ilce_id, magaza_adi, aracli_durum)).fetchone()
                
                if not kadro:
                    # Kadro yoksa oluÅŸtur
                    cursor.execute('''
                        INSERT INTO hedef_kadrolar
                        (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id,
                         magaza_adi, aracli_durum, hedef_kisi_sayisi, dolu_kisi_sayisi)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 1, 1)
                    ''', (proje_id, pozisyon_adi, mudurluk_id, direktorluk_id, il_id, ilce_id,
                          magaza_adi, aracli_durum))
                    kadro_id = cursor.lastrowid
                else:
                    kadro_id = kadro['id']
                    # Dolu sayÄ±sÄ±nÄ± artÄ±r
                    cursor.execute('''
                        UPDATE hedef_kadrolar
                        SET dolu_kisi_sayisi = dolu_kisi_sayisi + 1
                        WHERE id = %s
                    ''', (kadro_id,))
                
                # Opsiyonel alanlar
                telefon = str(row.get('Telefon', '')).strip()
                if telefon == 'nan':
                    telefon = ''
                
                email = str(row.get('Email', '')).strip()
                if email == 'nan':
                    email = ''
                
                # Ã‡alÄ±ÅŸanÄ± ekle (aday_id olmadan direkt ekleme)
                cursor.execute('''
                    INSERT INTO calisanlar 
                    (aday_id, kadro_id, ad_soyad, telefon, email, tc_kimlik, ise_baslama_tarihi)
                    VALUES (0, %s, %s, %s, %s, %s, %s)
                ''', (kadro_id, ad_soyad, telefon or None, email or None, tc_kimlik, ise_baslama_tarihi))
                
                basarili_sayisi += 1
                
            except Exception as e:
                hatalar.append(f"SatÄ±r {satir_no}: {str(e)}")
                continue
        
        conn.commit()
        conn.close()
        
        # Log kaydÄ±
        log_islem('TOPLU_EKLEME', 'calisanlar', 0,
                 f'{basarili_sayisi} Ã§alÄ±ÅŸan Excel ile yÃ¼klendi',
                 session.get('username'))
        
        # SonuÃ§ mesajÄ±
        if basarili_sayisi > 0 and len(hatalar) == 0:
            return jsonify({
                'success': True,
                'message': f'âœ… {basarili_sayisi} Ã§alÄ±ÅŸan baÅŸarÄ±yla eklendi!',
                'basarili_sayisi': basarili_sayisi,
                'hata_sayisi': 0
            })
        elif basarili_sayisi > 0 and len(hatalar) > 0:
            return jsonify({
                'success': True,
                'message': f'âš ï¸ {basarili_sayisi} Ã§alÄ±ÅŸan eklendi, {len(hatalar)} satÄ±rda hata oluÅŸtu',
                'basarili_sayisi': basarili_sayisi,
                'hata_sayisi': len(hatalar),
                'hatalar': hatalar[:10]
            })
        else:
            return jsonify({
                'success': False,
                'message': f'âŒ HiÃ§bir Ã§alÄ±ÅŸan eklenemedi. {len(hatalar)} hata bulundu.',
                'basarili_sayisi': 0,
                'hata_sayisi': len(hatalar),
                'hatalar': hatalar[:10]
            }), 400
            
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Dosya iÅŸlenirken hata oluÅŸtu: {str(e)}'
        }), 500            

@app.route('/kaynaklar')
@admin_required
def kaynaklar():
    """Kaynak listesi (Sadece Admin)"""
    conn = get_db()
    
    stats = {
        'toplam_kaynak': conn.execute('SELECT COUNT(*) as cnt FROM kaynaklar WHERE aktif = TRUE').fetchone()['cnt'],
        'kaynak_kullanan_aday': conn.execute('SELECT COUNT(DISTINCT kaynak_id) as cnt FROM adaylar WHERE kaynak_id IS NOT NULL').fetchone()['cnt']
    }
    
    kaynaklar = conn.execute('''
        SELECT k.*,
               COUNT(a.id) as aday_sayisi
        FROM kaynaklar k
        LEFT JOIN adaylar a ON k.id = a.kaynak_id
        GROUP BY k.id
        ORDER BY k.kaynak_adi
    ''').fetchall()
    
    # Manuel girilen diÄŸer kaynaklar
    diger_kaynaklar = conn.execute('''
        SELECT kaynak_diger, COUNT(*) as sayi
        FROM adaylar
        WHERE kaynak_diger IS NOT NULL AND kaynak_diger != ''
        GROUP BY kaynak_diger
        ORDER BY sayi DESC
    ''').fetchall()
    
    conn.close()
    
    return render_template('kaynaklar.html', 
                         kaynaklar=kaynaklar,
                         diger_kaynaklar=diger_kaynaklar,
                         stats=stats)

@app.route('/kaynak/ekle', methods=['POST'])
@admin_required
def kaynak_ekle():
    """Yeni kaynak ekle"""
    kaynak_adi = request.form['kaynak_adi']
    aciklama = request.form.get('aciklama', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO kaynaklar (kaynak_adi, aciklama) VALUES (%s, %s)', 
                      (kaynak_adi, aciklama))
        kaynak_id = cursor.lastrowid
        conn.commit()
        log_islem('EKLEME', 'kaynaklar', kaynak_id, f'{kaynak_adi} kaynaÄŸÄ± eklendi')
        flash(f'Kaynak "{kaynak_adi}" baÅŸarÄ±yla eklendi!', 'success')
    except sqlite3.IntegrityError:
        flash('Bu kaynak zaten mevcut!', 'danger')
    
    conn.close()
    return redirect(url_for('kaynaklar'))

@app.route('/kaynak/<int:kaynak_id>/sil', methods=['POST'])
@admin_required
def kaynak_sil(kaynak_id):
    """Kaynak sil"""
    conn = get_db()
    
    # Ã–nce bu kaynaÄŸÄ± kullanan aday var mÄ± kontrol et
    aday_sayisi = conn.execute('SELECT COUNT(*) as cnt FROM adaylar WHERE kaynak_id = %s', 
                               (kaynak_id,)).fetchone()['cnt']
    
    if aday_sayisi > 0:
        flash(f'Bu kaynak {aday_sayisi} aday tarafÄ±ndan kullanÄ±lÄ±yor, silinemez!', 'danger')
    else:
        kaynak = conn.execute('SELECT kaynak_adi FROM kaynaklar WHERE id = %s', (kaynak_id,)).fetchone()
        if kaynak:
            conn.execute('DELETE FROM kaynaklar WHERE id = %s', (kaynak_id,))
            conn.commit()
            log_islem('SÄ°LME', 'kaynaklar', kaynak_id, f'{kaynak["kaynak_adi"]} kaynaÄŸÄ± silindi')
            flash('Kaynak baÅŸarÄ±yla silindi!', 'success')
    
    conn.close()
    return redirect(url_for('kaynaklar'))

@app.route('/kaynak/<int:kaynak_id>/toggle', methods=['POST'])
@admin_required
def kaynak_toggle(kaynak_id):
    """Kaynak aktif/pasif durumu deÄŸiÅŸtir"""
    conn = get_db()
    
    kaynak = conn.execute('SELECT * FROM kaynaklar WHERE id = %s', (kaynak_id,)).fetchone()
    if kaynak:
        yeni_durum = 0 if kaynak['aktif'] else 1
        conn.execute('UPDATE kaynaklar SET aktif = %s WHERE id = %s', (yeni_durum, kaynak_id))
        conn.commit()
        
        durum_text = 'aktif' if yeni_durum else 'pasif'
        log_islem('GÃœNCELLEME', 'kaynaklar', kaynak_id, 
                 f'{kaynak["kaynak_adi"]} kaynaÄŸÄ± {durum_text} yapÄ±ldÄ±')
        flash(f'Kaynak {durum_text} yapÄ±ldÄ±!', 'success')
    
    conn.close()
    return redirect(url_for('kaynaklar'))   

@app.route('/calisma-sekilleri')
@admin_required
def calisma_sekilleri():
    """Ã‡alÄ±ÅŸma ÅŸekilleri listesi (Sadece Admin)"""
    conn = get_db()
    
    stats = {
        'toplam_sekil': conn.execute('SELECT COUNT(*) as cnt FROM calisma_sekilleri WHERE aktif = TRUE').fetchone()['cnt'],
        'kullanan_kadro': conn.execute('SELECT COUNT(DISTINCT calisma_sekli) as cnt FROM hedef_kadrolar WHERE calisma_sekli IS NOT NULL').fetchone()['cnt']
    }
    
    calisma_sekilleri = conn.execute('''
        SELECT cs.*,
               COUNT(hk.id) as kadro_sayisi
        FROM calisma_sekilleri cs
        LEFT JOIN hedef_kadrolar hk ON cs.calisma_sekli = hk.calisma_sekli
        GROUP BY cs.id
        ORDER BY cs.calisma_sekli
    ''').fetchall()
    
    conn.close()
    
    return render_template('calisma_sekilleri.html', 
                         calisma_sekilleri=calisma_sekilleri,
                         stats=stats)

@app.route('/calisma-sekli/ekle', methods=['POST'])
@admin_required
def calisma_sekli_ekle():
    """Yeni Ã§alÄ±ÅŸma ÅŸekli ekle"""
    calisma_sekli = request.form['calisma_sekli']
    aciklama = request.form.get('aciklama', '')
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO calisma_sekilleri (calisma_sekli, aciklama) VALUES (%s, %s)', 
                      (calisma_sekli, aciklama))
        sekil_id = cursor.lastrowid
        conn.commit()
        log_islem('EKLEME', 'calisma_sekilleri', sekil_id, f'{calisma_sekli} Ã§alÄ±ÅŸma ÅŸekli eklendi')
        flash(f'Ã‡alÄ±ÅŸma ÅŸekli "{calisma_sekli}" baÅŸarÄ±yla eklendi!', 'success')
    except sqlite3.IntegrityError:
        flash('Bu Ã§alÄ±ÅŸma ÅŸekli zaten mevcut!', 'danger')
    
    conn.close()
    return redirect(url_for('calisma_sekilleri'))

@app.route('/calisma-sekli/<int:sekil_id>/sil', methods=['POST'])
@admin_required
def calisma_sekli_sil(sekil_id):
    """Ã‡alÄ±ÅŸma ÅŸekli sil"""
    conn = get_db()
    
    # Ã–nce bu Ã§alÄ±ÅŸma ÅŸeklini kullanan kadro var mÄ± kontrol et
    kadro_sayisi = conn.execute('SELECT COUNT(*) as cnt FROM hedef_kadrolar WHERE calisma_sekli = (SELECT calisma_sekli FROM calisma_sekilleri WHERE id = %s)', 
                                (sekil_id,)).fetchone()['cnt']
    
    if kadro_sayisi > 0:
        flash(f'Bu Ã§alÄ±ÅŸma ÅŸekli {kadro_sayisi} kadro tarafÄ±ndan kullanÄ±lÄ±yor, silinemez!', 'danger')
    else:
        sekil = conn.execute('SELECT calisma_sekli FROM calisma_sekilleri WHERE id = %s', (sekil_id,)).fetchone()
        if sekil:
            conn.execute('DELETE FROM calisma_sekilleri WHERE id = %s', (sekil_id,))
            conn.commit()
            log_islem('SÄ°LME', 'calisma_sekilleri', sekil_id, f'{sekil["calisma_sekli"]} Ã§alÄ±ÅŸma ÅŸekli silindi')
            flash('Ã‡alÄ±ÅŸma ÅŸekli baÅŸarÄ±yla silindi!', 'success')
    
    conn.close()
    return redirect(url_for('calisma_sekilleri'))

@app.route('/calisma-sekli/<int:sekil_id>/toggle', methods=['POST'])
@admin_required
def calisma_sekli_toggle(sekil_id):
    """Ã‡alÄ±ÅŸma ÅŸekli aktif/pasif durumu deÄŸiÅŸtir"""
    conn = get_db()
    
    sekil = conn.execute('SELECT * FROM calisma_sekilleri WHERE id = %s', (sekil_id,)).fetchone()
    if sekil:
        yeni_durum = 0 if sekil['aktif'] else 1
        conn.execute('UPDATE calisma_sekilleri SET aktif = %s WHERE id = %s', (yeni_durum, sekil_id))
        conn.commit()
        
        durum_text = 'aktif' if yeni_durum else 'pasif'
        log_islem('GÃœNCELLEME', 'calisma_sekilleri', sekil_id, 
                 f'{sekil["calisma_sekli"]} Ã§alÄ±ÅŸma ÅŸekli {durum_text} yapÄ±ldÄ±')
        flash(f'Ã‡alÄ±ÅŸma ÅŸekli {durum_text} yapÄ±ldÄ±!', 'success')
    
    conn.close()
    return redirect(url_for('calisma_sekilleri'))    
    
@app.route('/cikis-kayitlari/excel-export')
@login_required
def cikis_kayitlari_excel_export():
    """Ã‡Ä±kÄ±ÅŸ kayÄ±tlarÄ±nÄ± Excel'e aktar"""
    conn = get_db()
    
    # Ã‡Ä±kÄ±ÅŸ kayÄ±tlarÄ±nÄ± getir
    kayitlar = conn.execute('''
        SELECT ck.*,
               c.ad_soyad,
               c.telefon,
               c.email,
               c.tc_kimlik,
               c.ise_baslama_tarihi,
               hk.pozisyon_adi,
               hk.magaza_adi,
               p.proje_adi,
               m.mudurluk_adi,
               d.direktorluk_adi,
               i.il_adi,
               ic.ilce_adi
        FROM cikis_kayitlari ck
        LEFT JOIN calisanlar c ON ck.calisan_id = c.id
        LEFT JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
        LEFT JOIN projeler p ON hk.proje_id = p.id
        LEFT JOIN mudurluker m ON hk.mudurluk_id = m.id
        LEFT JOIN direktorlukler d ON hk.direktorluk_id = d.id
        LEFT JOIN iller i ON hk.il_id = i.id
        LEFT JOIN ilceler ic ON hk.ilce_id = ic.id
        ORDER BY ck.cikis_tarihi DESC, ck.olusturma_tarihi DESC
    ''').fetchall()
    
    conn.close()
    
    # Excel workbook oluÅŸtur
    wb = Workbook()
    ws = wb.active
    ws.title = "Ã‡Ä±kÄ±ÅŸ KayÄ±tlarÄ±"
    
    # Stil tanÄ±mlamalarÄ±
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # BaÅŸlÄ±klar
    headers = [
        'KayÄ±t ID', 'Ad Soyad', 'TC Kimlik', 'Telefon', 'Email',
        'Proje', 'Pozisyon', 'MÃ¼dÃ¼rlÃ¼k', 'DirektÃ¶rlÃ¼k',
        'Ä°l', 'Ä°lÃ§e', 'MaÄŸaza',
        'Ä°ÅŸe BaÅŸlama', 'Ã‡Ä±kÄ±ÅŸ Tarihi', 'Ã‡Ä±kÄ±ÅŸ Nedeni',
        'Liste Durumu', 'Tekrar Ä°ÅŸe AlÄ±nabilir',
        'Zimmet Teslim', 'KÄ±yafet Teslim', 'Anahtar Teslim', 'Kimlik Teslim',
        'Ä°hbar Tazminat', 'KÄ±dem Tazminat',
        'YÃ¶netici Notu', 'Ä°K Notu', 'Genel DeÄŸerlendirme',
        'Ä°ÅŸlem Yapan', 'KayÄ±t Tarihi'
    ]
    
    # BaÅŸlÄ±k satÄ±rÄ±nÄ± yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Veri satÄ±rlarÄ±nÄ± yaz
    for row_num, kayit in enumerate(kayitlar, 2):
        # Temel bilgiler
        ws.cell(row=row_num, column=1, value=kayit['id']).alignment = center_alignment
        ws.cell(row=row_num, column=2, value=kayit['ad_soyad']).alignment = cell_alignment
        ws.cell(row=row_num, column=3, value=kayit['tc_kimlik'] or '-').alignment = center_alignment
        ws.cell(row=row_num, column=4, value=kayit['telefon'] or '-').alignment = center_alignment
        ws.cell(row=row_num, column=5, value=kayit['email'] or '-').alignment = cell_alignment
        
        # Ä°ÅŸ bilgileri
        ws.cell(row=row_num, column=6, value=kayit['proje_adi'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=7, value=kayit['pozisyon_adi'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=8, value=kayit['mudurluk_adi'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=9, value=kayit['direktorluk_adi'] or '-').alignment = cell_alignment
        
        # Lokasyon
        ws.cell(row=row_num, column=10, value=kayit['il_adi'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=11, value=kayit['ilce_adi'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=12, value=kayit['magaza_adi'] or '-').alignment = cell_alignment
        
        # Tarihler
        ws.cell(row=row_num, column=13, value=kayit['ise_baslama_tarihi'] or '-').alignment = center_alignment
        ws.cell(row=row_num, column=14, value=kayit['cikis_tarihi']).alignment = center_alignment
        
        # Ã‡Ä±kÄ±ÅŸ bilgileri
        ws.cell(row=row_num, column=15, value=kayit['cikis_nedeni']).alignment = cell_alignment
        ws.cell(row=row_num, column=16, value=kayit['liste_durumu']).alignment = center_alignment
        ws.cell(row=row_num, column=17, value='Evet' if kayit['tekrar_ise_alinabilir'] else 'HayÄ±r').alignment = center_alignment
        
        # Checklist
        ws.cell(row=row_num, column=18, value='âœ“' if kayit['zimmet_teslim'] else 'âœ—').alignment = center_alignment
        ws.cell(row=row_num, column=19, value='âœ“' if kayit['kiyafet_teslim'] else 'âœ—').alignment = center_alignment
        ws.cell(row=row_num, column=20, value='âœ“' if kayit['anahtar_teslim'] else 'âœ—').alignment = center_alignment
        ws.cell(row=row_num, column=21, value='âœ“' if kayit['kimlik_teslim'] else 'âœ—').alignment = center_alignment
        
        # Tazminatlar
        ws.cell(row=row_num, column=22, value=kayit['ihbar_tazminat_durumu'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=23, value=kayit['kidem_tazminat_durumu'] or '-').alignment = cell_alignment
        
        # Notlar
        ws.cell(row=row_num, column=24, value=kayit['yonetici_notu'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=25, value=kayit['ik_notu'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=26, value=kayit['genel_degerlendirme'] or '-').alignment = cell_alignment
        
        # Meta
        ws.cell(row=row_num, column=27, value=kayit['islem_yapan'] or '-').alignment = cell_alignment
        ws.cell(row=row_num, column=28, value=str(kayit['olusturma_tarihi']) if kayit['olusturma_tarihi'] else '-').alignment = center_alignment
        
        # Border ekle
        for col in range(1, 29):
            ws.cell(row=row_num, column=col).border = border
    
    # SÃ¼tun geniÅŸliklerini ayarla
    column_widths = {
        'A': 10, 'B': 25, 'C': 15, 'D': 15, 'E': 30,
        'F': 20, 'G': 20, 'H': 20, 'I': 20,
        'J': 15, 'K': 15, 'L': 25,
        'M': 15, 'N': 15, 'O': 20,
        'P': 15, 'Q': 20,
        'R': 12, 'S': 12, 'T': 12, 'U': 12,
        'V': 20, 'W': 20,
        'X': 40, 'Y': 40, 'Z': 40,
        'AA': 20, 'AB': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # SatÄ±r yÃ¼ksekliÄŸini ayarla
    ws.row_dimensions[1].height = 40  # BaÅŸlÄ±k
    for row in range(2, len(kayitlar) + 2):
        ws.row_dimensions[row].height = 30
    
    # Excel dosyasÄ±nÄ± memory'ye kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Dosya adÄ±
    from datetime import datetime
    filename = f"Cikis_Kayitlari_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )    


# =============================================================================
# FİLO YÖNETİMİ MODÜLÜ
# =============================================================================

# =============================================================================
# FİLO YÖNETİMİ MODÜLÜ - FLASK ROUTES
# =============================================================================
# Bu dosya app.py'ye eklenecek

# =============================================================================
# ARAÇLAR - TEMEL İŞLEMLER
# =============================================================================

@app.route('/filo/araclar')
@login_required
def filo_araclar():
    """Araç listesi"""
    conn = get_db()
    
    # Filtreler
    durum_filtre = request.args.get('durum', '')
    zimmet_filtre = request.args.get('zimmet', '')  # 'zimmetli', 'bosta'
    
    # Temel sorgu
    query = '''
        SELECT 
            a.*,
            z.calisan_id,
            c.ad_soyad as zimmetli_calisan,
            z.teslim_tarihi as zimmet_tarihi
        FROM araclar a
        LEFT JOIN (
            SELECT * FROM arac_zimmet WHERE aktif = TRUE
        ) z ON a.id = z.arac_id
        LEFT JOIN calisanlar c ON z.calisan_id = c.id
        WHERE 1=1
    '''
    
    params = []
    
    if durum_filtre:
        query += ' AND a.durum = %s'
        params.append(durum_filtre)
    
    if zimmet_filtre == 'zimmetli':
        query += ' AND z.id IS NOT NULL'
    elif zimmet_filtre == 'bosta':
        query += ' AND z.id IS NULL'
    
    query += ' ORDER BY a.plaka'
    
    araclar = conn.execute(query, params).fetchall()
    conn.close()
    
    return render_template('filo/araclar.html', araclar=araclar)


@app.route('/filo/arac/ekle', methods=['GET', 'POST'])
@login_required
def filo_arac_ekle():
    """Yeni araç ekle"""
    if request.method == 'POST':
        # Form verilerini al
        plaka = request.form['plaka'].upper().strip()
        marka = request.form['marka']
        model = request.form['model']
        yil = request.form.get('yil')
        renk = request.form.get('renk')
        
        # Teknik bilgiler
        sasi_no = request.form.get('sasi_no')
        motor_no = request.form.get('motor_no')
        yakit_tipi = request.form.get('yakit_tipi')
        
        # Sigorta bilgileri
        sigorta_sirket = request.form.get('sigorta_sirket')
        sigorta_police_no = request.form.get('sigorta_police_no')
        sigorta_bitis_tarihi = request.form.get('sigorta_bitis_tarihi') or None
        
        # Muayene
        muayene_tarihi = request.form.get('muayene_tarihi') or None
        sonraki_muayene_tarihi = request.form.get('sonraki_muayene_tarihi') or None
        
        # Kilometre
        baslangic_km = request.form.get('baslangic_km', 0)
        guncel_km = request.form.get('guncel_km', baslangic_km)
        
        # Notlar
        notlar = request.form.get('notlar')
        
        # Mülkiyet tipi
        mulkiyet_tipi = request.form.get('mulkiyet_tipi', 'Özmal')
        
        conn = get_db()
        
        # Plaka kontrolü
        mevcut = conn.execute('SELECT id FROM araclar WHERE plaka = %s', (plaka,)).fetchone()
        if mevcut:
            flash('Bu plaka ile kayıtlı bir araç zaten mevcut!', 'danger')
            conn.close()
            return redirect(url_for('filo_arac_ekle'))
        
        # Araç ekle
        conn.execute('''
            INSERT INTO araclar (
                plaka, marka, model, yil, renk,
                sasi_no, motor_no, yakit_tipi,
                sigorta_sirket, sigorta_police_no, sigorta_bitis_tarihi,
                muayene_tarihi, sonraki_muayene_tarihi,
                baslangic_km, guncel_km, mulkiyet_tipi, notlar
            ) VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s,
                %s, %s, %s, %s
            )
        ''', (plaka, marka, model, yil, renk,
              sasi_no, motor_no, yakit_tipi,
              sigorta_sirket, sigorta_police_no, sigorta_bitis_tarihi,
              muayene_tarihi, sonraki_muayene_tarihi,
              baslangic_km, guncel_km, mulkiyet_tipi, notlar))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'araclar', None, f'{plaka} plakalı araç eklendi', session.get('username'))
        flash(f'{plaka} plakalı araç başarıyla eklendi!', 'success')
        return redirect(url_for('filo_araclar'))
    
    return render_template('filo/arac_ekle.html')


@app.route('/filo/arac/<int:arac_id>')
@login_required
def filo_arac_detay(arac_id):
    """Araç detay sayfası"""
    conn = get_db()
    
    # Araç bilgileri
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    # Mevcut zimmet durumu
    zimmet = conn.execute('''
        SELECT z.*, c.ad_soyad, c.telefon
        FROM arac_zimmet z
        JOIN calisanlar c ON z.calisan_id = c.id
        WHERE z.arac_id = %s AND z.aktif = TRUE
    ''', (arac_id,)).fetchone()
    
    # Teslim/İade geçmişi
    teslim_iade_gecmis = conn.execute('''
        SELECT ti.*, c.ad_soyad, u.username as teslim_eden
        FROM arac_teslim_iade ti
        JOIN calisanlar c ON ti.calisan_id = c.id
        LEFT JOIN users u ON ti.teslim_eden_user_id = u.id
        WHERE ti.arac_id = %s
        ORDER BY ti.islem_tarihi DESC
        LIMIT 20
    ''', (arac_id,)).fetchall()
    
    # Bakım kayıtları
    bakim_kayitlari = conn.execute('''
        SELECT * FROM arac_bakim
        WHERE arac_id = %s
        ORDER BY bakim_tarihi DESC
        LIMIT 10
    ''', (arac_id,)).fetchall()
    
    # Yakıt kayıtları
    yakit_kayitlari = conn.execute('''
        SELECT y.*, c.ad_soyad
        FROM arac_yakit y
        LEFT JOIN calisanlar c ON y.calisan_id = c.id
        WHERE y.arac_id = %s
        ORDER BY y.yakit_tarihi DESC
        LIMIT 10
    ''', (arac_id,)).fetchall()
    
    # Sigorta/Kaza kayıtları
    sigorta_kaza = conn.execute('''
        SELECT * FROM arac_sigorta_kaza
        WHERE arac_id = %s
        ORDER BY kayit_tarihi DESC
        LIMIT 10
    ''', (arac_id,)).fetchall()
    
    # Kiralama bilgisi (eğer kiralık ise)
    kiralama = None
    if arac['mulkiyet_tipi'] == 'Kiralık':
        kiralama = conn.execute('''
            SELECT * FROM arac_kira_bilgileri
            WHERE arac_id = %s AND aktif = TRUE
        ''', (arac_id,)).fetchone()
    
    conn.close()
    
    return render_template('filo/arac_detay.html',
                         arac=dict(arac),
                         zimmet=dict(zimmet) if zimmet else None,
                         teslim_iade_gecmis=teslim_iade_gecmis,
                         bakim_kayitlari=bakim_kayitlari,
                         yakit_kayitlari=yakit_kayitlari,
                         sigorta_kaza=sigorta_kaza,
                         kiralama=dict(kiralama) if kiralama else None)


@app.route('/filo/arac/<int:arac_id>/duzenle', methods=['GET', 'POST'])
@login_required
def filo_arac_duzenle(arac_id):
    """Araç bilgilerini düzenle"""
    conn = get_db()
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    if request.method == 'POST':
        # Form verilerini al (arac_ekle ile benzer)
        plaka = request.form['plaka'].upper().strip()
        marka = request.form['marka']
        model = request.form['model']
        yil = request.form.get('yil')
        renk = request.form.get('renk')
        sasi_no = request.form.get('sasi_no')
        motor_no = request.form.get('motor_no')
        yakit_tipi = request.form.get('yakit_tipi')
        sigorta_sirket = request.form.get('sigorta_sirket')
        sigorta_police_no = request.form.get('sigorta_police_no')
        sigorta_bitis_tarihi = request.form.get('sigorta_bitis_tarihi') or None
        muayene_tarihi = request.form.get('muayene_tarihi') or None
        sonraki_muayene_tarihi = request.form.get('sonraki_muayene_tarihi') or None
        guncel_km = request.form.get('guncel_km')
        durum = request.form.get('durum')
        notlar = request.form.get('notlar')
        
        # Plaka kontrolü (kendi plakas hariç)
        mevcut = conn.execute('SELECT id FROM araclar WHERE plaka = %s AND id != %s', 
                            (plaka, arac_id)).fetchone()
        if mevcut:
            flash('Bu plaka ile kayıtlı başka bir araç mevcut!', 'danger')
            conn.close()
            return redirect(url_for('filo_arac_duzenle', arac_id=arac_id))
        
        # Güncelle
        conn.execute('''
            UPDATE araclar SET
                plaka = %s, marka = %s, model = %s, yil = %s, renk = %s,
                sasi_no = %s, motor_no = %s, yakit_tipi = %s,
                sigorta_sirket = %s, sigorta_police_no = %s, sigorta_bitis_tarihi = %s,
                muayene_tarihi = %s, sonraki_muayene_tarihi = %s,
                guncel_km = %s, durum = %s, notlar = %s,
                guncelleme_tarihi = CURRENT_TIMESTAMP
            WHERE id = %s
        ''', (plaka, marka, model, yil, renk,
              sasi_no, motor_no, yakit_tipi,
              sigorta_sirket, sigorta_police_no, sigorta_bitis_tarihi,
              muayene_tarihi, sonraki_muayene_tarihi,
              guncel_km, durum, notlar, arac_id))
        
        conn.commit()
        conn.close()
        
        log_islem('GÜNCELLEME', 'araclar', arac_id, f'{plaka} plakalı araç güncellendi', 
                 session.get('username'))
        flash('Araç bilgileri güncellendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    conn.close()
    return render_template('filo/arac_duzenle.html', arac=dict(arac))


# =============================================================================
# TESLİM / İADE İŞLEMLERİ
# =============================================================================

@app.route('/filo/teslim-iade', methods=['GET', 'POST'])
@login_required
def filo_teslim_iade():
    """Araç teslim/iade işlemi"""
    conn = get_db()
    
    if request.method == 'POST':
        islem_tipi = request.form['islem_tipi']  # Teslim veya İade
        arac_id = request.form['arac_id']
        calisan_id = request.form['calisan_id']
        islem_tarihi = request.form.get('islem_tarihi', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        kilometre = request.form['kilometre']
        yakit_durumu = request.form.get('yakit_durumu')
        hasar_var = request.form.get('hasar_var') == 'on'
        hasar_aciklama = request.form.get('hasar_aciklama')
        temizlik_durumu = request.form.get('temizlik_durumu')
        ic_durum_aciklama = request.form.get('ic_durum_aciklama')
        dis_durum_aciklama = request.form.get('dis_durum_aciklama')
        eksikler = request.form.get('eksikler')
        notlar = request.form.get('notlar')
        
        # Teslim/İade kaydını oluştur
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO arac_teslim_iade (
                arac_id, calisan_id, islem_tipi, islem_tarihi,
                kilometre, yakit_durumu, hasar_var, hasar_aciklama,
                temizlik_durumu, ic_durum_aciklama, dis_durum_aciklama,
                eksikler, teslim_eden_user_id, notlar
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s
            ) RETURNING id
        ''', (arac_id, calisan_id, islem_tipi, islem_tarihi,
              kilometre, yakit_durumu, hasar_var, hasar_aciklama,
              temizlik_durumu, ic_durum_aciklama, dis_durum_aciklama,
              eksikler, session['user_id'], notlar))
        
        teslim_iade_id = cursor.fetchone()[0]
        
        # Araç km güncelle
        conn.execute('UPDATE araclar SET guncel_km = %s WHERE id = %s', (kilometre, arac_id))
        
        if islem_tipi == 'Teslim':
            # Zimmet kaydı oluştur
            conn.execute('''
                INSERT INTO arac_zimmet (
                    arac_id, calisan_id, teslim_tarihi, teslim_km, teslim_kayit_id
                ) VALUES (%s, %s, %s, %s, %s)
            ''', (arac_id, calisan_id, islem_tarihi.split()[0], kilometre, teslim_iade_id))
            
            mesaj = 'Araç teslim edildi'
            
        else:  # İade
            # Mevcut zimmet kaydını kapat
            conn.execute('''
                UPDATE arac_zimmet SET
                    iade_tarihi = %s,
                    iade_km = %s,
                    iade_kayit_id = %s,
                    aktif = FALSE,
                    guncelleme_tarihi = CURRENT_TIMESTAMP
                WHERE arac_id = %s AND aktif = TRUE
            ''', (islem_tarihi.split()[0], kilometre, teslim_iade_id, arac_id))
            
            mesaj = 'Araç iade edildi'
        
        conn.commit()
        conn.close()
        
        log_islem(islem_tipi.upper(), 'arac_teslim_iade', teslim_iade_id, mesaj, session.get('username'))
        flash(f'{mesaj}!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    # GET request - Form göster
    # Boşta olan araçlar (teslim için)
    bosta_araclar = conn.execute('''
        SELECT a.id, a.plaka, a.marka, a.model
        FROM araclar a
        LEFT JOIN arac_zimmet z ON a.id = z.arac_id AND z.aktif = TRUE
        WHERE a.aktif = TRUE AND a.durum = 'Aktif' AND z.id IS NULL
        ORDER BY a.plaka
    ''').fetchall()
    
    # Zimmetli araçlar (iade için)
    zimmetli_araclar = conn.execute('''
        SELECT a.id, a.plaka, a.marka, a.model, c.id as calisan_id, c.ad_soyad
        FROM araclar a
        JOIN arac_zimmet z ON a.id = z.arac_id AND z.aktif = TRUE
        JOIN calisanlar c ON z.calisan_id = c.id
        WHERE a.aktif = TRUE
        ORDER BY a.plaka
    ''').fetchall()
    
    # Aktif çalışanlar
    calisanlar = conn.execute('''
        SELECT id, ad_soyad, telefon
        FROM calisanlar
        WHERE aktif = TRUE
        ORDER BY ad_soyad
    ''').fetchall()
    
    conn.close()
    
    return render_template('filo/teslim_iade.html',
                         bosta_araclar=bosta_araclar,
                         zimmetli_araclar=zimmetli_araclar,
                         calisanlar=calisanlar,
                         now=datetime.now())


# =============================================================================
# BAKIM YÖNETİMİ
# =============================================================================

@app.route('/filo/arac/<int:arac_id>/bakim/ekle', methods=['GET', 'POST'])
@login_required
def filo_bakim_ekle(arac_id):
    """Araç bakım kaydı ekle"""
    conn = get_db()
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    if request.method == 'POST':
        bakim_tipi = request.form['bakim_tipi']
        bakim_tarihi = request.form['bakim_tarihi']
        bakim_km = request.form.get('bakim_km')
        aciklama = request.form['aciklama']
        yapilan_islemler = request.form.get('yapilan_islemler')
        tutar = request.form.get('tutar')
        servis_adi = request.form.get('servis_adi')
        servis_telefon = request.form.get('servis_telefon')
        sonraki_bakim_km = request.form.get('sonraki_bakim_km')
        sonraki_bakim_tarihi = request.form.get('sonraki_bakim_tarihi') or None
        
        conn.execute('''
            INSERT INTO arac_bakim (
                arac_id, bakim_tipi, bakim_tarihi, bakim_km,
                aciklama, yapilan_islemler, tutar,
                servis_adi, servis_telefon,
                sonraki_bakim_km, sonraki_bakim_tarihi
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (arac_id, bakim_tipi, bakim_tarihi, bakim_km,
              aciklama, yapilan_islemler, tutar,
              servis_adi, servis_telefon,
              sonraki_bakim_km, sonraki_bakim_tarihi))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'arac_bakim', arac_id, f'{bakim_tipi} bakımı eklendi', session.get('username'))
        flash('Bakım kaydı eklendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    conn.close()
    return render_template('filo/bakim_ekle.html', arac=dict(arac))


# =============================================================================
# YAKIT YÖNETİMİ
# =============================================================================

@app.route('/filo/arac/<int:arac_id>/yakit/ekle', methods=['GET', 'POST'])
@login_required
def filo_yakit_ekle(arac_id):
    """Yakıt kaydı ekle"""
    conn = get_db()
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    # Araç zimmetli mi kontrol et
    zimmet = conn.execute('''
        SELECT calisan_id FROM arac_zimmet
        WHERE arac_id = %s AND aktif = TRUE
    ''', (arac_id,)).fetchone()
    
    if request.method == 'POST':
        calisan_id = zimmet['calisan_id'] if zimmet else None
        yakit_tarihi = request.form['yakit_tarihi']
        kilometre = request.form['kilometre']
        litre = request.form['litre']
        birim_fiyat = request.form.get('birim_fiyat')
        toplam_tutar = request.form.get('toplam_tutar')
        yakit_tipi = request.form.get('yakit_tipi')
        fatura_no = request.form.get('fatura_no')
        istasyon = request.form.get('istasyon')
        notlar = request.form.get('notlar')
        
        conn.execute('''
            INSERT INTO arac_yakit (
                arac_id, calisan_id, yakit_tarihi, kilometre,
                litre, birim_fiyat, toplam_tutar, yakit_tipi,
                fatura_no, istasyon, notlar
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (arac_id, calisan_id, yakit_tarihi, kilometre,
              litre, birim_fiyat, toplam_tutar, yakit_tipi,
              fatura_no, istasyon, notlar))
        
        # Araç km güncelle
        conn.execute('UPDATE araclar SET guncel_km = %s WHERE id = %s', (kilometre, arac_id))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'arac_yakit', arac_id, f'Yakıt kaydı eklendi: {litre} lt', session.get('username'))
        flash('Yakıt kaydı eklendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    conn.close()
    return render_template('filo/yakit_ekle.html', arac=dict(arac), zimmet=zimmet)


# =============================================================================
# SİGORTA/KAZA YÖNETİMİ
# =============================================================================

@app.route('/filo/arac/<int:arac_id>/sigorta-kaza/ekle', methods=['GET', 'POST'])
@login_required
def filo_sigorta_kaza_ekle(arac_id):
    """Sigorta/kaza kaydı ekle"""
    conn = get_db()
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    if request.method == 'POST':
        kayit_tipi = request.form['kayit_tipi']
        kayit_tarihi = request.form['kayit_tarihi']
        notlar = request.form.get('notlar')
        
        if kayit_tipi == 'Kaza':
            kaza_aciklama = request.form.get('kaza_aciklama')
            kusur_durumu = request.form.get('kusur_durumu')
            hasar_tutari = request.form.get('hasar_tutari')
            
            conn.execute('''
                INSERT INTO arac_sigorta_kaza (
                    arac_id, kayit_tipi, kayit_tarihi,
                    kaza_aciklama, kusur_durumu, hasar_tutari, notlar
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (arac_id, kayit_tipi, kayit_tarihi,
                  kaza_aciklama, kusur_durumu, hasar_tutari, notlar))
        else:  # Sigorta
            sigorta_sirket = request.form.get('sigorta_sirket')
            police_no = request.form.get('police_no')
            baslangic_tarihi = request.form.get('baslangic_tarihi')
            bitis_tarihi = request.form.get('bitis_tarihi')
            prim_tutari = request.form.get('prim_tutari')
            
            conn.execute('''
                INSERT INTO arac_sigorta_kaza (
                    arac_id, kayit_tipi, kayit_tarihi,
                    sigorta_sirket, police_no, baslangic_tarihi,
                    bitis_tarihi, prim_tutari, notlar
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (arac_id, kayit_tipi, kayit_tarihi,
                  sigorta_sirket, police_no, baslangic_tarihi,
                  bitis_tarihi, prim_tutari, notlar))
            
            # Araç sigorta bilgilerini güncelle
            conn.execute('''
                UPDATE araclar SET
                    sigorta_sirket = %s,
                    sigorta_police_no = %s,
                    sigorta_bitis_tarihi = %s
                WHERE id = %s
            ''', (sigorta_sirket, police_no, bitis_tarihi, arac_id))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'arac_sigorta_kaza', arac_id, f'{kayit_tipi} kaydı eklendi', session.get('username'))
        flash(f'{kayit_tipi} kaydı eklendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    conn.close()
    return render_template('filo/sigorta_kaza_ekle.html', arac=dict(arac))


# =============================================================================
# RAPORLAR
# =============================================================================

@app.route('/filo/raporlar')
@login_required
def filo_raporlar():
    """Filo raporları"""
    conn = get_db()
    
    # Genel özet
    ozet = conn.execute('''
        SELECT 
            COUNT(*) as toplam_arac,
            COUNT(CASE WHEN durum = 'Aktif' THEN 1 END) as aktif_arac,
            COUNT(CASE WHEN durum = 'Bakımda' THEN 1 END) as bakimda_arac,
            COUNT(CASE WHEN durum = 'Hasarlı' THEN 1 END) as hasarli_arac
        FROM araclar
        WHERE aktif = TRUE
    ''').fetchone()
    
    # Zimmet durumu
    zimmet_durum = conn.execute('''
        SELECT 
            COUNT(DISTINCT a.id) as toplam_arac,
            COUNT(DISTINCT z.arac_id) as zimmetli_arac,
            COUNT(DISTINCT a.id) - COUNT(DISTINCT z.arac_id) as bosta_arac
        FROM araclar a
        LEFT JOIN arac_zimmet z ON a.id = z.arac_id AND z.aktif = TRUE
        WHERE a.aktif = TRUE AND a.durum = 'Aktif'
    ''').fetchone()
    
    # Yakında dolacak sigortalar (30 gün içinde)
    yakin_sigortalar = conn.execute('''
        SELECT plaka, marka, model, sigorta_bitis_tarihi
        FROM araclar
        WHERE sigorta_bitis_tarihi IS NOT NULL
          AND sigorta_bitis_tarihi <= CURRENT_DATE + INTERVAL '30 days'
          AND sigorta_bitis_tarihi >= CURRENT_DATE
          AND aktif = TRUE
        ORDER BY sigorta_bitis_tarihi
    ''').fetchall()
    
    # Yakında gelecek muayeneler (30 gün içinde)
    yakin_muayeneler = conn.execute('''
        SELECT plaka, marka, model, sonraki_muayene_tarihi
        FROM araclar
        WHERE sonraki_muayene_tarihi IS NOT NULL
          AND sonraki_muayene_tarihi <= CURRENT_DATE + INTERVAL '30 days'
          AND sonraki_muayene_tarihi >= CURRENT_DATE
          AND aktif = TRUE
        ORDER BY sonraki_muayene_tarihi
    ''').fetchall()
    
    # Aylık yakıt tüketimi (son 6 ay)
    aylik_yakit = conn.execute('''
        SELECT 
            TO_CHAR(yakit_tarihi, 'YYYY-MM') as ay,
            SUM(litre) as toplam_litre,
            SUM(toplam_tutar) as toplam_tutar,
            COUNT(*) as islem_sayisi
        FROM arac_yakit
        WHERE yakit_tarihi >= CURRENT_DATE - INTERVAL '6 months'
        GROUP BY TO_CHAR(yakit_tarihi, 'YYYY-MM')
        ORDER BY ay DESC
    ''').fetchall()
    
    # Aylık bakım maliyeti (son 6 ay)
    aylik_bakim = conn.execute('''
        SELECT 
            TO_CHAR(bakim_tarihi, 'YYYY-MM') as ay,
            SUM(tutar) as toplam_tutar,
            COUNT(*) as islem_sayisi
        FROM arac_bakim
        WHERE bakim_tarihi >= CURRENT_DATE - INTERVAL '6 months'
          AND tutar IS NOT NULL
        GROUP BY TO_CHAR(bakim_tarihi, 'YYYY-MM')
        ORDER BY ay DESC
    ''').fetchall()
    
    conn.close()
    
    return render_template('filo/raporlar.html',
                         ozet=ozet,
                         zimmet_durum=zimmet_durum,
                         yakin_sigortalar=yakin_sigortalar,
                         yakin_muayeneler=yakin_muayeneler,
                         aylik_yakit=aylik_yakit,
                         aylik_bakim=aylik_bakim)

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5001)# =============================================================================
# FİLO YÖNETİMİ - KİRALAMA ÖZELLİKLERİ
# =============================================================================
# Bu route'lar mevcut filo_routes.py dosyasına eklenecek

# =============================================================================
# KİRALAMA YÖNETİMİ
# =============================================================================

@app.route('/filo/kiralamalar')
@login_required
def filo_kiralamalar():
    """Kiralık araçlar listesi"""
    conn = get_db()
    
    # Kiralık araçlar ve sözleşme bilgileri
    kiralamalar = conn.execute('''
        SELECT 
            a.id as arac_id,
            a.plaka,
            a.marka,
            a.model,
            k.*,
            CASE 
                WHEN k.sozlesme_bitis < CURRENT_DATE THEN 'Süresi Dolmuş'
                WHEN k.sozlesme_bitis <= CURRENT_DATE + INTERVAL '30 days' THEN 'Yakında Bitiyor'
                ELSE 'Aktif'
            END as sozlesme_durumu,
            (k.sozlesme_bitis - CURRENT_DATE) as kalan_gun
        FROM araclar a
        JOIN arac_kira_bilgileri k ON a.id = k.arac_id
        WHERE a.mulkiyet_tipi = 'Kiralık' AND k.aktif = TRUE
        ORDER BY k.sozlesme_bitis
    ''').fetchall()
    
    # Rol kontrolü - Finans bilgilerini gösterebilir mi?
    user_role = session.get('role', '')
    can_view_finance = user_role in ['admin', 'finans']
    
    conn.close()
    
    return render_template('filo/kiralamalar.html', 
                         kiralamalar=kiralamalar,
                         can_view_finance=can_view_finance)


@app.route('/filo/arac/<int:arac_id>/kiralama/ekle', methods=['GET', 'POST'])
@login_required
def filo_kiralama_ekle(arac_id):
    """Araç kiralama bilgisi ekle"""
    conn = get_db()
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    if request.method == 'POST':
        # Form verilerini al
        kiralama_sirket = request.form['kiralama_sirket']
        sirket_telefon = request.form.get('sirket_telefon')
        sirket_email = request.form.get('sirket_email')
        yetkili_kisi = request.form.get('yetkili_kisi')
        sozlesme_no = request.form.get('sozlesme_no')
        sozlesme_baslangic = request.form['sozlesme_baslangic']
        sozlesme_bitis = request.form['sozlesme_bitis']
        aylik_kira_bedeli = request.form.get('aylik_kira_bedeli')
        para_birimi = request.form.get('para_birimi', 'TRY')
        odeme_donemi = request.form.get('odeme_donemi')
        odeme_gunu = request.form.get('odeme_gunu')
        km_limiti = request.form.get('km_limiti')
        km_asim_ucreti = request.form.get('km_asim_ucreti')
        sigorta_dahil = request.form.get('sigorta_dahil') == 'on'
        bakim_dahil = request.form.get('bakim_dahil') == 'on'
        notlar = request.form.get('notlar')
        
        # Sözleşme dosyası upload (şimdilik yol sakla)
        sozlesme_dosya_yolu = None
        if 'sozlesme_dosya' in request.files:
            file = request.files['sozlesme_dosya']
            if file and file.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                # Dosyayı uploads/sozlesmeler/ klasörüne kaydet
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'sozlesmeler')
                os.makedirs(upload_path, exist_ok=True)
                file_path = os.path.join(upload_path, f"{arac_id}_{filename}")
                file.save(file_path)
                sozlesme_dosya_yolu = file_path
        
        # Kiralama kaydı oluştur
        conn.execute('''
            INSERT INTO arac_kira_bilgileri (
                arac_id, kiralama_sirket, sirket_telefon, sirket_email,
                yetkili_kisi, sozlesme_no, sozlesme_baslangic, sozlesme_bitis,
                aylik_kira_bedeli, para_birimi, odeme_donemi, odeme_gunu,
                sozlesme_dosya_yolu, km_limiti, km_asim_ucreti,
                sigorta_dahil, bakim_dahil, notlar
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
        ''', (arac_id, kiralama_sirket, sirket_telefon, sirket_email,
              yetkili_kisi, sozlesme_no, sozlesme_baslangic, sozlesme_bitis,
              aylik_kira_bedeli, para_birimi, odeme_donemi, odeme_gunu,
              sozlesme_dosya_yolu, km_limiti, km_asim_ucreti,
              sigorta_dahil, bakim_dahil, notlar))
        
        # Aracın mülkiyet tipini "Kiralık" yap
        conn.execute('UPDATE araclar SET mulkiyet_tipi = %s WHERE id = %s', ('Kiralık', arac_id))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'arac_kira_bilgileri', arac_id, 
                 f'{arac["plaka"]} kiralama bilgisi eklendi', session.get('username'))
        flash('Kiralama bilgisi eklendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=arac_id))
    
    conn.close()
    return render_template('filo/kiralama_ekle.html', arac=dict(arac))


@app.route('/filo/kiralama/<int:kiralama_id>/duzenle', methods=['GET', 'POST'])
@login_required
def filo_kiralama_duzenle(kiralama_id):
    """Kiralama bilgisini düzenle"""
    conn = get_db()
    
    kiralama = conn.execute('''
        SELECT k.*, a.plaka, a.marka, a.model
        FROM arac_kira_bilgileri k
        JOIN araclar a ON k.arac_id = a.id
        WHERE k.id = %s
    ''', (kiralama_id,)).fetchone()
    
    if not kiralama:
        flash('Kiralama bilgisi bulunamadı!', 'danger')
        return redirect(url_for('filo_kiralamalar'))
    
    if request.method == 'POST':
        # Form verilerini al (kiralama_ekle ile aynı)
        kiralama_sirket = request.form['kiralama_sirket']
        sirket_telefon = request.form.get('sirket_telefon')
        sirket_email = request.form.get('sirket_email')
        yetkili_kisi = request.form.get('yetkili_kisi')
        sozlesme_no = request.form.get('sozlesme_no')
        sozlesme_baslangic = request.form['sozlesme_baslangic']
        sozlesme_bitis = request.form['sozlesme_bitis']
        aylik_kira_bedeli = request.form.get('aylik_kira_bedeli')
        para_birimi = request.form.get('para_birimi', 'TRY')
        odeme_donemi = request.form.get('odeme_donemi')
        odeme_gunu = request.form.get('odeme_gunu')
        km_limiti = request.form.get('km_limiti')
        km_asim_ucreti = request.form.get('km_asim_ucreti')
        sigorta_dahil = request.form.get('sigorta_dahil') == 'on'
        bakim_dahil = request.form.get('bakim_dahil') == 'on'
        notlar = request.form.get('notlar')
        
        # Dosya upload varsa güncelle
        sozlesme_dosya_yolu = kiralama['sozlesme_dosya_yolu']
        if 'sozlesme_dosya' in request.files:
            file = request.files['sozlesme_dosya']
            if file and file.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'sozlesmeler')
                os.makedirs(upload_path, exist_ok=True)
                file_path = os.path.join(upload_path, f"{kiralama['arac_id']}_{filename}")
                file.save(file_path)
                sozlesme_dosya_yolu = file_path
        
        # Güncelle
        conn.execute('''
            UPDATE arac_kira_bilgileri SET
                kiralama_sirket = %s,
                sirket_telefon = %s,
                sirket_email = %s,
                yetkili_kisi = %s,
                sozlesme_no = %s,
                sozlesme_baslangic = %s,
                sozlesme_bitis = %s,
                aylik_kira_bedeli = %s,
                para_birimi = %s,
                odeme_donemi = %s,
                odeme_gunu = %s,
                sozlesme_dosya_yolu = %s,
                km_limiti = %s,
                km_asim_ucreti = %s,
                sigorta_dahil = %s,
                bakim_dahil = %s,
                notlar = %s,
                guncelleme_tarihi = CURRENT_TIMESTAMP
            WHERE id = %s
        ''', (kiralama_sirket, sirket_telefon, sirket_email,
              yetkili_kisi, sozlesme_no, sozlesme_baslangic, sozlesme_bitis,
              aylik_kira_bedeli, para_birimi, odeme_donemi, odeme_gunu,
              sozlesme_dosya_yolu, km_limiti, km_asim_ucreti,
              sigorta_dahil, bakim_dahil, notlar, kiralama_id))
        
        conn.commit()
        conn.close()
        
        log_islem('GÜNCELLEME', 'arac_kira_bilgileri', kiralama_id, 
                 'Kiralama bilgisi güncellendi', session.get('username'))
        flash('Kiralama bilgisi güncellendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=kiralama['arac_id']))
    
    conn.close()
    return render_template('filo/kiralama_duzenle.html', kiralama=dict(kiralama))


@app.route('/filo/kiralama/<int:kiralama_id>/odeme/ekle', methods=['GET', 'POST'])
@login_required
def filo_kiralama_odeme_ekle(kiralama_id):
    """Kiralama ödeme kaydı ekle"""
    conn = get_db()
    
    kiralama = conn.execute('''
        SELECT k.*, a.plaka, a.marka, a.model
        FROM arac_kira_bilgileri k
        JOIN araclar a ON k.arac_id = a.id
        WHERE k.id = %s
    ''', (kiralama_id,)).fetchone()
    
    if not kiralama:
        flash('Kiralama bilgisi bulunamadı!', 'danger')
        return redirect(url_for('filo_kiralamalar'))
    
    if request.method == 'POST':
        odeme_donemi = request.form['odeme_donemi']
        tutar = request.form['tutar']
        para_birimi = request.form.get('para_birimi', 'TRY')
        km_asim_tutari = request.form.get('km_asim_tutari', 0)
        toplam_tutar = float(tutar) + float(km_asim_tutari or 0)
        odeme_durumu = request.form.get('odeme_durumu', 'Bekliyor')
        planlanan_odeme_tarihi = request.form.get('planlanan_odeme_tarihi')
        gerceklesen_odeme_tarihi = request.form.get('gerceklesen_odeme_tarihi')
        fatura_no = request.form.get('fatura_no')
        notlar = request.form.get('notlar')
        
        conn.execute('''
            INSERT INTO arac_kira_odemeler (
                kira_bilgi_id, arac_id, odeme_donemi, tutar, para_birimi,
                km_asim_tutari, toplam_tutar, odeme_durumu,
                planlanan_odeme_tarihi, gerceklesen_odeme_tarihi,
                fatura_no, notlar
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (kiralama_id, kiralama['arac_id'], odeme_donemi, tutar, para_birimi,
              km_asim_tutari, toplam_tutar, odeme_durumu,
              planlanan_odeme_tarihi, gerceklesen_odeme_tarihi,
              fatura_no, notlar))
        
        conn.commit()
        conn.close()
        
        log_islem('EKLEME', 'arac_kira_odemeler', kiralama_id, 
                 f'{odeme_donemi} dönemi ödeme kaydı', session.get('username'))
        flash('Ödeme kaydı eklendi!', 'success')
        return redirect(url_for('filo_arac_detay', arac_id=kiralama['arac_id']))
    
    conn.close()
    return render_template('filo/odeme_ekle.html', kiralama=dict(kiralama))


@app.route('/filo/kiralama/odemeler')
@login_required
def filo_kiralama_odemeler():
    """Tüm kiralama ödemeleri listesi"""
    # Sadece admin ve finans görebilir
    if session.get('role') not in ['admin', 'finans']:
        flash('Bu sayfaya erişim yetkiniz yok!', 'danger')
        return redirect(url_for('filo_kiralamalar'))
    
    conn = get_db()
    
    # Ödeme listesi
    odemeler = conn.execute('''
        SELECT 
            o.*,
            a.plaka,
            a.marka,
            a.model,
            k.kiralama_sirket
        FROM arac_kira_odemeler o
        JOIN araclar a ON o.arac_id = a.id
        JOIN arac_kira_bilgileri k ON o.kira_bilgi_id = k.id
        ORDER BY o.odeme_donemi DESC, o.planlanan_odeme_tarihi DESC
    ''').fetchall()
    
    # Özet istatistikler
    ozet = conn.execute('''
        SELECT 
            COUNT(*) as toplam_odeme,
            COUNT(CASE WHEN odeme_durumu = 'Bekliyor' THEN 1 END) as bekleyen,
            COUNT(CASE WHEN odeme_durumu = 'Ödendi' THEN 1 END) as odenen,
            COUNT(CASE WHEN odeme_durumu = 'Gecikti' THEN 1 END) as geciken,
            SUM(CASE WHEN odeme_durumu = 'Bekliyor' THEN toplam_tutar ELSE 0 END) as bekleyen_tutar,
            SUM(CASE WHEN odeme_durumu = 'Ödendi' THEN toplam_tutar ELSE 0 END) as odenen_tutar
        FROM arac_kira_odemeler
    ''').fetchone()
    
    conn.close()
    
    return render_template('filo/kiralama_odemeler.html', 
                         odemeler=odemeler,
                         ozet=ozet)


# ARAÇ EKLE/DÜZENLE route'larını güncelle - mulkiyet_tipi ekle
# Bu kısım mevcut arac_ekle ve arac_duzenle fonksiyonlarına eklenecek

# Araç ekle formunda:
# mulkiyet_tipi = request.form.get('mulkiyet_tipi', 'Özmal')
# INSERT sorgusuna ekle

# Araç düzenle formunda:
# mulkiyet_tipi = request.form.get('mulkiyet_tipi')
# UPDATE sorgusuna ekle
# =============================================================================
# TRAFİK CEZASI YÖNETİMİ
# =============================================================================

@app.route('/filo/trafik-cezalari')
@login_required
def filo_trafik_cezalari():
    """Tüm trafik cezaları listesi"""
    conn = get_db()
    
    # Filtreler
    durum_filtre = request.args.get('durum')
    sorumlu_filtre = request.args.get('sorumlu')
    
    query = '''
        SELECT 
            tc.*,
            a.plaka,
            a.marka,
            a.model,
            c.ad_soyad as calisan_adi
        FROM arac_trafik_cezalari tc
        JOIN araclar a ON tc.arac_id = a.id
        LEFT JOIN calisanlar c ON tc.calisan_id = c.id
        WHERE 1=1
    '''
    
    params = []
    
    if durum_filtre:
        query += ' AND tc.odeme_durumu = %s'
        params.append(durum_filtre)
    
    if sorumlu_filtre:
        query += ' AND tc.sorumlu = %s'
        params.append(sorumlu_filtre)
    
    query += ' ORDER BY tc.ceza_tarihi DESC'
    
    cezalar = conn.execute(query, params).fetchall()
    
    # Özet istatistikler
    ozet = conn.execute('''
        SELECT 
            COUNT(*) as toplam_ceza,
            COUNT(CASE WHEN odeme_durumu = 'Ödenmedi' THEN 1 END) as odenmemis,
            COUNT(CASE WHEN odeme_durumu = 'Ödendi' THEN 1 END) as odenen,
            COUNT(CASE WHEN odeme_durumu = 'İtiraz Edildi' THEN 1 END) as itiraz,
            SUM(ceza_tutari) as toplam_tutar,
            SUM(CASE WHEN odeme_durumu = 'Ödenmedi' THEN ceza_tutari ELSE 0 END) as odenmemis_tutar,
            SUM(CASE WHEN sorumlu = 'Çalışan' AND calisan_odemesi_kesildi = TRUE THEN kesinti_tutari ELSE 0 END) as kesilen_tutar
        FROM arac_trafik_cezalari
    ''').fetchone()
    
    conn.close()
    
    return render_template('filo/trafik_cezalari.html', 
                         cezalar=cezalar,
                         ozet=ozet)


@app.route('/filo/trafik-cezasi/ekle', methods=['GET', 'POST'])
@login_required
def filo_trafik_cezasi_ekle():
    """Yeni trafik cezası ekle"""
    conn = get_db()
    
    if request.method == 'POST':
        # Form verilerini al
        arac_id = request.form['arac_id']
        calisan_id = request.form.get('calisan_id')
        ceza_tarihi = request.form['ceza_tarihi']
        ceza_saati = request.form.get('ceza_saati')
        ceza_yeri = request.form.get('ceza_yeri')
        ceza_turu = request.form['ceza_turu']
        ceza_aciklama = request.form.get('ceza_aciklama')
        ceza_no = request.form.get('ceza_no')
        teblig_tarihi = request.form.get('teblig_tarihi')
        ceza_tutari = request.form['ceza_tutari']
        indirimli_tutar = request.form.get('indirimli_tutar')
        para_birimi = request.form.get('para_birimi', 'TRY')
        sorumlu = request.form.get('sorumlu', 'Çalışan')
        notlar = request.form.get('notlar')
        
        # Belge upload
        ceza_belgesi_yolu = None
        if 'ceza_belgesi' in request.files:
            file = request.files['ceza_belgesi']
            if file and file.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ceza_belgeleri')
                os.makedirs(upload_path, exist_ok=True)
                file_path = os.path.join(upload_path, f"{arac_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                file.save(file_path)
                ceza_belgesi_yolu = file_path
        
        # Ceza kaydı oluştur
        conn.execute('''
            INSERT INTO arac_trafik_cezalari (
                arac_id, calisan_id, ceza_tarihi, ceza_saati, ceza_yeri,
                ceza_turu, ceza_aciklama, ceza_no, teblig_tarihi,
                ceza_tutari, indirimli_tutar, para_birimi,
                sorumlu, ceza_belgesi_yolu, notlar,
                olusturan_user_id
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
        ''', (arac_id, calisan_id, ceza_tarihi, ceza_saati, ceza_yeri,
              ceza_turu, ceza_aciklama, ceza_no, teblig_tarihi,
              ceza_tutari, indirimli_tutar, para_birimi,
              sorumlu, ceza_belgesi_yolu, notlar,
              session.get('user_id')))
        
        conn.commit()
        
        # Araç bilgisi
        arac = conn.execute('SELECT plaka FROM araclar WHERE id = %s', (arac_id,)).fetchone()
        conn.close()
        
        log_islem('EKLEME', 'arac_trafik_cezalari', None, 
                 f'{arac["plaka"]} - {ceza_turu} cezası eklendi', session.get('username'))
        flash(f'Trafik cezası kaydedildi!', 'success')
        return redirect(url_for('filo_trafik_cezalari'))
    
    # GET isteği - Form göster
    araclar = conn.execute('''
        SELECT id, plaka, marka, model
        FROM araclar
        WHERE aktif = TRUE
        ORDER BY plaka
    ''').fetchall()
    
    calisanlar = conn.execute('''
        SELECT id, ad_soyad, telefon
        FROM calisanlar
        WHERE aktif = TRUE
        ORDER BY ad_soyad
    ''').fetchall()
    
    # Ceza türleri
    ceza_turleri = conn.execute('''
        SELECT ceza_turu, varsayilan_tutar
        FROM trafik_ceza_turleri
        WHERE aktif = TRUE
        ORDER BY ceza_turu
    ''').fetchall()
    
    conn.close()
    
    return render_template('filo/trafik_cezasi_ekle.html',
                         araclar=araclar,
                         calisanlar=calisanlar,
                         ceza_turleri=ceza_turleri)


@app.route('/filo/trafik-cezasi/<int:ceza_id>/duzenle', methods=['GET', 'POST'])
@login_required
def filo_trafik_cezasi_duzenle(ceza_id):
    """Trafik cezası düzenle"""
    conn = get_db()
    
    ceza = conn.execute('''
        SELECT tc.*, a.plaka
        FROM arac_trafik_cezalari tc
        JOIN araclar a ON tc.arac_id = a.id
        WHERE tc.id = %s
    ''', (ceza_id,)).fetchone()
    
    if not ceza:
        flash('Ceza kaydı bulunamadı!', 'danger')
        return redirect(url_for('filo_trafik_cezalari'))
    
    if request.method == 'POST':
        # Form verilerini al ve güncelle
        calisan_id = request.form.get('calisan_id')
        ceza_tarihi = request.form['ceza_tarihi']
        ceza_saati = request.form.get('ceza_saati')
        ceza_yeri = request.form.get('ceza_yeri')
        ceza_turu = request.form['ceza_turu']
        ceza_aciklama = request.form.get('ceza_aciklama')
        ceza_no = request.form.get('ceza_no')
        teblig_tarihi = request.form.get('teblig_tarihi')
        ceza_tutari = request.form['ceza_tutari']
        indirimli_tutar = request.form.get('indirimli_tutar')
        sorumlu = request.form.get('sorumlu')
        odeme_durumu = request.form.get('odeme_durumu')
        odeme_tarihi = request.form.get('odeme_tarihi')
        odeme_sekli = request.form.get('odeme_sekli')
        notlar = request.form.get('notlar')
        
        # Belge güncelleme
        ceza_belgesi_yolu = ceza['ceza_belgesi_yolu']
        if 'ceza_belgesi' in request.files:
            file = request.files['ceza_belgesi']
            if file and file.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ceza_belgeleri')
                os.makedirs(upload_path, exist_ok=True)
                file_path = os.path.join(upload_path, f"{ceza['arac_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                file.save(file_path)
                ceza_belgesi_yolu = file_path
        
        odeme_belgesi_yolu = ceza['odeme_belgesi_yolu']
        if 'odeme_belgesi' in request.files:
            file = request.files['odeme_belgesi']
            if file and file.filename:
                from werkzeug.utils import secure_filename
                filename = secure_filename(file.filename)
                upload_path = os.path.join(app.config['UPLOAD_FOLDER'], 'ceza_belgeleri')
                os.makedirs(upload_path, exist_ok=True)
                file_path = os.path.join(upload_path, f"odeme_{ceza['arac_id']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")
                file.save(file_path)
                odeme_belgesi_yolu = file_path
        
        # Güncelle
        conn.execute('''
            UPDATE arac_trafik_cezalari SET
                calisan_id = %s,
                ceza_tarihi = %s,
                ceza_saati = %s,
                ceza_yeri = %s,
                ceza_turu = %s,
                ceza_aciklama = %s,
                ceza_no = %s,
                teblig_tarihi = %s,
                ceza_tutari = %s,
                indirimli_tutar = %s,
                sorumlu = %s,
                odeme_durumu = %s,
                odeme_tarihi = %s,
                odeme_sekli = %s,
                ceza_belgesi_yolu = %s,
                odeme_belgesi_yolu = %s,
                notlar = %s
            WHERE id = %s
        ''', (calisan_id, ceza_tarihi, ceza_saati, ceza_yeri, ceza_turu,
              ceza_aciklama, ceza_no, teblig_tarihi, ceza_tutari, indirimli_tutar,
              sorumlu, odeme_durumu, odeme_tarihi, odeme_sekli,
              ceza_belgesi_yolu, odeme_belgesi_yolu, notlar, ceza_id))
        
        conn.commit()
        conn.close()
        
        log_islem('GÜNCELLEME', 'arac_trafik_cezalari', ceza_id, 
                 f'{ceza["plaka"]} ceza kaydı güncellendi', session.get('username'))
        flash('Ceza kaydı güncellendi!', 'success')
        return redirect(url_for('filo_trafik_cezalari'))
    
    # GET - Form göster
    calisanlar = conn.execute('''
        SELECT id, ad_soyad FROM calisanlar WHERE aktif = TRUE ORDER BY ad_soyad
    ''').fetchall()
    
    ceza_turleri = conn.execute('''
        SELECT ceza_turu FROM trafik_ceza_turleri WHERE aktif = TRUE ORDER BY ceza_turu
    ''').fetchall()
    
    conn.close()
    
    return render_template('filo/trafik_cezasi_duzenle.html',
                         ceza=dict(ceza),
                         calisanlar=calisanlar,
                         ceza_turleri=ceza_turleri)


@app.route('/filo/trafik-cezasi/<int:ceza_id>/odeme', methods=['POST'])
@login_required
def filo_trafik_cezasi_odeme(ceza_id):
    """Ceza ödeme kaydı"""
    conn = get_db()
    
    odeme_tarihi = request.form['odeme_tarihi']
    odeme_sekli = request.form['odeme_sekli']
    
    conn.execute('''
        UPDATE arac_trafik_cezalari SET
            odeme_durumu = 'Ödendi',
            odeme_tarihi = %s,
            odeme_sekli = %s
        WHERE id = %s
    ''', (odeme_tarihi, odeme_sekli, ceza_id))
    
    conn.commit()
    conn.close()
    
    flash('Ödeme kaydedildi!', 'success')
    return redirect(url_for('filo_trafik_cezalari'))


@app.route('/filo/trafik-cezasi/<int:ceza_id>/kesinti', methods=['POST'])
@login_required
def filo_trafik_cezasi_kesinti(ceza_id):
    """Çalışandan kesinti yap"""
    conn = get_db()
    
    kesinti_tutari = request.form['kesinti_tutari']
    kesinti_tarihi = request.form['kesinti_tarihi']
    
    conn.execute('''
        UPDATE arac_trafik_cezalari SET
            calisan_odemesi_kesildi = TRUE,
            kesinti_tarihi = %s,
            kesinti_tutari = %s
        WHERE id = %s
    ''', (kesinti_tarihi, kesinti_tutari, ceza_id))
    
    conn.commit()
    conn.close()
    
    flash('Kesinti kaydedildi!', 'success')
    return redirect(url_for('filo_trafik_cezalari'))


@app.route('/filo/trafik-cezalari/raporlar')
@login_required
def filo_trafik_cezalari_raporlar():
    """Trafik cezaları raporları"""
    conn = get_db()
    
    # Araç bazlı ceza sayısı
    arac_bazli = conn.execute('''
        SELECT 
            a.plaka,
            a.marka,
            a.model,
            COUNT(tc.id) as ceza_sayisi,
            SUM(tc.ceza_tutari) as toplam_tutar
        FROM araclar a
        LEFT JOIN arac_trafik_cezalari tc ON a.id = tc.arac_id
        GROUP BY a.id, a.plaka, a.marka, a.model
        HAVING COUNT(tc.id) > 0
        ORDER BY ceza_sayisi DESC
        LIMIT 10
    ''').fetchall()
    
    # Çalışan bazlı ceza sayısı
    calisan_bazli = conn.execute('''
        SELECT 
            c.ad_soyad,
            COUNT(tc.id) as ceza_sayisi,
            SUM(tc.ceza_tutari) as toplam_tutar,
            SUM(CASE WHEN tc.calisan_odemesi_kesildi THEN tc.kesinti_tutari ELSE 0 END) as kesilen_tutar
        FROM calisanlar c
        LEFT JOIN arac_trafik_cezalari tc ON c.id = tc.calisan_id
        WHERE c.aktif = TRUE
        GROUP BY c.id, c.ad_soyad
        HAVING COUNT(tc.id) > 0
        ORDER BY ceza_sayisi DESC
        LIMIT 10
    ''').fetchall()
    
    # Ceza türü bazlı
    tur_bazli = conn.execute('''
        SELECT 
            ceza_turu,
            COUNT(*) as adet,
            SUM(ceza_tutari) as toplam_tutar
        FROM arac_trafik_cezalari
        GROUP BY ceza_turu
        ORDER BY adet DESC
    ''').fetchall()
    
    # Aylık ceza grafiği (son 12 ay)
    aylik = conn.execute('''
        SELECT 
            TO_CHAR(ceza_tarihi, 'YYYY-MM') as ay,
            COUNT(*) as adet,
            SUM(ceza_tutari) as tutar
        FROM arac_trafik_cezalari
        WHERE ceza_tarihi >= CURRENT_DATE - INTERVAL '12 months'
        GROUP BY TO_CHAR(ceza_tarihi, 'YYYY-MM')
        ORDER BY ay
    ''').fetchall()
    
    conn.close()
    
    return render_template('filo/trafik_cezalari_raporlar.html',
                         arac_bazli=arac_bazli,
                         calisan_bazli=calisan_bazli,
                         tur_bazli=tur_bazli,
                         aylik=aylik)


@app.route('/filo/arac/<int:arac_id>/trafik-cezalari')
@login_required
def filo_arac_trafik_cezalari(arac_id):
    """Belirli bir aracın trafik cezaları"""
    conn = get_db()
    
    arac = conn.execute('SELECT * FROM araclar WHERE id = %s', (arac_id,)).fetchone()
    
    if not arac:
        flash('Araç bulunamadı!', 'danger')
        return redirect(url_for('filo_araclar'))
    
    cezalar = conn.execute('''
        SELECT tc.*, c.ad_soyad
        FROM arac_trafik_cezalari tc
        LEFT JOIN calisanlar c ON tc.calisan_id = c.id
        WHERE tc.arac_id = %s
        ORDER BY tc.ceza_tarihi DESC
    ''', (arac_id,)).fetchall()
    
    conn.close()
    
    return render_template('filo/arac_trafik_cezalari.html',
                         arac=dict(arac),
                         cezalar=cezalar)