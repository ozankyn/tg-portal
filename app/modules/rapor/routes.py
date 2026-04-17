# -*- coding: utf-8 -*-
"""
TG Portal - Raporlama Modülü
Tüm modüller için merkezi raporlama
"""

from datetime import datetime, date, timedelta
from decimal import Decimal
import io

from flask import Blueprint, render_template, request, jsonify, Response, current_app
from flask_login import login_required, current_user
from sqlalchemy import func, extract, case, text

from app import db
from app.utils import permission_required

# Modelleri import et
from app.models.ik import Calisan, Departman
from app.models.masraf import Masraf, MasrafKategorisi
from app.models.sozlesme import Sozlesme, SozlesmeTipi
from app.models.satinalma import SatinAlmaTalebi, SatinAlmaSiparisi, SatinAlmaKategorisi
from app.models.talep import Talep, TalepKategorisi
from app.models.core import User

rapor_bp = Blueprint('rapor', __name__)


# ============================================================
# ANA DASHBOARD
# ============================================================

@rapor_bp.route('/')
@login_required
@permission_required('rapor.view')
def dashboard():
    """Raporlama ana sayfası"""
    return render_template('rapor/dashboard.html')


# ============================================================
# İK RAPORLARI
# ============================================================

@rapor_bp.route('/ik')
@login_required
@permission_required('rapor.view')
def ik_rapor():
    """İK raporları"""
    from app.models.base import CalisanDurumu
    
    # Genel istatistikler
    stats = {
        'toplam': Calisan.query.filter_by(is_deleted=False).count(),
        'aktif': Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).count(),
        'izinli': Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.IZINLI).count(),
        'ayrildi': Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AYRILDI).count(),
    }
    
    # Departman dağılımı
    departman_dagilim = db.session.query(
        Departman.ad,
        func.count(Calisan.id).label('sayi')
    ).join(Calisan, Calisan.departman_id == Departman.id).filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF
    ).group_by(Departman.ad).all()
    
    # Aylık işe alım (son 12 ay)
    aylik_ise_alim = db.session.query(
        extract('year', Calisan.ise_baslama).label('yil'),
        extract('month', Calisan.ise_baslama).label('ay'),
        func.count(Calisan.id).label('sayi')
    ).filter(
        Calisan.is_deleted == False,
        Calisan.ise_baslama >= date.today() - timedelta(days=365)
    ).group_by('yil', 'ay').order_by('yil', 'ay').all()
    
    return render_template('rapor/ik_rapor.html',
                          stats=stats,
                          departman_dagilim=departman_dagilim,
                          aylik_ise_alim=aylik_ise_alim)


# ============================================================
# MASRAF RAPORLARI
# ============================================================

@rapor_bp.route('/masraf')
@login_required
@permission_required('rapor.view')
def masraf_rapor():
    """Masraf raporları"""
    yil = request.args.get('yil', date.today().year, type=int)
    ay = request.args.get('ay', type=int)
    
    # Genel istatistikler
    query = Masraf.query.filter_by(is_deleted=False)
    if yil:
        query = query.filter(extract('year', Masraf.masraf_tarihi) == yil)
    if ay:
        query = query.filter(extract('month', Masraf.masraf_tarihi) == ay)
    
    stats = {
        'toplam_tutar': query.with_entities(func.sum(Masraf.tl_karsiligi)).scalar() or 0,
        'toplam_adet': query.count(),
        'onaylanan': query.filter(Masraf.durum == 'onaylandi').with_entities(func.sum(Masraf.tl_karsiligi)).scalar() or 0,
        'bekleyen': query.filter(Masraf.durum == 'onay_bekliyor').count(),
    }
    
    # Kategori dağılımı
    kategori_dagilim = db.session.query(
        MasrafKategorisi.ad,
        func.sum(Masraf.tl_karsiligi).label('toplam'),
        func.count(Masraf.id).label('adet')
    ).join(Masraf, Masraf.kategori_id == MasrafKategorisi.id).filter(
        Masraf.is_deleted == False,
        extract('year', Masraf.masraf_tarihi) == yil
    ).group_by(MasrafKategorisi.ad).all()
    
    # Aylık trend
    aylik_trend = db.session.query(
        extract('month', Masraf.masraf_tarihi).label('ay'),
        func.sum(Masraf.tl_karsiligi).label('toplam')
    ).filter(
        Masraf.is_deleted == False,
        extract('year', Masraf.masraf_tarihi) == yil
    ).group_by('ay').order_by('ay').all()
    
    return render_template('rapor/masraf_rapor.html',
                          stats=stats,
                          kategori_dagilim=kategori_dagilim,
                          aylik_trend=aylik_trend,
                          yil=yil,
                          ay=ay)


# ============================================================
# SÖZLEŞME RAPORLARI
# ============================================================

@rapor_bp.route('/sozlesme')
@login_required
@permission_required('rapor.view')
def sozlesme_rapor():
    """Sözleşme raporları"""
    # Genel istatistikler
    stats = {
        'toplam': Sozlesme.query.filter_by(is_deleted=False).count(),
        'aktif': Sozlesme.query.filter_by(is_deleted=False, durum='aktif').count(),
        'sona_eren': Sozlesme.query.filter_by(is_deleted=False, durum='sona_erdi').count(),
        'toplam_deger': Sozlesme.query.filter_by(is_deleted=False, durum='aktif').with_entities(
            func.sum(Sozlesme.tutar)
        ).scalar() or 0,
    }
    
    # Tip dağılımı
    tip_dagilim = db.session.query(
        SozlesmeTipi.ad,
        func.count(Sozlesme.id).label('adet'),
        func.sum(Sozlesme.tutar).label('toplam')
    ).join(Sozlesme, Sozlesme.tip_id == SozlesmeTipi.id).filter(
        Sozlesme.is_deleted == False,
        Sozlesme.durum == 'aktif'
    ).group_by(SozlesmeTipi.ad).all()
    
    # Yaklaşan bitiş (30 gün)
    yaklasan = Sozlesme.query.filter(
        Sozlesme.is_deleted == False,
        Sozlesme.durum == 'aktif',
        Sozlesme.bitis_tarihi <= date.today() + timedelta(days=30),
        Sozlesme.bitis_tarihi >= date.today()
    ).order_by(Sozlesme.bitis_tarihi).all()
    
    return render_template('rapor/sozlesme_rapor.html',
                          stats=stats,
                          tip_dagilim=tip_dagilim,
                          yaklasan=yaklasan)


# ============================================================
# SATIN ALMA RAPORLARI
# ============================================================

@rapor_bp.route('/satinalma')
@login_required
@permission_required('rapor.view')
def satinalma_rapor():
    """Satın alma raporları"""
    yil = request.args.get('yil', date.today().year, type=int)
    
    # Genel istatistikler
    stats = {
        'toplam_talep': SatinAlmaTalebi.query.filter_by(is_deleted=False).filter(
            extract('year', SatinAlmaTalebi.talep_tarihi) == yil
        ).count(),
        'toplam_siparis': SatinAlmaSiparisi.query.filter_by(is_deleted=False).filter(
            extract('year', SatinAlmaSiparisi.siparis_tarihi) == yil
        ).count(),
        'toplam_harcama': SatinAlmaSiparisi.query.filter_by(is_deleted=False).filter(
            extract('year', SatinAlmaSiparisi.siparis_tarihi) == yil
        ).with_entities(func.sum(SatinAlmaSiparisi.toplam_tutar)).scalar() or 0,
        'bekleyen': SatinAlmaTalebi.query.filter_by(is_deleted=False, durum='onay_bekliyor').count(),
    }
    
    # Kategori dağılımı
    kategori_dagilim = db.session.query(
        SatinAlmaKategorisi.ad,
        func.count(SatinAlmaTalebi.id).label('adet')
    ).join(SatinAlmaTalebi, SatinAlmaTalebi.kategori_id == SatinAlmaKategorisi.id).filter(
        SatinAlmaTalebi.is_deleted == False,
        extract('year', SatinAlmaTalebi.talep_tarihi) == yil
    ).group_by(SatinAlmaKategorisi.ad).all()
    
    # Aylık harcama
    aylik_harcama = db.session.query(
        extract('month', SatinAlmaSiparisi.siparis_tarihi).label('ay'),
        func.sum(SatinAlmaSiparisi.toplam_tutar).label('toplam')
    ).filter(
        SatinAlmaSiparisi.is_deleted == False,
        extract('year', SatinAlmaSiparisi.siparis_tarihi) == yil
    ).group_by('ay').order_by('ay').all()
    
    return render_template('rapor/satinalma_rapor.html',
                          stats=stats,
                          kategori_dagilim=kategori_dagilim,
                          aylik_harcama=aylik_harcama,
                          yil=yil)


# ============================================================
# TALEP RAPORLARI
# ============================================================

@rapor_bp.route('/talep')
@login_required
@permission_required('rapor.view')
def talep_rapor():
    """Talep raporları"""
    yil = request.args.get('yil', date.today().year, type=int)
    ay = request.args.get('ay', type=int)
    
    query = Talep.query.filter_by(is_deleted=False)
    if yil:
        query = query.filter(extract('year', Talep.created_at) == yil)
    if ay:
        query = query.filter(extract('month', Talep.created_at) == ay)
    
    # Genel istatistikler
    stats = {
        'toplam': query.count(),
        'acik': query.filter(Talep.durum.in_(['acik', 'atandi', 'devam_ediyor', 'beklemede'])).count(),
        'cozuldu': query.filter(Talep.durum == 'cozuldu').count(),
        'kapatildi': query.filter(Talep.durum == 'kapatildi').count(),
    }
    
    # Kategori dağılımı
    kategori_dagilim = db.session.query(
        TalepKategorisi.ad,
        func.count(Talep.id).label('adet')
    ).join(Talep, Talep.kategori_id == TalepKategorisi.id).filter(
        Talep.is_deleted == False,
        extract('year', Talep.created_at) == yil
    ).group_by(TalepKategorisi.ad).all()
    
    # Öncelik dağılımı
    oncelik_dagilim = db.session.query(
        Talep.oncelik,
        func.count(Talep.id).label('adet')
    ).filter(
        Talep.is_deleted == False,
        extract('year', Talep.created_at) == yil
    ).group_by(Talep.oncelik).all()
    
    # Aylık trend
    aylik_trend = db.session.query(
        extract('month', Talep.created_at).label('ay'),
        func.count(Talep.id).label('toplam')
    ).filter(
        Talep.is_deleted == False,
        extract('year', Talep.created_at) == yil
    ).group_by('ay').order_by('ay').all()

        # Öncelik dict (template bunu bekliyor)
    oncelik = {"dusuk": 0, "normal": 0, "yuksek": 0, "kritik": 0}
    for onc, adet in oncelik_dagilim:
        if not onc:
            continue
        if onc in oncelik:
            oncelik[onc] = int(adet or 0)

    # SLA aşımı (modelde sla_asim bool varsayımı)
    stats["sla_asim"] = query.filter(Talep.sla_asim == True).count()

    # Template uyumluluğu: "cozen"
    # (istersen sadece cozuldu da yapabiliriz)
    stats["cozen"] = int(stats.get("cozuldu", 0) + stats.get("kapatildi", 0))

    # Kategori detay (kategori adı, talep adedi, SLA aşımı)
    kategori_detay = db.session.query(
        TalepKategorisi.ad,
        func.count(Talep.id).label('adet'),
        func.sum(case((Talep.sla_asim == True, 1), else_=0)).label('sla_asim')
    ).join(Talep, Talep.kategori_id == TalepKategorisi.id).filter(
        Talep.is_deleted == False,
        extract('year', Talep.created_at) == yil
    ).group_by(TalepKategorisi.ad).order_by(func.count(Talep.id).desc()).all()

    # Aylık trend'i dict'e çevir (tojson rahat serialize etsin)
    aylik_trend_dict = {int(ay): int(toplam) for ay, toplam in aylik_trend}
    
    return render_template('rapor/talep_rapor.html',
                          stats=stats,
                          kategori_dagilim=kategori_dagilim,
                          kategori_detay=kategori_detay,
                          oncelik=oncelik,
                          aylik_trend=aylik_trend_dict,
                          yil=yil,
                          ay=ay)



# ============================================================
# EXCEL EXPORT
# ============================================================

@rapor_bp.route('/export/<modul>')
@login_required
@permission_required('rapor.export')
def export_excel(modul):
    """Excel export"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl yüklü değil", 500
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    if modul == 'masraf':
        ws.title = 'Masraflar'
        headers = ['Tarih', 'Başlık', 'Kategori', 'Tutar', 'Durum', 'Çalışan']
        ws.append(headers)
        
        masraflar = Masraf.query.filter_by(is_deleted=False).order_by(Masraf.masraf_tarihi.desc()).all()
        for m in masraflar:
            ws.append([
                m.masraf_tarihi.strftime('%d.%m.%Y') if m.masraf_tarihi else '',
                m.baslik,
                m.kategori.ad if m.kategori else '',
                float(m.tl_karsiligi or 0),
                m.durum_text,
                m.calisan.ad_soyad if m.calisan else ''
            ])
    
    elif modul == 'sozlesme':
        ws.title = 'Sözleşmeler'
        headers = ['No', 'Başlık', 'Tip', 'Taraf', 'Başlangıç', 'Bitiş', 'Tutar', 'Durum']
        ws.append(headers)
        
        sozlesmeler = Sozlesme.query.filter_by(is_deleted=False).order_by(Sozlesme.bitis_tarihi).all()
        for s in sozlesmeler:
            ws.append([
                s.sozlesme_no or str(s.id),
                s.baslik,
                s.tip.ad if s.tip else '',
                s.taraf_adi,
                s.baslangic_tarihi.strftime('%d.%m.%Y'),
                s.bitis_tarihi.strftime('%d.%m.%Y'),
                float(s.tutar or 0),
                s.durum_text
            ])
    
    elif modul == 'talep':
        ws.title = 'Talepler'
        headers = ['No', 'Konu', 'Kategori', 'Öncelik', 'Durum', 'Oluşturan', 'Tarih']
        ws.append(headers)
        
        talepler = Talep.query.filter_by(is_deleted=False).order_by(Talep.created_at.desc()).all()
        for t in talepler:
            ws.append([
                t.talep_no,
                t.konu,
                t.kategori.ad if t.kategori else '',
                t.oncelik_text,
                t.durum_text,
                t.olusturan.full_name if t.olusturan else '',
                t.created_at.strftime('%d.%m.%Y %H:%M')
            ])
    
    else:
        return "Geçersiz modül", 400
    
    # Sütun genişliklerini ayarla
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={modul}_rapor.xlsx'}
    )


# ============================================================
# API - Grafik verileri
# ============================================================

@rapor_bp.route('/api/ozet')
@login_required
@permission_required('rapor.view')
def api_ozet():
    """Genel özet API"""
    from app.models.base import CalisanDurumu
    
    data = {
        'calisan': Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).count(),
        'sozlesme': Sozlesme.query.filter_by(is_deleted=False, durum='aktif').count(),
        'acik_talep': Talep.query.filter(
            Talep.is_deleted == False,
            Talep.durum.in_(['acik', 'atandi', 'devam_ediyor', 'beklemede'])
        ).count(),
        'bu_ay_masraf': Masraf.query.filter(
            Masraf.is_deleted == False,
            extract('month', Masraf.masraf_tarihi) == date.today().month,
            extract('year', Masraf.masraf_tarihi) == date.today().year
        ).with_entities(func.sum(Masraf.tl_karsiligi)).scalar() or 0,
    }
    
    return jsonify(data)


# ============================================================
# AI RAPORLAMA ASISTANI
# ============================================================

AI_SYSTEM_PROMPT = """Sen TG Portal (Team Guerilla ERP) raporlama asistanısın.
Kullanıcının doğal dilde sorduğu soruları PostgreSQL sorguları ve Türkçe açıklamalarla yanıtlıyorsun.

## GÖREV
1. Kullanıcının sorusunu analiz et
2. Doğru SQL sorgusu oluştur (READ-ONLY: sadece SELECT)
3. Sorgu sonuçlarını Türkçe tablolar/özetlerle sun

## VERİTABANI ŞEMASI

### TEMEL TABLOLAR

**calisanlar** (Çalışan kayıtları — ~340 kayıt)
- id, sicil_no (unique), ad, soyad, tc_kimlik (unique)
- departman_id → departmanlar.id
- pozisyon_id → pozisyonlar.id
- yonetici_id → calisanlar.id (self-ref, direkt yönetici)
- kadro_id → hedef_kadrolar.id (proje ataması)
- durum: enum 'ADAY','AKTIF','IZINLI','ASKIYA_ALINDI','AYRILDI' (PostgreSQL enum, büyük harf)
- ise_baslama (date), isten_ayrilma (date), ayrilma_nedeni (text)
- kidem_tarihi, dogum_tarihi, cinsiyet, egitim_durumu
- email, telefon, adres, il, ilce
- yemek_karti, beden, kargo_subesi
- is_deleted (bool), created_at, updated_at

**departmanlar** (Departmanlar)
- id, ad, kod, ust_departman_id → departmanlar.id
- yonetici_id → calisanlar.id
- aktif (bool), is_deleted

**pozisyonlar** (Pozisyonlar)
- id, ad, kod, departman_id → departmanlar.id
- seviye (int), aktif (bool), is_deleted

**users** (Portal kullanıcıları)
- id, email (unique), ad, soyad, is_active, is_admin
- calisan_id → calisanlar.id (1-1 bağlantı)
- last_login, is_deleted

**roles / permissions** (Yetki sistemi)
- roles: id, name, display_name
- permissions: id, code, name, module
- user_roles (M2M): user_id, role_id
- role_permissions (M2M): role_id, permission_id

### PROJE & MÜŞTERİ

**musteriler** (Müşteriler)
- id, ad, kisa_ad, aktif, is_deleted
- Mevcut: Efes(1), Beylerbeyi(2), Kemer Gıda(3), PMI(4), Brown Forman(5)

**projeler** (Projeler)
- id, musteri_id → musteriler.id, ad, kod
- durum: 'aktif','tamamlandi','askida','iptal'
- aktif (bool), baslangic_tarihi, bitis_tarihi, butce

**hedef_kadrolar** (Proje kadroları — konum bazlı pozisyonlar)
- id, proje_id → projeler.id
- pozisyon_adi (varchar — "SSE - İzmir", "Merchandiser - Ankara" gibi lokasyonlu)
- il_id → iller.id, ilce_id → ilceler.id, bolge
- hedef_sayi, aktif, is_deleted
- İLİŞKİ: calisanlar.kadro_id → hedef_kadrolar.id

### FİLO (ARAÇ YÖNETİMİ)

**araclar** (Şirket araçları)
- id, plaka (unique), marka, model, model_yili, renk
- yakit_tipi: enum 'BENZIN','DIZEL','LPG','ELEKTRIK','HIBRIT'
- durum: enum 'AKTIF','BAKIM','ARIZALI','SATILDI','HURDA'
- km, sahiplik_tipi ('sirket','kiralama','leasing'), aylik_kira
- atanan_calisan_id → calisanlar.id, proje_id → projeler.id

**filo_islemler** (Bakım/tamir/sigorta/muayene/kaza/yakıt kayıtları)
- id, arac_id → araclar.id
- islem_tipi: enum 'BAKIM','TAMIR','SIGORTA','MUAYENE','KAZA','YAKIT','DIGER'
- tarih, km, tutar, kdv, toplam, aciklama, fatura_no

**yakit_kayitlari** (Yakıt alımları)
- id, arac_id → araclar.id, tarih, km, litre, birim_fiyat, tutar, istasyon_adi, tuketim (lt/100km)

**sigortalar** (Araç sigortaları)
- id, arac_id, sigorta_tipi ('kasko','trafik','ihtiyari'), sirket, police_no
- baslangic, bitis, prim, iptal (bool)

**muayeneler** (Araç muayeneleri)
- id, arac_id, tarih, sonuc ('gecti','kaldi','sartli_gecti'), sonraki_muayene

**kazalar** (Kaza kayıtları)
- id, arac_id, surucu_id → calisanlar.id, tarih, kusur_orani, hasar_tutari, durum ('acik','kapandi','dava')

### MASRAF

**masraflar** (Masraf girişleri)
- id, calisan_id → calisanlar.id, baslik, masraf_tarihi
- kategori_id → masraf_kategorileri.id
- tutar, kdv_orani, toplam_tutar, tl_karsiligi
- proje_id → projeler.id
- durum: 'taslak','onay_bekliyor','onaylandi','reddedildi','odendi'
- firma_adi, fatura_no, is_deleted

**masraf_kategorileri**: id, ad, kod (ULASIM, YEMEK, KONAKLAMA vb.)

### EĞİTİM

**egitim_tipleri**: id, ad, kod, kategori ('zorunlu','teknik','soft_skill','urun')
**egitimler** (Eğitim seansları)
- id, egitim_tipi_id, baslik, proje_id, baslangic_tarihi, bitis_tarihi
- durum: 'planli','devam_ediyor','tamamlandi','iptal'
- egitmen_id → calisanlar.id

**egitim_katilimcilar** (Katılımcılar)
- id, egitim_id → egitimler.id, calisan_id → calisanlar.id
- durum: 'davetli','katildi','gecti','kaldi','iptal','mazeret'
- puan (0-100)

### TEDARİKÇİ

**tedarikciler**: id, unvan, kisa_ad, tip (enum SERVIS,YAKIT,SIGORTA,YEDEK_PARCA,GENEL,KIRA,DIGER), aktif

### İK EK TABLOLAR

**izinler**: id, calisan_id, izin_tipi ('yillik','mazeret','hastalik','ucretsiz','dogum'), baslangic, bitis, gun_sayisi, durum ('beklemede','onaylandi','reddedildi')
**isten_cikislar**: id, calisan_id, cikis_tipi, planlanan_cikis_tarihi, gerceklesen_cikis_tarihi, durum
**zimmetler**: id, calisan_id, zimmet_tipi_id, tanim, teslim_tarihi, iade_tarihi, durum ('teslim_edildi','iade_edildi','kayip','hasarli')
**disiplin_kayitlari**: id, calisan_id, tarih, tur ('uyari','ihtar','fesih_uyarisi'), konu, durum
**davalar**: id, calisan_id, dosya_no, mahkeme, dava_turu, durum, talep_tutari
**icra_dosyalari**: id, calisan_id, dosya_no, toplam_borc, kalan_borc, durum

### ONAY WORKFLOW

**onay_talepleri**: id, onay_tipi_id, referans_tablo, referans_id, talep_eden_id → users.id, durum ('bekliyor','onaylandi','reddedildi','iptal')
**onay_kayitlari**: id, talep_id, onaylayici_id → users.id, durum, islem_tarihi

### ŞİRKET

**tuzel_kisiler**: id, ad, kisa_ad, vergi_no
**sgk_dosyalari**: id, tuzel_kisi_id, dosya_no, il, aktif

### COĞRAFYA

**iller**: id, ad, plaka_kodu
**ilceler**: id, il_id → iller.id, ad

## ÖNEMLİ KURALLAR

1. **SADECE SELECT** sorguları yaz. Asla INSERT/UPDATE/DELETE yapma.
2. **is_deleted = false** filtresi her zaman ekle (soft delete).
3. **CalisanDurumu** enum büyük harf: WHERE durum = 'AKTIF' (string olarak karşılaştır).
4. **AracDurumu/YakitTipi/IslemTipi** enum'ları da büyük harf.
5. Sonuçları Türkçe açıkla, tablo formatında sun.
6. Tarih formatı: DD.MM.YYYY (Türk formatı).
7. Para formatı: ₺1.234,56 (Türk formatı).
8. Çalışan tam adı: ad || ' ' || soyad
9. Proje-çalışan ilişkisi: calisanlar.kadro_id → hedef_kadrolar.id → projeler.id
10. Sorgu çok büyükse LIMIT 50 ekle.

## ÖRNEK SORGULAR

**"Kaç aktif çalışan var?"**
```sql
SELECT COUNT(*) FROM calisanlar WHERE durum = 'AKTIF' AND is_deleted = false;
```

**"Proje bazlı personel dağılımı"**
```sql
SELECT p.ad AS proje, COUNT(c.id) AS calisan_sayisi
FROM calisanlar c
JOIN hedef_kadrolar hk ON c.kadro_id = hk.id
JOIN projeler p ON hk.proje_id = p.id
WHERE c.durum = 'AKTIF' AND c.is_deleted = false
GROUP BY p.ad ORDER BY calisan_sayisi DESC;
```

**"Bu ay ayrılanlar"**
```sql
SELECT c.ad || ' ' || c.soyad AS calisan, c.sicil_no, c.isten_ayrilma, c.ayrilma_nedeni
FROM calisanlar c
WHERE c.durum = 'AYRILDI' AND c.is_deleted = false
  AND EXTRACT(MONTH FROM c.isten_ayrilma) = EXTRACT(MONTH FROM CURRENT_DATE)
  AND EXTRACT(YEAR FROM c.isten_ayrilma) = EXTRACT(YEAR FROM CURRENT_DATE)
ORDER BY c.isten_ayrilma DESC;
```

**"Departman bazlı dağılım"**
```sql
SELECT d.ad AS departman, COUNT(c.id) AS calisan_sayisi
FROM calisanlar c
JOIN departmanlar d ON c.departman_id = d.id
WHERE c.durum = 'AKTIF' AND c.is_deleted = false AND d.is_deleted = false
GROUP BY d.ad ORDER BY calisan_sayisi DESC;
```

**"Araç filosu özeti"**
```sql
SELECT durum, COUNT(*) FROM araclar WHERE is_deleted = false GROUP BY durum;
```

## YANIT FORMATI

Her yanıtta:
1. Sorunun kısa analizi
2. Çalıştırılan SQL sorgusu (```sql bloku)
3. Sonuç tablosu veya özet
4. Varsa ek notlar/öneriler

Eğer soru belirsizse, ne anladığını belirt ve en olası yorumu uygula.
Eğer soruya SQL ile yanıt verilemiyorsa (örn: tahmin, yorum), bunu belirt.
"""


@rapor_bp.route('/ai-asistan')
@login_required
@permission_required('rapor.view')
def ai_asistan():
    """AI Raporlama Asistanı sayfası"""
    return render_template('rapor/ai_asistan.html')


@rapor_bp.route('/ai-asistan/sorgula', methods=['POST'])
@login_required
@permission_required('rapor.view')
def ai_asistan_sorgula():
    """AI Asistan'a soru sor, SQL çalıştır, sonuçla birlikte yanıtla"""
    import anthropic
    import json
    import re
    from decimal import Decimal

    data = request.get_json()
    soru = data.get('soru', '').strip()
    if not soru:
        return jsonify({'success': False, 'error': 'Soru boş olamaz'}), 400

    api_key = current_app.config.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'success': False, 'error': 'AI servisi yapılandırılmamış'}), 500

    def serialize_value(val):
        """DB değerlerini JSON-safe tiplere dönüştür."""
        if val is None:
            return None
        if isinstance(val, Decimal):
            return float(val)
        if isinstance(val, datetime):
            return val.strftime('%d.%m.%Y %H:%M')
        if isinstance(val, date):
            return val.strftime('%d.%m.%Y')
        if isinstance(val, timedelta):
            return str(val)
        return val

    def extract_sql_blocks(text):
        """Claude yanıtından SQL bloklarını çıkar (```sql ... ``` veya ``` ... ```)."""
        # Önce ```sql ... ``` bloklarını dene
        pattern = r'```(?:sql)?\s*\n([\s\S]*?)```'
        blocks = re.findall(pattern, text)
        return [b.strip() for b in blocks if b.strip()]

    try:
        client = anthropic.Anthropic(api_key=api_key)

        # Adım 1: Claude'dan SQL + açıklama iste
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2048,
            system=AI_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": soru}
            ]
        )
        ai_yanit = message.content[0].text
        print(f"[AI Asistan] Soru: {soru}")
        print(f"[AI Asistan] Yanıt uzunluk: {len(ai_yanit)} karakter")

        # Adım 2: SQL sorgularını çıkar
        sql_blocks = extract_sql_blocks(ai_yanit)
        print(f"[AI Asistan] Bulunan SQL blokları: {len(sql_blocks)}")

        sql_sonuclar = []
        for i, sql in enumerate(sql_blocks):
            print(f"[AI Asistan] SQL #{i+1}: {sql[:200]}...")

            # Güvenlik: sadece SELECT/WITH
            sql_upper = sql.upper().strip()
            if not sql_upper.startswith('SELECT') and not sql_upper.startswith('WITH'):
                print(f"[AI Asistan] SQL #{i+1} atlandı: SELECT/WITH ile başlamıyor")
                continue

            yasakli = ['INSERT ', 'UPDATE ', 'DELETE ', 'DROP ', 'ALTER ', 'TRUNCATE ', 'CREATE ', 'GRANT ', 'REVOKE ']
            if any(k in sql_upper for k in yasakli):
                print(f"[AI Asistan] SQL #{i+1} atlandı: yasaklı komut içeriyor")
                continue

            try:
                result = db.session.execute(text(sql))
                kolonlar = list(result.keys())
                raw_rows = result.fetchall()
                satirlar = []
                for row in raw_rows[:100]:
                    satirlar.append({k: serialize_value(v) for k, v in zip(kolonlar, row)})
                sql_sonuclar.append({
                    'sql': sql,
                    'kolonlar': kolonlar,
                    'satirlar': satirlar,
                    'toplam': len(raw_rows),
                })
                print(f"[AI Asistan] SQL #{i+1} başarılı: {len(raw_rows)} satır")
            except Exception as e:
                print(f"[AI Asistan] SQL #{i+1} HATA: {e}")
                db.session.rollback()
                sql_sonuclar.append({
                    'sql': sql,
                    'hata': str(e),
                })

        if not sql_blocks:
            print(f"[AI Asistan] SQL bulunamadı. Yanıt ilk 500 kar: {ai_yanit[:500]}")

        return jsonify({
            'success': True,
            'yanit': ai_yanit,
            'sonuclar': sql_sonuclar,
        })

    except anthropic.APIError as e:
        print(f"[AI Asistan] API HATA: {e}")
        return jsonify({'success': False, 'error': f'AI API hatası: {str(e)}'}), 500
    except Exception as e:
        print(f"[AI Asistan] GENEL HATA: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@rapor_bp.route('/ai-asistan/export', methods=['POST'])
@login_required
@permission_required('rapor.view')
def ai_asistan_export():
    """AI Asistan sorgu sonucunu Excel'e aktar"""
    from decimal import Decimal

    data = request.get_json()
    sql = data.get('sql', '').strip()
    if not sql:
        return jsonify({'success': False, 'error': 'SQL boş'}), 400

    sql_upper = sql.upper().strip()
    if not sql_upper.startswith('SELECT') and not sql_upper.startswith('WITH'):
        return jsonify({'success': False, 'error': 'Sadece SELECT sorguları export edilebilir'}), 400

    yasakli = ['INSERT ', 'UPDATE ', 'DELETE ', 'DROP ', 'ALTER ', 'TRUNCATE ', 'CREATE ', 'GRANT ', 'REVOKE ']
    if any(k in sql_upper for k in yasakli):
        return jsonify({'success': False, 'error': 'Yasaklı SQL komutu'}), 400

    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        result = db.session.execute(text(sql))
        kolonlar = list(result.keys())
        rows = result.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'AI Rapor'
        ws.append(kolonlar)

        for row in rows:
            ws.append([float(v) if isinstance(v, Decimal) else v for v in row])

        for col in range(1, len(kolonlar) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=ai_rapor.xlsx'}
        )
    except Exception as e:
        print(f"[AI Asistan Export] HATA: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
