# -*- coding: utf-8 -*-
"""
TG Portal - Masraf Modülü Routes
"""

from datetime import datetime, date
from decimal import Decimal
import os
import uuid

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from app import db
from app.models.masraf import Masraf, MasrafKategorisi, MasrafKalemi, MasrafAvans, MasrafRaporu, get_calisan_masraf_ozeti
from app.models.ik import Calisan
from app.models.proje import Proje
from app.models.onay import OnayServisi
from app.utils import permission_required, paginate_query

masraf_bp = Blueprint('masraf', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# DASHBOARD
# ============================================================

@masraf_bp.route('/')
@login_required
def dashboard():
    """Masraf dashboard - özet ve son masraflar"""
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    
    # Admin ise admin dashboard'a yönlendir
    if not calisan:
        if current_user.has_permission('masraf.admin'):
            return redirect(url_for('masraf.admin_dashboard'))
        flash('Çalışan kaydınız bulunamadı.', 'warning')
        return redirect(url_for('core.dashboard'))
    
    # Özet bilgiler
    bugun = date.today()
    ozet = get_calisan_masraf_ozeti(calisan.id, bugun.year, bugun.month)
    
    # Son masraflar
    son_masraflar = Masraf.query.filter_by(
        calisan_id=calisan.id,
        is_deleted=False
    ).order_by(Masraf.created_at.desc()).limit(10).all()
    
    # Kategoriler (hızlı ekleme için)
    kategoriler = MasrafKategorisi.query.filter_by(aktif=True).order_by(MasrafKategorisi.sira).all()
    
    return render_template('masraf/dashboard.html',
                          calisan=calisan,
                          ozet=ozet,
                          son_masraflar=son_masraflar,
                          kategoriler=kategoriler,
                          bugun=bugun)


# ============================================================
# MASRAF LİSTESİ
# ============================================================

@masraf_bp.route('/liste')
@login_required
def liste():
    """Masraf listesi"""
    page = request.args.get('page', 1, type=int)
    durum = request.args.get('durum')
    kategori_id = request.args.get('kategori_id', type=int)
    ay = request.args.get('ay', type=int)
    yil = request.args.get('yil', type=int)
    
    # Kullanıcının çalışan kaydı
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    
    # Admin tüm masrafları görebilir
    if current_user.has_permission('masraf.admin'):
        query = Masraf.query.filter_by(is_deleted=False)
    elif calisan:
        query = Masraf.query.filter_by(calisan_id=calisan.id, is_deleted=False)
    else:
        flash('Çalışan kaydınız bulunamadı.', 'warning')
        return redirect(url_for('core.dashboard'))
    
    # Filtreler
    if durum:
        query = query.filter(Masraf.durum == durum)
    if kategori_id:
        query = query.filter(Masraf.kategori_id == kategori_id)
    if ay and yil:
        query = query.filter(Masraf.donem_ay == ay, Masraf.donem_yil == yil)
    elif yil:
        query = query.filter(Masraf.donem_yil == yil)
    
    query = query.order_by(Masraf.masraf_tarihi.desc())
    pagination = paginate_query(query, page, 20)
    
    kategoriler = MasrafKategorisi.query.filter_by(aktif=True).order_by(MasrafKategorisi.ad).all()
    
    return render_template('masraf/liste.html',
                          masraflar=pagination.items,
                          pagination=pagination,
                          kategoriler=kategoriler)


# ============================================================
# MASRAF EKLE/DÜZENLE
# ============================================================

@masraf_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
def ekle():
    """Yeni masraf ekle"""
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    
    if not calisan:
        flash('Çalışan kaydınız bulunamadı.', 'warning')
        return redirect(url_for('core.dashboard'))
    
    if request.method == 'POST':
        masraf = Masraf(
            calisan_id=calisan.id,
            baslik=request.form.get('baslik', '').strip(),
            aciklama=request.form.get('aciklama', '').strip() or None,
            masraf_tarihi=datetime.strptime(request.form['masraf_tarihi'], '%Y-%m-%d').date(),
            kategori_id=int(request.form['kategori_id']) if request.form.get('kategori_id') else None,
            tutar=Decimal(request.form.get('tutar', '0').replace(',', '.')),
            kdv_orani=int(request.form.get('kdv_orani', 20)),
            para_birimi=request.form.get('para_birimi', 'TRY'),
            proje_id=int(request.form['proje_id']) if request.form.get('proje_id') else None,
            firma_adi=request.form.get('firma_adi', '').strip() or None,
            fatura_no=request.form.get('fatura_no', '').strip() or None,
            durum='taslak'
        )
        
        # Dönem bilgisi
        masraf.donem_ay = masraf.masraf_tarihi.month
        masraf.donem_yil = masraf.masraf_tarihi.year
        
        # Döviz kuru
        if masraf.para_birimi != 'TRY' and request.form.get('kur'):
            masraf.kur = Decimal(request.form.get('kur', '1').replace(',', '.'))
        
        # KDV hesapla
        masraf.hesapla_kdv()
        
        db.session.add(masraf)
        db.session.flush()
        
        # Dosya yükleme
        if 'dosya' in request.files:
            file = request.files['dosya']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                unique_name = f"{uuid.uuid4().hex}_{filename}"
                
                upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'masraf', str(masraf.id))
                os.makedirs(upload_folder, exist_ok=True)
                
                filepath = os.path.join(upload_folder, unique_name)
                file.save(filepath)
                
                masraf.dosya_adi = filename
                masraf.dosya_yolu = filepath
                masraf.dosya_tipi = file.content_type
        
        db.session.commit()
        
        flash('Masraf kaydedildi.', 'success')
        return redirect(url_for('masraf.detay', id=masraf.id))
    
    kategoriler = MasrafKategorisi.query.filter_by(aktif=True).order_by(MasrafKategorisi.sira).all()
    projeler = Proje.query.filter_by(is_deleted=False, durum='aktif').order_by(Proje.ad).all()
    
    return render_template('masraf/form.html',
                          masraf=None,
                          kategoriler=kategoriler,
                          projeler=projeler)


@masraf_bp.route('/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
def duzenle(id):
    """Masraf düzenle"""
    masraf = Masraf.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    if not calisan or (masraf.calisan_id != calisan.id and not current_user.has_permission('masraf.admin')):
        flash('Bu masrafı düzenleme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    if not masraf.duzenlenebilir:
        flash('Bu masraf düzenlenemez.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    if request.method == 'POST':
        masraf.baslik = request.form.get('baslik', '').strip()
        masraf.aciklama = request.form.get('aciklama', '').strip() or None
        masraf.masraf_tarihi = datetime.strptime(request.form['masraf_tarihi'], '%Y-%m-%d').date()
        masraf.kategori_id = int(request.form['kategori_id']) if request.form.get('kategori_id') else None
        masraf.tutar = Decimal(request.form.get('tutar', '0').replace(',', '.'))
        masraf.kdv_orani = int(request.form.get('kdv_orani', 20))
        masraf.para_birimi = request.form.get('para_birimi', 'TRY')
        masraf.proje_id = int(request.form['proje_id']) if request.form.get('proje_id') else None
        masraf.firma_adi = request.form.get('firma_adi', '').strip() or None
        masraf.fatura_no = request.form.get('fatura_no', '').strip() or None
        
        # Dönem güncelle
        masraf.donem_ay = masraf.masraf_tarihi.month
        masraf.donem_yil = masraf.masraf_tarihi.year
        
        # Döviz kuru
        if masraf.para_birimi != 'TRY' and request.form.get('kur'):
            masraf.kur = Decimal(request.form.get('kur', '1').replace(',', '.'))
        
        # KDV hesapla
        masraf.hesapla_kdv()
        
        # Reddedilmişse taslağa çevir
        if masraf.durum == 'reddedildi':
            masraf.durum = 'taslak'
        
        # Yeni dosya yükleme
        if 'dosya' in request.files:
            file = request.files['dosya']
            if file and file.filename and allowed_file(file.filename):
                # Eski dosyayı sil
                if masraf.dosya_yolu and os.path.exists(masraf.dosya_yolu):
                    os.remove(masraf.dosya_yolu)
                
                filename = secure_filename(file.filename)
                unique_name = f"{uuid.uuid4().hex}_{filename}"
                
                upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'masraf', str(masraf.id))
                os.makedirs(upload_folder, exist_ok=True)
                
                filepath = os.path.join(upload_folder, unique_name)
                file.save(filepath)
                
                masraf.dosya_adi = filename
                masraf.dosya_yolu = filepath
                masraf.dosya_tipi = file.content_type
        
        db.session.commit()
        
        flash('Masraf güncellendi.', 'success')
        return redirect(url_for('masraf.detay', id=id))
    
    kategoriler = MasrafKategorisi.query.filter_by(aktif=True).order_by(MasrafKategorisi.sira).all()
    projeler = Proje.query.filter_by(is_deleted=False, durum='aktif').order_by(Proje.ad).all()
    
    return render_template('masraf/form.html',
                          masraf=masraf,
                          kategoriler=kategoriler,
                          projeler=projeler)


# ============================================================
# MASRAF DETAY
# ============================================================

@masraf_bp.route('/<int:id>')
@login_required
def detay(id):
    """Masraf detayı"""
    masraf = Masraf.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_owner = calisan and masraf.calisan_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    if not is_owner and not is_admin:
        flash('Bu masrafı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    return render_template('masraf/detay.html',
                          masraf=masraf,
                          is_owner=is_owner)


# ============================================================
# ONAYA GÖNDER
# ============================================================

@masraf_bp.route('/<int:id>/onaya-gonder', methods=['POST'])
@login_required
def onaya_gonder(id):
    """Masrafı onaya gönder"""
    masraf = Masraf.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    if not calisan or masraf.calisan_id != calisan.id:
        flash('Bu masrafı onaya gönderme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    if masraf.durum not in ['taslak', 'reddedildi']:
        flash('Bu masraf onaya gönderilemez.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    # Fatura zorunlu mu kontrol
    if masraf.kategori and masraf.kategori.fatura_zorunlu and not masraf.dosya_yolu:
        flash('Bu kategori için fatura/fiş yüklemek zorunludur.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    # Onay talebi oluştur
    talep, hata = OnayServisi.talep_olustur(
        onay_tipi_kod='MASRAF',
        referans_tablo='masraflar',
        referans_id=masraf.id,
        talep_eden_id=current_user.id
    )
    
    if hata:
        flash(f'Onay talebi oluşturulamadı: {hata}', 'danger')
        return redirect(url_for('masraf.detay', id=id))
    
    masraf.durum = 'onay_bekliyor'
    masraf.onay_talebi_id = talep.id
    db.session.commit()
    
    flash('Masraf onaya gönderildi.', 'success')
    return redirect(url_for('masraf.detay', id=id))




# ============================================================
# ÖDENDİ İŞARETLE
# ============================================================
@masraf_bp.route('/<int:id>/odendi-isaretle', methods=['POST'])
@login_required
@permission_required('masraf.admin')
def odendi_isaretle(id):
    """Masrafı ödendi olarak işaretle"""
    masraf = Masraf.query.get_or_404(id)
    
    if masraf.durum != 'onaylandi':
        flash('Sadece onaylanmış masraflar ödendi olarak işaretlenebilir.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    masraf.durum = 'odendi'
    masraf.odeme_tarihi = date.today()
    db.session.commit()
    
    flash('Masraf ödendi olarak işaretlendi.', 'success')
    return redirect(url_for('masraf.detay', id=id))

# ============================================================
# DOSYA İNDİR
# ============================================================

@masraf_bp.route('/<int:id>/dosya')
@login_required
def dosya_indir(id):
    """Masraf faturası/fişini indir"""
    masraf = Masraf.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_owner = calisan and masraf.calisan_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    if not is_owner and not is_admin:
        flash('Bu dosyaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    if not masraf.dosya_yolu or not os.path.exists(masraf.dosya_yolu):
        flash('Dosya bulunamadı.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    return send_file(masraf.dosya_yolu, 
                     download_name=masraf.dosya_adi,
                     as_attachment=True)


# ============================================================
# SİL
# ============================================================

@masraf_bp.route('/<int:id>/sil', methods=['POST'])
@login_required
def sil(id):
    """Masrafı sil (soft delete)"""
    masraf = Masraf.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    if not calisan or masraf.calisan_id != calisan.id:
        flash('Bu masrafı silme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    if masraf.durum not in ['taslak', 'reddedildi']:
        flash('Sadece taslak veya reddedilmiş masraflar silinebilir.', 'warning')
        return redirect(url_for('masraf.detay', id=id))
    
    masraf.is_deleted = True
    masraf.deleted_at = datetime.now()
    db.session.commit()
    
    flash('Masraf silindi.', 'success')
    return redirect(url_for('masraf.liste'))


# ============================================================
# KATEGORİ YÖNETİMİ (Admin)
# ============================================================

@masraf_bp.route('/kategoriler')
@login_required
@permission_required('masraf.admin')
def kategori_liste():
    """Masraf kategorileri"""
    kategoriler = MasrafKategorisi.query.order_by(MasrafKategorisi.sira, MasrafKategorisi.ad).all()
    return render_template('masraf/kategori_liste.html', kategoriler=kategoriler)


@masraf_bp.route('/kategori/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('masraf.admin')
def kategori_ekle():
    """Yeni kategori ekle"""
    if request.method == 'POST':
        kategori = MasrafKategorisi(
            ad=request.form.get('ad', '').strip(),
            kod=request.form.get('kod', '').strip().upper() or None,
            aciklama=request.form.get('aciklama', '').strip() or None,
            ikon=request.form.get('ikon', 'bi-receipt'),
            renk=request.form.get('renk', 'secondary'),
            fatura_zorunlu=request.form.get('fatura_zorunlu') == 'on',
            muhasebe_kodu=request.form.get('muhasebe_kodu', '').strip() or None
        )
        
        if request.form.get('gunluk_limit'):
            kategori.gunluk_limit = Decimal(request.form['gunluk_limit'].replace(',', '.'))
        if request.form.get('aylik_limit'):
            kategori.aylik_limit = Decimal(request.form['aylik_limit'].replace(',', '.'))
        
        db.session.add(kategori)
        db.session.commit()
        
        flash('Kategori eklendi.', 'success')
        return redirect(url_for('masraf.kategori_liste'))
    
    return render_template('masraf/kategori_form.html', kategori=None)


# ============================================================
# RAPORLAR
# ============================================================


@masraf_bp.route('/admin')
@login_required
@permission_required('masraf.admin')
def admin_dashboard():
    """Admin masraf dashboard"""
    from sqlalchemy import func
    from datetime import date
    
    bugun = date.today()
    
    # Genel istatistikler
    stats = {
        'bekleyen': Masraf.query.filter_by(durum='onay_bekliyor', is_deleted=False).count(),
        'bu_ay_toplam': db.session.query(func.sum(Masraf.tl_karsiligi)).filter(
            Masraf.is_deleted == False,
            Masraf.donem_ay == bugun.month,
            Masraf.donem_yil == bugun.year
        ).scalar() or 0,
        'onaylanan': Masraf.query.filter_by(durum='onaylandi', is_deleted=False).count(),
    }
    
    # Son masraflar (tümü)
    son_masraflar = Masraf.query.filter_by(is_deleted=False).order_by(Masraf.created_at.desc()).limit(20).all()
    
    # Kategoriler
    kategoriler = MasrafKategorisi.query.filter_by(aktif=True).order_by(MasrafKategorisi.sira).all()
    
    return render_template('masraf/admin_dashboard.html',
                          stats=stats,
                          son_masraflar=son_masraflar,
                          kategoriler=kategoriler,
                          bugun=bugun)


@masraf_bp.route('/rapor')
@login_required
@permission_required('masraf.admin')
def rapor():
    """Masraf raporları"""
    from sqlalchemy import func
    
    yil = request.args.get('yil', date.today().year, type=int)
    ay = request.args.get('ay', type=int)
    
    # Genel özet
    query = Masraf.query.filter(
        Masraf.is_deleted == False,
        Masraf.donem_yil == yil
    )
    
    if ay:
        query = query.filter(Masraf.donem_ay == ay)
    
    # Kategori bazlı
    kategori_ozet = db.session.query(
        MasrafKategorisi.ad,
        func.count(Masraf.id).label('adet'),
        func.sum(Masraf.tl_karsiligi).label('toplam')
    ).join(Masraf, Masraf.kategori_id == MasrafKategorisi.id).filter(
        Masraf.is_deleted == False,
        Masraf.donem_yil == yil
    )
    
    if ay:
        kategori_ozet = kategori_ozet.filter(Masraf.donem_ay == ay)
    
    kategori_ozet = kategori_ozet.group_by(MasrafKategorisi.ad).all()
    
    # Aylık trend
    aylik_trend = db.session.query(
        Masraf.donem_ay,
        func.sum(Masraf.tl_karsiligi).label('toplam')
    ).filter(
        Masraf.is_deleted == False,
        Masraf.donem_yil == yil
    ).group_by(Masraf.donem_ay).order_by(Masraf.donem_ay).all()
    
    return render_template('masraf/rapor.html',
                          yil=yil,
                          ay=ay,
                          kategori_ozet=kategori_ozet,
                          aylik_trend=aylik_trend)


# ============================================================
# API - AJAX için
# ============================================================

@masraf_bp.route('/api/kategori/<int:id>')
@login_required
def api_kategori_detay(id):
    """Kategori detayı (form için)"""
    kategori = MasrafKategorisi.query.get_or_404(id)
    return jsonify({
        'id': kategori.id,
        'ad': kategori.ad,
        'fatura_zorunlu': kategori.fatura_zorunlu,
        'gunluk_limit': float(kategori.gunluk_limit) if kategori.gunluk_limit else None,
        'aylik_limit': float(kategori.aylik_limit) if kategori.aylik_limit else None
    })


# ============================================================
# AI FİŞ TANIMA
# ============================================================

@masraf_bp.route('/fis-tani', methods=['POST'])
@login_required
def fis_tani():
    """Fiş/fatura görselinden bilgi çıkar (Claude AI)"""
    import anthropic
    import json
    import base64
    
    if 'dosya' not in request.files:
        return jsonify({'success': False, 'error': 'Dosya yüklenmedi'}), 400
    
    dosya = request.files['dosya']
    if dosya.filename == '':
        return jsonify({'success': False, 'error': 'Dosya seçilmedi'}), 400
    
    # Dosya tipini kontrol et
    allowed = {'jpg', 'jpeg', 'png', 'gif', 'webp'}
    ext = dosya.filename.rsplit('.', 1)[1].lower() if '.' in dosya.filename else ''
    if ext not in allowed:
        return jsonify({'success': False, 'error': 'Sadece görsel dosyalar desteklenir (JPG, PNG)'}), 400
    
    # Dosyayı base64'e çevir
    image_data = base64.standard_b64encode(dosya.read()).decode('utf-8')
    
    # MIME type belirle
    mime_types = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'webp': 'image/webp'
    }
    media_type = mime_types.get(ext, 'image/jpeg')
    
    # Claude API çağrısı
    try:
        client = anthropic.Anthropic(api_key=current_app.config.get('ANTHROPIC_API_KEY'))
        
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": media_type,
                                "data": image_data
                            }
                        },
                        {
                            "type": "text",
                            "text": """Bu bir Türk fişi veya faturası görseli. Lütfen aşağıdaki bilgileri JSON formatında çıkar:

{
    "firma_adi": "Faturayı kesen firma/işletme adı",
    "tarih": "YYYY-MM-DD formatında tarih",
    "tutar": "Fişte yazan TOPLAM tutar (KDV dahil genel toplam, sadece sayı, nokta ile ondalık ayracı)",
    "kdv_orani": "0",
    "fatura_no": "Fatura veya fiş numarası",
    "aciklama": "Satın alınan ürün/hizmetin kısa açıklaması (örn: Yemek, Taksi, Ofis malzemesi)"
}

KURALLAR:
- tutar alanına her zaman fişin EN ALTINDAKİ TOPLAM tutarı yaz (KDV dahil)
- kdv_orani her zaman 0 olsun (kullanıcı manuel seçecek)
- Türk lirası formatında virgül ondalık ayracı olabilir, bunu noktaya çevir (örn: 5.670,50 -> 5670.50)
- Eğer bir bilgiyi okuyamıyorsan null yaz
- Sadece JSON döndür, başka açıklama ekleme"""
                        }
                    ]
                }
            ]
        )
        
        # Response'u parse et
        response_text = message.content[0].text.strip()
        
        # JSON'u çıkar (bazen markdown code block içinde gelebilir)
        if response_text.startswith('```'):
            response_text = response_text.split('```')[1]
            if response_text.startswith('json'):
                response_text = response_text[4:]
        response_text = response_text.strip()
        
        import json
        result = json.loads(response_text)
        
        return jsonify({'success': True, 'data': result})
        
    except anthropic.APIError as e:
        return jsonify({'success': False, 'error': f'API hatası: {str(e)}'}), 500
    except json.JSONDecodeError:
        return jsonify({'success': False, 'error': 'AI yanıtı işlenemedi'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500



# ============================================================
# MASRAF RAPORU
# ============================================================

@masraf_bp.route('/raporlarim')
@login_required
def rapor_liste():
    """Çalışanın masraf raporları listesi"""
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    
    if not calisan:
        flash('Çalışan kaydınız bulunamadı.', 'warning')
        return redirect(url_for('core.dashboard'))
    
    raporlar = MasrafRaporu.query.filter_by(
        calisan_id=calisan.id,
        is_deleted=False
    ).order_by(MasrafRaporu.donem_yil.desc(), MasrafRaporu.donem_ay.desc()).all()
    
    return render_template('masraf/rapor_liste.html', raporlar=raporlar, calisan=calisan)


@masraf_bp.route('/rapor/olustur', methods=['GET', 'POST'])
@login_required
def rapor_olustur():
    """Yeni masraf raporu oluştur"""
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    
    if not calisan:
        flash('Çalışan kaydınız bulunamadı.', 'warning')
        return redirect(url_for('core.dashboard'))
    
    if request.method == 'POST':
        donem_yil = int(request.form.get('donem_yil'))
        donem_ay = int(request.form.get('donem_ay'))
        avans_tutari = request.form.get('avans_tutari', '0')
        avans_tutari = Decimal(avans_tutari.replace(',', '.')) if avans_tutari else Decimal('0')
        
        # Bu dönem için rapor var mı kontrol et
        mevcut = MasrafRaporu.query.filter_by(
            calisan_id=calisan.id,
            donem_yil=donem_yil,
            donem_ay=donem_ay,
            is_deleted=False
        ).first()
        
        if mevcut:
            flash('Bu dönem için zaten bir rapor mevcut.', 'warning')
            return redirect(url_for('masraf.rapor_detay', id=mevcut.id))
        
        # Onaylanmış masrafları bul
        masraflar = Masraf.query.filter(
            Masraf.calisan_id == calisan.id,
            Masraf.donem_yil == donem_yil,
            Masraf.donem_ay == donem_ay,
            Masraf.durum == 'onaylandi',
            Masraf.is_deleted == False
        ).all()
        
        if not masraflar:
            flash('Bu dönemde onaylanmış masrafınız bulunmuyor.', 'warning')
            return redirect(url_for('masraf.rapor_olustur'))
        
        # Rapor oluştur
        aylar = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
                 'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
        
        rapor = MasrafRaporu(
            calisan_id=calisan.id,
            donem_yil=donem_yil,
            donem_ay=donem_ay,
            baslik=f"{aylar[donem_ay]} {donem_yil} Masraf Raporu",
            avans_tutari=avans_tutari
        )
        
        # Masrafları rapora ekle
        toplam = Decimal('0')
        for masraf in masraflar:
            rapor.masraflar.append(masraf)
            toplam += masraf.tl_karsiligi or Decimal('0')
        
        rapor.toplam_masraf = toplam
        rapor.net_odeme = toplam - avans_tutari
        
        db.session.add(rapor)
        db.session.commit()
        
        flash('Masraf raporu oluşturuldu.', 'success')
        return redirect(url_for('masraf.rapor_detay', id=rapor.id))
    
    # GET - Form göster
    bugun = date.today()
    
    # Mevcut avansları getir
    avanslar = MasrafAvans.query.filter_by(
        calisan_id=calisan.id,
        durum='aktif',
        is_deleted=False
    ).all()
    
    toplam_avans = sum(a.kalan_tutar for a in avanslar)
    
    return render_template('masraf/rapor_olustur.html',
                          calisan=calisan,
                          bugun=bugun,
                          avanslar=avanslar,
                          toplam_avans=toplam_avans)


@masraf_bp.route('/rapor/<int:id>')
@login_required
def rapor_detay(id):
    """Masraf raporu detayı"""
    rapor = MasrafRaporu.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_owner = calisan and rapor.calisan_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    # Yönetici mi kontrol (onaylama yetkisi için)
    is_yonetici = False
    if calisan and rapor.calisan.yonetici_id == calisan.id:
        is_yonetici = True
    
    if not is_owner and not is_admin and not is_yonetici:
        flash('Bu raporu görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    return render_template('masraf/rapor_detay.html',
                          rapor=rapor,
                          is_owner=is_owner,
                          is_admin=is_admin,
                          is_yonetici=is_yonetici)


@masraf_bp.route('/rapor/<int:id>/onaya-gonder', methods=['POST'])
@login_required
def rapor_onaya_gonder(id):
    """Masraf raporunu onaya gönder"""
    rapor = MasrafRaporu.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    if not calisan or rapor.calisan_id != calisan.id:
        flash('Bu raporu onaya gönderme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.rapor_liste'))
    
    if rapor.durum not in ['taslak', 'reddedildi']:
        flash('Bu rapor onaya gönderilemez.', 'warning')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    rapor.durum = 'onay_bekliyor'
    db.session.commit()
    
    # Yöneticiye email gönder
    try:
        from app.services.notification import notify_masraf_raporu_onay
        print(f"DEBUG: Email gönderiliyor rapor #{rapor.id}")
        result = notify_masraf_raporu_onay(rapor)
        print(f"DEBUG: Email sonucu: {result}")
        if result:
            rapor.yonetici_mail_gonderildi = True
            db.session.commit()
            flash('Email gönderildi!', 'info')
        else:
            flash('Email gönderilemedi!', 'warning')
    except Exception as e:
        print(f"DEBUG: Email hatası: {str(e)}")
        flash(f'Email hatası: {str(e)}', 'danger')
    
    flash('Masraf raporu onaya gönderildi.', 'success')
    return redirect(url_for('masraf.rapor_detay', id=id))


@masraf_bp.route('/rapor/<int:id>/onayla', methods=['POST'])
@login_required
def rapor_onayla(id):
    """Masraf raporunu onayla (yönetici)"""
    rapor = MasrafRaporu.query.get_or_404(id)
    
    # Yetki kontrolü - yönetici veya admin
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_yonetici = calisan and rapor.calisan.yonetici_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    if not is_yonetici and not is_admin:
        flash('Bu raporu onaylama yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    if rapor.durum != 'onay_bekliyor':
        flash('Bu rapor onaylanamaz.', 'warning')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    rapor.durum = 'onaylandi'
    rapor.onaylayan_id = current_user.id
    rapor.onay_tarihi = datetime.now()
    rapor.onay_notu = request.form.get('not', '').strip() or None
    db.session.commit()
    
    # Muhasebeye ve çalışana email gönder
    from app.services.notification import notify_masraf_raporu_muhasebe, notify_masraf_raporu_sonuc
    notify_masraf_raporu_muhasebe(rapor)
    notify_masraf_raporu_sonuc(rapor, onaylandi=True)
    rapor.muhasebe_mail_gonderildi = True
    db.session.commit()
    
    flash('Masraf raporu onaylandı.', 'success')
    return redirect(url_for('masraf.rapor_detay', id=id))


@masraf_bp.route('/rapor/<int:id>/reddet', methods=['POST'])
@login_required
def rapor_reddet(id):
    """Masraf raporunu reddet (yönetici)"""
    rapor = MasrafRaporu.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_yonetici = calisan and rapor.calisan.yonetici_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    if not is_yonetici and not is_admin:
        flash('Bu raporu reddetme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    if rapor.durum != 'onay_bekliyor':
        flash('Bu rapor reddedilemez.', 'warning')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    rapor.durum = 'reddedildi'
    rapor.onaylayan_id = current_user.id
    rapor.onay_tarihi = datetime.now()
    rapor.onay_notu = request.form.get('not', '').strip() or None
    db.session.commit()
    
    # Çalışana email gönder
    from app.services.notification import notify_masraf_raporu_sonuc
    notify_masraf_raporu_sonuc(rapor, onaylandi=False, aciklama=rapor.onay_notu)
    db.session.commit()
    
    flash('Masraf raporu reddedildi.', 'success')
    return redirect(url_for('masraf.rapor_detay', id=id))


@masraf_bp.route('/rapor/<int:id>/excel')
@login_required
def rapor_excel(id):
    """Masraf raporu Excel export"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    rapor = MasrafRaporu.query.get_or_404(id)
    
    # Yetki kontrolü
    calisan = Calisan.query.filter_by(email=current_user.email, is_deleted=False).first()
    is_owner = calisan and rapor.calisan_id == calisan.id
    is_admin = current_user.has_permission('masraf.admin')
    
    if not is_owner and not is_admin:
        flash('Bu raporu indirme yetkiniz yok.', 'danger')
        return redirect(url_for('masraf.liste'))
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Masraf Raporu"
    
    # Stiller
    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    
    # Başlık
    ws.merge_cells('A1:G1')
    ws['A1'] = f"MASRAF RAPORU - {rapor.donem_text}"
    ws['A1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Çalışan bilgileri
    ws['A3'] = "Çalışan:"
    ws['B3'] = rapor.calisan.full_name
    ws['A4'] = "Departman:"
    ws['B4'] = rapor.calisan.departman.ad if rapor.calisan.departman else "-"
    ws['A5'] = "Dönem:"
    ws['B5'] = rapor.donem_text
    ws['A3'].font = bold_font
    ws['A4'].font = bold_font
    ws['A5'].font = bold_font
    
    # Tablo başlıkları
    headers = ['#', 'Tarih', 'Kategori', 'Açıklama', 'Firma', 'Fatura No', 'Tutar (₺)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Masraf satırları
    row = 8
    for idx, masraf in enumerate(rapor.masraflar, 1):
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=masraf.masraf_tarihi.strftime('%d.%m.%Y') if masraf.masraf_tarihi else '-').border = thin_border
        ws.cell(row=row, column=3, value=masraf.kategori.ad if masraf.kategori else '-').border = thin_border
        ws.cell(row=row, column=4, value=masraf.baslik or masraf.aciklama or '-').border = thin_border
        ws.cell(row=row, column=5, value=masraf.firma_adi or '-').border = thin_border
        ws.cell(row=row, column=6, value=masraf.fatura_no or '-').border = thin_border
        tutar_cell = ws.cell(row=row, column=7, value=float(masraf.tl_karsiligi or 0))
        tutar_cell.border = thin_border
        tutar_cell.number_format = '#,##0.00'
        row += 1
    
    # Özet satırları
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="TOPLAM MASRAF").font = bold_font
    ws.cell(row=row, column=7, value=float(rapor.toplam_masraf or 0)).font = bold_font
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="Avans Mahsup").font = bold_font
    ws.cell(row=row, column=7, value=float(rapor.avans_tutari or 0))
    ws.cell(row=row, column=7).number_format = '#,##0.00'
    
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="NET ÖDEME").font = Font(bold=True, size=12)
    net_cell = ws.cell(row=row, column=7, value=float(rapor.net_odeme or 0))
    net_cell.font = Font(bold=True, size=12, color="006600")
    net_cell.number_format = '#,##0.00'
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # Dosyayı kaydet
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"masraf_raporu_{rapor.calisan.full_name.replace(' ', '_')}_{rapor.donem_yil}_{rapor.donem_ay:02d}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@masraf_bp.route('/rapor/<int:id>/odendi', methods=['POST'])
@login_required
@permission_required('masraf.admin')
def rapor_odendi(id):
    """Masraf raporunu ödendi olarak işaretle ve masrafları güncelle"""
    rapor = MasrafRaporu.query.get_or_404(id)
    
    if rapor.durum != 'onaylandi':
        flash('Sadece onaylanmış raporlar ödendi olarak işaretlenebilir.', 'warning')
        return redirect(url_for('masraf.rapor_detay', id=id))
    
    # Raporu ödendi yap
    rapor.durum = 'odendi'
    rapor.odeme_tarihi = date.today()
    rapor.odeme_referans = request.form.get('referans', '').strip() or None
    
    # Rapordaki tüm masrafları ödendi yap
    for masraf in rapor.masraflar:
        masraf.durum = 'odendi'
        masraf.odeme_tarihi = date.today()
    
    db.session.commit()
    
    flash(f'Rapor ve {len(rapor.masraflar)} masraf ödendi olarak işaretlendi.', 'success')
    return redirect(url_for('masraf.rapor_detay', id=id))


# ============================================================
# TOPLU ÖDEME (ADMIN)
# ============================================================

@masraf_bp.route('/toplu-odeme', methods=['GET', 'POST'])
@login_required
@permission_required('masraf.admin')
def toplu_odeme():
    """Toplu masraf ödeme - onaylanmış raporları listele ve öde"""
    
    if request.method == 'POST':
        rapor_ids = request.form.getlist('rapor_ids')
        odeme_referans = request.form.get('odeme_referans', '').strip()
        
        if not rapor_ids:
            flash('Lütfen en az bir rapor seçin.', 'warning')
            return redirect(url_for('masraf.toplu_odeme'))
        
        odenen_sayisi = 0
        for rapor_id in rapor_ids:
            rapor = MasrafRaporu.query.get(int(rapor_id))
            if rapor and rapor.durum == 'onaylandi':
                rapor.durum = 'odendi'
                rapor.odeme_tarihi = date.today()
                rapor.odeme_referans = odeme_referans
                
                # Masrafları da ödendi yap
                for masraf in rapor.masraflar:
                    masraf.durum = 'odendi'
                    masraf.odeme_tarihi = date.today()
                
                odenen_sayisi += 1
        
        db.session.commit()
        flash(f'{odenen_sayisi} rapor ödendi olarak işaretlendi.', 'success')
        return redirect(url_for('masraf.toplu_odeme'))
    
    # GET - Onaylanmış raporları listele
    raporlar = MasrafRaporu.query.filter_by(
        durum='onaylandi',
        is_deleted=False
    ).order_by(MasrafRaporu.donem_yil.desc(), MasrafRaporu.donem_ay.desc()).all()
    
    # Özet hesapla
    toplam_odeme = sum(r.net_odeme or 0 for r in raporlar)
    
    return render_template('masraf/toplu_odeme.html',
                          raporlar=raporlar,
                          toplam_odeme=toplam_odeme)
