# -*- coding: utf-8 -*-
"""
TG Portal - İK (Human Resources) Routes
Güncellenmiş versiyon: Evrak yönetimi, İşten çıkış, Aday→Çalışan dönüşümü
"""

from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import (Blueprint, render_template, redirect, url_for, flash, request,
                   current_app, send_file, jsonify, abort)
from flask_login import login_required, current_user
from app.models.ik import ZimmetTipi, Zimmet, ZimmetLog, SgkCikisKodu
from app.models.sirket import SgkDosya
from app.models.proje import HedefKadro, Proje
from app.models.base import CalisanDurumu, ListeDurumu
from werkzeug.utils import secure_filename
from sqlalchemy.exc import IntegrityError
import os
import io
import re
import zipfile

from app import db
from app.models.ik import (
    Departman, Pozisyon, Calisan, Izin, Aday,
    EvrakTipi, AdayEvrak, AdayMedya, CalisanEvrak, IstenCikis,
    IstenCikisBildirimi,
    SozlesmeSablonu, AdayIslemGecmisi, ADAY_DURUM_AKISI, EHLIYET_SINIFLARI
)
from app.models.base import CalisanDurumu
from app.utils import (
    permission_required, paginate_query,
    apply_calisan_scope, calisan_in_scope,
    apply_aday_scope, aday_in_scope, user_scoped_projeler,
    normalize_telefon,
)

ik_bp = Blueprint('ik', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# DASHBOARD
# ============================================================

@ik_bp.route('/dashboard')
@login_required
@permission_required('ik.view')
def dashboard():
    """İK Dashboard"""
    # İstatistikler
    aktif_calisan = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).count()
    bekleyen_aday = Aday.query.filter_by(is_deleted=False).filter(
        Aday.durum.in_(['basvurdu', 'inceleniyor', 'onaylandi', 'sgk_giris_talebi',
                        'sgk_girisi_yapildi', 'degerlendiriliyor', 'mulakat'])
    ).count()
    bekleyen_izin = Izin.query.filter_by(durum='beklemede').count()
    eksik_evrak_aday = 0
    
    # Eksik evraklı aday sayısı
    for aday in Aday.query.filter(Aday.is_deleted==False, Aday.durum.notin_(['red', 'iptal', 'ise_alindi'])).all():
        zorunlu = EvrakTipi.query.filter_by(zorunlu=True, aktif=True).count()
        yuklenen = aday.evraklar.join(EvrakTipi).filter(
            EvrakTipi.zorunlu == True,
            AdayEvrak.durum == 'onaylandi'
        ).count()
        if yuklenen < zorunlu:
            eksik_evrak_aday += 1
    
    # Departman bazlı dağılım
    departman_stats = db.session.query(
    Departman.ad,
    db.func.count(Calisan.id)
    ).join(Calisan, Calisan.departman_id == Departman.id).filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF
    ).group_by(Departman.ad).all()
    
    # Son başvurular
    son_adaylar = Aday.query.filter_by(is_deleted=False)\
        .order_by(Aday.created_at.desc()).limit(5).all()
    
    # İşten çıkış bekleyenler
    cikis_bekleyen = IstenCikis.query.filter(
        IstenCikis.durum.in_(['basladi', 'devam_ediyor'])
    ).count()
    
    return render_template('ik/dashboard.html',
                          aktif_calisan=aktif_calisan,
                          bekleyen_aday=bekleyen_aday,
                          bekleyen_izin=bekleyen_izin,
                          eksik_evrak_aday=eksik_evrak_aday,
                          departman_stats=departman_stats,
                          son_adaylar=son_adaylar,
                          cikis_bekleyen=cikis_bekleyen)


# ============================================================
# CANLI İŞE ALIM DASHBOARD
# ============================================================

# Aday durumlarının dashboard bucket'larına eşlenmesi (legacy değerler dahil)
_ISE_ALIM_BASVURU_DURUMLAR = ['basvurdu', 'inceleniyor', 'degerlendiriliyor', 'mulakat']
_ISE_ALIM_RED_DURUMLAR = ['red', 'reddedildi', 'aday_reddetti', 'iptal']


def _bos_ozet():
    return {
        'hedef': 0, 'basvuru': 0, 'onaylanan': 0, 'sgk_talebi': 0,
        'sgk_yapildi': 0, 'donusturuldu': 0, 'reddedilen': 0,
        'havuzda': 0, 'kadin': 0, 'erkek': 0, 'toplam_aday': 0,
    }


def _oran(part, total):
    """Yüzde hesapla (sıfıra bölme korumalı)."""
    return round((part / total) * 100, 1) if total else 0


def _aktif_aday(o):
    """Reddedilen hariç süreçteki toplam aday (başvuru..işe başlayan)."""
    return (o['basvuru'] + o['onaylanan'] + o['sgk_talebi'] +
            o['sgk_yapildi'] + o['donusturuldu'])


def _ise_alim_dashboard_data(secili_proje_id=None):
    """İşe alım dashboard veri toplama - hem sayfa hem Excel export için ortak.

    Döner: dict(genel, proje_satirlari, kadro_detay, projeler, secili_proje_id)
    """
    # Kullanıcının görebildiği projeler (scope) + dropdown için
    projeler = user_scoped_projeler()
    proje_ids = [p.id for p in projeler]

    # Aktif kadrolar (silinmemiş) - scope'a göre filtreli
    kadro_q = HedefKadro.query.filter_by(is_deleted=False).filter(
        HedefKadro.proje_id.in_(proje_ids) if proje_ids else False
    )
    # Proje filtresi - seçiliyse tüm dashboard o projeye göre filtrelenir
    if secili_proje_id:
        kadro_q = kadro_q.filter(HedefKadro.proje_id == secili_proje_id)
    kadrolar = kadro_q.all()

    # Aday sayılarını tek sorguda topla: kadro_id + durum + cinsiyet bazında
    aday_q = db.session.query(
        Aday.kadro_id,
        Aday.durum,
        Aday.cinsiyet,
        db.func.count(Aday.id),
    ).filter(
        Aday.is_deleted == False,
        Aday.kadro_id.isnot(None),
    ).group_by(Aday.kadro_id, Aday.durum, Aday.cinsiyet)
    aday_q = apply_aday_scope(aday_q)

    # kadro_id -> özet dict
    kadro_ozet = {}
    for kadro_id, durum, cinsiyet, adet in aday_q.all():
        ozet = kadro_ozet.setdefault(kadro_id, _bos_ozet())
        ozet['toplam_aday'] += adet
        if durum in _ISE_ALIM_BASVURU_DURUMLAR:
            ozet['basvuru'] += adet
        elif durum == 'onaylandi':
            ozet['onaylanan'] += adet
        elif durum == 'sgk_giris_talebi':
            ozet['sgk_talebi'] += adet
        elif durum == 'sgk_girisi_yapildi':
            ozet['sgk_yapildi'] += adet
        elif durum == 'calisana_donusturuldu':
            ozet['donusturuldu'] += adet
        elif durum == 'havuzda':
            ozet['havuzda'] += adet
        elif durum in _ISE_ALIM_RED_DURUMLAR:
            ozet['reddedilen'] += adet
        # cinsiyet (red/iptal hariç toplam kadın-erkek dağılımı)
        if cinsiyet == 'kadin':
            ozet['kadin'] += adet
        elif cinsiyet == 'erkek':
            ozet['erkek'] += adet

    # "İşe Başlayan" = kadroya bağlı aktif çalışan sayısı (durum AKTIF/IZINLI).
    # Aday tablosundaki 'calisana_donusturuldu' yerine çalışan tablosundan
    # sayılır ki proje detay sayfasıyla tutarlı olsun.
    kadro_ids = [k.id for k in kadrolar]
    calisan_sayilari = {}
    if kadro_ids:
        calisan_q = db.session.query(
            Calisan.kadro_id,
            db.func.count(Calisan.id),
        ).filter(
            Calisan.is_deleted == False,
            Calisan.kadro_id.in_(kadro_ids),
            Calisan.durum.in_([CalisanDurumu.AKTIF, CalisanDurumu.IZINLI]),
        ).group_by(Calisan.kadro_id)
        calisan_sayilari = {kid: adet for kid, adet in calisan_q.all()}

    # Kadro bazlı detay satırları + proje bazlı toplama
    kadro_detay = []
    proje_ozet = {}  # proje_id -> {ad, ozet}
    for k in kadrolar:
        o = kadro_ozet.get(k.id, _bos_ozet())
        hedef = k.hedef_sayi or 0
        # İşe başlayan: aday değil, kadrodaki aktif/izinli çalışan sayısı.
        # o['donusturuldu']'yu da güncelle ki proje ve genel toplamlar tutarlı olsun.
        o['donusturuldu'] = calisan_sayilari.get(k.id, 0)
        donusturuldu = o['donusturuldu']
        aktif_aday = _aktif_aday(o)

        kadro_detay.append({
            'id': k.id,
            'proje_id': k.proje_id,
            'proje_ad': k.proje.ad if k.proje else '-',
            'baslik': k.full_title,
            'il': (k.il if k.il else k.il_adi) or '-',
            'hedef': hedef,
            'aktif_aday': aktif_aday,
            'toplam_aday': o['toplam_aday'],
            'basvuru': o['basvuru'],
            'onaylanan': o['onaylanan'],
            'sgk_talebi': o['sgk_talebi'],
            'sgk_yapildi': o['sgk_yapildi'],
            'donusturuldu': donusturuldu,
            'reddedilen': o['reddedilen'],
            'havuzda': o['havuzda'],
            'kadin': o['kadin'],
            'erkek': o['erkek'],
            'kadin_oran': _oran(o['kadin'], o['kadin'] + o['erkek']),
            'erkek_oran': _oran(o['erkek'], o['kadin'] + o['erkek']),
            'doluluk_calisan': _oran(donusturuldu, hedef),
            'doluluk_aday': _oran(aktif_aday, hedef),
        })

        # Proje toplama
        po = proje_ozet.setdefault(k.proje_id, {
            'ad': k.proje.ad if k.proje else '-',
            'ozet': _bos_ozet(),
        })
        for key in ('hedef',):
            po['ozet']['hedef'] += hedef
        for key in ('basvuru', 'onaylanan', 'sgk_talebi', 'sgk_yapildi',
                    'donusturuldu', 'reddedilen', 'havuzda', 'kadin', 'erkek', 'toplam_aday'):
            po['ozet'][key] += o[key]

    # Proje özet listesi (doluluk hesaplı)
    proje_satirlari = []
    for pid, pdata in proje_ozet.items():
        oz = pdata['ozet']
        aktif_aday = _aktif_aday(oz)
        proje_satirlari.append({
            'id': pid,
            'ad': pdata['ad'],
            'hedef': oz['hedef'],
            'aktif_aday': aktif_aday,
            'toplam_aday': oz['toplam_aday'],
            'reddedilen': oz['reddedilen'],
            'basvuru': oz['basvuru'],
            'onaylanan': oz['onaylanan'],
            'sgk_talebi': oz['sgk_talebi'],
            'sgk_yapildi': oz['sgk_yapildi'],
            'donusturuldu': oz['donusturuldu'],
            'kadin': oz['kadin'],
            'erkek': oz['erkek'],
            'kadin_oran': _oran(oz['kadin'], oz['kadin'] + oz['erkek']),
            'erkek_oran': _oran(oz['erkek'], oz['kadin'] + oz['erkek']),
            'doluluk_calisan': _oran(oz['donusturuldu'], oz['hedef']),
            'doluluk_aday': _oran(aktif_aday, oz['hedef']),
        })
    proje_satirlari.sort(key=lambda x: x['ad'])

    # Genel özet kartları (tüm scope - proje filtresinden bağımsız)
    genel = _bos_ozet()
    for d in kadro_detay:
        genel['hedef'] += d['hedef']
        genel['basvuru'] += d['basvuru']
        genel['onaylanan'] += d['onaylanan']
        genel['sgk_talebi'] += d['sgk_talebi']
        genel['sgk_yapildi'] += d['sgk_yapildi']
        genel['donusturuldu'] += d['donusturuldu']
        genel['reddedilen'] += d['reddedilen']
        genel['havuzda'] += d['havuzda']
        genel['kadin'] += d['kadin']
        genel['erkek'] += d['erkek']
    # Toplam başvuru = scope'taki aktif kadrolara bağlı tüm adaylar
    genel['toplam_basvuru'] = sum(kadro_ozet.get(k.id, _bos_ozet())['toplam_aday'] for k in kadrolar)
    genel['aktif_aday'] = _aktif_aday(genel)
    genel['doluluk'] = _oran(genel['donusturuldu'], genel['hedef'])
    genel['doluluk_aday'] = _oran(genel['aktif_aday'], genel['hedef'])
    genel['kadin_oran'] = _oran(genel['kadin'], genel['kadin'] + genel['erkek'])
    genel['erkek_oran'] = _oran(genel['erkek'], genel['kadin'] + genel['erkek'])

    kadro_detay.sort(key=lambda x: (x['proje_ad'], x['baslik']))

    return {
        'genel': genel,
        'proje_satirlari': proje_satirlari,
        'kadro_detay': kadro_detay,
        'projeler': projeler,
        'secili_proje_id': secili_proje_id,
    }


@ik_bp.route('/ise-alim-dashboard')
@login_required
@permission_required('ik.view')
def ise_alim_dashboard():
    """Canlı işe alım dashboard'ı - kadro/proje bazlı başvuru ve doluluk özeti."""
    secili_proje_id = request.args.get('proje', type=int)
    data = _ise_alim_dashboard_data(secili_proje_id)
    return render_template('ik/ise_alim_dashboard.html', active='ik-ise-alim', **data)


@ik_bp.route('/ise-alim-dashboard/export')
@login_required
@permission_required('ik.view')
def ise_alim_dashboard_export():
    """İşe alım dashboard'ını 3 sheet'li Excel olarak indir (seçili proje filtresine göre)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import Response

    secili_proje_id = request.args.get('proje', type=int)
    data = _ise_alim_dashboard_data(secili_proje_id)
    genel = data['genel']
    proje_satirlari = data['proje_satirlari']
    kadro_detay = data['kadro_detay']

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    baslik_font = Font(bold=True, size=12)

    def _stil_basliklar(ws, satir=1):
        for cell in ws[satir]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _genislikler(ws, widths):
        for idx, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    wb = Workbook()

    # ---- Sheet 1: Genel Özet ----
    ws1 = wb.active
    ws1.title = 'Genel Özet'
    ws1['A1'] = 'İşe Alım Genel Özet'
    ws1['A1'].font = baslik_font
    ws1.append([])
    ws1.append(['Metrik', 'Değer'])
    _stil_basliklar(ws1, 3)
    ozet_satirlar = [
        ('Toplam Kadro Hedefi', genel['hedef']),
        ('Toplam Aday', genel.get('toplam_basvuru', 0)),
        ('Onaylanan', genel['onaylanan']),
        ('SGK Bekleyen', genel['sgk_talebi']),
        ('İşe Başlayan', genel['donusturuldu']),
        ('Doluluk (Çalışan) %', genel['doluluk']),
        ('Doluluk (Aday) %', genel['doluluk_aday']),
        ('Kadın Sayısı', genel['kadin']),
        ('Kadın %', genel['kadin_oran']),
        ('Erkek Sayısı', genel['erkek']),
        ('Erkek %', genel['erkek_oran']),
    ]
    for ad, deger in ozet_satirlar:
        ws1.append([ad, deger])
    _genislikler(ws1, [28, 16])

    # ---- Sheet 2: Proje Bazlı ----
    ws2 = wb.create_sheet('Proje Bazlı')
    proje_headers = ['Proje Adı', 'Hedef', 'Toplam Aday', 'Başvuru', 'Onaylanan',
                     'SGK Talebi', 'SGK Yapıldı', 'İşe Başlayan', 'Reddedilen',
                     'Kadın', 'Kadın %', 'Erkek', 'Erkek %',
                     'Doluluk (Çalışan) %', 'Doluluk (Aday) %']
    ws2.append(proje_headers)
    _stil_basliklar(ws2)
    for p in proje_satirlari:
        ws2.append([
            p['ad'], p['hedef'], p['toplam_aday'],
            p['basvuru'], p['onaylanan'], p['sgk_talebi'], p['sgk_yapildi'],
            p['donusturuldu'], p['reddedilen'],
            p['kadin'], p['kadin_oran'], p['erkek'], p['erkek_oran'],
            p['doluluk_calisan'], p['doluluk_aday'],
        ])
    _genislikler(ws2, [24, 8, 12, 10, 11, 11, 12, 13, 11, 8, 9, 8, 9, 18, 16])

    # ---- Sheet 3: Kadro Bazlı ----
    ws3 = wb.create_sheet('Kadro Bazlı')
    kadro_headers = ['Proje', 'Kadro', 'İl', 'Hedef', 'Toplam Aday', 'Başvuru',
                     'Onaylanan', 'SGK Talebi', 'SGK Yapıldı', 'İşe Başlayan',
                     'Reddedilen', 'Kadın', 'Kadın %', 'Erkek', 'Erkek %',
                     'Doluluk (Çalışan) %', 'Doluluk (Aday) %']
    ws3.append(kadro_headers)
    _stil_basliklar(ws3)
    for d in kadro_detay:
        ws3.append([
            d['proje_ad'], d['baslik'], d['il'], d['hedef'], d['toplam_aday'],
            d['basvuru'], d['onaylanan'], d['sgk_talebi'], d['sgk_yapildi'],
            d['donusturuldu'], d['reddedilen'],
            d['kadin'], d['kadin_oran'], d['erkek'], d['erkek_oran'],
            d['doluluk_calisan'], d['doluluk_aday'],
        ])
    _genislikler(ws3, [22, 26, 12, 8, 12, 10, 11, 11, 12, 13, 11, 8, 9, 8, 9, 18, 16])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ise_alim_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ============================================================
# ÇALIŞAN YÖNETİMİ
# ============================================================

def _calisan_liste_query():
    """Çalışan listesi query builder - liste ve export için ortak filtre mantığı"""
    departman_id = request.args.get('departman_id', type=int)
    proje_id = request.args.get('proje_id', type=int)
    durum = request.args.get('durum')
    ehliyet = request.args.get('ehliyet', '').strip()
    search = request.args.get('search', '').strip()

    query = Calisan.query.filter_by(is_deleted=False)
    query = apply_calisan_scope(query)

    if departman_id:
        query = query.filter(Calisan.departman_id == departman_id)
    if proje_id:
        query = query.join(HedefKadro, Calisan.kadro_id == HedefKadro.id).filter(HedefKadro.proje_id == proje_id)
    if durum:
        query = query.filter(Calisan.durum == CalisanDurumu(durum))
    if ehliyet == 'var':
        query = query.filter(Calisan.ehliyet_sinifi.isnot(None), Calisan.ehliyet_sinifi != '')
    elif ehliyet == 'yok':
        query = query.filter(db.or_(Calisan.ehliyet_sinifi.is_(None), Calisan.ehliyet_sinifi == ''))
    if search:
        search_filter = f'%{search}%'
        query = query.filter(
            db.or_(
                Calisan.ad.ilike(search_filter),
                Calisan.soyad.ilike(search_filter),
                Calisan.sicil_no.ilike(search_filter),
                Calisan.email.ilike(search_filter)
            )
        )

    return query.order_by(Calisan.ad, Calisan.soyad)


@ik_bp.route('/')
@login_required
@permission_required('ik.view')
def liste():
    """Çalışan listesi"""
    page = request.args.get('page', 1, type=int)
    query = _calisan_liste_query()
    pagination = paginate_query(query, page, 20)

    departmanlar = Departman.query.filter_by(aktif=True).order_by(Departman.ad).all()
    projeler = Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all()

    return render_template('ik/liste.html',
                          calisanlar=pagination.items,
                          pagination=pagination,
                          departmanlar=departmanlar,
                          projeler=projeler,
                          durumlar=CalisanDurumu)


@ik_bp.route('/export')
@login_required
@permission_required('ik.view')
def calisanlar_export():
    """Filtrelenmiş çalışan listesini Excel olarak indir"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    calisanlar = _calisan_liste_query().all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Çalışanlar'

    headers = ['Sicil No', 'Ad Soyad', 'TC Kimlik', 'Proje', 'Kadro', 'Pozisyon',
               'Departman', 'Durum', 'İşe Başlama', 'Telefon', 'Email', 'IBAN',
               'Ehliyet']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    durum_etiket = {
        'aktif': 'Aktif', 'izinli': 'İzinli', 'ayrildi': 'Ayrıldı',
        'askiya_alindi': 'Askıda', 'aday': 'Aday',
    }

    for c in calisanlar:
        ws.append([
            c.sicil_no or '',
            f'{c.ad} {c.soyad}',
            c.tc_kimlik or '',
            c.kadro.proje.ad if c.kadro and c.kadro.proje else '',
            c.kadro.pozisyon_adi if c.kadro else '',
            c.pozisyon.ad if c.pozisyon else '',
            c.departman.ad if c.departman else '',
            durum_etiket.get(c.durum.value if c.durum else '', c.durum.value if c.durum else ''),
            c.ise_baslama.strftime('%d.%m.%Y') if c.ise_baslama else '',
            c.telefon or '',
            c.email or '',
            c.iban or '',
            c.ehliyet_sinifi or 'Yok',
        ])

    widths = [12, 28, 13, 22, 22, 22, 20, 12, 14, 16, 28, 28, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"calisanlar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@ik_bp.route('/<int:id>')
@login_required
@permission_required('ik.view')
def detay(id):
    """Çalışan detay sayfası"""
    calisan = Calisan.query.get_or_404(id)

    if not calisan_in_scope(calisan):
        flash('Bu çalışanı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    izinler = calisan.izinler.order_by(Izin.baslangic.desc()).limit(10).all()
    evraklar = calisan.evraklar.all() if hasattr(calisan, 'evraklar') else []

    # Adayken yüklenen bilgiler - bağlı Aday kaydı (varsa)
    aday = Aday.query.filter_by(calisan_id=calisan.id, is_deleted=False).first()
    aday_gecmis = aday.islem_gecmisi.all() if aday else []
    aday_evraklar = aday.evraklar.all() if aday else []

    # Tekrar işe alım modalı için aktif kadrolar (yalnızca gerekince yükle)
    kadrolar = []
    if calisan.tekrar_ise_alinabilir:
        kadrolar = (HedefKadro.query
                    .filter_by(is_deleted=False, aktif=True)
                    .join(Proje, HedefKadro.proje_id == Proje.id)
                    .order_by(Proje.ad, HedefKadro.pozisyon_adi).all())

    # İşten çıkış bildirimleri (SPV/koordinatör ön bildirimleri)
    cikis_bildirimleri = calisan.cikis_bildirimleri.all()

    return render_template('ik/detay.html',
                          calisan=calisan,
                          izinler=izinler,
                          evraklar=evraklar,
                          aday=aday,
                          aday_gecmis=aday_gecmis,
                          aday_evraklar=aday_evraklar,
                          cikis_bildirimleri=cikis_bildirimleri,
                          cikis_nedenleri=IstenCikisBildirimi.CIKIS_NEDENLERI,
                          kadrolar=kadrolar)


@ik_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.create')
def ekle():
    """Yeni çalışan ekle"""
    if request.method == 'POST':
        tc = request.form.get('tc_kimlik', '').strip()
        if tc and Calisan.query.filter_by(tc_kimlik=tc, is_deleted=False).first():
            flash('Bu TC Kimlik numarası zaten kayıtlı.', 'danger')
            return redirect(url_for('ik.ekle'))

        telefon_ham = request.form.get('telefon', '').strip()
        telefon = normalize_telefon(telefon_ham)
        if telefon_ham and not telefon:
            flash('Geçersiz telefon numarası. Örnek format: 05XX XXX XX XX', 'danger')
            return redirect(url_for('ik.ekle'))

        calisan = Calisan(
            sicil_no=None if request.form.get('sicil_no', '').strip() in ('', 'None', 'none') else request.form.get('sicil_no', '').strip(),
            ad=request.form.get('ad', '').strip(),
            soyad=request.form.get('soyad', '').strip(),
            tc_kimlik=tc or None,
            dogum_tarihi=datetime.strptime(request.form.get('dogum_tarihi'), '%Y-%m-%d').date() if request.form.get('dogum_tarihi') else None,
            cinsiyet=request.form.get('cinsiyet') or None,
            email=request.form.get('email', '').strip() or None,
            telefon=telefon,
            adres=request.form.get('adres', '').strip() or None,
            il=request.form.get('il', '').strip() or None,
            ilce=request.form.get('ilce', '').strip() or None,
            departman_id=int(request.form.get('departman_id')) if request.form.get('departman_id') else None,
            pozisyon_id=int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None,
            sgk_dosya_id=int(request.form.get('sgk_dosya_id')) if request.form.get('sgk_dosya_id') else None,
            kadro_id=int(request.form.get('kadro_id')) if request.form.get('kadro_id') else None,
            ise_baslama=datetime.strptime(request.form.get('ise_baslama'), '%Y-%m-%d').date() if request.form.get('ise_baslama') else None,
            calisma_tipi=request.form.get('calisma_tipi') or None,
            durum=CalisanDurumu(request.form.get('durum')) if request.form.get('durum') else CalisanDurumu.AKTIF,
            notlar=request.form.get('notlar', '').strip() or None,
            yonetici_id=int(request.form.get('yonetici_id')) if request.form.get('yonetici_id') else None,
            ehliyet_sinifi=request.form.get('ehliyet_sinifi', '').strip() or None,
            created_by=current_user.id
        )
        
        # Sözleşme bilgileri
        calisan.sozlesme_sablon_id = int(request.form.get('sozlesme_sablon_id')) if request.form.get('sozlesme_sablon_id', '').strip() else None
        calisan.sozlesme_baslangic = datetime.strptime(request.form.get('sozlesme_baslangic'), '%Y-%m-%d').date() if request.form.get('sozlesme_baslangic', '').strip() else None
        calisan.sozlesme_bitis = datetime.strptime(request.form.get('sozlesme_bitis'), '%Y-%m-%d').date() if request.form.get('sozlesme_bitis', '').strip() else None

        db.session.add(calisan)
        db.session.commit()
        print(f">>> CALISAN KAYDEDILDI id={calisan.id} {calisan.ad} {calisan.soyad}", flush=True)

        # Sözleşme PDF yükleme
        if 'sozlesme_pdf' in request.files:
            spdf = request.files['sozlesme_pdf']
            if spdf and spdf.filename:
                import uuid
                ext = spdf.filename.rsplit('.', 1)[-1].lower() if '.' in spdf.filename else 'pdf'
                filename = f"sozlesme_{calisan.id}_{uuid.uuid4().hex[:8]}.{ext}"
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesmeler')
                os.makedirs(upload_folder, exist_ok=True)
                spdf.save(os.path.join(upload_folder, filename))
                calisan.sozlesme_pdf = f"sozlesmeler/{filename}"
                db.session.commit()

        # Fotoğraf yükleme
        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and foto.filename:
                from werkzeug.utils import secure_filename
                import uuid

                ext = foto.filename.rsplit('.', 1)[-1].lower() if '.' in foto.filename else 'jpg'
                filename = f"calisan_{calisan.id}_{uuid.uuid4().hex[:8]}.{ext}"

                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos')
                os.makedirs(upload_folder, exist_ok=True)

                filepath = os.path.join(upload_folder, filename)
                foto.save(filepath)
                calisan.foto = f"photos/{filename}"
                db.session.commit()

        # İşe giriş bildirimi gönder
        print(f">>> NOTIFY_ISE_GIRIS CAGRILIYOR: {calisan.ad} {calisan.soyad}", flush=True)
        try:
            from app.services.notification import notify_ise_giris
            sonuc = notify_ise_giris(calisan)
            print(f">>> NOTIFY SONUC: {sonuc}", flush=True)
        except Exception as e:
            print(f">>> NOTIFY HATA: {e}", flush=True)
            import traceback
            traceback.print_exc()

        flash(f'{calisan.full_name} çalışanı oluşturuldu.', 'success')
        return redirect(url_for('ik.detay', id=calisan.id))

    departmanlar = Departman.query.filter_by(aktif=True).order_by(Departman.ad).all()
    pozisyonlar = Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all()
    yoneticiler = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).order_by(Calisan.ad).all()
    
    sgk_dosyalari = SgkDosya.query.filter_by(is_deleted=False, aktif=True).all()
    kadrolar = HedefKadro.query.filter_by(is_deleted=False, aktif=True).all()
    try:
        sablonlar = SozlesmeSablonu.query.filter_by(aktif=True, is_deleted=False).order_by(SozlesmeSablonu.ad).all()
    except Exception:
        db.session.rollback()
        sablonlar = []
    return render_template('ik/form.html',
                          calisan=None,
                          departmanlar=departmanlar,
                          pozisyonlar=pozisyonlar,
                          yoneticiler=yoneticiler,
                          sgk_dosyalari=sgk_dosyalari,
                          kadrolar=kadrolar,
                          sablonlar=sablonlar,
                          ehliyet_siniflari=EHLIYET_SINIFLARI,
                          durumlar=CalisanDurumu)


@ik_bp.route('/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def duzenle(id):
    """Çalışan düzenle"""
    calisan = Calisan.query.get_or_404(id)

    if not calisan_in_scope(calisan):
        flash('Bu çalışanı düzenleme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    if request.method == 'POST':
        tc = request.form.get('tc_kimlik', '').strip()
        if tc:
            existing = Calisan.query.filter(
                Calisan.tc_kimlik == tc,
                Calisan.id != id,
                Calisan.is_deleted == False
            ).first()
            if existing:
                flash('Bu TC Kimlik numarası başka bir çalışanda kayıtlı.', 'danger')
                return redirect(url_for('ik.duzenle', id=id))

        telefon_ham = request.form.get('telefon', '').strip()
        telefon = normalize_telefon(telefon_ham)
        if telefon_ham and not telefon:
            flash('Geçersiz telefon numarası. Örnek format: 05XX XXX XX XX', 'danger')
            return redirect(url_for('ik.duzenle', id=id))

        calisan.sicil_no = None if request.form.get('sicil_no', '').strip() in ('', 'None', 'none') else request.form.get('sicil_no', '').strip()
        calisan.ad = request.form.get('ad', '').strip()
        calisan.soyad = request.form.get('soyad', '').strip()
        calisan.tc_kimlik = tc or None
        calisan.dogum_tarihi = datetime.strptime(request.form.get('dogum_tarihi'), '%Y-%m-%d').date() if request.form.get('dogum_tarihi') else None
        calisan.cinsiyet = request.form.get('cinsiyet') or None
        calisan.email = request.form.get('email', '').strip() or None
        calisan.telefon = telefon
        calisan.iban = request.form.get('iban', '').strip().replace(' ', '').upper() or None
        calisan.adres = request.form.get('adres', '').strip() or None
        calisan.il = request.form.get('il', '').strip() or None
        calisan.ilce = request.form.get('ilce', '').strip() or None
        calisan.departman_id = int(request.form.get('departman_id')) if request.form.get('departman_id') else None
        calisan.pozisyon_id = int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None
        calisan.yonetici_id = int(request.form.get('yonetici_id')) if request.form.get('yonetici_id') else None
        calisan.ise_baslama = datetime.strptime(request.form.get('ise_baslama'), '%Y-%m-%d').date() if request.form.get('ise_baslama') else None
        calisan.sgk_dosya_id = int(request.form.get('sgk_dosya_id')) if request.form.get('sgk_dosya_id') else None
        calisan.kadro_id = int(request.form.get('kadro_id')) if request.form.get('kadro_id') else None
        calisan.calisma_tipi = request.form.get('calisma_tipi') or None
        calisan.durum = CalisanDurumu(request.form.get('durum')) if request.form.get('durum') else CalisanDurumu.AKTIF
        calisan.notlar = request.form.get('notlar', '').strip() or None
        calisan.ehliyet_sinifi = request.form.get('ehliyet_sinifi', '').strip() or None
        calisan.updated_by = current_user.id

        # Sözleşme bilgileri
        calisan.sozlesme_sablon_id = int(request.form.get('sozlesme_sablon_id')) if request.form.get('sozlesme_sablon_id', '').strip() else None
        calisan.sozlesme_baslangic = datetime.strptime(request.form.get('sozlesme_baslangic'), '%Y-%m-%d').date() if request.form.get('sozlesme_baslangic', '').strip() else None
        calisan.sozlesme_bitis = datetime.strptime(request.form.get('sozlesme_bitis'), '%Y-%m-%d').date() if request.form.get('sozlesme_bitis', '').strip() else None

        # Sözleşme PDF yükleme
        if 'sozlesme_pdf' in request.files:
            spdf = request.files['sozlesme_pdf']
            if spdf and spdf.filename:
                import uuid
                ext = spdf.filename.rsplit('.', 1)[-1].lower() if '.' in spdf.filename else 'pdf'
                filename = f"sozlesme_{calisan.id}_{uuid.uuid4().hex[:8]}.{ext}"
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesmeler')
                os.makedirs(upload_folder, exist_ok=True)
                spdf.save(os.path.join(upload_folder, filename))
                calisan.sozlesme_pdf = f"sozlesmeler/{filename}"

        # Fotoğraf yükleme
        if 'foto' in request.files:
            foto = request.files['foto']
            if foto and foto.filename:
                from werkzeug.utils import secure_filename
                import uuid
                
                # Dosya adı oluştur
                ext = foto.filename.rsplit('.', 1)[-1].lower() if '.' in foto.filename else 'jpg'
                filename = f"calisan_{calisan.id}_{uuid.uuid4().hex[:8]}.{ext}"
                
                # Klasör oluştur
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'photos')
                os.makedirs(upload_folder, exist_ok=True)
                
                # Eski fotoğrafı sil
                if calisan.foto:
                    old_path = os.path.join(current_app.config['UPLOAD_FOLDER'], calisan.foto)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                
                # Yeni fotoğrafı kaydet
                filepath = os.path.join(upload_folder, filename)
                foto.save(filepath)
                calisan.foto = f"photos/{filename}"
        
        db.session.commit()
        
        flash('Çalışan güncellendi.', 'success')
        return redirect(url_for('ik.detay', id=id))
    
    departmanlar = Departman.query.filter_by(aktif=True).order_by(Departman.ad).all()
    pozisyonlar = Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all()
    yoneticiler = Calisan.query.filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        Calisan.id != calisan.id
    ).order_by(Calisan.ad).all()
    
    sgk_dosyalari = SgkDosya.query.filter_by(is_deleted=False, aktif=True).all()
    kadrolar = HedefKadro.query.filter_by(is_deleted=False, aktif=True).all()
    try:
        sablonlar = SozlesmeSablonu.query.filter_by(aktif=True, is_deleted=False).order_by(SozlesmeSablonu.ad).all()
    except Exception:
        db.session.rollback()
        sablonlar = []
    return render_template('ik/form.html',
                          calisan=calisan,
                          departmanlar=departmanlar,
                          pozisyonlar=pozisyonlar,
                          sgk_dosyalari=sgk_dosyalari,
                          kadrolar=kadrolar,
                          sablonlar=sablonlar,
                          yoneticiler=yoneticiler,
                          ehliyet_siniflari=EHLIYET_SINIFLARI,
                          durumlar=CalisanDurumu)


# ============================================================
# ADAY YÖNETİMİ
# ============================================================

def _aday_liste_query():
    """Aday listesi query builder - liste ve export icin ortak filtre mantigi"""
    durum = request.args.get('durum')
    kaynak = request.args.get('kaynak')
    cinsiyet = request.args.get('cinsiyet')
    proje_id = request.args.get('proje_id', type=int)
    il = request.args.get('il', '').strip()
    ilce = request.args.get('ilce', '').strip()
    ehliyet = request.args.get('ehliyet', '').strip()
    search = request.args.get('search', '').strip()
    iletisim = request.args.get('iletisim', '').strip()

    query = Aday.query.filter_by(is_deleted=False)
    query = apply_aday_scope(query)

    if iletisim:
        query = _iletisim_filtre_uygula(query, iletisim)

    if durum:
        query = query.filter(Aday.durum == durum)
    if kaynak:
        query = query.filter(Aday.kaynak == kaynak)
    if cinsiyet:
        query = query.filter(Aday.cinsiyet == cinsiyet)
    if proje_id:
        query = query.join(HedefKadro, Aday.kadro_id == HedefKadro.id).filter(HedefKadro.proje_id == proje_id)
    if il:
        query = query.filter(Aday.il == il)
    if ilce:
        query = query.filter(Aday.ilce == ilce)
    # Ehliyet: eski kayıtlarda yalnızca ehliyet_var işaretli olabilir,
    # yeni kayıtlarda sınıf da doluyor — ikisi birlikte değerlendirilir.
    _ehliyet_var_kosul = db.or_(
        Aday.ehliyet_var.is_(True),
        db.and_(Aday.ehliyet_sinifi.isnot(None), Aday.ehliyet_sinifi != ''),
    )
    if ehliyet == 'var':
        query = query.filter(_ehliyet_var_kosul)
    elif ehliyet == 'yok':
        query = query.filter(db.not_(_ehliyet_var_kosul))
    if search:
        search_filter = f'%{search}%'
        query = query.filter(
            db.or_(
                Aday.ad.ilike(search_filter),
                Aday.soyad.ilike(search_filter),
                Aday.tc_kimlik.ilike(search_filter),
                Aday.telefon.ilike(search_filter),
                Aday.email.ilike(search_filter)
            )
        )

    return query.order_by(Aday.created_at.desc())


@ik_bp.route('/adaylar')
@login_required
@permission_required('ik.view')
def aday_liste():
    """Aday listesi"""
    page = request.args.get('page', 1, type=int)
    query = _aday_liste_query()
    pagination = paginate_query(query, page, 20)

    # "Son İletişim" kolonu için toplu yükleme (sadece bu sayfadaki adaylar)
    son_iletisim = _aday_son_iletisim([a.id for a in pagination.items])

    # İstatistikler (scope'a gore)
    scoped_base = apply_aday_scope(Aday.query.filter_by(is_deleted=False))
    stats = {
        'toplam': scoped_base.count(),
        'basvurdu': scoped_base.filter(Aday.durum == 'basvurdu').count(),
        'degerlendiriliyor': scoped_base.filter(Aday.durum == 'degerlendiriliyor').count(),
        'mulakat': scoped_base.filter(Aday.durum == 'mulakat').count(),
        'teklif': scoped_base.filter(Aday.durum == 'teklif').count(),
        'ise_alindi': scoped_base.filter(Aday.durum == 'ise_alindi').count(),
    }

    projeler = user_scoped_projeler()

    # İl / İlçe filtre seçenekleri (scope'a göre, dolu olan adaylardan)
    il_ilce_rows = apply_aday_scope(
        Aday.query.filter(Aday.is_deleted == False, Aday.il.isnot(None), Aday.il != '')
    ).with_entities(Aday.il, Aday.ilce).distinct().all()
    iller = sorted({r.il for r in il_ilce_rows if r.il})
    il_ilce_map = {}
    for r in il_ilce_rows:
        if r.il and r.ilce:
            il_ilce_map.setdefault(r.il, set()).add(r.ilce)
    il_ilce_map = {k: sorted(v) for k, v in il_ilce_map.items()}

    return render_template('ik/aday_liste.html',
                          adaylar=pagination.items,
                          pagination=pagination,
                          stats=stats,
                          projeler=projeler,
                          iller=iller,
                          il_ilce_map=il_ilce_map,
                          son_iletisim=son_iletisim)


def _aday_ehliyet_text(aday):
    """Aday ehliyet bilgisini tek metne indirger: 'B', 'Var' veya 'Yok'.

    Eski kayıtlarda sınıf girilmeden yalnızca ehliyet_var işaretlenmiş olabilir.
    """
    if aday.ehliyet_sinifi:
        return aday.ehliyet_sinifi
    return 'Var' if aday.ehliyet_var else 'Yok'


def _aday_org_bilgisi(aday):
    """Adayın kadrosundan (Direktörlük, Müdürlük) bilgisini çıkarır.
    Önce HedefKadro FK alanları, yoksa pozisyon_adi içinden parse edilir.
    Örn: "Trakya Md. - Tekirdağ - P.T Sniper" -> Müdürlük: "Trakya Md."
    """
    direktorluk = ''
    mudurluk = ''
    kadro = aday.kadro
    if not kadro:
        return direktorluk, mudurluk

    if kadro.direktorluk:
        direktorluk = kadro.direktorluk.ad
    if kadro.mudurluk:
        mudurluk = kadro.mudurluk.ad

    # FK yoksa pozisyon_adi içinden müdürlük bilgisini parse et
    if not mudurluk and kadro.pozisyon_adi:
        for part in kadro.pozisyon_adi.split(' - '):
            p = part.strip()
            if 'Md.' in p or 'Müdürlü' in p or 'Mudurlu' in p:
                mudurluk = p
                break

    # Direktörlük hâlâ boşsa projenin müşterisinden al
    if not direktorluk and kadro.proje and kadro.proje.musteri:
        direktorluk = kadro.proje.musteri.display_name or ''

    return direktorluk, mudurluk


@ik_bp.route('/adaylar/export')
@login_required
@permission_required('ik.view')
def adaylar_export():
    """Filtrelenmis aday listesini Excel olarak indir"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    adaylar = _aday_liste_query().all()
    son_iletisim = _aday_son_iletisim([a.id for a in adaylar])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Adaylar'

    headers = ['Ad Soyad', 'TC Kimlik', 'IBAN', 'Cinsiyet', 'Telefon', 'Email', 'Başvuru Tarihi',
               'Proje', 'Direktörlük', 'Müdürlük', 'Kadro/Pozisyon', 'Durum',
               'Planlı Başlangıç Tarihi',
               'İl', 'İlçe', 'Üst Beden', 'Alt Beden', 'Ayakkabı No',
               'Kargo Şubesi', 'TG\'de Çalıştı', 'Seyahat Engeli',
               'Askerlik', 'Ehliyet', 'Başvuru Kaynağı',
               'Son İletişim Tarihi', 'Son İletişim Tipi', 'Arayan']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    durum_etiket = {
        'davet_gonderildi': 'Davet Gönderildi',
        'kvkk_bekleniyor': 'KVKK Bekleniyor',
        'form_bekleniyor': 'Form Bekleniyor',
        'basvurdu': 'Başvurdu',
        'degerlendiriliyor': 'Değerlendiriliyor',
        'mulakat': 'Mülakat',
        'teklif': 'Teklif',
        'ise_alindi': 'İşe Alındı',
        'red': 'Reddedildi',
        'reddedildi': 'Reddedildi',
        'havuzda': 'Havuzda',
        'aday_reddetti': 'Aday Reddetti',
        'iptal': 'İptal',
    }

    cinsiyet_etiket = {'erkek': 'Erkek', 'kadin': 'Kadın'}

    for a in adaylar:
        proje_ad = a.kadro.proje.ad if a.kadro and a.kadro.proje else ''
        if a.kadro and a.kadro.pozisyon_adi:
            kadro_pozisyon = a.kadro.pozisyon_adi
        elif a.pozisyon:
            kadro_pozisyon = a.pozisyon.ad
        else:
            kadro_pozisyon = ''
        direktorluk, mudurluk = _aday_org_bilgisi(a)
        son_log = son_iletisim.get(a.id)
        son_iletisim_tarihi = son_log.created_at.strftime('%d.%m.%Y %H:%M') if son_log else ''
        son_iletisim_tipi = son_log.islem_etiket if son_log else ''
        arayan = (son_log.kullanici.full_name if son_log and son_log.kullanici else '') if son_log else ''
        ws.append([
            f'{a.ad or ""} {a.soyad or ""}'.strip(),
            a.tc_kimlik or '',
            a.iban or '',
            cinsiyet_etiket.get(a.cinsiyet, a.cinsiyet or ''),
            a.telefon or '',
            a.email or '',
            a.basvuru_tarihi.strftime('%d.%m.%Y') if a.basvuru_tarihi else '',
            proje_ad,
            direktorluk,
            mudurluk,
            kadro_pozisyon,
            durum_etiket.get(a.durum, a.durum or ''),
            a.planlanan_baslangic.strftime('%d.%m.%Y') if a.planlanan_baslangic else '',
            a.il or '',
            a.ilce or '',
            a.ust_beden or '',
            a.alt_beden or '',
            a.ayakkabi_no or '',
            a.kargo_subesi or '',
            'Evet' if a.tg_calistimi else 'Hayır',
            'Var' if a.seyahat_engeli else 'Yok',
            a.askerlik_durumu or '',
            _aday_ehliyet_text(a),
            a.basvuru_kaynak_text if a.basvuru_kaynak else '',
            son_iletisim_tarihi,
            son_iletisim_tipi,
            arayan,
        ])

    widths = [28, 13, 28, 9, 16, 28, 14, 22, 22, 22, 24, 18, 18,
              14, 14, 10, 10, 10, 26, 12, 14, 12, 10, 18,
              18, 18, 20]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"adaylar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@ik_bp.route('/aday/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def aday_duzenle(id):
    """Aday düzenle"""
    aday = Aday.query.get_or_404(id)

    if not aday_in_scope(aday):
        flash('Bu adayı düzenleme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if request.method == 'POST':
        telefon_ham = request.form.get('telefon', '').strip()
        telefon = normalize_telefon(telefon_ham)
        if telefon_ham and not telefon:
            flash('Geçersiz telefon numarası. Örnek format: 05XX XXX XX XX', 'danger')
            return redirect(url_for('ik.aday_duzenle', id=id))
        if telefon:
            existing = Aday.query.filter(
                Aday.telefon == telefon,
                Aday.id != id,
                Aday.is_deleted == False
            ).first()
            if existing:
                flash('Bu telefon numarası başka bir adayda kayıtlı.', 'danger')
                return redirect(url_for('ik.aday_duzenle', id=id))

        aday.ad = request.form.get('ad', '').strip()
        aday.soyad = request.form.get('soyad', '').strip()
        aday.telefon = telefon
        aday.email = request.form.get('email', '').strip() or None
        aday.iban = request.form.get('iban', '').strip().replace(' ', '').upper() or None
        aday.tc_kimlik = request.form.get('tc_kimlik', '').strip() or None
        aday.dogum_tarihi = datetime.strptime(request.form.get('dogum_tarihi'), '%Y-%m-%d').date() if request.form.get('dogum_tarihi') else None
        aday.cinsiyet = request.form.get('cinsiyet') or None
        _ehliyet_sec = request.form.get('ehliyet_sinifi', '').strip()
        # 'VAR' = ehliyeti var ama sınıfı bilinmiyor (başvuru formunda sınıf
        # seçmeden işaretlemiş adaylar); boş = ehliyet yok
        aday.ehliyet_var = bool(_ehliyet_sec)
        aday.ehliyet_sinifi = None if _ehliyet_sec in ('', 'VAR') else _ehliyet_sec
        aday.adres = request.form.get('adres', '').strip() or None
        aday.il = request.form.get('il', '').strip() or None
        aday.ilce = request.form.get('ilce', '').strip() or None
        aday.pozisyon_id = int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None
        aday.kaynak = request.form.get('kaynak') or aday.kaynak
        aday.degerlendirme_notu = request.form.get('degerlendirme_notu', '').strip() or None
        aday.updated_by = current_user.id
        
        db.session.commit()
        
        flash('Aday bilgileri güncellendi.', 'success')
        return redirect(url_for('ik.aday_detay', id=id))
    
    pozisyonlar = Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all()
    
    return render_template('ik/aday_form.html',
                          aday=aday,
                          pozisyonlar=pozisyonlar,
                          ehliyet_siniflari=EHLIYET_SINIFLARI)

# ============================================================
# ADAY EVRAKLARI (sabit alanlar + aday_evraklar tablosu)
# ============================================================

# Adaylar tablosundaki sabit dosya alanları: (alan, etiket).
# Başvuru formundan gelen dosyalar bu alanlara yazılır; aday_evraklar
# tablosuna kayıt düşmez. Bu yüzden evrak listesinde iki kaynak birleştirilir.
ADAY_DOSYA_ALANLARI = [
    ('foto', 'Fotoğraf'),
    ('cv_dosya', 'CV'),
    ('kimlik_on', 'Kimlik (Ön)'),
    ('kimlik_arka', 'Kimlik (Arka)'),
    ('diploma_foto', 'Diploma'),
    ('ehliyet_foto', 'Ehliyet'),
    ('src_foto', 'SRC Belgesi'),
    ('ikametgah', 'İkametgah'),
    ('adli_sicil', 'Adli Sicil Kaydı'),
    ('kargo_barkod_foto', 'Kargo Barkodu'),
]
ADAY_DOSYA_ETIKETLERI = dict(ADAY_DOSYA_ALANLARI)


def _dosya_yol_adaylari(deger):
    """Bir dosya yolu değeri için diskteki olası tam yollar.

    Başvuru/kariyer formları 'uploads/adaylar/<id>/<dosya>' (static klasörüne
    göre), evrak akışı ise UPLOAD_FOLDER'a göre relatif yol kaydediyor; eski
    kayıtlarda mutlak yol da olabilir.
    """
    yol = (deger or '').strip()
    if not yol:
        return []
    if os.path.isabs(yol):
        return [yol]
    adaylar = [
        os.path.join(current_app.config['UPLOAD_FOLDER'], yol),
        os.path.join(current_app.static_folder, yol),
    ]
    if yol.startswith('uploads/'):
        # 'uploads/adaylar/..' değeri UPLOAD_FOLDER'ın kendisine göre de denenir
        adaylar.append(os.path.join(current_app.config['UPLOAD_FOLDER'],
                                    yol[len('uploads/'):]))
    return adaylar


def _aday_dosya_tam_yol(deger):
    """Dosya değerinin diskte var olan tam yolu (bulunamazsa None).

    Relatif değerler yalnızca izinli köklerin (UPLOAD_FOLDER / static) altında
    çözülür — DB'deki bozuk bir yol ('../..') klasör dışına çıkamasın.
    """
    mutlak = os.path.isabs((deger or '').strip())
    kokler = [os.path.realpath(current_app.config['UPLOAD_FOLDER']),
              os.path.realpath(current_app.static_folder)]
    for yol in _dosya_yol_adaylari(deger):
        if not (os.path.exists(yol) and os.path.isfile(yol)):
            continue
        if mutlak:
            return yol
        gercek = os.path.realpath(yol)
        if any(gercek.startswith(kok + os.sep) for kok in kokler):
            return yol
    return None


def _aday_evrak_listesi(aday):
    """Evrak Takibi listesi: sabit alanlar + aday_evraklar kayıtları.

    Aynı fiziksel dosya iki kaynakta da varsa bir kez gösterilir.
    """
    liste = []
    gorulen_yollar = set()

    for alan, etiket in ADAY_DOSYA_ALANLARI:
        deger = getattr(aday, alan, None)
        if not deger:
            continue
        tam = _aday_dosya_tam_yol(deger)
        if tam:
            gorulen_yollar.add(os.path.realpath(tam))
        liste.append({
            'kaynak': 'alan',
            'alan': alan,
            'evrak_id': None,
            'ad': etiket,
            'dosya_adi': os.path.basename(deger),
            'dosya_var': tam is not None,
            'durum': 'yuklendi' if tam else 'kayip',
            'tarih': None,
        })

    for ev in aday.evraklar:
        tam = _evrak_tam_yol(ev.dosya_yolu)
        var = bool(tam) and os.path.exists(tam) and os.path.isfile(tam)
        if var and os.path.realpath(tam) in gorulen_yollar:
            continue
        liste.append({
            'kaynak': 'evrak',
            'alan': None,
            'evrak_id': ev.id,
            'ad': (ev.evrak_tipi.ad if ev.evrak_tipi else None) or ev.dosya_adi or 'Evrak',
            'dosya_adi': ev.dosya_adi or (os.path.basename(ev.dosya_yolu) if ev.dosya_yolu else None),
            'dosya_var': var,
            'durum': ev.durum,
            'tarih': ev.created_at,
        })

    return liste


@ik_bp.route('/aday/<int:id>/dosya/<alan>')
@login_required
@permission_required('ik.view')
def aday_dosya_indir(id, alan):
    """Adaylar tablosundaki sabit alandaki dosyayı indir / görüntüle (?goster=1)."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adayın evraklarını görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if alan not in ADAY_DOSYA_ETIKETLERI:
        abort(404)

    inline = request.args.get('goster') == '1'
    tam = _aday_dosya_tam_yol(getattr(aday, alan, None))
    if not tam:
        if inline:
            # <img>/yeni sekme isteklerinde flash mesajı bırakmadan 404 dön
            abort(404)
        flash('Evrak dosyası bulunamadı.', 'danger')
        return redirect(url_for('ik.aday_detay', id=aday.id))

    return send_file(tam, as_attachment=not inline,
                     download_name=os.path.basename(tam))


@ik_bp.route('/aday/<int:id>')
@login_required
@permission_required('ik.view')
def aday_detay(id):
    """Aday detay sayfası"""
    aday = Aday.query.get_or_404(id)

    if not aday_in_scope(aday):
        flash('Bu adayı görüntüleme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    evrak_tipleri = EvrakTipi.query.filter_by(aktif=True).order_by(EvrakTipi.sira).all()
    
    # Evrak tamamlanma oranı hesapla
    zorunlu_evraklar = EvrakTipi.query.filter_by(zorunlu=True, aktif=True).count()
    if zorunlu_evraklar == 0:
        evrak_tamamlanma = 100
    else:
        yuklenen = aday.evraklar.join(EvrakTipi).filter(
            EvrakTipi.zorunlu == True,
            AdayEvrak.durum == 'onaylandi'
        ).count()
        evrak_tamamlanma = int((yuklenen / zorunlu_evraklar) * 100)
    
    # Eksik evraklar
    yuklenen_tipler = [e.evrak_tipi_id for e in aday.evraklar.filter(
        AdayEvrak.durum.in_(['yuklendi', 'onaylandi'])
    ).all()]
    eksik_evraklar = [t for t in EvrakTipi.query.filter_by(zorunlu=True, aktif=True).all() if t.id not in yuklenen_tipler]
    
    ise_alim_hazir = len(eksik_evraklar) == 0 and aday.kvkk_onay
    
    return render_template('ik/aday_detay.html',
                          aday=aday,
                          evrak_tipleri=evrak_tipleri,
                          evrak_listesi=_aday_evrak_listesi(aday),
                          aday_foto_url=(url_for('ik.aday_dosya_indir', id=aday.id,
                                                 alan='foto', goster=1)
                                         if _aday_dosya_tam_yol(aday.foto) else None),
                          evrak_tamamlanma=evrak_tamamlanma,
                          eksik_evraklar=eksik_evraklar,
                          ise_alim_hazir=ise_alim_hazir,
                          has_iletisim=_aday_has_iletisim(aday))


def _aday_has_iletisim(aday):
    """Adayın en az bir iletişim (arama/SMS/WhatsApp/not) kaydı var mı?"""
    return db.session.query(AdayIslemGecmisi.id).filter(
        AdayIslemGecmisi.aday_id == aday.id,
        AdayIslemGecmisi.islem.in_(AdayIslemGecmisi.ILETISIM_ISLEMLER)
    ).first() is not None


def _aday_log(aday, islem, aciklama=None, yeni_durum=None):
    """Aday işlem geçmişine kayıt ekler. aday.durum DEĞİŞTİRİLMEDEN önce çağrılmalı
    (onceki_durum doğru yakalansın diye)."""
    db.session.add(AdayIslemGecmisi(
        aday_id=aday.id,
        islem=islem,
        aciklama=aciklama,
        onceki_durum=aday.durum,
        yeni_durum=yeni_durum or aday.durum,
        kullanici_id=current_user.id,
    ))


@ik_bp.route('/aday/<int:id>/iletisim-ekle', methods=['POST'])
@login_required
@permission_required('ik.view')
def aday_iletisim_ekle(id):
    """Aday ile iletişim (arama / SMS / WhatsApp / not) kaydı ekler.

    ik.view yetkisi yeterlidir; saha koordinatörleri (SPV) de kendi
    adaylarını arayıp not düşebilsin. Adayın DURUMU değişmez; sadece süreç
    geçmişine iletişim logu eklenir."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    tip = (request.form.get('iletisim_tipi') or '').strip()
    if tip not in AdayIslemGecmisi.ILETISIM_ISLEMLER:
        flash('Geçersiz iletişim tipi.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    not_metni = (request.form.get('not') or '').strip()

    hatirlatma = None
    if tip == 'geri_aranacak':
        h_str = (request.form.get('hatirlatma_tarihi') or '').strip()
        if not h_str:
            flash('"Geri Aranacak" için tarih/saat seçmelisiniz.', 'danger')
            return redirect(url_for('ik.aday_detay', id=id))
        for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
            try:
                hatirlatma = datetime.strptime(h_str, fmt)
                break
            except ValueError:
                continue
        if hatirlatma is None:
            flash('Geçersiz hatırlatma tarihi.', 'danger')
            return redirect(url_for('ik.aday_detay', id=id))

    db.session.add(AdayIslemGecmisi(
        aday_id=aday.id,
        islem=tip,
        aciklama=not_metni or None,
        onceki_durum=aday.durum,
        yeni_durum=aday.durum,
        hatirlatma_tarihi=hatirlatma,
        kullanici_id=current_user.id,
    ))
    db.session.commit()

    flash('İletişim kaydı eklendi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/durum', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_durum_degistir(id):
    """Aday durumunu manuel değiştir (genel)"""
    aday = Aday.query.get_or_404(id)
    yeni = request.form.get('durum')

    # Aktif çalışana dönüştürülmüş adayın durumu geriye alınamaz.
    # Çalışan ayrılmış/askıya alınmışsa (tekrar işe alım) durum değiştirilebilir.
    if aday.donusum_kilitli and yeni != 'calisana_donusturuldu':
        flash('Bu aday aktif çalışana dönüştürülmüş, durumu değiştirilemez.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    if request.form.get('degerlendirme_notu'):
        aday.degerlendirme_notu = request.form.get('degerlendirme_notu')
    if yeni and yeni != aday.durum:
        _aday_log(aday, 'durum', 'Durum manuel güncellendi.', yeni)
        aday.durum = yeni
    db.session.commit()

    flash('Aday durumu güncellendi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/planli-tarih-degistir', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_planli_tarih_degistir(id):
    """Planlı başlangıç tarihini değiştir.

    Bu bir DURUM değişikliği DEĞİLDİR; sadece planlı tarih güncellenir. Bu yüzden
    aday çalışana dönüştürülmüş (aktif çalışan) olsa bile değiştirilebilir."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    # Değişiklik sonrası nereye dönülecek (çalışan kartından çağrılabilir)
    next_url = request.form.get('next')
    if not (next_url and next_url.startswith('/')):
        next_url = url_for('ik.aday_detay', id=id)

    yeni_str = request.form.get('planlanan_baslangic')
    if not yeni_str:
        flash('Yeni planlı başlangıç tarihi zorunludur.', 'danger')
        return redirect(next_url)

    try:
        yeni_tarih = datetime.strptime(yeni_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Geçersiz tarih formatı.', 'danger')
        return redirect(next_url)

    eski_tarih = aday.planlanan_baslangic
    if eski_tarih == yeni_tarih:
        flash('Planlı başlangıç tarihi zaten aynı.', 'info')
        return redirect(next_url)

    neden = (request.form.get('degisiklik_nedeni') or '').strip()

    eski_g = eski_tarih.strftime('%d.%m.%Y') if eski_tarih else '-'
    yeni_g = yeni_tarih.strftime('%d.%m.%Y')
    aciklama = f'Planlı başlangıç: {eski_g} → {yeni_g}'
    if neden:
        aciklama += f' (Neden: {neden})'

    # Durum değişmez; sadece tarih güncellenir.
    _aday_log(aday, 'planli_tarih', aciklama)
    aday.planlanan_baslangic = yeni_tarih
    db.session.commit()

    # İlgili birimlere bildirim (SGK giriş talebi ile aynı alıcı listesi)
    try:
        from app.services.notification import notify_planli_tarih_degisikligi
        notify_planli_tarih_degisikligi(aday, eski_tarih, yeni_tarih, neden)
    except Exception as e:
        current_app.logger.warning(f"Planlı tarih değişikliği bildirimi gönderilemedi: {e}")

    flash('Planlı başlangıç tarihi güncellendi ve ilgili birimlere bildirim gönderildi.', 'success')
    return redirect(next_url)


@ik_bp.route('/aday/<int:id>/incele', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_incele(id):
    """Adayı incelemeye al (Başvurdu → İnceleniyor)"""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))
    if aday.durum == 'inceleniyor':
        flash('Aday zaten incelemeye alınmış.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))
    _aday_log(aday, 'incele', 'İncelemeye alındı.', 'inceleniyor')
    aday.durum = 'inceleniyor'
    db.session.commit()
    flash('Aday incelemeye alındı.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/onayla', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_onayla(id):
    """Adayı onayla. Planlı başlangıç tarihi zorunlu (İnceleniyor → Onaylandı)"""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if aday.durum == 'onaylandi':
        flash('Aday zaten onaylanmış.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    # Onay öncesi en az bir iletişim kaydı zorunlu (sadece onay adımında)
    if not _aday_has_iletisim(aday):
        flash('Bu adayı onaylamak için önce en az bir iletişim kaydı '
              'eklemelisiniz (arama, SMS vb.).', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    planlanan = request.form.get('planlanan_baslangic')
    if not planlanan:
        flash('Onay için "Planlı Başlangıç Tarihi" zorunludur.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    aday.planlanan_baslangic = datetime.strptime(planlanan, '%Y-%m-%d').date()
    _aday_log(aday, 'onayla',
              f'Onaylandı. Planlı başlangıç: {aday.planlanan_baslangic.strftime("%d.%m.%Y")}',
              'onaylandi')
    aday.durum = 'onaylandi'
    db.session.commit()

    # Adaya evrak yükleme linkli onay SMS'i gönder (başarısız olsa bile onay akışı devam eder)
    sms_uyari = None
    try:
        from app.services.notification import notify_aday_onay_sms
        sonuc = notify_aday_onay_sms(aday)
        if not sonuc.get('success'):
            sms_uyari = sonuc.get('error', 'Bilinmeyen hata')
    except Exception as e:
        current_app.logger.warning(f"Aday onay SMS gönderilemedi (aday_id={aday.id}): {e}")
        sms_uyari = str(e)

    # Yeni işe giriş eğitimi daveti SMS'i. Yalnızca adayın kadrosu
    # EGITIM_DAVET_PROJE_IDS içindeki bir projedeyse ve uygun (aktif,
    # kontenjanı dolmamış) bir yeni giriş eğitimi varsa gönderilir.
    # Onay işlemi zaten commit edildi; SMS hatası onay akışını bozmamalı.
    egitim_daveti_gitti = False
    try:
        from app.services.notification import notify_egitim_davet_sms
        egitim_sonuc = notify_egitim_davet_sms(aday, kaynak='otomatik')
        egitim_daveti_gitti = bool(egitim_sonuc and egitim_sonuc.get('success'))
    except Exception as e:
        current_app.logger.warning(
            f"Eğitim davet SMS gönderilemedi (aday_id={aday.id}): {e}")

    if sms_uyari:
        flash(f'Aday onaylandı ancak onay SMS\'i gönderilemedi: {sms_uyari}', 'warning')
    else:
        flash('Aday onaylandı ve evrak yükleme SMS\'i gönderildi.', 'success')
    if egitim_daveti_gitti:
        flash('Yeni işe giriş eğitimi davet SMS\'i de gönderildi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/sgk-talep', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_sgk_talep(id):
    """SGK giriş talebi oluştur. Planlı başlangıç yoksa formdan al (Onaylandı → SGK Giriş Talebi)"""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    # Zaten talep oluşturulmuşsa (veya sonraki aşamadaysa) tekrar işlem yapma:
    # mail tekrar gitmesin, log tekrar eklenmesin.
    if aday.durum in ('sgk_giris_talebi', 'sgk_girisi_yapildi', 'calisana_donusturuldu'):
        flash('SGK giriş talebi zaten oluşturulmuş.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    # Planlı başlangıç tarihi formdan da gelebilir
    if request.form.get('planlanan_baslangic'):
        aday.planlanan_baslangic = datetime.strptime(request.form['planlanan_baslangic'], '%Y-%m-%d').date()

    if not aday.planlanan_baslangic:
        flash('SGK giriş talebi için önce "Planlı Başlangıç Tarihi" girilmelidir.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    _aday_log(aday, 'sgk_talep', 'SGK giriş talebi oluşturuldu; bordro/muhasebe bilgilendirildi.',
              'sgk_giris_talebi')
    aday.durum = 'sgk_giris_talebi'
    db.session.commit()

    # Bordro/muhasebe ekibine bildirim
    try:
        from app.services.notification import notify_sgk_giris_talebi
        notify_sgk_giris_talebi(aday)
    except Exception as e:
        current_app.logger.warning(f"SGK giriş talebi bildirimi gönderilemedi: {e}")

    flash('SGK giriş talebi oluşturuldu ve bordro/muhasebe ekibine bildirim gönderildi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/sgk-girisi-yapildi', methods=['POST'])
@login_required
def aday_sgk_girisi_yapildi(id):
    """SGK girişi yapıldı + giriş bildirgesi (PDF) yükle. Muhasebe/bordro veya İK yapar."""
    # Bu adım bordro/muhasebe tarafından yapılabilir → ik.edit VEYA masraf.edit yeter
    if not (current_user.has_permission('ik.edit') or current_user.has_permission('masraf.edit')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    aday = Aday.query.get_or_404(id)

    if aday.durum in ('sgk_girisi_yapildi', 'calisana_donusturuldu'):
        flash('SGK girişi zaten kaydedilmiş.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    dosya = request.files.get('sgk_bildirgesi')
    if not dosya or not dosya.filename:
        flash('SGK giriş bildirgesi (PDF) yüklemek zorunludur.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    ext = dosya.filename.rsplit('.', 1)[1].lower() if '.' in dosya.filename else ''
    if ext not in ('pdf', 'jpg', 'jpeg', 'png'):
        flash('Geçersiz format. PDF, JPG veya PNG yükleyiniz.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    # Tekrar işe alımda eski bildirge dosyası varsa sil (yenisi yükleniyor)
    if aday.sgk_bildirgesi:
        try:
            eski_yol = os.path.join(current_app.config['UPLOAD_FOLDER'], aday.sgk_bildirgesi)
            if os.path.isfile(eski_yol):
                os.remove(eski_yol)
        except Exception as e:
            current_app.logger.warning(f"Eski SGK bildirgesi silinemedi (aday_id={aday.id}): {e}")

    sgk_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'adaylar', str(aday.id), 'sgk')
    os.makedirs(sgk_dir, exist_ok=True)
    fname = f"sgk_bildirge_{aday.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    dosya.save(os.path.join(sgk_dir, fname))
    aday.sgk_bildirgesi = f"adaylar/{aday.id}/sgk/{fname}"

    _aday_log(aday, 'sgk_giris', 'SGK girişi yapıldı, giriş bildirgesi yüklendi.', 'sgk_girisi_yapildi')
    aday.durum = 'sgk_girisi_yapildi'
    db.session.commit()

    # İK'ya bilgilendirme bildirimi
    try:
        from app.services.notification import notify_sgk_girisi_yapildi
        notify_sgk_girisi_yapildi(aday)
    except Exception as e:
        current_app.logger.warning(f"SGK girişi yapıldı bildirimi gönderilemedi: {e}")

    flash('SGK girişi kaydedildi. Aday çalışana dönüştürülmeye hazır.', 'success')
    if request.form.get('next') == 'sgk_bekleyen':
        return redirect(url_for('ik.sgk_bekleyen'))
    return redirect(url_for('ik.aday_detay', id=id))


# ============================================================
# SGK GİRİŞ BEKLEYEN ADAYLAR (Muhasebe/Bordro)
# ============================================================

def _sgk_talep_tarihleri(aday_ids):
    """aday_id -> SGK giriş talebinin oluşturulduğu tarih (islem_gecmisi sgk_talep)."""
    if not aday_ids:
        return {}
    rows = db.session.query(
        AdayIslemGecmisi.aday_id,
        db.func.max(AdayIslemGecmisi.created_at)
    ).filter(
        AdayIslemGecmisi.aday_id.in_(aday_ids),
        AdayIslemGecmisi.islem == 'sgk_talep'
    ).group_by(AdayIslemGecmisi.aday_id).all()
    return {aday_id: tarih for aday_id, tarih in rows}


def _aday_son_iletisim(aday_ids):
    """aday_id -> en son iletişim (arama/SMS/WhatsApp/not) logu.

    Aday listesinde "Son İletişim" kolonu için toplu yükleme yapar."""
    if not aday_ids:
        return {}
    ILETISIM = AdayIslemGecmisi.ILETISIM_ISLEMLER
    # Her aday için en son iletişim logunun id'si
    alt = db.session.query(
        AdayIslemGecmisi.aday_id.label('aday_id'),
        db.func.max(AdayIslemGecmisi.id).label('max_id')
    ).filter(
        AdayIslemGecmisi.aday_id.in_(aday_ids),
        AdayIslemGecmisi.islem.in_(ILETISIM)
    ).group_by(AdayIslemGecmisi.aday_id).subquery()

    loglar = db.session.query(AdayIslemGecmisi).join(
        alt, AdayIslemGecmisi.id == alt.c.max_id
    ).all()
    return {log.aday_id: log for log in loglar}


def _iletisim_filtre_uygula(query, iletisim):
    """Aday listesine iletişim durumuna göre filtre uygular.

    - hic_aranmamis: hiç iletişim logu olmayan adaylar
    - ulasilamadi: en son iletişimi "Ulaşılamadı" olan adaylar
    - geri_aranacak: en son iletişimi "Geri Aranacak" olan adaylar (bekleyen)
    """
    ILETISIM = AdayIslemGecmisi.ILETISIM_ISLEMLER

    if iletisim == 'hic_aranmamis':
        var_olan = db.session.query(AdayIslemGecmisi.aday_id).filter(
            AdayIslemGecmisi.islem.in_(ILETISIM)
        )
        return query.filter(~Aday.id.in_(var_olan))

    if iletisim in ('ulasilamadi', 'geri_aranacak'):
        # Her adayın en son iletişim logu bu tip mi?
        son = db.session.query(
            AdayIslemGecmisi.aday_id.label('aday_id'),
            db.func.max(AdayIslemGecmisi.id).label('max_id')
        ).filter(AdayIslemGecmisi.islem.in_(ILETISIM)).group_by(
            AdayIslemGecmisi.aday_id
        ).subquery()
        eslesen = db.session.query(son.c.aday_id).join(
            AdayIslemGecmisi, AdayIslemGecmisi.id == son.c.max_id
        ).filter(AdayIslemGecmisi.islem == iletisim)
        return query.filter(Aday.id.in_(eslesen))

    return query


def geri_aranacak_adaylar(scoped=True):
    """Bekleyen "geri aranacak" hatırlatmaları.

    Bir aday, EN SON iletişim logu 'geri_aranacak' (ve hatırlatma tarihi dolu)
    ise "beklemede" sayılır. Sonradan başka bir iletişim kaydı (ör. arama
    yapıldı) eklenirse otomatik olarak listeden düşer.

    Dönüş: (aday, hatirlatma_log) tuple listesi, hatırlatma tarihine göre artan.
    """
    ILETISIM = AdayIslemGecmisi.ILETISIM_ISLEMLER
    son = db.session.query(
        AdayIslemGecmisi.aday_id.label('aday_id'),
        db.func.max(AdayIslemGecmisi.id).label('max_id')
    ).filter(AdayIslemGecmisi.islem.in_(ILETISIM)).group_by(
        AdayIslemGecmisi.aday_id
    ).subquery()

    q = db.session.query(Aday, AdayIslemGecmisi).join(
        son, Aday.id == son.c.aday_id
    ).join(
        AdayIslemGecmisi, AdayIslemGecmisi.id == son.c.max_id
    ).filter(
        AdayIslemGecmisi.islem == 'geri_aranacak',
        AdayIslemGecmisi.hatirlatma_tarihi.isnot(None),
        Aday.is_deleted == False,
        # Süreci kapanmış adaylar hatırlatma üretmesin
        ~Aday.durum.in_([
            'calisana_donusturuldu', 'reddedildi', 'red',
            'aday_reddetti', 'havuzda', 'iptal',
        ]),
    )
    if scoped:
        q = apply_aday_scope(q)
    rows = q.order_by(AdayIslemGecmisi.hatirlatma_tarihi.asc()).all()
    return [(aday, log) for aday, log in rows]


def _sgk_bekleyen_query():
    """Durumu 'sgk_giris_talebi' olan adaylar - scope filtreli."""
    q = Aday.query.filter_by(is_deleted=False, durum='sgk_giris_talebi')
    return apply_aday_scope(q)


def _sgk_bekleyen_calisan_query():
    """Tekrar işe alımda SGK girişi bekleyen çalışanlar - scope filtreli."""
    q = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.SGK_BEKLIYOR)
    return apply_calisan_scope(q)


# SGK çıkış bildiriminin yasal bildirim süresi (son çalışma gününden itibaren)
SGK_CIKIS_YASAL_GUN = 10


def _sgk_cikis_bekleyen_query():
    """SGK çıkışı bekleyen: resmi işten çıkışı tamamlanmış (IstenCikis.durum='tamamlandi')
    ama SGK çıkış bildirgesi henüz yüklenmemiş çalışanların IstenCikis kayıtları.
    Scope filtreli (çalışan kapsamına göre)."""
    scoped_ids = apply_calisan_scope(db.session.query(Calisan.id))
    return IstenCikis.query.join(Calisan, IstenCikis.calisan_id == Calisan.id).filter(
        IstenCikis.durum == 'tamamlandi',
        Calisan.is_deleted == False,
        Calisan.sgk_cikis_bildirgesi.is_(None),
        IstenCikis.calisan_id.in_(scoped_ids),
    ).order_by(IstenCikis.created_at.desc())


def _sgk_cikis_bekleyen_rows():
    """SGK çıkış bekleyen kayıtları - çalışan başına en yeni çıkış, en acil (en eski
    son çalışma günü) üstte."""
    seen = set()
    rows = []
    for cikis in _sgk_cikis_bekleyen_query().all():
        if cikis.calisan_id in seen:
            continue
        seen.add(cikis.calisan_id)
        rows.append(cikis)
    rows.sort(key=lambda c: (c.gerceklesen_cikis_tarihi is None, c.gerceklesen_cikis_tarihi or date.max))
    return rows


def _cikis_bildiren_map(calisan_ids):
    """calisan_id -> en son IstenCikisBildirimi (bildiren SPV / son çalışma günü için)."""
    if not calisan_ids:
        return {}
    alt = db.session.query(
        IstenCikisBildirimi.calisan_id.label('calisan_id'),
        db.func.max(IstenCikisBildirimi.id).label('max_id')
    ).filter(
        IstenCikisBildirimi.calisan_id.in_(calisan_ids)
    ).group_by(IstenCikisBildirimi.calisan_id).subquery()
    kayitlar = db.session.query(IstenCikisBildirimi).join(
        alt, IstenCikisBildirimi.id == alt.c.max_id
    ).all()
    return {b.calisan_id: b for b in kayitlar}


@ik_bp.route('/sgk-bekleyen')
@login_required
@permission_required('ik.view')
def sgk_bekleyen():
    """SGK girişi bekleyen adaylar - bordro/muhasebe iş listesi."""
    adaylar = _sgk_bekleyen_query().all()
    # En yakın planlı başlangıç en üstte; tarihsizler en sona
    adaylar.sort(key=lambda a: (a.planlanan_baslangic is None, a.planlanan_baslangic or date.max))

    talep_tarihleri = _sgk_talep_tarihleri([a.id for a in adaylar])

    bugun = date.today()
    hafta_sonu = bugun + timedelta(days=7)
    toplam = len(adaylar)
    bugun_baslamasi = sum(
        1 for a in adaylar
        if a.planlanan_baslangic and a.planlanan_baslangic <= bugun
    )
    bu_hafta = sum(
        1 for a in adaylar
        if a.planlanan_baslangic and bugun <= a.planlanan_baslangic <= hafta_sonu
    )

    # Tekrar işe alım - SGK bekleyen çalışanlar
    calisanlar = _sgk_bekleyen_calisan_query().all()
    calisanlar.sort(key=lambda c: (c.ise_baslama is None, c.ise_baslama or date.max))

    # SGK ÇIKIŞ bekleyen (çıkışı tamamlanmış ama bildirge yüklenmemiş)
    cikis_bekleyen = _sgk_cikis_bekleyen_rows()
    cikis_bildiren = _cikis_bildiren_map([c.calisan_id for c in cikis_bekleyen])
    cikis_toplam = len(cikis_bekleyen)
    cikis_geciken = sum(
        1 for c in cikis_bekleyen
        if c.gerceklesen_cikis_tarihi
        and (bugun - c.gerceklesen_cikis_tarihi).days > SGK_CIKIS_YASAL_GUN
    )

    return render_template('ik/sgk_bekleyen.html',
                           adaylar=adaylar,
                           calisanlar=calisanlar,
                           talep_tarihleri=talep_tarihleri,
                           cikis_bekleyen=cikis_bekleyen,
                           cikis_bildiren=cikis_bildiren,
                           cikis_toplam=cikis_toplam,
                           cikis_geciken=cikis_geciken,
                           sgk_cikis_yasal_gun=SGK_CIKIS_YASAL_GUN,
                           bugun=bugun,
                           toplam=toplam,
                           bugun_baslamasi=bugun_baslamasi,
                           bu_hafta=bu_hafta,
                           active='ik-sgk-bekleyen')


@ik_bp.route('/sgk-bekleyen/export')
@login_required
@permission_required('ik.view')
def sgk_bekleyen_export():
    """SGK bekleyen adayları Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    adaylar = _sgk_bekleyen_query().all()
    adaylar.sort(key=lambda a: (a.planlanan_baslangic is None, a.planlanan_baslangic or date.max))
    talep_tarihleri = _sgk_talep_tarihleri([a.id for a in adaylar])

    wb = Workbook()
    ws = wb.active
    ws.title = 'SGK Bekleyen'

    headers = ['Ad Soyad', 'TC Kimlik', 'Eğitim Durumu', 'Proje', 'Kadro', 'İl',
               'Planlanan Başlangıç', 'SGK Talep Tarihi', 'Telefon', 'Email']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for a in adaylar:
        talep = talep_tarihleri.get(a.id)
        ws.append([
            a.full_name,
            a.tc_kimlik or '',
            a.egitim_durumu_label or '',
            a.kadro.proje.ad if a.kadro and a.kadro.proje else '',
            a.kadro.pozisyon_adi if a.kadro else '',
            (a.kadro.il if a.kadro and a.kadro.il else (a.il or '')),
            a.planlanan_baslangic.strftime('%d.%m.%Y') if a.planlanan_baslangic else '',
            talep.strftime('%d.%m.%Y %H:%M') if talep else '',
            a.telefon or '',
            a.email or '',
        ])

    widths = [26, 14, 14, 22, 24, 14, 18, 18, 16, 28]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"sgk_bekleyen_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@ik_bp.app_context_processor
def inject_sgk_bekleyen_count():
    """Sidebar rozeti için SGK bekleyen aday sayısı (scope filtreli)."""
    def sgk_bekleyen_count():
        # Sidebar rozeti: SGK giriş + tekrar işe alım + SGK çıkış bekleyen toplamı
        if not current_user.is_authenticated:
            return 0
        try:
            return (_sgk_bekleyen_query().count()
                    + _sgk_bekleyen_calisan_query().count()
                    + len(_sgk_cikis_bekleyen_rows()))
        except Exception:
            db.session.rollback()
            return 0

    def sgk_cikis_bekleyen_count():
        if not current_user.is_authenticated:
            return 0
        try:
            return len(_sgk_cikis_bekleyen_rows())
        except Exception:
            db.session.rollback()
            return 0

    return dict(sgk_bekleyen_count=sgk_bekleyen_count,
                sgk_cikis_bekleyen_count=sgk_cikis_bekleyen_count)


@ik_bp.route('/<int:id>/sgk-cikis-bildirge-yukle', methods=['POST'])
@login_required
def sgk_cikis_bildirge_yukle(id):
    """SGK çıkışı yapıldı: çıkış bildirgesini (PDF/JPG/PNG) yükle.
    Bordro/muhasebe (masraf.edit) veya İK (ik.edit) yapabilir."""
    if not (current_user.has_permission('ik.edit') or current_user.has_permission('masraf.edit')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    calisan = Calisan.query.get_or_404(id)
    if not calisan_in_scope(calisan):
        flash('Bu çalışana erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    dosya = request.files.get('sgk_cikis_bildirgesi')
    if not dosya or not dosya.filename:
        flash('SGK çıkış bildirgesi yüklemek zorunludur.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    ext = dosya.filename.rsplit('.', 1)[1].lower() if '.' in dosya.filename else ''
    if ext not in ('pdf', 'jpg', 'jpeg', 'png'):
        flash('Geçersiz format. PDF, JPG veya PNG yükleyiniz.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    # Eski bildirge varsa sil (yenisi yükleniyor)
    if calisan.sgk_cikis_bildirgesi:
        try:
            eski = os.path.join(current_app.config['UPLOAD_FOLDER'], calisan.sgk_cikis_bildirgesi)
            if os.path.isfile(eski):
                os.remove(eski)
        except Exception as e:
            current_app.logger.warning(f"Eski SGK çıkış bildirgesi silinemedi (calisan_id={id}): {e}")

    hedef_dir = os.path.join(current_app.config['UPLOAD_FOLDER'], 'calisanlar', str(id), 'sgk_cikis')
    os.makedirs(hedef_dir, exist_ok=True)
    fname = f"sgk_cikis_{id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    dosya.save(os.path.join(hedef_dir, fname))
    calisan.sgk_cikis_bildirgesi = f"calisanlar/{id}/sgk_cikis/{fname}"

    # Akış: SPV ön bildirimini "SGK çıkışı yapıldı" durumuna al
    bildirim = IstenCikisBildirimi.query.filter_by(calisan_id=id).order_by(
        IstenCikisBildirimi.created_at.desc()).first()
    if bildirim:
        bildirim.durum = 'sgk_cikis_yapildi'

    db.session.commit()

    # Çıkış nedeni/kodu için en güncel resmi çıkış kaydı
    cikis = IstenCikis.query.filter_by(calisan_id=id).order_by(
        IstenCikis.created_at.desc()).first()

    # İK/Bordro ekibine bilgilendirme
    try:
        from app.services.notification import notify_sgk_cikis_yapildi
        notify_sgk_cikis_yapildi(calisan, cikis=cikis, yukleyen=current_user)
    except Exception as e:
        current_app.logger.warning(f"SGK çıkışı yapıldı bildirimi gönderilemedi (calisan_id={id}): {e}")

    flash('SGK çıkış bildirgesi yüklendi ve ilgili ekiplere bildirim gönderildi.', 'success')
    if request.form.get('next') == 'sgk_bekleyen':
        return redirect(url_for('ik.sgk_bekleyen'))
    return redirect(url_for('ik.detay', id=id))


@ik_bp.route('/sgk-bekleyen/cikis-export')
@login_required
@permission_required('ik.view')
def sgk_bekleyen_cikis_export():
    """SGK çıkış bekleyen listesini Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    rows = _sgk_cikis_bekleyen_rows()
    bildiren_map = _cikis_bildiren_map([c.calisan_id for c in rows])

    wb = Workbook()
    ws = wb.active
    ws.title = 'SGK Çıkış Bekleyen'

    headers = ['Ad Soyad', 'TC Kimlik', 'Telefon', 'Proje', 'Kadro',
               'Son Çalışma Günü', 'Çıkış Nedeni', 'SGK Çıkış Kodu',
               'Bildirim Tarihi', 'Bildiren SPV']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in rows:
        calisan = c.calisan
        proje = calisan.kadro.proje.ad if calisan.kadro and calisan.kadro.proje else ''
        kadro = calisan.kadro.pozisyon_adi if calisan.kadro else (calisan.pozisyon.ad if calisan.pozisyon else '')
        bildirim = bildiren_map.get(c.calisan_id)
        cikis_nedeni = f"{c.cikis_tipi or ''}: {c.cikis_sebebi or ''}".strip(': ') or (calisan.ayrilma_nedeni or '')
        ws.append([
            calisan.full_name,
            calisan.tc_kimlik or '',
            calisan.telefon or '',
            proje,
            kadro,
            c.gerceklesen_cikis_tarihi.strftime('%d.%m.%Y') if c.gerceklesen_cikis_tarihi else '',
            cikis_nedeni,
            str(c.sgk_cikis_kodu.kod) if c.sgk_cikis_kodu else '',
            bildirim.created_at.strftime('%d.%m.%Y') if bildirim else '',
            bildirim.bildiren.full_name if bildirim and bildirim.bildiren else '',
        ])

    widths = [26, 13, 16, 22, 24, 16, 24, 14, 16, 22]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"sgk_cikis_bekleyen_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@ik_bp.route('/aday/<int:id>/reddet', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_reddet(id):
    """Adayı reddet (her aşamadan). Red nedeni zorunlu."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if aday.donusum_kilitli:
        flash('Bu aday aktif çalışana dönüştürülmüş, reddedilemez.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    if aday.durum == 'reddedildi':
        flash('Aday zaten reddedilmiş.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    # Red öncesi en az bir iletişim kaydı zorunlu (sadece şirketin reddettiği durumda)
    if not _aday_has_iletisim(aday):
        flash('Bu adayı reddetmek için önce en az bir iletişim kaydı '
              'eklemelisiniz (arama, SMS vb.).', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    red_nedeni = (request.form.get('red_nedeni') or request.form.get('red_sebebi') or '').strip()
    if not red_nedeni:
        flash('Red nedeni zorunludur.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    aday.red_nedeni = red_nedeni
    aday.red_tarihi = datetime.now()
    _aday_log(aday, 'reddet', red_nedeni, 'reddedildi')
    aday.durum = 'reddedildi'
    db.session.commit()
    flash('Aday reddedildi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/havuza-al', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_havuza_al(id):
    """Adayı havuza al (bu kadro için değil ama potansiyeli var). Not zorunlu, SMS yok.
    Kadro bağlantısı korunur (nereden geldiği belli olsun)."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if aday.donusum_kilitli:
        flash('Bu aday aktif çalışana dönüştürülmüş, havuza alınamaz.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    if aday.durum == 'havuzda':
        flash('Aday zaten havuzda.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    havuz_notu = (request.form.get('havuz_notu') or '').strip()
    if not havuz_notu:
        flash('Havuza alma notu zorunludur (hangi tür pozisyona uygun, neden havuza alındı).', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    aday.havuz_notu = havuz_notu
    aday.havuza_alinma_tarihi = datetime.now()
    _aday_log(aday, 'havuza_al', havuz_notu, 'havuzda')
    aday.durum = 'havuzda'
    db.session.commit()
    flash('Aday havuza alındı.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday/<int:id>/aday-reddetti', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_kendisi_reddetti(id):
    """Aday işi kendisi reddetti. Not opsiyonel, SMS yok."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if aday.donusum_kilitli:
        flash('Bu aday aktif çalışana dönüştürülmüş, "aday reddetti" olarak işaretlenemez.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    if aday.durum == 'aday_reddetti':
        flash('Aday zaten "aday reddetti" olarak işaretlenmiş.', 'info')
        return redirect(url_for('ik.aday_detay', id=id))

    aciklama = (request.form.get('red_nedeni') or '').strip()
    aday.red_nedeni = aciklama or None
    aday.red_tarihi = datetime.now()
    _aday_log(aday, 'aday_reddetti', aciklama or 'Aday işi reddetti', 'aday_reddetti')
    aday.durum = 'aday_reddetti'
    db.session.commit()
    flash('Aday "aday reddetti" olarak işaretlendi.', 'success')
    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/aday-havuzu')
@login_required
@permission_required('ik.view')
def aday_havuzu():
    """Havuzdaki (rezerve) adaylar listesi."""
    adaylar = apply_aday_scope(
        Aday.query.filter_by(is_deleted=False, durum='havuzda')
    ).order_by(Aday.havuza_alinma_tarihi.desc().nullslast(), Aday.created_at.desc()).all()
    # Kadroya atama için scope'taki aktif kadrolar
    proje_ids = [p.id for p in user_scoped_projeler()]
    kadrolar = HedefKadro.query.filter_by(is_deleted=False, aktif=True).filter(
        HedefKadro.proje_id.in_(proje_ids) if proje_ids else False
    ).all()
    kadrolar.sort(key=lambda k: (k.proje.ad if k.proje else '', k.full_title))
    return render_template('ik/aday_havuzu.html', adaylar=adaylar,
                           kadrolar=kadrolar, active='ik-aday-havuzu')


@ik_bp.route('/aday-havuzu/export')
@login_required
@permission_required('ik.view')
def aday_havuzu_export():
    """Havuzdaki adayları Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    adaylar = apply_aday_scope(
        Aday.query.filter_by(is_deleted=False, durum='havuzda')
    ).order_by(Aday.havuza_alinma_tarihi.desc().nullslast(), Aday.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Aday Havuzu'

    headers = ['Ad Soyad', 'TC Kimlik', 'Telefon', 'Önceki Kadro/Proje',
               'Not', 'Havuza Alınma Tarihi']
    ws.append(headers)

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for a in adaylar:
        if a.kadro:
            proje_ad = a.kadro.proje.ad if a.kadro.proje else ''
            kadro_pozisyon = a.kadro.pozisyon_adi or ''
            onceki = ' / '.join(filter(None, [kadro_pozisyon, proje_ad]))
        else:
            onceki = ''
        ws.append([
            f'{a.ad or ""} {a.soyad or ""}'.strip(),
            a.tc_kimlik or '',
            a.telefon or '',
            onceki,
            a.havuz_notu or '',
            a.havuza_alinma_tarihi.strftime('%d.%m.%Y %H:%M') if a.havuza_alinma_tarihi else '',
        ])

    widths = [28, 13, 16, 34, 40, 18]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"aday_havuzu_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@ik_bp.route('/aday/<int:id>/havuzdan-ata', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_havuzdan_ata(id):
    """Havuzdaki adayı bir kadroya ata: durum 'basvurdu', yeni kadro_id."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adaya erişim yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    if aday.durum != 'havuzda':
        flash('Bu aday havuzda değil.', 'warning')
        return redirect(url_for('ik.aday_havuzu'))

    kadro_id = request.form.get('kadro_id', type=int)
    if not kadro_id:
        flash('Kadro seçimi zorunludur.', 'danger')
        return redirect(url_for('ik.aday_havuzu'))

    kadro = HedefKadro.query.filter_by(id=kadro_id, is_deleted=False).first()
    if not kadro:
        flash('Seçilen kadro bulunamadı.', 'danger')
        return redirect(url_for('ik.aday_havuzu'))

    aciklama = f'Havuzdan kadroya atandı: {kadro.full_title}'
    _aday_log(aday, 'havuzdan_ata', aciklama, 'basvurdu')
    aday.kadro_id = kadro.id
    aday.durum = 'basvurdu'
    db.session.commit()
    flash(f'Aday "{kadro.full_title}" kadrosuna atandı.', 'success')
    return redirect(url_for('ik.aday_detay', id=aday.id))


@ik_bp.route('/aday/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def aday_sil(id):
    """Aday sil (soft delete)"""
    aday = Aday.query.get_or_404(id)
    aday.soft_delete(current_user.id)
    db.session.commit()
    
    flash('Aday silindi.', 'success')
    return redirect(url_for('ik.aday_liste'))


# ============================================================
# EVRAK YÖNETİMİ
# ============================================================

def _is_iban_evrak_tipi(evrak_tipi):
    """Evrak tipinin IBAN / hesap bilgisi belgesi olup olmadığını belirler.
    Kod 'IBAN' ise veya ad/aciklama içinde 'iban' geçiyorsa True.
    """
    if not evrak_tipi:
        return False
    if (evrak_tipi.kod or '').strip().upper() == 'IBAN':
        return True
    metin = f"{evrak_tipi.ad or ''} {evrak_tipi.kod or ''}".lower()
    return 'iban' in metin


@ik_bp.route('/aday/<int:id>/evrak', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_evrak_yukle(id):
    """Aday evrak yükle"""
    aday = Aday.query.get_or_404(id)
    
    if 'dosya' not in request.files:
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))
    
    dosya = request.files['dosya']
    if dosya.filename == '':
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))
    
    if dosya and allowed_file(dosya.filename):
        evrak_tipi_id = int(request.form['evrak_tipi_id'])
        
        # Dosya adı oluştur
        filename = secure_filename(dosya.filename)
        ext = filename.rsplit('.', 1)[1].lower()
        new_filename = f"aday_{id}_{evrak_tipi_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        
        # Klasör oluştur
        upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'evraklar', 'adaylar', str(id))
        os.makedirs(upload_folder, exist_ok=True)
        
        # Dosyayı kaydet
        filepath = os.path.join(upload_folder, new_filename)
        dosya.save(filepath)

        # dosya_yolu UPLOAD_FOLDER'a göre RELATİF saklanır (uploaded_file / send_file ile uyumlu)
        rel_path = f"evraklar/adaylar/{id}/{new_filename}"

        # Veritabanına ekle
        evrak = AdayEvrak(
            aday_id=id,
            evrak_tipi_id=evrak_tipi_id,
            dosya_adi=filename,
            dosya_yolu=rel_path,
            dosya_boyut=os.path.getsize(filepath),
            mime_type=dosya.content_type,
            yukleyen_id=current_user.id
        )
        db.session.add(evrak)
        db.session.commit()

        flash('Evrak başarıyla yüklendi.', 'success')

        # IBAN / Hesap bilgisi evrağı ise IBAN'ı otomatik oku ve adaya kaydet
        evrak_tipi = EvrakTipi.query.get(evrak_tipi_id)
        if _is_iban_evrak_tipi(evrak_tipi):
            try:
                from app.services.ocr_service import extract_iban_from_image
                okunan = extract_iban_from_image(filepath)
                if okunan:
                    if aday.iban and aday.iban != okunan:
                        flash(f'Görselden IBAN okundu: {okunan} '
                              f'(Kayıtlı IBAN farklı: {aday.iban} - manuel kontrol edin.)', 'warning')
                    else:
                        aday.iban = okunan
                        db.session.commit()
                        flash(f'IBAN otomatik okundu ve kaydedildi: {okunan}', 'success')
                else:
                    flash('IBAN belgesinden numara okunamadı, lütfen manuel girin.', 'warning')
            except Exception as e:
                current_app.logger.warning(f"IBAN otomatik okuma hatası (aday_id={id}): {e}")
    else:
        flash('Geçersiz dosya formatı. (PDF, JPG, PNG, DOC, DOCX)', 'danger')

    return redirect(url_for('ik.aday_detay', id=id))


@ik_bp.route('/evrak/<int:id>/onayla', methods=['POST'])
@login_required
@permission_required('ik.edit')
def evrak_onayla(id):
    """Evrak onayla"""
    evrak = AdayEvrak.query.get_or_404(id)
    evrak.durum = 'onaylandi'
    evrak.onaylayan_id = current_user.id
    evrak.onay_tarihi = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Evrak onaylandı'})


@ik_bp.route('/evrak/<int:id>/reddet', methods=['POST'])
@login_required
@permission_required('ik.edit')
def evrak_reddet(id):
    """Evrak reddet"""
    evrak = AdayEvrak.query.get_or_404(id)
    data = request.get_json()
    evrak.durum = 'reddedildi'
    evrak.red_sebebi = data.get('sebep', '')
    evrak.onaylayan_id = current_user.id
    evrak.onay_tarihi = datetime.now()
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Evrak reddedildi'})


@ik_bp.route('/evrak/<int:id>/indir')
@login_required
@permission_required('ik.view')
def evrak_indir(id):
    """Evrak indir / görüntüle.

    dosya_yolu hem yeni (UPLOAD_FOLDER'a göre relatif) hem de eski
    (mutlak yol) kayıtlarla uyumlu şekilde çözümlenir.
    ?goster=1 → tarayıcıda inline aç (indirme yerine).
    """
    evrak = AdayEvrak.query.get_or_404(id)

    yol = evrak.dosya_yolu or ''
    if os.path.isabs(yol):
        full_path = yol
    else:
        full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], yol)

    if not os.path.exists(full_path):
        flash('Evrak dosyası bulunamadı.', 'danger')
        return redirect(url_for('ik.aday_detay', id=evrak.aday_id))

    inline = request.args.get('goster') == '1'
    return send_file(full_path, as_attachment=not inline, download_name=evrak.dosya_adi)


# ============================================================
# TOPLU IBAN TARAMA (Mevcut IBAN evraklarından IBAN okuma)
# ============================================================

def _evrak_tam_yol(dosya_yolu):
    """AdayEvrak/CalisanEvrak dosya_yolu değerini tam dosya yoluna çözer.

    Farklı yükleme akışları yolu UPLOAD_FOLDER'a ya da static klasörüne göre
    kaydettiği için diskte var olan ilk aday yol döner; hiçbiri yoksa (geriye
    dönük uyumluluk) UPLOAD_FOLDER'a göre çözülmüş yol döner.
    """
    yol = dosya_yolu or ''
    if not yol:
        return None
    if os.path.isabs(yol):
        return yol
    return _aday_dosya_tam_yol(yol) or os.path.join(current_app.config['UPLOAD_FOLDER'], yol)


def _iban_evrak_tipi_idleri():
    """IBAN tipli evrak tiplerinin id listesini döndürür."""
    return [t.id for t in EvrakTipi.query.all() if _is_iban_evrak_tipi(t)]


def _iban_evrak_coz(evrak_id):
    """Kodlu evrak_id ('A123' aday, 'C123' çalışan) -> (tip_adi, evrak, sahip) veya None."""
    if not evrak_id or len(evrak_id) < 2:
        return None
    onek, ham = evrak_id[0].upper(), evrak_id[1:]
    if not ham.isdigit():
        return None
    eid = int(ham)
    if onek == 'A':
        evrak = AdayEvrak.query.get(eid)
        return ('Aday', evrak, evrak.aday) if evrak else None
    if onek == 'C':
        evrak = CalisanEvrak.query.get(eid)
        return ('Çalışan', evrak, evrak.calisan) if evrak else None
    return None


@ik_bp.route('/iban-tara')
@login_required
@permission_required('ik.edit')
def iban_tara():
    """IBAN toplu tarama sayfası (tarama JavaScript ile tek tek yapılır)."""
    return render_template('ik/iban_tara.html')


@ik_bp.route('/iban-tara/liste')
@login_required
@permission_required('ik.edit')
def iban_tara_liste():
    """Taranacak IBAN evraklarının id listesini JSON döndürür."""
    sadece_eksik = request.args.get('sadece_eksik', '1') == '1'
    tip_idleri = _iban_evrak_tipi_idleri()
    evraklar = []
    if tip_idleri:
        for evrak in AdayEvrak.query.filter(AdayEvrak.evrak_tipi_id.in_(tip_idleri)).all():
            aday = evrak.aday
            if not aday or aday.is_deleted:
                continue
            if sadece_eksik and aday.iban:
                continue
            evraklar.append({
                'id': f'A{evrak.id}',
                'tip': 'Aday',
                'sahip_ad': aday.full_name,
                'dosya': evrak.dosya_adi or '',
            })
        for evrak in CalisanEvrak.query.filter(CalisanEvrak.evrak_tipi_id.in_(tip_idleri)).all():
            calisan = evrak.calisan
            if not calisan or calisan.is_deleted:
                continue
            if sadece_eksik and calisan.iban:
                continue
            evraklar.append({
                'id': f'C{evrak.id}',
                'tip': 'Çalışan',
                'sahip_ad': calisan.full_name,
                'dosya': evrak.dosya_adi or '',
            })
    return jsonify({'evraklar': evraklar, 'toplam': len(evraklar)})


@ik_bp.route('/iban-tara/tek', methods=['POST'])
@login_required
@permission_required('ik.edit')
def iban_tara_tek():
    """Tek bir IBAN evrağını tarar, IBAN'ı çıkarır ve sahibine kaydeder."""
    from app.services.ocr_service import extract_iban_from_image

    data = request.get_json(silent=True) or {}
    evrak_id = (data.get('evrak_id') or '').strip()

    cozum = _iban_evrak_coz(evrak_id)
    if not cozum:
        return jsonify({'basarili': False, 'durum': 'hatali',
                        'hata': 'Evrak bulunamadı', 'mesaj': 'Evrak bulunamadı',
                        'evrak_adi': '', 'ad_soyad': '', 'tip': '', 'iban': ''}), 404

    tip_adi, evrak, sahip = cozum
    if not sahip or getattr(sahip, 'is_deleted', False):
        return jsonify({'basarili': False, 'durum': 'hatali',
                        'hata': 'Kayıt bulunamadı', 'mesaj': 'Kayıt bulunamadı',
                        'evrak_adi': evrak.dosya_adi or '', 'ad_soyad': '',
                        'tip': tip_adi, 'iban': ''})

    sonuc = {
        'tip': tip_adi,
        'kayit_id': sahip.id,
        'ad_soyad': sahip.full_name,
        'evrak_adi': evrak.dosya_adi or '',
        'iban': sahip.iban or '',
    }

    tam_yol = _evrak_tam_yol(evrak.dosya_yolu)
    if not tam_yol or not os.path.exists(tam_yol):
        sonuc.update({'basarili': False, 'durum': 'hatali',
                      'hata': 'Dosya bulunamadı', 'mesaj': 'Dosya bulunamadı'})
        return jsonify(sonuc)

    try:
        okunan = extract_iban_from_image(tam_yol)
    except Exception as e:
        current_app.logger.warning(f"IBAN tek tarama hatası (evrak={evrak_id}): {e}")
        sonuc.update({'basarili': False, 'durum': 'hatali',
                      'hata': 'Okuma hatası', 'mesaj': str(e)})
        return jsonify(sonuc)

    if okunan:
        sonuc['iban'] = okunan
        if not sahip.iban:
            sahip.iban = okunan
            db.session.commit()
            mesaj = 'Okundu ve kaydedildi'
        elif sahip.iban == okunan:
            mesaj = 'Zaten kayıtlı (aynı)'
        else:
            mesaj = f'Farklı IBAN kayıtlı ({sahip.iban}) - kaydedilmedi'
        sonuc.update({'basarili': True, 'durum': 'basarili', 'mesaj': mesaj})
    else:
        sonuc.update({'basarili': False, 'durum': 'bulunamadi',
                      'hata': 'IBAN bulunamadı', 'mesaj': 'IBAN okunamadı'})
    return jsonify(sonuc)


@ik_bp.route('/iban-tara/export', methods=['POST'])
@login_required
@permission_required('ik.view')
def iban_tara_export():
    """Tarama sonuçlarını Excel olarak dışa aktar (yeniden tarama yapmaz)."""
    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    try:
        sonuclar = _json.loads(request.form.get('sonuc_json', '[]'))
    except (ValueError, TypeError):
        sonuclar = []

    wb = Workbook()
    ws = wb.active
    ws.title = 'IBAN Tarama'

    basliklar = ['Tip', 'Kayıt ID', 'Ad Soyad', 'Evrak', 'IBAN', 'Durum', 'Açıklama']
    ws.append(basliklar)
    baslik_font = Font(bold=True, color='FFFFFF')
    baslik_dolgu = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    for cell in ws[1]:
        cell.font = baslik_font
        cell.fill = baslik_dolgu

    durum_etiket = {'basarili': 'Başarılı', 'bulunamadi': 'Bulunamadı', 'hatali': 'Hatalı'}
    for s in sonuclar:
        ws.append([
            s.get('tip', ''),
            s.get('kayit_id', ''),
            s.get('ad_soyad', ''),
            s.get('evrak_adi', ''),
            s.get('iban', ''),
            durum_etiket.get(s.get('durum', ''), s.get('durum', '')),
            s.get('mesaj', ''),
        ])

    for i, genislik in enumerate([10, 10, 28, 30, 30, 14, 40], start=1):
        ws.column_dimensions[chr(64 + i)].width = genislik

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"iban_tarama_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ============================================================
# ADAY MEDYA (Foto/Video) YÖNETİMİ
# ============================================================

FOTO_EXTS = {'jpg', 'jpeg', 'png', 'webp'}
VIDEO_EXTS = {'mp4', 'mov', 'webm'}
FOTO_MAX = 10 * 1024 * 1024     # 10MB
VIDEO_MAX = 100 * 1024 * 1024   # 100MB


def _ext(filename):
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def _medya_to_dict(m):
    return {
        'id': m.id,
        'tip': m.tip,
        'dosya_adi': m.dosya_adi,
        'url': url_for('uploaded_file', filename=m.dosya_yolu),
        'boyut': m.dosya_boyut,
        'mime_type': m.mime_type,
        'tarih': m.created_at.strftime('%d.%m.%Y %H:%M') if m.created_at else '',
    }


@ik_bp.route('/aday/<int:id>/medya/foto', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_foto_yukle(id):
    """Aday fotoğrafı yükle (çoklu)."""
    aday = Aday.query.get_or_404(id)

    files = request.files.getlist('fotos')
    if not files:
        return jsonify({'success': False, 'message': 'Dosya seçilmedi.'}), 400

    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'adaylar', str(id), 'fotos')
    os.makedirs(upload_folder, exist_ok=True)

    eklenen = []
    hatalar = []
    for dosya in files:
        if not dosya or not dosya.filename:
            continue
        ext = _ext(dosya.filename)
        if ext not in FOTO_EXTS:
            hatalar.append(f"{dosya.filename}: geçersiz format")
            continue

        # Boyut kontrolü
        dosya.seek(0, os.SEEK_END)
        boyut = dosya.tell()
        dosya.seek(0)
        if boyut > FOTO_MAX:
            hatalar.append(f"{dosya.filename}: 10MB'dan büyük")
            continue

        orijinal = secure_filename(dosya.filename)
        new_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}.{ext}"
        filepath = os.path.join(upload_folder, new_filename)
        dosya.save(filepath)

        # uploads/'a göre relatif yol
        rel_path = os.path.join('adaylar', str(id), 'fotos', new_filename).replace('\\', '/')

        medya = AdayMedya(
            aday_id=aday.id,
            tip='foto',
            dosya_adi=orijinal,
            dosya_yolu=rel_path,
            dosya_boyut=boyut,
            mime_type=dosya.mimetype or f'image/{ext}',
            yukleyen_id=current_user.id
        )
        db.session.add(medya)
        db.session.flush()
        eklenen.append(_medya_to_dict(medya))

    db.session.commit()

    return jsonify({
        'success': True,
        'eklenen': eklenen,
        'hatalar': hatalar,
        'message': f"{len(eklenen)} fotoğraf yüklendi" + (f", {len(hatalar)} hata" if hatalar else "")
    })


@ik_bp.route('/aday/<int:id>/medya/video', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_video_yukle(id):
    """Aday videosu yükle (tek)."""
    aday = Aday.query.get_or_404(id)

    dosya = request.files.get('video')
    if not dosya or not dosya.filename:
        return jsonify({'success': False, 'message': 'Dosya seçilmedi.'}), 400

    ext = _ext(dosya.filename)
    if ext not in VIDEO_EXTS:
        return jsonify({'success': False, 'message': 'Geçersiz video formatı (mp4, mov, webm).'}), 400

    dosya.seek(0, os.SEEK_END)
    boyut = dosya.tell()
    dosya.seek(0)
    if boyut > VIDEO_MAX:
        return jsonify({'success': False, 'message': 'Video 100MB\'dan büyük olamaz.'}), 400

    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'adaylar', str(id), 'videos')
    os.makedirs(upload_folder, exist_ok=True)

    orijinal = secure_filename(dosya.filename)
    new_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}.{ext}"
    filepath = os.path.join(upload_folder, new_filename)
    dosya.save(filepath)

    rel_path = os.path.join('adaylar', str(id), 'videos', new_filename).replace('\\', '/')

    medya = AdayMedya(
        aday_id=aday.id,
        tip='video',
        dosya_adi=orijinal,
        dosya_yolu=rel_path,
        dosya_boyut=boyut,
        mime_type=dosya.mimetype or f'video/{ext}',
        yukleyen_id=current_user.id
    )
    db.session.add(medya)
    db.session.commit()

    return jsonify({'success': True, 'eklenen': _medya_to_dict(medya), 'message': 'Video yüklendi'})


@ik_bp.route('/aday/medya/<int:medya_id>/sil', methods=['POST'])
@login_required
@permission_required('ik.edit')
def aday_medya_sil(medya_id):
    """Aday foto/video sil."""
    medya = AdayMedya.query.get_or_404(medya_id)

    # Dosyayı diskten sil
    try:
        full_path = os.path.join(current_app.config['UPLOAD_FOLDER'], medya.dosya_yolu)
        if os.path.exists(full_path):
            os.remove(full_path)
    except Exception as e:
        current_app.logger.warning(f"Medya dosya silinemedi: {e}")

    db.session.delete(medya)
    db.session.commit()

    return jsonify({'success': True, 'message': 'Silindi'})


@ik_bp.route('/evrak-tipleri')
@login_required
@permission_required('ik.view')
def evrak_tipleri():
    """Evrak tipleri listesi"""
    tipleri = EvrakTipi.query.order_by(EvrakTipi.sira).all()
    return render_template('ik/evrak_tipleri.html', evrak_tipleri=tipleri)


@ik_bp.route('/evrak-tipi/ekle', methods=['POST'])
@login_required
@permission_required('ik.edit')
def evrak_tipi_ekle():
    """Yeni evrak tipi ekle"""
    evrak_tipi = EvrakTipi(
        ad=request.form.get('ad'),
        kod=request.form.get('kod'),
        kategori=request.form.get('kategori'),
        zorunlu=request.form.get('zorunlu') == 'on',
        aciklama=request.form.get('aciklama'),
        sira=int(request.form.get('sira', 0))
    )
    db.session.add(evrak_tipi)
    db.session.commit()
    
    flash('Evrak tipi eklendi.', 'success')
    return redirect(url_for('ik.evrak_tipleri'))


@ik_bp.route('/eksik-evraklar')
@login_required
@permission_required('ik.view')
def eksik_evraklar():
    """Eksik evrakları olan adaylar"""
    adaylar_data = []
    
    for aday in Aday.query.filter(Aday.is_deleted==False, Aday.durum.notin_(['red', 'iptal', 'ise_alindi'])).all():
        zorunlu_tipler = EvrakTipi.query.filter_by(zorunlu=True, aktif=True).all()
        yuklenen_tipler = [e.evrak_tipi_id for e in aday.evraklar.filter(
            AdayEvrak.durum.in_(['yuklendi', 'onaylandi'])
        ).all()]
        eksik = [t for t in zorunlu_tipler if t.id not in yuklenen_tipler]
        
        if eksik:
            zorunlu_count = len(zorunlu_tipler)
            yuklenen_count = zorunlu_count - len(eksik)
            oran = int((yuklenen_count / zorunlu_count) * 100) if zorunlu_count > 0 else 0
            
            adaylar_data.append({
                'aday': aday,
                'eksik': eksik,
                'oran': oran
            })
    
    return render_template('ik/eksik_evraklar.html', adaylar=adaylar_data)


# ============================================================
# ADAY → ÇALIŞAN DÖNÜŞÜMÜ
# ============================================================

def _aday_donustur_eksikler(aday):
    """Çalışana dönüştürme için zorunlu alanların eksik olup olmadığını kontrol eder.
    Her eksik alan için spesifik (kullanıcı dostu) mesaj döndürür."""
    eksikler = []
    if not aday.kadro_id:
        eksikler.append('Adayın kadro bilgisi eksik, lütfen önce kadro atayın.')
    if not (aday.tc_kimlik and aday.tc_kimlik.strip()):
        eksikler.append('TC kimlik numarası eksik.')
    if not (aday.ad and aday.ad.strip()) or not (aday.soyad and aday.soyad.strip()):
        eksikler.append('Ad veya soyad bilgisi eksik.')
    return eksikler


def _tc_cakisan_calisan(aday):
    """Aynı TC ile mevcut (silinmemiş) çalışan kaydı varsa döndürür, yoksa None."""
    if not (aday.tc_kimlik and aday.tc_kimlik.strip()):
        return None
    return Calisan.query.filter(
        Calisan.tc_kimlik == aday.tc_kimlik.strip(),
        Calisan.is_deleted == False
    ).first()


def _integrity_cakisan_alan(e):
    """IntegrityError mesajından hangi benzersiz alanın çakıştığını tespit eder."""
    msg = str(getattr(e, 'orig', e)).lower()
    if 'tc_kimlik' in msg:
        return 'TC kimlik numarası'
    if 'sicil_no' in msg:
        return 'Sicil numarası'
    if 'email' in msg:
        return 'E-posta adresi'
    return None


def _aday_evraklarini_kopyala(aday, calisan):
    """Adayın onaylı evraklarını çalışan kaydına kopyalar."""
    for aday_evrak in aday.evraklar.filter_by(durum='onaylandi').all():
        db.session.add(CalisanEvrak(
            calisan_id=calisan.id,
            evrak_tipi_id=aday_evrak.evrak_tipi_id,
            dosya_adi=aday_evrak.dosya_adi,
            dosya_yolu=aday_evrak.dosya_yolu,
            dosya_boyut=aday_evrak.dosya_boyut,
            mime_type=aday_evrak.mime_type,
            gecerlilik_bitis=aday_evrak.gecerlilik_bitis
        ))


@ik_bp.route('/aday/<int:id>/calisana-donustur', methods=['GET', 'POST'])
@login_required
@permission_required('ik.create')
def aday_calisana_donustur(id):
    """Adayı çalışana dönüştür - akıllı hata yönetimi ile"""
    aday = Aday.query.get_or_404(id)

    # 1) Durum zaten 'calisana_donusturuldu' ise -> tekrar dönüştürmeyi engelle
    if aday.durum == 'calisana_donusturuldu':
        flash('Bu aday zaten çalışana dönüştürülmüş.', 'warning')
        return redirect(url_for('ik.aday_detay', id=id))

    # 2) calisan_id zaten doluysa:
    #    - Bağlı çalışan AKTIF/İZİNLİ ise -> tekrarı engelle
    #    - Bağlı çalışan AYRILDI/ASKIYA_ALINDI ise -> tekrar işe alım: mevcut kayıt
    #      yeniden aktifleştirilecek (aşağıda), yeni kayıt oluşturulmayacak.
    tekrar_calisan = None
    if aday.calisan_id:
        mevcut = aday.donusen_calisan
        if mevcut and mevcut.durum in (CalisanDurumu.AKTIF, CalisanDurumu.IZINLI):
            flash(f'Bu aday zaten aktif çalışan kaydına bağlı: {mevcut.full_name}', 'warning')
            return redirect(url_for('ik.aday_detay', id=id))
        tekrar_calisan = mevcut  # ayrılmış/askıdaki kayıt (varsa) -> yeniden aktifleştirilecek

    # Faz 3 akışı: yalnızca SGK girişi yapıldıktan sonra çalışana dönüştürülebilir
    if aday.durum != 'sgk_girisi_yapildi':
        flash('Aday yalnızca "SGK Girişi Yapıldı" aşamasında çalışana dönüştürülebilir.', 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    # 2) Eksik bilgi kontrolü - her eksik için spesifik flash
    eksikler = _aday_donustur_eksikler(aday)
    if eksikler:
        for m in eksikler:
            flash(m, 'danger')
        return redirect(url_for('ik.aday_detay', id=id))

    # 1) TC çakışması var mı? (panelde gösterilir; GET'te de görünür)
    tc_cakisan = _tc_cakisan_calisan(aday)
    tc_cakisma_tipi = None
    if tc_cakisan:
        tc_cakisma_tipi = 'ayrildi' if tc_cakisan.durum == CalisanDurumu.AYRILDI else 'aktif'

    if request.method == 'POST':
        tc_action = request.form.get('tc_action', '').strip()

        # 4) Her dönüştürme girişimini logla (kim, hangi aday, çakışma durumu)
        current_app.logger.info(
            "Çalışana dönüştürme girişimi: aday_id=%s (%s), kullanici_id=%s (%s), "
            "tc_action=%r, tc_cakisan_id=%s",
            aday.id, aday.full_name, current_user.id, current_user.email,
            tc_action or None, (tc_cakisan.id if tc_cakisan else None)
        )

        # İptal seçildiyse
        if tc_action == 'iptal':
            flash('İşlem iptal edildi.', 'info')
            return redirect(url_for('ik.aday_detay', id=id))

        # TC çakışması varsa kullanıcı bir çözüm seçmek zorunda.
        # Tekrar işe alımda (tekrar_calisan) çözüm otomatiktir: mevcut kayıt yeniden aktifleştirilir.
        if tc_cakisan and tekrar_calisan is None and tc_action not in ('baglan', 'yeni_tcsiz'):
            flash('Aynı TC ile mevcut bir çalışan kaydı var. Lütfen nasıl devam edileceğini seçin.', 'warning')
        else:
            ise_baslama_form = request.form.get('ise_baslama')
            ise_baslama = (datetime.strptime(ise_baslama_form, '%Y-%m-%d').date()
                           if ise_baslama_form else aday.planlanan_baslangic)
            sicil_raw = request.form.get('sicil_no', '').strip()
            sicil_no = None if sicil_raw in ('', 'None', 'none') else sicil_raw

            try:
                if tekrar_calisan is not None or (tc_cakisan and tc_action == 'baglan'):
                    # Mevcut kaydı bağla + güncelle (yeniden işe alım).
                    # Tekrar işe alımda aday zaten bu kayda bağlı; yeni kayıt açılmaz.
                    calisan = tekrar_calisan if tekrar_calisan is not None else tc_cakisan
                    if aday.kadro_id:
                        calisan.kadro_id = aday.kadro_id
                    if aday.pozisyon_id:
                        calisan.pozisyon_id = aday.pozisyon_id
                    if ise_baslama:
                        calisan.ise_baslama = ise_baslama
                    if sicil_no:
                        calisan.sicil_no = sicil_no
                    if aday.iban and not calisan.iban:
                        calisan.iban = aday.iban
                    if aday.ehliyet_sinifi and not calisan.ehliyet_sinifi:
                        calisan.ehliyet_sinifi = aday.ehliyet_sinifi
                    # Ayrılış bilgilerini temizle, aktif yap
                    calisan.isten_ayrilma = None
                    calisan.ayrilma_nedeni = None
                    calisan.durum = CalisanDurumu.AKTIF
                    db.session.flush()
                else:
                    # Yeni çalışan oluştur. Çakışma varsa "TC'siz" seçeneği TC'yi boş bırakır.
                    yeni_tc = None if (tc_cakisan and tc_action == 'yeni_tcsiz') else aday.tc_kimlik
                    calisan = Calisan(
                        ad=aday.ad,
                        soyad=aday.soyad,
                        tc_kimlik=yeni_tc,
                        dogum_tarihi=aday.dogum_tarihi,
                        dogum_yeri=aday.dogum_yeri,
                        cinsiyet=aday.cinsiyet,
                        medeni_durum=aday.medeni_durum,
                        telefon=normalize_telefon(aday.telefon),
                        email=aday.email,
                        adres=aday.adres,
                        il=aday.il,
                        ilce=aday.ilce,
                        iban=aday.iban,
                        egitim_durumu=aday.egitim_durumu,
                        beden=aday.ust_beden,
                        kargo_subesi=aday.kargo_subesi,
                        ehliyet_sinifi=aday.ehliyet_sinifi,
                        pozisyon_id=aday.pozisyon_id,
                        kadro_id=aday.kadro_id,
                        sicil_no=sicil_no,
                        ise_baslama=ise_baslama,
                        calisma_tipi=request.form.get('calisma_tipi', 'tam_zamanli'),
                        durum=CalisanDurumu.AKTIF,
                        created_by=current_user.id
                    )
                    db.session.add(calisan)
                    db.session.flush()  # ID almak için

                # Aday durumunu güncelle + çalışana bağla + logla
                _aday_log(aday, 'donustur',
                          f'Çalışana dönüştürüldü (Çalışan #{calisan.id}).', 'calisana_donusturuldu')
                aday.durum = 'calisana_donusturuldu'
                aday.calisan_id = calisan.id

                _aday_evraklarini_kopyala(aday, calisan)
                db.session.commit()
                current_app.logger.info(
                    "Çalışana dönüştürme başarılı: aday_id=%s -> calisan_id=%s, "
                    "kullanici_id=%s, tc_action=%r",
                    aday.id, calisan.id, current_user.id, tc_action or None
                )

            except IntegrityError as e:
                db.session.rollback()
                current_app.logger.error(
                    "Çalışana dönüştürme IntegrityError: aday_id=%s, kullanici_id=%s, hata=%s",
                    id, current_user.id, e
                )
                alan = _integrity_cakisan_alan(e)
                if alan:
                    flash(f'Kayıt oluşturulamadı: "{alan}" başka bir kayıtla çakışıyor. '
                          f'Lütfen bu bilgiyi kontrol edin veya farklı bir değer girin.', 'danger')
                else:
                    flash('Kayıt oluşturulurken bir veritabanı kısıtı ihlal edildi. '
                          'Lütfen girdiğiniz bilgileri kontrol edin.', 'danger')
                return redirect(url_for('ik.aday_detay', id=id))
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Aday->Çalışan dönüşüm hatası (aday_id={id}): {e}")
                flash('Çalışana dönüştürme sırasında beklenmeyen bir hata oluştu. İşlem geri alındı.', 'danger')
                return redirect(url_for('ik.aday_detay', id=id))

            # İşe giriş bildirimi gönder (başarısız olsa bile akış devam eder)
            try:
                from app.services.notification import notify_ise_giris
                notify_ise_giris(calisan)
            except Exception as e:
                current_app.logger.warning(f"İşe giriş bildirimi gönderilemedi (calisan_id={calisan.id}): {e}")

            # NOT: Eğitim davet SMS'i artık burada değil, aday onaylandığında
            # (ik.aday_onayla) gönderiliyor.

            if tekrar_calisan is not None or (tc_cakisan and tc_action == 'baglan'):
                flash(f'{calisan.full_name} mevcut çalışan kaydına bağlandı ve yeniden işe alındı.', 'success')
            else:
                flash(f'{calisan.full_name} başarıyla çalışan olarak kaydedildi.', 'success')

            if not aday.kvkk_onay:
                flash('Çalışan oluşturuldu. Dikkat: KVKK onayı henüz alınmamış.', 'warning')
                current_app.logger.warning(
                    "KVKK onayı olmadan çalışana dönüştürüldü: aday_id=%s -> calisan_id=%s, kullanici_id=%s",
                    aday.id, calisan.id, current_user.id
                )
            return redirect(url_for('ik.detay', id=calisan.id))

    departmanlar = Departman.query.filter_by(aktif=True).order_by(Departman.ad).all()
    pozisyonlar = Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all()

    return render_template('ik/aday_calisana_donustur.html',
                          aday=aday,
                          departmanlar=departmanlar,
                          pozisyonlar=pozisyonlar,
                          tc_cakisan=tc_cakisan,
                          tc_cakisma_tipi=tc_cakisma_tipi)


# ============================================================
# İŞTEN ÇIKIŞ YÖNETİMİ
# ============================================================

@ik_bp.route('/isten-cikislar')
@login_required
@permission_required('ik.view')
def isten_cikis_liste():
    """İşten çıkış listesi"""
    page = request.args.get('page', 1, type=int)
    durum = request.args.get('durum')
    
    query = IstenCikis.query
    
    if durum:
        query = query.filter(IstenCikis.durum == durum)
    
    query = query.order_by(IstenCikis.created_at.desc())
    pagination = paginate_query(query, page, 20)
    
    return render_template('ik/isten_cikis_liste.html',
                          cikislar=pagination.items,
                          pagination=pagination)


@ik_bp.route('/calisan/<int:id>/cikis-baslat', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def isten_cikis_baslat(id):
    """İşten çıkış süreci başlat"""
    calisan = Calisan.query.get_or_404(id)

    if request.method == 'POST':
        sgk_kodu_id = request.form.get('sgk_cikis_kodu_id', type=int)

        cikis = IstenCikis(
            calisan_id=id,
            planlanan_cikis_tarihi=datetime.strptime(request.form['planlanan_cikis_tarihi'], '%Y-%m-%d').date(),
            cikis_tipi=request.form.get('cikis_tipi'),
            cikis_sebebi=request.form.get('cikis_sebebi'),
            detay_notu=request.form.get('detay_notu'),
            sgk_cikis_kodu_id=sgk_kodu_id,
            olusturan_id=current_user.id
        )

        # Çalışan durumunu güncelle
        calisan.durum = CalisanDurumu.ASKIYA_ALINDI

        # Kara/Gri liste
        liste_val = (request.form.get('liste_durumu') or '').strip().lower()
        if liste_val in ('temiz', 'gri_liste', 'kara_liste'):
            yeni_liste = ListeDurumu(liste_val)
            if calisan.liste_durumu != yeni_liste or yeni_liste != ListeDurumu.TEMIZ:
                calisan.liste_durumu = yeni_liste
                calisan.liste_nedeni = request.form.get('liste_nedeni', '').strip() or None
                calisan.liste_tarihi = datetime.now()
                calisan.listeye_alan_id = current_user.id

        db.session.add(cikis)
        db.session.commit()

        flash('İşten çıkış süreci başlatıldı.', 'success')
        return redirect(url_for('ik.isten_cikis_detay', id=cikis.id))

    # SPV/Koordinatör işten çıkış bildirimi varsa son çalışma gününü forma öner
    onerilen_cikis_tarihi = None
    son_bildirim = calisan.cikis_bildirimleri.filter(
        IstenCikisBildirimi.durum != 'tamamlandi').first() or calisan.cikis_bildirimleri.first()
    if son_bildirim:
        onerilen_cikis_tarihi = son_bildirim.son_calisma_gunu

    sgk_kodlari = SgkCikisKodu.query.filter_by(aktif=True).order_by(SgkCikisKodu.kod).all()
    return render_template('ik/isten_cikis_baslat.html',
                           calisan=calisan,
                           sgk_kodlari=sgk_kodlari,
                           onerilen_cikis_tarihi=onerilen_cikis_tarihi)


# ============================================================
# TEKRAR İŞE ALIM (Ayrılmış/askıdaki çalışanı yeniden işe al)
# ============================================================

def _sgk_giris_evrak_tipi():
    """SGK Giriş Bildirgesi evrak tipini bulur, yoksa oluşturur."""
    tip = EvrakTipi.query.filter_by(kod='SGK_GIRIS_BILDIRGESI').first()
    if not tip:
        tip = EvrakTipi(
            ad='SGK Giriş Bildirgesi',
            kod='SGK_GIRIS_BILDIRGESI',
            kategori='sozlesme',
            aciklama='SGK işe giriş bildirgesi',
            zorunlu=False,
            aktif=True,
        )
        db.session.add(tip)
        db.session.flush()
    return tip


@ik_bp.route('/<int:id>/tekrar-ise-al', methods=['POST'])
@login_required
@permission_required('ik.edit')
def tekrar_ise_al(id):
    """Ayrılmış/askıdaki çalışan için tekrar işe alım başlat.
    Planlı başlangıç + kadro alınır, durum SGK_BEKLIYOR'a çekilir,
    bordro/muhasebeye SGK giriş talebi maili gönderilir."""
    calisan = Calisan.query.get_or_404(id)
    if not calisan_in_scope(calisan):
        flash('Bu çalışan için işlem yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    if not calisan.tekrar_ise_alinabilir:
        flash('Yalnızca ayrılmış veya askıya alınmış çalışanlar tekrar işe alınabilir.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    tarih_str = (request.form.get('planlanan_baslangic') or '').strip()
    if not tarih_str:
        flash('Planlı başlangıç tarihi zorunludur.', 'danger')
        return redirect(url_for('ik.detay', id=id))
    try:
        planlanan = datetime.strptime(tarih_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Geçersiz tarih formatı.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    # Kadro seçimi (opsiyonel; mevcut kadro korunur veya değiştirilir)
    kadro_id = request.form.get('kadro_id', type=int)
    if kadro_id:
        calisan.kadro_id = kadro_id

    not_metni = (request.form.get('not') or '').strip()
    if not_metni:
        onceki = (calisan.notlar or '').strip()
        etiket = f"[Tekrar işe alım {date.today().strftime('%d.%m.%Y')}] {not_metni}"
        calisan.notlar = f"{onceki}\n{etiket}".strip() if onceki else etiket

    # Planlı başlangıcı işe başlama olarak yaz, ara duruma çek
    calisan.ise_baslama = planlanan
    calisan.durum = CalisanDurumu.SGK_BEKLIYOR
    db.session.commit()

    # Bordro/muhasebe ekibine SGK giriş talebi bildirimi
    try:
        from app.services.notification import notify_sgk_giris_talebi_calisan
        notify_sgk_giris_talebi_calisan(calisan, planlanan_baslangic=planlanan)
    except Exception as e:
        current_app.logger.warning(f"Tekrar işe alım SGK talebi bildirimi gönderilemedi (calisan_id={id}): {e}")

    flash('Tekrar işe alım başlatıldı; SGK giriş talebi bordro/muhasebe ekibine iletildi.', 'success')
    return redirect(url_for('ik.detay', id=id))


@ik_bp.route('/<int:id>/sgk-girisi-yaptim', methods=['POST'])
@login_required
def sgk_girisi_yaptim(id):
    """Tekrar işe alımda SGK girişi yapıldı: bildirge yükle, çalışanı AKTIF yap.
    Bordro/muhasebe (masraf.edit) veya İK (ik.edit) yapabilir."""
    if not (current_user.has_permission('ik.edit') or current_user.has_permission('masraf.edit')):
        flash('Bu işlem için yetkiniz yok.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    calisan = Calisan.query.get_or_404(id)

    if calisan.durum != CalisanDurumu.SGK_BEKLIYOR:
        flash('Bu çalışan SGK giriş bekleme aşamasında değil.', 'info')
        return redirect(url_for('ik.detay', id=id))

    dosya = request.files.get('sgk_bildirgesi')
    if not dosya or not dosya.filename:
        flash('SGK giriş bildirgesi yüklemek zorunludur.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    ext = dosya.filename.rsplit('.', 1)[1].lower() if '.' in dosya.filename else ''
    if ext not in ('pdf', 'jpg', 'jpeg', 'png'):
        flash('Geçersiz format. PDF, JPG veya PNG yükleyiniz.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    # Dosyayı çalışan evrak klasörüne kaydet
    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'evraklar', 'calisanlar', str(id))
    os.makedirs(upload_folder, exist_ok=True)
    new_filename = f"calisan_{id}_sgk_giris_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
    filepath = os.path.join(upload_folder, new_filename)
    dosya.save(filepath)

    # Çalışan evraklarına kaydet
    tip = _sgk_giris_evrak_tipi()
    db.session.add(CalisanEvrak(
        calisan_id=id,
        evrak_tipi_id=tip.id,
        dosya_adi=secure_filename(dosya.filename),
        dosya_yolu=filepath,
        dosya_boyut=os.path.getsize(filepath),
        mime_type=dosya.content_type,
    ))

    # İşe başlama tarihi güncelle (opsiyonel override)
    if request.form.get('ise_baslama'):
        try:
            calisan.ise_baslama = datetime.strptime(request.form['ise_baslama'], '%Y-%m-%d').date()
        except ValueError:
            pass

    # Aktifleştir, ayrılış bilgilerini temizle
    calisan.durum = CalisanDurumu.AKTIF
    calisan.isten_ayrilma = None
    calisan.ayrilma_nedeni = None
    db.session.commit()

    # İşe giriş bildirimi gönder
    try:
        from app.services.notification import notify_ise_giris
        notify_ise_giris(calisan)
    except Exception as e:
        current_app.logger.warning(f"Tekrar işe alım - işe giriş bildirimi gönderilemedi (calisan_id={id}): {e}")

    flash('SGK girişi kaydedildi, çalışan yeniden aktifleştirildi.', 'success')
    if request.form.get('next') == 'sgk_bekleyen':
        return redirect(url_for('ik.sgk_bekleyen'))
    return redirect(url_for('ik.detay', id=id))


@ik_bp.route('/isten-cikis/<int:id>')
@login_required
@permission_required('ik.view')
def isten_cikis_detay(id):
    """İşten çıkış detay"""
    cikis = IstenCikis.query.get_or_404(id)
    return render_template('ik/isten_cikis_detay.html', cikis=cikis)


@ik_bp.route('/isten-cikis/<int:id>/guncelle', methods=['POST'])
@login_required
@permission_required('ik.edit')
def isten_cikis_guncelle(id):
    """İşten çıkış checklist güncelle"""
    cikis = IstenCikis.query.get_or_404(id)
    
    cikis.zimmet_teslim = request.form.get('zimmet_teslim') == 'on'
    cikis.zimmet_notu = request.form.get('zimmet_notu')
    cikis.sgk_cikis_bildirimi = request.form.get('sgk_cikis_bildirimi') == 'on'
    cikis.cikis_mulakati_yapildi = request.form.get('cikis_mulakati_yapildi') == 'on'
    cikis.cikis_mulakat_notu = request.form.get('cikis_mulakat_notu')
    
    if request.form.get('kidem_tazminati'):
        cikis.kidem_tazminati = Decimal(request.form.get('kidem_tazminati'))
    if request.form.get('ihbar_tazminati'):
        cikis.ihbar_tazminati = Decimal(request.form.get('ihbar_tazminati'))
    
    # Tüm adımlar tamamlandıysa
    if cikis.zimmet_teslim and cikis.sgk_cikis_bildirimi:
        cikis.durum = 'devam_ediyor'
    
    db.session.commit()

    flash('İşten çıkış bilgileri güncellendi.', 'success')
    return redirect(url_for('ik.isten_cikis_detay', id=id))


# ============================================================
# İŞTEN ÇIKIŞ BİLDİRİMİ (SPV / Koordinatör -> İK + Bordro)
# ============================================================

@ik_bp.route('/<int:id>/isten-cikis-bildirimi', methods=['POST'])
@login_required
@permission_required('ik.view')
def isten_cikis_bildirimi_gonder(id):
    """SPV/Koordinatör çalışan detayından işten çıkış bildirimi gönderir.
    ik.view yeterli; ekip (scope) kontrolü yapılır. İK+Bordro'ya mail gider."""
    calisan = Calisan.query.get_or_404(id)

    if not calisan_in_scope(calisan):
        flash('Bu çalışan için işlem yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    if calisan.durum != CalisanDurumu.AKTIF:
        flash('İşten çıkış bildirimi yalnızca aktif çalışanlar için gönderilebilir.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    cikis_nedeni = (request.form.get('cikis_nedeni') or '').strip()
    gecerli_nedenler = [k for k, _ in IstenCikisBildirimi.CIKIS_NEDENLERI]
    if cikis_nedeni not in gecerli_nedenler:
        flash('Geçerli bir çıkış nedeni seçiniz.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    tarih_str = (request.form.get('son_calisma_gunu') or '').strip()
    if not tarih_str:
        flash('Son çalışma günü zorunludur.', 'danger')
        return redirect(url_for('ik.detay', id=id))
    try:
        son_calisma_gunu = datetime.strptime(tarih_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Geçersiz tarih formatı.', 'danger')
        return redirect(url_for('ik.detay', id=id))

    bildirim = IstenCikisBildirimi(
        calisan_id=calisan.id,
        bildiren_user_id=current_user.id,
        cikis_nedeni=cikis_nedeni,
        son_calisma_gunu=son_calisma_gunu,
        aciklama=(request.form.get('aciklama') or '').strip() or None,
        durum='beklemede',
    )
    db.session.add(bildirim)
    db.session.commit()

    # İK + Bordro ekibine bildirim maili
    try:
        from app.services.notification import notify_isten_cikis_bildirimi
        notify_isten_cikis_bildirimi(bildirim)
    except Exception as e:
        current_app.logger.warning(f"İşten çıkış bildirimi maili gönderilemedi (calisan_id={id}): {e}")

    flash('İşten çıkış bildirimi İK ve bordro ekibine iletildi.', 'success')
    return redirect(url_for('ik.detay', id=id))


@ik_bp.route('/isten-cikis-bildirimleri')
@login_required
@permission_required('ik.view')
def isten_cikis_bildirimleri():
    """Tüm işten çıkış bildirimleri listesi (scope filtreli). Filtre: durum, proje."""
    durum = request.args.get('durum')
    proje_id = request.args.get('proje_id', type=int)

    # Scope: kullanıcı yalnızca erişebildiği çalışanların bildirimlerini görür
    scoped_calisan_ids = apply_calisan_scope(db.session.query(Calisan.id))

    query = IstenCikisBildirimi.query.filter(
        IstenCikisBildirimi.calisan_id.in_(scoped_calisan_ids)
    )

    if durum in ('beklemede', 'isleme_alindi', 'tamamlandi'):
        query = query.filter(IstenCikisBildirimi.durum == durum)

    if proje_id:
        kadro_ids = db.session.query(HedefKadro.id).filter(HedefKadro.proje_id == proje_id)
        query = query.join(Calisan, IstenCikisBildirimi.calisan_id == Calisan.id).filter(
            Calisan.kadro_id.in_(kadro_ids)
        )

    bildirimler = query.order_by(IstenCikisBildirimi.created_at.desc()).all()
    projeler = user_scoped_projeler()

    # Durum bazlı sayaçlar (scope'lu, filtresiz temel üzerinden)
    temel = IstenCikisBildirimi.query.filter(
        IstenCikisBildirimi.calisan_id.in_(apply_calisan_scope(db.session.query(Calisan.id)))
    )
    sayilar = {
        'toplam': temel.count(),
        'beklemede': temel.filter(IstenCikisBildirimi.durum == 'beklemede').count(),
        'isleme_alindi': temel.filter(IstenCikisBildirimi.durum == 'isleme_alindi').count(),
        'tamamlandi': temel.filter(IstenCikisBildirimi.durum == 'tamamlandi').count(),
    }

    return render_template('ik/isten_cikis_bildirimleri.html',
                           bildirimler=bildirimler,
                           projeler=projeler,
                           sayilar=sayilar,
                           secili_durum=durum,
                           secili_proje_id=proje_id,
                           active='ik-isten-cikis-bildirimleri')


@ik_bp.route('/isten-cikis-bildirimi/<int:id>/durum', methods=['POST'])
@login_required
@permission_required('ik.edit')
def isten_cikis_bildirimi_durum(id):
    """İK bildirimin durumunu günceller: beklemede -> isleme_alindi -> tamamlandi."""
    bildirim = IstenCikisBildirimi.query.get_or_404(id)

    yeni_durum = (request.form.get('durum') or '').strip()
    if yeni_durum not in IstenCikisBildirimi.DURUMLAR:
        flash('Geçersiz durum.', 'danger')
        return redirect(url_for('ik.isten_cikis_bildirimleri'))

    bildirim.durum = yeni_durum

    # Bildirim "Tamamlandı" yapıldığında, SPV'nin bildirdiği son çalışma gününü
    # çalışanın ayrılış tarihine yaz (yalnızca alan boşsa - mevcut resmi tarihi ezmez).
    if yeni_durum == 'tamamlandi' and bildirim.son_calisma_gunu:
        calisan = bildirim.calisan
        if calisan and not calisan.isten_ayrilma:
            calisan.isten_ayrilma = bildirim.son_calisma_gunu

    db.session.commit()

    flash('Bildirim durumu güncellendi.', 'success')
    next_url = request.form.get('next')
    if next_url == 'detay':
        return redirect(url_for('ik.detay', id=bildirim.calisan_id))
    return redirect(url_for('ik.isten_cikis_bildirimleri'))


@ik_bp.app_context_processor
def inject_cikis_bildirimi_count():
    """Sidebar rozeti için beklemedeki işten çıkış bildirimi sayısı (scope filtreli)."""
    def cikis_bildirimi_bekleyen_count():
        if not current_user.is_authenticated:
            return 0
        try:
            scoped_ids = apply_calisan_scope(db.session.query(Calisan.id))
            return IstenCikisBildirimi.query.filter(
                IstenCikisBildirimi.calisan_id.in_(scoped_ids),
                IstenCikisBildirimi.durum == 'beklemede',
            ).count()
        except Exception:
            db.session.rollback()
            return 0
    return dict(cikis_bildirimi_bekleyen_count=cikis_bildirimi_bekleyen_count)


@ik_bp.route('/isten-cikis/<int:id>/tamamla', methods=['POST'])
@login_required
@permission_required('ik.edit')
def isten_cikis_tamamla(id):
    """İşten çıkışı tamamla"""
    cikis = IstenCikis.query.get_or_404(id)

    cikis.durum = 'tamamlandi'

    # Gerçekleşen çıkış (SGK çıkış) tarihi = personelin SON ÇALIŞMA GÜNÜ.
    # Öncelik: SPV/koordinatör işten çıkış bildirimindeki son_calisma_gunu,
    # yoksa İK'nın formda girdiği planlanan çıkış tarihi, o da yoksa bugün.
    bildirim = IstenCikisBildirimi.query.filter_by(
        calisan_id=cikis.calisan_id
    ).order_by(IstenCikisBildirimi.created_at.desc()).first()

    if bildirim and bildirim.son_calisma_gunu:
        cikis.gerceklesen_cikis_tarihi = bildirim.son_calisma_gunu
    elif cikis.planlanan_cikis_tarihi:
        cikis.gerceklesen_cikis_tarihi = cikis.planlanan_cikis_tarihi
    else:
        cikis.gerceklesen_cikis_tarihi = date.today()

    # Çalışan durumunu güncelle — ayrılma tarihi de son çalışma günü ile aynı
    calisan = cikis.calisan
    calisan.durum = CalisanDurumu.AYRILDI
    calisan.isten_ayrilma = cikis.gerceklesen_cikis_tarihi
    calisan.ayrilma_nedeni = f"{cikis.cikis_tipi}: {cikis.cikis_sebebi}"

    # Akış: resmi çıkış tamamlandı -> SPV ön bildirimi (varsa) SGK çıkışı bekliyor
    # durumuna geçer (bordro SGK çıkışını yapıp bildirge yükleyene kadar).
    if bildirim and bildirim.durum != 'sgk_cikis_yapildi':
        bildirim.durum = 'sgk_cikis_bekleniyor'

    db.session.commit()

    # İşten çıkış bildirimi gönder
    print(f"[İşten Çıkış] Çalışan: {calisan.ad} {calisan.soyad}, tarih={cikis.gerceklesen_cikis_tarihi}")
    try:
        from app.services.notification import notify_isten_cikis
        sonuc = notify_isten_cikis(
            calisan=calisan,
            cikis_tarihi=cikis.gerceklesen_cikis_tarihi,
            cikis_nedeni=f"{cikis.cikis_tipi}: {cikis.cikis_sebebi}",
            zimmet_teslim=cikis.zimmet_teslim,
            sgk_cikis_kodu=cikis.sgk_cikis_kodu,
            liste_durumu=calisan.liste_durumu,
            bildirim_tarihi=date.today(),
        )
        print(f"[İşten Çıkış] Bildirim sonucu: {sonuc}")
    except Exception as e:
        print(f"[İşten Çıkış] Bildirim HATA: {e}")
        import traceback
        traceback.print_exc()

    flash('İşten çıkış tamamlandı.', 'success')
    return redirect(url_for('ik.isten_cikis_detay', id=id))


# ============================================================
# SAHA DASHBOARD (Müdürlük Bazlı Canlı Ekran / "Borsa Ekranı")
# ============================================================

def _parse_mudurluk(pozisyon_adi):
    """HedefKadro.pozisyon_adi'ndaki ilk ' - ' öncesi kısmı müdürlük olarak döndürür.
    Örn: 'Akdeniz Md. - Antalya - P.T Sniper' -> 'Akdeniz Md.'
    ' - ' yoksa None."""
    if not pozisyon_adi:
        return None
    parts = pozisyon_adi.split(' - ')
    if len(parts) >= 2:
        return parts[0].strip() or None
    return None


def _cinsiyet_kisa(deger):
    return {'kadin': 'K', 'erkek': 'E'}.get((deger or '').lower(), (deger or '')[:1].upper() or '-')


def _saha_dashboard_verileri():
    """Saha dashboard için tüm veriyi hesaplar (sayfa + Excel export ortak kullanır).
    request.args'tan proje_id / mudurluk / tarih okur, scope uygular."""
    from sqlalchemy import func
    from app.models.proje import AKTIF_SUREC_DURUMLARI

    projeler = user_scoped_projeler()
    proje_ids = [p.id for p in projeler]

    # Seçili proje - varsayılan 500 Ek Sniper (proje 12), scope'ta değilse ilk proje
    secili_proje_id = request.args.get('proje_id', type=int)
    if secili_proje_id not in proje_ids:
        secili_proje_id = 12 if 12 in proje_ids else (proje_ids[0] if proje_ids else None)

    secili_mudurluk = (request.args.get('mudurluk') or 'Tümü').strip() or 'Tümü'

    tarih_str = request.args.get('tarih')
    try:
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d').date() if tarih_str else date.today()
    except ValueError:
        tarih = date.today()

    secili_proje = next((p for p in projeler if p.id == secili_proje_id), None)

    # Projenin tüm kadroları (müdürlük dropdown'ı bunlardan doldurulur)
    tum_kadrolar = []
    if secili_proje_id:
        tum_kadrolar = HedefKadro.query.filter_by(
            proje_id=secili_proje_id, is_deleted=False, aktif=True
        ).all()

    mudurlukler = sorted({m for m in (_parse_mudurluk(k.pozisyon_adi) for k in tum_kadrolar) if m})

    # Müdürlük filtresi
    if secili_mudurluk != 'Tümü':
        kadrolar = [k for k in tum_kadrolar if _parse_mudurluk(k.pozisyon_adi) == secili_mudurluk]
    else:
        kadrolar = tum_kadrolar

    kadro_ids = [k.id for k in kadrolar]
    kadro_by_id = {k.id: k for k in kadrolar}

    # Aktif süreç aday durumları (sgk_giris_talebi hariç -> ayrı gösterilir)
    aday_durumlari = list(AKTIF_SUREC_DURUMLARI)
    aday_durumlari_sgk_haric = [d for d in aday_durumlari if d != 'sgk_giris_talebi']

    def _grup(query):
        return dict(query.all()) if kadro_ids else {}

    mevcut_map, giris_map, cikis_map, aday_map = {}, {}, {}, {}
    if kadro_ids:
        mevcut_map = _grup(db.session.query(Calisan.kadro_id, func.count(Calisan.id)).filter(
            Calisan.kadro_id.in_(kadro_ids), Calisan.is_deleted == False,
            Calisan.durum.in_([CalisanDurumu.AKTIF, CalisanDurumu.IZINLI])
        ).group_by(Calisan.kadro_id))

        giris_map = _grup(db.session.query(Calisan.kadro_id, func.count(Calisan.id)).filter(
            Calisan.kadro_id.in_(kadro_ids), Calisan.is_deleted == False,
            Calisan.ise_baslama == tarih, Calisan.durum == CalisanDurumu.AKTIF
        ).group_by(Calisan.kadro_id))

        cikis_map = _grup(db.session.query(Calisan.kadro_id, func.count(Calisan.id)).filter(
            Calisan.kadro_id.in_(kadro_ids), Calisan.is_deleted == False,
            Calisan.isten_ayrilma == tarih, Calisan.durum == CalisanDurumu.AYRILDI
        ).group_by(Calisan.kadro_id))

        aday_map = _grup(db.session.query(Aday.kadro_id, func.count(Aday.id)).filter(
            Aday.kadro_id.in_(kadro_ids), Aday.is_deleted == False,
            Aday.durum.in_(aday_durumlari)
        ).group_by(Aday.kadro_id))

    # Kadro bazlı satırlar
    kadro_rows = []
    for k in kadrolar:
        hedef = k.hedef_sayi or 0
        mevcut = mevcut_map.get(k.id, 0)
        doluluk = round((mevcut / hedef) * 100) if hedef else 0
        kadro_rows.append({
            'pozisyon': k.pozisyon_adi,
            'il': k.il_adi or '-',
            'mudurluk': _parse_mudurluk(k.pozisyon_adi) or '-',
            'hedef': hedef,
            'mevcut': mevcut,
            'giris': giris_map.get(k.id, 0),
            'cikis': cikis_map.get(k.id, 0),
            'aday': aday_map.get(k.id, 0),
            'doluluk': doluluk,
        })
    # En düşük doluluk üstte (dikkat çekmesi için)
    kadro_rows.sort(key=lambda r: r['doluluk'])

    # SGK bekleyen (kadro bazlı sayı - özet için)
    sgk_bekleyen_sayi = 0
    if kadro_ids:
        sgk_bekleyen_sayi = Aday.query.filter(
            Aday.kadro_id.in_(kadro_ids), Aday.is_deleted == False,
            Aday.durum == 'sgk_giris_talebi'
        ).count()

    hedef_top = sum(r['hedef'] for r in kadro_rows)
    mevcut_top = sum(r['mevcut'] for r in kadro_rows)
    ozet = {
        'hedef': hedef_top,
        'mevcut': mevcut_top,
        'doluluk': round((mevcut_top / hedef_top) * 100) if hedef_top else 0,
        'bugun_giris': sum(r['giris'] for r in kadro_rows),
        'bugun_cikis': sum(r['cikis'] for r in kadro_rows),
        'toplam_aday': sum(r['aday'] for r in kadro_rows),
        'sgk_bekleyen': sgk_bekleyen_sayi,
    }

    # ---- Detay satırları (bugünkü hareketler + adaylar) ----
    detay_rows = []

    def _il_of(rec, kadro):
        return (getattr(rec, 'il', None) or (kadro.il_adi if kadro else None) or '-')

    if kadro_ids:
        # 🟢 GİRİŞ
        for c in Calisan.query.filter(
            Calisan.kadro_id.in_(kadro_ids), Calisan.is_deleted == False,
            Calisan.ise_baslama == tarih, Calisan.durum == CalisanDurumu.AKTIF
        ).all():
            kadro = kadro_by_id.get(c.kadro_id)
            detay_rows.append({
                'order': 0, 'tip': 'giris', 'tip_text': 'GİRİŞ',
                'ad_soyad': c.full_name, 'telefon': c.telefon or '-',
                'pozisyon': kadro.pozisyon_adi if kadro else (c.pozisyon.ad if c.pozisyon else '-'),
                'il': _il_of(c, kadro), 'cinsiyet': _cinsiyet_kisa(c.cinsiyet),
                'tarih': c.ise_baslama, 'durum': 'İşe Başladı',
            })
        # 🔴 ÇIKIŞ
        for c in Calisan.query.filter(
            Calisan.kadro_id.in_(kadro_ids), Calisan.is_deleted == False,
            Calisan.isten_ayrilma == tarih, Calisan.durum == CalisanDurumu.AYRILDI
        ).all():
            kadro = kadro_by_id.get(c.kadro_id)
            detay_rows.append({
                'order': 1, 'tip': 'cikis', 'tip_text': 'ÇIKIŞ',
                'ad_soyad': c.full_name, 'telefon': c.telefon or '-',
                'pozisyon': kadro.pozisyon_adi if kadro else (c.pozisyon.ad if c.pozisyon else '-'),
                'il': _il_of(c, kadro), 'cinsiyet': _cinsiyet_kisa(c.cinsiyet),
                'tarih': c.isten_ayrilma, 'durum': 'Ayrıldı',
            })
        # 🔵 SGK BEKLİYOR
        for a in Aday.query.filter(
            Aday.kadro_id.in_(kadro_ids), Aday.is_deleted == False,
            Aday.durum == 'sgk_giris_talebi'
        ).all():
            kadro = kadro_by_id.get(a.kadro_id)
            detay_rows.append({
                'order': 2, 'tip': 'sgk', 'tip_text': 'SGK BEKLİYOR',
                'ad_soyad': a.full_name, 'telefon': a.telefon or '-',
                'pozisyon': kadro.pozisyon_adi if kadro else '-',
                'il': _il_of(a, kadro), 'cinsiyet': _cinsiyet_kisa(a.cinsiyet),
                'tarih': a.planlanan_baslangic, 'durum': a.basvuru_durumu_text,
            })
        # 🟡 ADAY (aktif süreç, sgk hariç)
        for a in Aday.query.filter(
            Aday.kadro_id.in_(kadro_ids), Aday.is_deleted == False,
            Aday.durum.in_(aday_durumlari_sgk_haric)
        ).all():
            kadro = kadro_by_id.get(a.kadro_id)
            detay_rows.append({
                'order': 3, 'tip': 'aday', 'tip_text': 'ADAY',
                'ad_soyad': a.full_name, 'telefon': a.telefon or '-',
                'pozisyon': kadro.pozisyon_adi if kadro else '-',
                'il': _il_of(a, kadro), 'cinsiyet': _cinsiyet_kisa(a.cinsiyet),
                'tarih': a.planlanan_baslangic, 'durum': a.basvuru_durumu_text,
            })

    detay_rows.sort(key=lambda r: (r['order'], r['ad_soyad']))

    return {
        'projeler': projeler,
        'secili_proje': secili_proje,
        'secili_proje_id': secili_proje_id,
        'mudurlukler': mudurlukler,
        'secili_mudurluk': secili_mudurluk,
        'tarih': tarih,
        'ozet': ozet,
        'kadro_rows': kadro_rows,
        'detay_rows': detay_rows,
    }


@ik_bp.route('/saha-dashboard')
@login_required
@permission_required('ik.view')
def saha_dashboard():
    """Müdürlük bazlı canlı saha dashboard (borsa ekranı). 30sn otomatik yenilenir."""
    veri = _saha_dashboard_verileri()
    return render_template('ik/saha_dashboard.html', active='ik-saha-dashboard', **veri)


@ik_bp.route('/saha-dashboard/export')
@login_required
@permission_required('ik.view')
def saha_dashboard_export():
    """Saha dashboard - kadro özeti + hareketler Excel export."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    veri = _saha_dashboard_verileri()

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='137FEC')

    # Sayfa 1: Kadro Özet
    ws1 = wb.active
    ws1.title = 'Kadro Özet'
    h1 = ['Müdürlük', 'Pozisyon', 'İl', 'Hedef', 'Mevcut', 'Bugün Giriş', 'Bugün Çıkış', 'Aktif Aday', 'Doluluk %']
    ws1.append(h1)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in veri['kadro_rows']:
        ws1.append([r['mudurluk'], r['pozisyon'], r['il'], r['hedef'], r['mevcut'],
                    r['giris'], r['cikis'], r['aday'], r['doluluk']])
    for idx, w in enumerate([18, 30, 14, 8, 8, 12, 12, 11, 10], start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # Sayfa 2: Hareketler
    ws2 = wb.create_sheet('Hareketler')
    h2 = ['Tip', 'Ad Soyad', 'Telefon', 'Pozisyon', 'İl', 'Cinsiyet', 'Planlı Tarih', 'Durum']
    ws2.append(h2)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for r in veri['detay_rows']:
        ws2.append([r['tip_text'], r['ad_soyad'], r['telefon'], r['pozisyon'], r['il'],
                    r['cinsiyet'], r['tarih'].strftime('%d.%m.%Y') if r['tarih'] else '',
                    r['durum']])
    for idx, w in enumerate([14, 26, 16, 30, 14, 9, 14, 20], start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    proje_ad = (veri['secili_proje'].ad if veri['secili_proje'] else 'saha').replace(' ', '_')
    filename = f"saha_dashboard_{proje_ad}_{veri['tarih'].strftime('%Y%m%d')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


# ============================================================
# İZİN YÖNETİMİ
# ============================================================

@ik_bp.route('/izinler')
@login_required
@permission_required('ik.view')
def izin_liste():
    """İzin talepleri listesi"""
    page = request.args.get('page', 1, type=int)
    durum = request.args.get('durum')
    
    query = Izin.query
    
    if durum:
        query = query.filter(Izin.durum == durum)
    
    query = query.order_by(Izin.created_at.desc())
    pagination = paginate_query(query, page, 20)
    
    # Bekleyen sayısı
    bekleyen_count = Izin.query.filter_by(durum='beklemede').count()
    
    return render_template('ik/izin_liste.html',
                          izinler=pagination.items,
                          pagination=pagination,
                          bekleyen_count=bekleyen_count)


@ik_bp.route('/izin/<int:id>/onayla', methods=['POST'])
@login_required
@permission_required('ik.edit')
def izin_onayla(id):
    """İzin talebi onayla"""
    izin = Izin.query.get_or_404(id)
    izin.durum = 'onaylandi'
    izin.onaylayan_id = current_user.id
    izin.onay_tarihi = datetime.now()
    db.session.commit()
    
    flash('İzin talebi onaylandı.', 'success')
    return redirect(url_for('ik.izin_liste'))


@ik_bp.route('/izin/<int:id>/reddet', methods=['POST'])
@login_required
@permission_required('ik.edit')
def izin_reddet(id):
    """İzin talebi reddet"""
    izin = Izin.query.get_or_404(id)
    izin.durum = 'reddedildi'
    izin.red_nedeni = request.form.get('red_nedeni')
    izin.onaylayan_id = current_user.id
    izin.onay_tarihi = datetime.now()
    db.session.commit()
    
    flash('İzin talebi reddedildi.', 'success')
    return redirect(url_for('ik.izin_liste'))

# ============================================================
# ZİMMET YÖNETİMİ
# ============================================================

# NOT: Önce import'lara şunları ekleyin:
# from app.models.ik import ZimmetTipi, Zimmet, ZimmetLog

@ik_bp.route('/zimmetler')
@login_required
@permission_required('ik.view')
def zimmet_liste():
    """Zimmet listesi"""
    page = request.args.get('page', 1, type=int)
    durum = request.args.get('durum')
    tip_id = request.args.get('tip_id', type=int)
    search = request.args.get('search', '').strip()
    
    query = Zimmet.query.filter_by(is_deleted=False)
    
    if durum:
        query = query.filter(Zimmet.durum == durum)
    if tip_id:
        query = query.filter(Zimmet.zimmet_tipi_id == tip_id)
    if search:
        search_filter = f'%{search}%'
        query = query.join(Calisan).filter(
            db.or_(
                Zimmet.tanim.ilike(search_filter),
                Zimmet.seri_no.ilike(search_filter),
                Zimmet.demirbas_no.ilike(search_filter),
                Calisan.ad.ilike(search_filter),
                Calisan.soyad.ilike(search_filter)
            )
        )
    
    query = query.order_by(Zimmet.created_at.desc())
    pagination = paginate_query(query, page, 20)
    
    zimmet_tipleri = ZimmetTipi.query.filter_by(aktif=True).order_by(ZimmetTipi.ad).all()
    
    # İstatistikler
    stats = {
        'toplam': Zimmet.query.filter_by(is_deleted=False).count(),
        'teslim_edildi': Zimmet.query.filter_by(is_deleted=False, durum='teslim_edildi').count(),
        'iade_edildi': Zimmet.query.filter_by(is_deleted=False, durum='iade_edildi').count(),
        'kayip': Zimmet.query.filter_by(is_deleted=False, durum='kayip').count(),
    }
    
    return render_template('ik/zimmet_liste.html',
                          zimmetler=pagination.items,
                          pagination=pagination,
                          zimmet_tipleri=zimmet_tipleri,
                          stats=stats)


@ik_bp.route('/zimmet/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.create')
def zimmet_ekle():
    """Yeni zimmet ekle"""
    if request.method == 'POST':
        zimmet = Zimmet(
            calisan_id=int(request.form['calisan_id']),
            zimmet_tipi_id=int(request.form['zimmet_tipi_id']),
            tanim=request.form.get('tanim', '').strip(),
            seri_no=request.form.get('seri_no', '').strip() or None,
            demirbas_no=request.form.get('demirbas_no', '').strip() or None,
            marka=request.form.get('marka', '').strip() or None,
            model=request.form.get('model', '').strip() or None,
            teslim_tarihi=datetime.strptime(request.form['teslim_tarihi'], '%Y-%m-%d').date(),
            teslim_eden_id=current_user.id,
            teslim_notu=request.form.get('teslim_notu', '').strip() or None,
            deger=Decimal(request.form['deger']) if request.form.get('deger') else None,
            durum='teslim_edildi'
        )
        db.session.add(zimmet)
        db.session.flush()
        
        # Log kaydı
        log = ZimmetLog(
            zimmet_id=zimmet.id,
            islem='teslim',
            aciklama=f'Zimmet {zimmet.calisan.full_name} adlı çalışana teslim edildi.',
            islem_yapan_id=current_user.id,
            yeni_calisan_id=zimmet.calisan_id
        )
        db.session.add(log)
        db.session.commit()
        
        flash('Zimmet başarıyla eklendi.', 'success')
        return redirect(url_for('ik.zimmet_detay', id=zimmet.id))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).order_by(Calisan.ad).all()
    zimmet_tipleri = ZimmetTipi.query.filter_by(aktif=True).order_by(ZimmetTipi.ad).all()
    
    # URL'den gelen calisan_id varsa
    calisan_id = request.args.get('calisan_id', type=int)
    
    return render_template('ik/zimmet_form.html',
                          zimmet=None,
                          calisanlar=calisanlar,
                          zimmet_tipleri=zimmet_tipleri,
                          secili_calisan_id=calisan_id)


@ik_bp.route('/zimmet/<int:id>')
@login_required
@permission_required('ik.view')
def zimmet_detay(id):
    """Zimmet detay"""
    zimmet = Zimmet.query.get_or_404(id)
    loglar = zimmet.loglar.order_by(ZimmetLog.created_at.desc()).all()
    
    return render_template('ik/zimmet_detay.html',
                          zimmet=zimmet,
                          loglar=loglar)


@ik_bp.route('/zimmet/<int:id>/iade', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def zimmet_iade(id):
    """Zimmet iade al"""
    zimmet = Zimmet.query.get_or_404(id)
    
    if zimmet.durum != 'teslim_edildi':
        flash('Bu zimmet zaten iade edilmiş veya durumu değiştirilmiş.', 'warning')
        return redirect(url_for('ik.zimmet_detay', id=id))
    
    if request.method == 'POST':
        zimmet.iade_tarihi = datetime.strptime(request.form['iade_tarihi'], '%Y-%m-%d').date()
        zimmet.iade_alan_id = current_user.id
        zimmet.iade_notu = request.form.get('iade_notu', '').strip() or None
        zimmet.iade_durumu = request.form.get('iade_durumu', 'saglam')
        zimmet.durum = 'iade_edildi'
        
        # Log kaydı
        log = ZimmetLog(
            zimmet_id=zimmet.id,
            islem='iade',
            aciklama=f'Zimmet {zimmet.calisan.full_name} tarafından iade edildi. Durum: {zimmet.iade_durumu}',
            islem_yapan_id=current_user.id,
            eski_calisan_id=zimmet.calisan_id
        )
        db.session.add(log)
        db.session.commit()
        
        flash('Zimmet başarıyla iade alındı.', 'success')
        return redirect(url_for('ik.zimmet_detay', id=id))
    
    return render_template('ik/zimmet_iade.html', zimmet=zimmet)


@ik_bp.route('/zimmet/<int:id>/transfer', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def zimmet_transfer(id):
    """Zimmet başka çalışana transfer et"""
    zimmet = Zimmet.query.get_or_404(id)
    
    if zimmet.durum != 'teslim_edildi':
        flash('Sadece teslim edilmiş zimmetler transfer edilebilir.', 'warning')
        return redirect(url_for('ik.zimmet_detay', id=id))
    
    if request.method == 'POST':
        eski_calisan_id = zimmet.calisan_id
        yeni_calisan_id = int(request.form['yeni_calisan_id'])
        
        if eski_calisan_id == yeni_calisan_id:
            flash('Zimmet zaten bu çalışanda.', 'warning')
            return redirect(url_for('ik.zimmet_transfer', id=id))
        
        zimmet.calisan_id = yeni_calisan_id
        zimmet.teslim_tarihi = datetime.strptime(request.form['transfer_tarihi'], '%Y-%m-%d').date()
        
        # Log kaydı
        eski_calisan = Calisan.query.get(eski_calisan_id)
        yeni_calisan = Calisan.query.get(yeni_calisan_id)
        
        log = ZimmetLog(
            zimmet_id=zimmet.id,
            islem='transfer',
            aciklama=f'Zimmet {eski_calisan.full_name} → {yeni_calisan.full_name} transfer edildi.',
            islem_yapan_id=current_user.id,
            eski_calisan_id=eski_calisan_id,
            yeni_calisan_id=yeni_calisan_id
        )
        db.session.add(log)
        db.session.commit()
        
        flash(f'Zimmet {yeni_calisan.full_name} adlı çalışana transfer edildi.', 'success')
        return redirect(url_for('ik.zimmet_detay', id=id))
    
    calisanlar = Calisan.query.filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        Calisan.id != zimmet.calisan_id
    ).order_by(Calisan.ad).all()
    
    return render_template('ik/zimmet_transfer.html',
                          zimmet=zimmet,
                          calisanlar=calisanlar)


@ik_bp.route('/zimmet/<int:id>/kayip', methods=['POST'])
@login_required
@permission_required('ik.edit')
def zimmet_kayip_bildir(id):
    """Zimmet kayıp bildirimi"""
    zimmet = Zimmet.query.get_or_404(id)
    zimmet.durum = 'kayip'
    
    log = ZimmetLog(
        zimmet_id=zimmet.id,
        islem='kayip_bildirimi',
        aciklama=request.form.get('aciklama', 'Zimmet kayıp olarak bildirildi.'),
        islem_yapan_id=current_user.id
    )
    db.session.add(log)
    db.session.commit()
    
    flash('Zimmet kayıp olarak işaretlendi.', 'warning')
    return redirect(url_for('ik.zimmet_detay', id=id))


# ============================================================
# ZİMMET TİPLERİ YÖNETİMİ
# ============================================================

@ik_bp.route('/zimmet-tipleri')
@login_required
@permission_required('ik.view')
def zimmet_tipleri():
    """Zimmet tipleri listesi"""
    tipler = ZimmetTipi.query.order_by(ZimmetTipi.ad).all()
    return render_template('ik/zimmet_tipleri.html', zimmet_tipleri=tipler)


@ik_bp.route('/zimmet-tipi/ekle', methods=['POST'])
@login_required
@permission_required('ik.edit')
def zimmet_tipi_ekle():
    """Yeni zimmet tipi ekle"""
    tip = ZimmetTipi(
        ad=request.form.get('ad'),
        kod=request.form.get('kod'),
        kategori=request.form.get('kategori'),
        aciklama=request.form.get('aciklama'),
        seri_no_zorunlu=request.form.get('seri_no_zorunlu') == 'on',
        iade_zorunlu=request.form.get('iade_zorunlu') == 'on'
    )
    db.session.add(tip)
    db.session.commit()
    
    flash('Zimmet tipi eklendi.', 'success')
    return redirect(url_for('ik.zimmet_tipleri'))


# ============================================================
# ÇALIŞAN ZİMMETLERİ
# ============================================================

@ik_bp.route('/calisan/<int:id>/zimmetler')
@login_required
@permission_required('ik.view')
def calisan_zimmetler(id):
    """Çalışanın zimmetleri"""
    calisan = Calisan.query.get_or_404(id)
    
    aktif_zimmetler = calisan.zimmetler.filter(
        Zimmet.durum == 'teslim_edildi',
        Zimmet.is_deleted == False
    ).all()
    
    gecmis_zimmetler = calisan.zimmetler.filter(
        Zimmet.durum != 'teslim_edildi',
        Zimmet.is_deleted == False
    ).order_by(Zimmet.iade_tarihi.desc()).all()
    
    return render_template('ik/calisan_zimmetler.html',
                          calisan=calisan,
                          aktif_zimmetler=aktif_zimmetler,
                          gecmis_zimmetler=gecmis_zimmetler)


# ============================================================
# ADAY EKLE (Manuel - İK Panelinden)
# ============================================================

@ik_bp.route('/aday/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.create')
def aday_ekle():
    """Yeni aday ekle (manuel)"""
    if request.method == 'POST':
        # Telefon kontrolü
        telefon_ham = request.form.get('telefon', '').strip()
        telefon = normalize_telefon(telefon_ham)
        if telefon_ham and not telefon:
            flash('Geçersiz telefon numarası. Örnek format: 05XX XXX XX XX', 'danger')
            return redirect(url_for('ik.aday_ekle'))
        if telefon and Aday.query.filter_by(telefon=telefon, is_deleted=False).first():
            flash('Bu telefon numarası zaten kayıtlı.', 'danger')
            return redirect(url_for('ik.aday_ekle'))

        # 'VAR' = ehliyeti var ama sınıfı bilinmiyor; boş = ehliyet yok
        ehliyet_sec = request.form.get('ehliyet_sinifi', '').strip()

        aday = Aday(
            ad=request.form.get('ad', '').strip(),
            soyad=request.form.get('soyad', '').strip(),
            telefon=telefon,
            email=request.form.get('email', '').strip() or None,
            tc_kimlik=request.form.get('tc_kimlik', '').strip() or None,
            dogum_tarihi=datetime.strptime(request.form.get('dogum_tarihi'), '%Y-%m-%d').date() if request.form.get('dogum_tarihi') else None,
            cinsiyet=request.form.get('cinsiyet') or None,
            ehliyet_var=bool(ehliyet_sec),
            ehliyet_sinifi=None if ehliyet_sec in ('', 'VAR') else ehliyet_sec,
            adres=request.form.get('adres', '').strip() or None,
            il=request.form.get('il', '').strip() or None,
            ilce=request.form.get('ilce', '').strip() or None,
            iban=request.form.get('iban', '').strip().replace(' ', '').upper() or None,
            pozisyon_id=int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None,
            kaynak=request.form.get('kaynak', 'manuel'),
            durum='basvurdu',
            basvuru_tarihi=date.today(),
            kvkk_onay=True,  # Manuel girişte KVKK zaten alınmış kabul edilir
            kvkk_onay_tarihi=datetime.now(),
            davet_eden_id=current_user.id
        )
        
        db.session.add(aday)
        db.session.commit()
        
        flash(f'{aday.full_name} adayı oluşturuldu.', 'success')
        return redirect(url_for('ik.aday_detay', id=aday.id))
    
    pozisyonlar = Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all()
    
    return render_template('ik/aday_form.html',
                          aday=None,
                          pozisyonlar=pozisyonlar,
                          ehliyet_siniflari=EHLIYET_SINIFLARI)




# ============================================================
# ÇALIŞAN KULLANICI OLUŞTUR
# ============================================================
@ik_bp.route('/<int:id>/kullanici-olustur', methods=['POST'])
@login_required
@permission_required('ik.edit')
def kullanici_olustur(id):
    """Çalışan için portal kullanıcısı oluştur"""
    import secrets
    from app.models.core import User
    from app.services.notification import send_notification
    from flask import render_template as rt
    
    calisan = Calisan.query.get_or_404(id)
    
    # Zaten kullanıcısı var mı?
    if calisan.user_account:
        flash('Bu çalışanın zaten bir portal kullanıcısı var.', 'warning')
        return redirect(url_for('ik.detay', id=id))
    
    # Email zorunlu
    if not calisan.email:
        flash('Kullanıcı oluşturmak için çalışanın email adresi gerekli.', 'error')
        return redirect(url_for('ik.detay', id=id))
    
    # Email zaten kullanılıyor mu?
    existing_user = User.query.filter_by(email=calisan.email, is_deleted=False).first()
    if existing_user:
        flash('Bu email adresi ile zaten bir kullanıcı mevcut.', 'error')
        return redirect(url_for('ik.detay', id=id))
    
    # Rastgele şifre oluştur
    password = secrets.token_urlsafe(12)
    
    # Kullanıcı oluştur
    user = User(
        email=calisan.email,
        ad=calisan.ad,
        soyad=calisan.soyad,
        telefon=normalize_telefon(calisan.telefon),
        is_active=True,
        calisan_id=calisan.id
    )
    user.set_password(password)
    
    db.session.add(user)
    db.session.commit()
    
    # Email gönder
    try:
        html_body = rt('emails/kullanici_olusturuldu.html', calisan=calisan, email=calisan.email, password=password)
        send_notification(
            to=calisan.email,
            subject='TG Portal - Kullanıcı Hesabınız Oluşturuldu',
            html_body=html_body
        )
        flash(f'Kullanıcı oluşturuldu ve giriş bilgileri {calisan.email} adresine gönderildi.', 'success')
    except Exception as e:
        flash(f'Kullanıcı oluşturuldu ama email gönderilemedi. Şifre: {password}', 'warning')

    return redirect(url_for('ik.detay', id=id))


# ============================================================
# ÇALIŞAN EVRAK YÖNETİMİ
# ============================================================

@ik_bp.route('/calisan/<int:id>/evraklar')
@login_required
@permission_required('ik.view')
def calisan_evraklar(id):
    """Çalışan evrakları listesi"""
    calisan = Calisan.query.get_or_404(id)
    evrak_tipleri = EvrakTipi.query.filter_by(aktif=True).order_by(EvrakTipi.sira).all()
    evraklar = calisan.evraklar.all() if hasattr(calisan, 'evraklar') else []
    
    from datetime import date as dt_date
    return render_template('ik/calisan_evraklar.html',
                          today=dt_date.today(),
                          calisan=calisan,
                          evraklar=evraklar,
                          evrak_tipleri=evrak_tipleri)


@ik_bp.route('/calisan/<int:id>/evrak', methods=['POST'])
@login_required
@permission_required('ik.edit')
def calisan_evrak_yukle(id):
    """Çalışan evrak yükle"""
    calisan = Calisan.query.get_or_404(id)
    
    if 'dosya' not in request.files:
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('ik.calisan_evraklar', id=id))
    
    dosya = request.files['dosya']
    if dosya.filename == '':
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('ik.calisan_evraklar', id=id))
    
    if dosya and allowed_file(dosya.filename):
        evrak_tipi_id = int(request.form['evrak_tipi_id'])
        
        # Dosya adı oluştur
        filename = secure_filename(dosya.filename)
        ext = filename.rsplit('.', 1)[1].lower()
        new_filename = f"calisan_{id}_{evrak_tipi_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}"
        
        # Klasör oluştur
        upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'evraklar', 'calisanlar', str(id))
        os.makedirs(upload_folder, exist_ok=True)
        
        # Dosyayı kaydet
        filepath = os.path.join(upload_folder, new_filename)
        dosya.save(filepath)
        
        # Geçerlilik tarihi
        gecerlilik = None
        if request.form.get('gecerlilik_bitis'):
            gecerlilik = datetime.strptime(request.form['gecerlilik_bitis'], '%Y-%m-%d').date()
        
        # Veritabanına ekle
        evrak = CalisanEvrak(
            calisan_id=id,
            evrak_tipi_id=evrak_tipi_id,
            dosya_adi=filename,
            dosya_yolu=filepath,
            dosya_boyut=os.path.getsize(filepath),
            mime_type=dosya.content_type,
            gecerlilik_bitis=gecerlilik
        )
        db.session.add(evrak)
        db.session.commit()
        
        flash('Evrak başarıyla yüklendi.', 'success')
    else:
        flash('Geçersiz dosya formatı. (PDF, JPG, PNG, DOC, DOCX)', 'danger')
    
    return redirect(url_for('ik.calisan_evraklar', id=id))


@ik_bp.route('/calisan/evrak/<int:id>/indir')
@login_required
@permission_required('ik.view')
def calisan_evrak_indir(id):
    """Çalışan evrak indir"""
    evrak = CalisanEvrak.query.get_or_404(id)
    return send_file(evrak.dosya_yolu, as_attachment=True, download_name=evrak.dosya_adi)


@ik_bp.route('/calisan/evrak/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def calisan_evrak_sil(id):
    """Çalışan evrak sil"""
    evrak = CalisanEvrak.query.get_or_404(id)
    calisan_id = evrak.calisan_id
    
    # Dosyayı sil
    if os.path.exists(evrak.dosya_yolu):
        os.remove(evrak.dosya_yolu)
    
    db.session.delete(evrak)
    db.session.commit()
    
    flash('Evrak silindi.', 'success')
    return redirect(url_for('ik.calisan_evraklar', id=calisan_id))


@ik_bp.route('/aday/<int:id>/evraklar')
@login_required
@permission_required('ik.view')
def aday_evraklar(id):
    """Aday evrakları sayfası"""
    aday = Aday.query.get_or_404(id)
    evrak_tipleri = EvrakTipi.query.filter_by(aktif=True).order_by(EvrakTipi.sira).all()
    evraklar = aday.evraklar.all()
    
    # Eksik evraklar
    yuklenen_tipler = [e.evrak_tipi_id for e in aday.evraklar.filter(
        AdayEvrak.durum.in_(['yuklendi', 'onaylandi'])
    ).all()]
    eksik_evraklar = [t for t in EvrakTipi.query.filter_by(zorunlu=True, aktif=True).all() if t.id not in yuklenen_tipler]
    
    # Tamamlanma oranı
    zorunlu_count = EvrakTipi.query.filter_by(zorunlu=True, aktif=True).count()
    if zorunlu_count == 0:
        evrak_tamamlanma = 100
    else:
        onaylanan = aday.evraklar.join(EvrakTipi).filter(
            EvrakTipi.zorunlu == True,
            AdayEvrak.durum == 'onaylandi'
        ).count()
        evrak_tamamlanma = int((onaylanan / zorunlu_count) * 100)
    
    return render_template('ik/aday_evraklar.html',
                          aday=aday,
                          evraklar=evraklar,
                          evrak_tipleri=evrak_tipleri,
                          eksik_evraklar=eksik_evraklar,
                          evrak_tamamlanma=evrak_tamamlanma)



# ============================================================
# ORGANİZASYON ŞEMASI
# ============================================================

@ik_bp.route('/organizasyon')
@login_required
@permission_required('ik.view')
def organizasyon():
    """Organizasyon şeması"""
    from sqlalchemy import func
    
    # Tüm çalışanları seviyeye göre grupla
    calisanlar = Calisan.query.filter_by(is_deleted=False).all()
    
    # Seviyeye göre grupla (pozisyon seviyesi)
    seviye_gruplari = {}
    for calisan in calisanlar:
        seviye = calisan.pozisyon.seviye if calisan.pozisyon and calisan.pozisyon.seviye else 10
        if seviye not in seviye_gruplari:
            seviye_gruplari[seviye] = []
        seviye_gruplari[seviye].append(calisan)
    
    # Sıralı seviyeler
    seviyeler = sorted(seviye_gruplari.keys())
    
    # Hiyerarşi ağacı oluştur (yönetici-ast ilişkisi)
    def build_tree(calisan_id=None):
        children = []
        for c in calisanlar:
            if c.yonetici_id == calisan_id:
                children.append({
                    'calisan': c,
                    'children': build_tree(c.id)
                })
        return children
    
    # Kök düğümler (yöneticisi olmayan en üst seviyedekiler)
    tree = build_tree(None)
    
    # Departman bazlı gruplama (pozisyon seviyesine göre sıralı)
    departman_gruplari = {}
    for calisan in calisanlar:
        dept = calisan.departman.ad if calisan.departman else 'Diğer'
        if dept not in departman_gruplari:
            departman_gruplari[dept] = []
        departman_gruplari[dept].append(calisan)
    for dept in departman_gruplari:
        departman_gruplari[dept].sort(key=lambda c: (c.pozisyon.seviye if c.pozisyon and c.pozisyon.seviye else 999))
    
    return render_template('ik/organizasyon.html',
                          calisanlar=calisanlar,
                          seviye_gruplari=seviye_gruplari,
                          seviyeler=seviyeler,
                          tree=tree,
                          departman_gruplari=departman_gruplari)



# ============================================================
# DİSİPLİN YÖNETİMİ
# ============================================================

@ik_bp.route('/disiplin')
@login_required
@permission_required('ik.view')
def disiplin_liste():
    """Disiplin kayıtları listesi"""
    from app.models.ik import DisiplinKaydi
    
    kayitlar = DisiplinKaydi.query.filter_by(is_deleted=False).order_by(DisiplinKaydi.tarih.desc()).all()
    return render_template('ik/disiplin_liste.html', kayitlar=kayitlar)


@ik_bp.route('/disiplin/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def disiplin_ekle():
    """Yeni disiplin kaydı"""
    from app.models.ik import DisiplinKaydi
    
    if request.method == 'POST':
        kayit = DisiplinKaydi(
            calisan_id=int(request.form.get('calisan_id')),
            tarih=datetime.strptime(request.form.get('tarih'), '%Y-%m-%d').date(),
            tur=request.form.get('tur'),
            seviye=int(request.form.get('seviye', 1)),
            konu=request.form.get('konu'),
            aciklama=request.form.get('aciklama'),
            durum='onaylandi',
            olusturan_id=current_user.id
        )
        
        # Belge yükleme
        if 'belge' in request.files:
            belge = request.files['belge']
            if belge and belge.filename:
                import uuid
                ext = belge.filename.rsplit('.', 1)[-1].lower()
                filename = f"disiplin_{uuid.uuid4().hex[:8]}.{ext}"
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'disiplin')
                os.makedirs(upload_folder, exist_ok=True)
                belge.save(os.path.join(upload_folder, filename))
                kayit.belge_path = f"disiplin/{filename}"
        
        db.session.add(kayit)
        db.session.commit()
        
        flash('Disiplin kaydı oluşturuldu.', 'success')
        return redirect(url_for('ik.disiplin_liste'))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False).order_by(Calisan.ad).all()
    return render_template('ik/disiplin_form.html', kayit=None, calisanlar=calisanlar)


@ik_bp.route('/disiplin/<int:id>')
@login_required
@permission_required('ik.view')
def disiplin_detay(id):
    """Disiplin kaydı detay"""
    from app.models.ik import DisiplinKaydi
    
    kayit = DisiplinKaydi.query.get_or_404(id)
    return render_template('ik/disiplin_detay.html', kayit=kayit)


@ik_bp.route('/disiplin/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def disiplin_sil(id):
    """Disiplin kaydı sil"""
    from app.models.ik import DisiplinKaydi
    
    kayit = DisiplinKaydi.query.get_or_404(id)
    kayit.is_deleted = True
    db.session.commit()
    
    flash('Disiplin kaydı silindi.', 'success')
    return redirect(url_for('ik.disiplin_liste'))


# ============================================================
# DAVA YÖNETİMİ
# ============================================================

@ik_bp.route('/davalar')
@login_required
@permission_required('ik.view')
def dava_liste():
    """Dava listesi"""
    from app.models.ik import Dava
    
    davalar = Dava.query.filter_by(is_deleted=False).order_by(Dava.acilis_tarihi.desc()).all()
    
    # İstatistikler
    devam_eden = Dava.query.filter_by(is_deleted=False, durum='devam_ediyor').count()
    kapanan = Dava.query.filter_by(is_deleted=False, durum='kapandi').count()
    
    return render_template('ik/dava_liste.html', davalar=davalar, devam_eden=devam_eden, kapanan=kapanan)


@ik_bp.route('/dava/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def dava_ekle():
    """Yeni dava kaydı"""
    from app.models.ik import Dava
    
    if request.method == 'POST':
        dava = Dava(
            dosya_no=request.form.get('dosya_no'),
            esas_no=request.form.get('esas_no'),
            mahkeme=request.form.get('mahkeme'),
            dava_turu=request.form.get('dava_turu'),
            davaci=request.form.get('davaci'),
            davali=request.form.get('davali'),
            calisan_id=int(request.form.get('calisan_id')) if request.form.get('calisan_id') else None,
            acilis_tarihi=datetime.strptime(request.form.get('acilis_tarihi'), '%Y-%m-%d').date(),
            sonraki_durusma=datetime.strptime(request.form.get('sonraki_durusma'), '%Y-%m-%d').date() if request.form.get('sonraki_durusma') else None,
            talep_tutari=float(request.form.get('talep_tutari')) if request.form.get('talep_tutari') else None,
            konu_ozeti=request.form.get('konu_ozeti'),
            avukat=request.form.get('avukat'),
            avukat_telefon=request.form.get('avukat_telefon'),
            durum='devam_ediyor',
            sorumlu_id=current_user.id
        )
        
        db.session.add(dava)
        db.session.commit()
        
        flash('Dava kaydı oluşturuldu.', 'success')
        return redirect(url_for('ik.dava_liste'))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False).order_by(Calisan.ad).all()
    return render_template('ik/dava_form.html', dava=None, calisanlar=calisanlar)


@ik_bp.route('/dava/<int:id>')
@login_required
@permission_required('ik.view')
def dava_detay(id):
    """Dava detay"""
    from app.models.ik import Dava
    
    dava = Dava.query.get_or_404(id)
    return render_template('ik/dava_detay.html', dava=dava)


@ik_bp.route('/dava/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def dava_duzenle(id):
    """Dava düzenle"""
    from app.models.ik import Dava
    
    dava = Dava.query.get_or_404(id)
    
    if request.method == 'POST':
        dava.dosya_no = request.form.get('dosya_no')
        dava.esas_no = request.form.get('esas_no')
        dava.mahkeme = request.form.get('mahkeme')
        dava.dava_turu = request.form.get('dava_turu')
        dava.davaci = request.form.get('davaci')
        dava.davali = request.form.get('davali')
        dava.calisan_id = int(request.form.get('calisan_id')) if request.form.get('calisan_id') else None
        dava.acilis_tarihi = datetime.strptime(request.form.get('acilis_tarihi'), '%Y-%m-%d').date()
        dava.sonraki_durusma = datetime.strptime(request.form.get('sonraki_durusma'), '%Y-%m-%d').date() if request.form.get('sonraki_durusma') else None
        dava.son_durusma = datetime.strptime(request.form.get('son_durusma'), '%Y-%m-%d').date() if request.form.get('son_durusma') else None
        dava.karar_tarihi = datetime.strptime(request.form.get('karar_tarihi'), '%Y-%m-%d').date() if request.form.get('karar_tarihi') else None
        dava.talep_tutari = float(request.form.get('talep_tutari')) if request.form.get('talep_tutari') else None
        dava.karar_tutari = float(request.form.get('karar_tutari')) if request.form.get('karar_tutari') else None
        dava.durum = request.form.get('durum')
        dava.sonuc = request.form.get('sonuc') if request.form.get('sonuc') else None
        dava.konu_ozeti = request.form.get('konu_ozeti')
        dava.notlar = request.form.get('notlar')
        dava.avukat = request.form.get('avukat')
        dava.avukat_telefon = request.form.get('avukat_telefon')
        
        db.session.commit()
        
        flash('Dava güncellendi.', 'success')
        return redirect(url_for('ik.dava_detay', id=id))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False).order_by(Calisan.ad).all()
    return render_template('ik/dava_form.html', dava=dava, calisanlar=calisanlar)


@ik_bp.route('/dava/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def dava_sil(id):
    """Dava sil"""
    from app.models.ik import Dava
    
    dava = Dava.query.get_or_404(id)
    dava.is_deleted = True
    db.session.commit()
    
    flash('Dava kaydı silindi.', 'success')
    return redirect(url_for('ik.dava_liste'))



# ============================================================
# İCRA DOSYASI YÖNETİMİ
# ============================================================

@ik_bp.route('/icra')
@login_required
@permission_required('ik.view')
def icra_liste():
    """İcra dosyaları listesi"""
    from app.models.ik import IcraDosyasi
    
    dosyalar = IcraDosyasi.query.filter_by(is_deleted=False).order_by(IcraDosyasi.created_at.desc()).all()
    
    # İstatistikler
    aktif = sum(1 for d in dosyalar if d.durum == 'aktif')
    toplam_borc = sum(float(d.toplam_borc or 0) for d in dosyalar if d.durum == 'aktif')
    toplam_kesilen = sum(float(d.toplam_kesilen or 0) for d in dosyalar)
    
    return render_template('ik/icra_liste.html', 
                          dosyalar=dosyalar,
                          aktif=aktif,
                          toplam_borc=toplam_borc,
                          toplam_kesilen=toplam_kesilen)


@ik_bp.route('/icra/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def icra_ekle():
    """Yeni icra dosyası"""
    from app.models.ik import IcraDosyasi
    
    if request.method == 'POST':
        dosya = IcraDosyasi(
            calisan_id=int(request.form.get('calisan_id')),
            dosya_no=request.form.get('dosya_no'),
            icra_dairesi=request.form.get('icra_dairesi'),
            alacakli=request.form.get('alacakli'),
            toplam_borc=float(request.form.get('toplam_borc')),
            kalan_borc=float(request.form.get('toplam_borc')),
            taksit_sayisi=int(request.form.get('taksit_sayisi')) if request.form.get('taksit_sayisi') else None,
            taksit_tutari=float(request.form.get('taksit_tutari')) if request.form.get('taksit_tutari') else None,
            kesinti_orani=float(request.form.get('kesinti_orani')) if request.form.get('kesinti_orani') else None,
            baslangic_tarihi=datetime.strptime(request.form.get('baslangic_tarihi'), '%Y-%m-%d').date() if request.form.get('baslangic_tarihi') else None,
            notlar=request.form.get('notlar'),
            durum='aktif'
        )
        
        db.session.add(dosya)
        db.session.commit()
        
        flash('İcra dosyası oluşturuldu.', 'success')
        return redirect(url_for('ik.icra_detay', id=dosya.id))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False).order_by(Calisan.ad).all()
    return render_template('ik/icra_form.html', dosya=None, calisanlar=calisanlar)


@ik_bp.route('/icra/<int:id>')
@login_required
@permission_required('ik.view')
def icra_detay(id):
    """İcra dosyası detay"""
    from app.models.ik import IcraDosyasi
    
    dosya = IcraDosyasi.query.get_or_404(id)
    kesintiler = dosya.kesintiler.filter_by(is_deleted=False).order_by(IcraDosyasi.created_at.desc()).all()
    
    return render_template('ik/icra_detay.html', dosya=dosya, kesintiler=kesintiler)


@ik_bp.route('/icra/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def icra_duzenle(id):
    """İcra dosyası düzenle"""
    from app.models.ik import IcraDosyasi
    
    dosya = IcraDosyasi.query.get_or_404(id)
    
    if request.method == 'POST':
        dosya.dosya_no = request.form.get('dosya_no')
        dosya.icra_dairesi = request.form.get('icra_dairesi')
        dosya.alacakli = request.form.get('alacakli')
        dosya.toplam_borc = float(request.form.get('toplam_borc'))
        dosya.taksit_sayisi = int(request.form.get('taksit_sayisi')) if request.form.get('taksit_sayisi') else None
        dosya.taksit_tutari = float(request.form.get('taksit_tutari')) if request.form.get('taksit_tutari') else None
        dosya.kesinti_orani = float(request.form.get('kesinti_orani')) if request.form.get('kesinti_orani') else None
        dosya.baslangic_tarihi = datetime.strptime(request.form.get('baslangic_tarihi'), '%Y-%m-%d').date() if request.form.get('baslangic_tarihi') else None
        dosya.durum = request.form.get('durum')
        dosya.notlar = request.form.get('notlar')
        
        # Kalan borcu güncelle
        dosya.kalan_borc = float(dosya.toplam_borc) - float(dosya.toplam_kesilen)
        
        db.session.commit()
        
        flash('İcra dosyası güncellendi.', 'success')
        return redirect(url_for('ik.icra_detay', id=id))
    
    calisanlar = Calisan.query.filter_by(is_deleted=False).order_by(Calisan.ad).all()
    return render_template('ik/icra_form.html', dosya=dosya, calisanlar=calisanlar)


@ik_bp.route('/icra/<int:id>/kesinti-ekle', methods=['POST'])
@login_required
@permission_required('ik.edit')
def icra_kesinti_ekle(id):
    """İcra kesintisi ekle"""
    from app.models.ik import IcraDosyasi, IcraKesinti
    
    dosya = IcraDosyasi.query.get_or_404(id)
    
    kesinti = IcraKesinti(
        icra_dosyasi_id=id,
        donem=request.form.get('donem'),
        tutar=float(request.form.get('tutar')),
        kesinti_tarihi=datetime.strptime(request.form.get('kesinti_tarihi'), '%Y-%m-%d').date() if request.form.get('kesinti_tarihi') else None,
        durum=request.form.get('durum', 'kesildi'),
        notlar=request.form.get('notlar')
    )
    
    db.session.add(kesinti)
    
    # Kalan borcu güncelle
    dosya.kalan_borc = float(dosya.toplam_borc) - float(dosya.toplam_kesilen) - float(kesinti.tutar)
    if dosya.kalan_borc <= 0:
        dosya.kalan_borc = 0
        dosya.durum = 'tamamlandi'
    
    db.session.commit()
    
    flash('Kesinti kaydedildi.', 'success')
    return redirect(url_for('ik.icra_detay', id=id))


@ik_bp.route('/icra/kesinti/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def icra_kesinti_sil(id):
    """İcra kesintisi sil"""
    from app.models.ik import IcraKesinti
    
    kesinti = IcraKesinti.query.get_or_404(id)
    dosya_id = kesinti.icra_dosyasi_id
    kesinti.is_deleted = True
    
    # Kalan borcu güncelle
    dosya = kesinti.icra_dosyasi
    dosya.kalan_borc = float(dosya.toplam_borc) - float(dosya.toplam_kesilen)
    
    db.session.commit()
    
    flash('Kesinti silindi.', 'success')
    return redirect(url_for('ik.icra_detay', id=dosya_id))


@ik_bp.route('/icra/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def icra_sil(id):
    """İcra dosyası sil"""
    from app.models.ik import IcraDosyasi
    
    dosya = IcraDosyasi.query.get_or_404(id)
    dosya.is_deleted = True
    db.session.commit()
    
    flash('İcra dosyası silindi.', 'success')
    return redirect(url_for('ik.icra_liste'))


# ============================================================
# SÖZLEŞME ŞABLONLARI
# ============================================================

@ik_bp.route('/sozlesme-sablonlari')
@login_required
@permission_required('ik.view')
def sablon_liste():
    """Sözleşme şablonları listesi"""
    sablonlar = SozlesmeSablonu.query.filter_by(is_deleted=False).order_by(SozlesmeSablonu.sira, SozlesmeSablonu.ad).all()
    return render_template('ik/sablon_liste.html', sablonlar=sablonlar, active='ik-sablonlar')


@ik_bp.route('/sozlesme-sablonlari/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def sablon_ekle():
    """Yeni sözleşme şablonu ekle"""
    if request.method == 'POST':
        sablon = SozlesmeSablonu(
            ad=request.form.get('ad', '').strip(),
            tip=request.form.get('tip'),
            musteri_id=int(request.form.get('musteri_id')) if request.form.get('musteri_id') else None,
            proje_id=int(request.form.get('proje_id')) if request.form.get('proje_id') else None,
            pozisyon_id=int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None,
            departman_id=int(request.form.get('departman_id')) if request.form.get('departman_id') else None,
            aciklama=request.form.get('aciklama', '').strip() or None,
            aktif='aktif' in request.form
        )

        # Şablon PDF yükleme
        if 'sablon_dosya' in request.files:
            dosya = request.files['sablon_dosya']
            if dosya and dosya.filename and allowed_file(dosya.filename):
                import uuid
                ext = dosya.filename.rsplit('.', 1)[-1].lower()
                filename = f"sablon_{uuid.uuid4().hex[:8]}.{ext}"
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesme_sablonlari')
                os.makedirs(upload_folder, exist_ok=True)
                dosya.save(os.path.join(upload_folder, filename))
                sablon.sablon_dosya = f"sozlesme_sablonlari/{filename}"

        db.session.add(sablon)
        db.session.commit()
        flash(f'"{sablon.ad}" şablonu oluşturuldu.', 'success')
        return redirect(url_for('ik.sablon_liste'))

    from app.models.proje import Musteri, Proje
    return render_template('ik/sablon_form.html',
                          sablon=None,
                          tipler=SozlesmeSablonu.TIPLER,
                          musteriler=Musteri.query.filter_by(is_deleted=False, aktif=True).order_by(Musteri.ad).all(),
                          projeler=Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all(),
                          pozisyonlar=Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all(),
                          departmanlar=Departman.query.filter_by(aktif=True).order_by(Departman.ad).all())


@ik_bp.route('/sozlesme-sablonlari/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def sablon_duzenle(id):
    """Sözleşme şablonu düzenle"""
    sablon = SozlesmeSablonu.query.get_or_404(id)

    if request.method == 'POST':
        sablon.ad = request.form.get('ad', '').strip()
        sablon.tip = request.form.get('tip')
        sablon.musteri_id = int(request.form.get('musteri_id')) if request.form.get('musteri_id') else None
        sablon.proje_id = int(request.form.get('proje_id')) if request.form.get('proje_id') else None
        sablon.pozisyon_id = int(request.form.get('pozisyon_id')) if request.form.get('pozisyon_id') else None
        sablon.departman_id = int(request.form.get('departman_id')) if request.form.get('departman_id') else None
        sablon.aciklama = request.form.get('aciklama', '').strip() or None
        sablon.aktif = 'aktif' in request.form

        if 'sablon_dosya' in request.files:
            dosya = request.files['sablon_dosya']
            if dosya and dosya.filename and allowed_file(dosya.filename):
                import uuid
                ext = dosya.filename.rsplit('.', 1)[-1].lower()
                filename = f"sablon_{uuid.uuid4().hex[:8]}.{ext}"
                upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesme_sablonlari')
                os.makedirs(upload_folder, exist_ok=True)
                dosya.save(os.path.join(upload_folder, filename))
                sablon.sablon_dosya = f"sozlesme_sablonlari/{filename}"

        db.session.commit()
        flash(f'"{sablon.ad}" şablonu güncellendi.', 'success')
        return redirect(url_for('ik.sablon_liste'))

    from app.models.proje import Musteri, Proje
    return render_template('ik/sablon_form.html',
                          sablon=sablon,
                          tipler=SozlesmeSablonu.TIPLER,
                          musteriler=Musteri.query.filter_by(is_deleted=False, aktif=True).order_by(Musteri.ad).all(),
                          projeler=Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all(),
                          pozisyonlar=Pozisyon.query.filter_by(aktif=True).order_by(Pozisyon.ad).all(),
                          departmanlar=Departman.query.filter_by(aktif=True).order_by(Departman.ad).all())


@ik_bp.route('/sozlesme-sablonlari/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('ik.delete')
def sablon_sil(id):
    """Sözleşme şablonu soft-delete"""
    sablon = SozlesmeSablonu.query.get_or_404(id)
    sablon.is_deleted = True
    db.session.commit()
    flash(f'"{sablon.ad}" şablonu silindi.', 'success')
    return redirect(url_for('ik.sablon_liste'))


@ik_bp.route('/api/sozlesme-sablonlari')
@login_required
@permission_required('ik.view')
def api_sozlesme_sablonlari():
    """Sözleşme şablonlarını filtrele (AJAX)"""
    musteri_id = request.args.get('musteri_id', type=int)
    proje_id = request.args.get('proje_id', type=int)
    pozisyon_id = request.args.get('pozisyon_id', type=int)
    departman_id = request.args.get('departman_id', type=int)

    sablonlar = SozlesmeSablonu.sablonlari_filtrele(
        musteri_id=musteri_id, proje_id=proje_id,
        pozisyon_id=pozisyon_id, departman_id=departman_id
    )

    return jsonify([{
        'id': s.id,
        'ad': s.ad,
        'tip': s.tip,
        'tip_text': s.tip_text,
        'kapsam': s.kapsam_text
    } for s in sablonlar])


# ============================================================
# SÖZLEŞME DOCX JENERATÖR
# ============================================================

def _read_sablon_docx(sablon):
    """Şablon dosyasını disk'ten okur. Bytes döner."""
    if not sablon.sablon_dosya:
        return None
    fpath = os.path.join(current_app.config['UPLOAD_FOLDER'], sablon.sablon_dosya)
    if not os.path.exists(fpath):
        return None
    with open(fpath, 'rb') as f:
        return f.read()


@ik_bp.route('/calisan/<int:id>/sozlesme-olustur', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def calisan_sozlesme_olustur(id):
    """Çalışana sözleşme .docx'i oluştur (otomatik + manuel değişkenler)."""
    from app.services.sozlesme_generator import (
        OTOMATIK_DEGISKENLER, MANUEL_DEGISKENLER,
        calisan_degiskenleri, degiskenleri_doldur_docx
    )

    calisan = Calisan.query.get_or_404(id)
    if not calisan_in_scope(calisan):
        flash('Bu çalışan için sözleşme oluşturma yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    sablonlar = SozlesmeSablonu.query.filter(
        SozlesmeSablonu.aktif == True,
        SozlesmeSablonu.is_deleted == False,
        SozlesmeSablonu.sablon_dosya.isnot(None),
        SozlesmeSablonu.sablon_dosya != ''
    ).order_by(SozlesmeSablonu.ad).all()

    if request.method == 'POST':
        sablon_id = int(request.form.get('sablon_id') or 0)
        sablon = SozlesmeSablonu.query.get_or_404(sablon_id)

        template_bytes = _read_sablon_docx(sablon)
        if not template_bytes:
            flash('Bu şablonun .docx dosyası bulunamadı.', 'danger')
            return redirect(url_for('ik.calisan_sozlesme_olustur', id=id))

        degerler = calisan_degiskenleri(calisan)
        for md in MANUEL_DEGISKENLER:
            degerler[md['kod']] = (request.form.get(md['kod']) or '').strip()

        try:
            docx_bytes = degiskenleri_doldur_docx(template_bytes, degerler)
        except Exception as e:
            current_app.logger.exception('Sözleşme DOCX hatası')
            flash(f'Sözleşme oluşturulamadı: {e}', 'danger')
            return redirect(url_for('ik.calisan_sozlesme_olustur', id=id))

        upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesmeler', str(calisan.id))
        os.makedirs(upload_folder, exist_ok=True)
        fname = f"sozlesme_{calisan.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.docx"
        with open(os.path.join(upload_folder, fname), 'wb') as f:
            f.write(docx_bytes)

        calisan.sozlesme_pdf = f"sozlesmeler/{calisan.id}/{fname}"
        calisan.sozlesme_sablon_id = sablon.id
        db.session.commit()

        flash(f'Sözleşme oluşturuldu: {fname}', 'success')
        return redirect(url_for('ik.detay', id=calisan.id))

    return render_template(
        'ik/sozlesme_olustur.html',
        calisan=calisan,
        sablonlar=sablonlar,
        otomatik_degerler=calisan_degiskenleri(calisan),
        otomatik_degiskenler=OTOMATIK_DEGISKENLER,
        manuel_degiskenler=MANUEL_DEGISKENLER,
    )


@ik_bp.route('/proje/<int:proje_id>/sozlesme-toplu', methods=['GET', 'POST'])
@login_required
@permission_required('ik.edit')
def sozlesme_toplu_olustur(proje_id):
    """Proje bazlı toplu sözleşme oluşturma (.docx)."""
    from app.models.proje import Proje
    from app.models.base import CalisanDurumu
    from app.services.sozlesme_generator import (
        MANUEL_DEGISKENLER, calisan_degiskenleri, degiskenleri_doldur_docx
    )

    proje = Proje.query.get_or_404(proje_id)
    calisanlar = Calisan.query.filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        Calisan.kadro.has(proje_id=proje_id)
    ).order_by(Calisan.ad, Calisan.soyad).all()

    sablonlar = SozlesmeSablonu.query.filter(
        SozlesmeSablonu.aktif == True,
        SozlesmeSablonu.is_deleted == False,
        SozlesmeSablonu.sablon_dosya.isnot(None),
        SozlesmeSablonu.sablon_dosya != ''
    ).order_by(SozlesmeSablonu.ad).all()

    if request.method == 'POST':
        sablon_id = int(request.form.get('sablon_id') or 0)
        sablon = SozlesmeSablonu.query.get_or_404(sablon_id)
        secili_ids = request.form.getlist('calisan_ids', type=int)

        if not secili_ids:
            flash('En az bir çalışan seçmelisiniz.', 'warning')
            return redirect(url_for('ik.sozlesme_toplu_olustur', proje_id=proje_id))

        template_bytes = _read_sablon_docx(sablon)
        if not template_bytes:
            flash('Bu şablonun .docx dosyası bulunamadı.', 'danger')
            return redirect(url_for('ik.sozlesme_toplu_olustur', proje_id=proje_id))

        ortak_manuel = {md['kod']: (request.form.get(md['kod']) or '').strip() for md in MANUEL_DEGISKENLER}

        basarili, hatali = 0, 0
        for cid in secili_ids:
            c = Calisan.query.get(cid)
            if not c or not calisan_in_scope(c):
                continue
            try:
                degerler = calisan_degiskenleri(c)
                degerler.update(ortak_manuel)
                docx_bytes = degiskenleri_doldur_docx(template_bytes, degerler)

                folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'sozlesmeler', str(c.id))
                os.makedirs(folder, exist_ok=True)
                fname = f"sozlesme_{c.id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.docx"
                with open(os.path.join(folder, fname), 'wb') as f:
                    f.write(docx_bytes)

                c.sozlesme_pdf = f"sozlesmeler/{c.id}/{fname}"
                c.sozlesme_sablon_id = sablon.id
                basarili += 1
            except Exception:
                current_app.logger.exception(f'Toplu sözleşme hatası: calisan_id={cid}')
                hatali += 1

        db.session.commit()
        flash(f'Toplu sözleşme tamamlandı. Başarılı: {basarili}, Hata: {hatali}', 'success' if hatali == 0 else 'warning')
        return redirect(url_for('proje.proje_detay', id=proje_id))

    return render_template(
        'ik/sozlesme_toplu.html',
        proje=proje,
        calisanlar=calisanlar,
        sablonlar=sablonlar,
        manuel_degiskenler=MANUEL_DEGISKENLER,
    )


# ============================================================
# TOPLU EVRAK ZIP İNDİRME
# Çalışan/Aday evraklarını klasörlü ZIP olarak indirir.
# Yetki: ik.view (detay sayfalarıyla aynı, scope kontrollü)
# ============================================================

def _zip_ad_temizle(s):
    """Klasör/dosya adını ZIP için güvenli hale getirir (Türkçe -> ASCII)."""
    s = (s or '').strip()
    tr = str.maketrans('çÇğĞıİöÖşŞüÜ', 'cCgGiIoOsSuU')
    s = s.translate(tr)
    s = re.sub(r'[^A-Za-z0-9._-]+', '_', s).strip('_')
    return s or 'dosya'


def _evrak_klasoru(evrak_tipi):
    """EvrakTipi'ni ZIP alt klasör adına eşler."""
    if not evrak_tipi:
        return 'Diger_Evraklar'
    ad = f"{evrak_tipi.kod or ''} {evrak_tipi.ad or ''} {evrak_tipi.kategori or ''}".lower()
    if 'iban' in ad:
        return 'IBAN'
    if 'cv' in ad or 'ozgec' in ad or 'özgeç' in ad:
        return 'CV'
    if 'sgk' in ad and ('cik' in ad or 'çık' in ad or 'ayril' in ad):
        return 'SGK_Cikis_Bildirgesi'
    if 'sgk' in ad:
        return 'SGK_Giris_Bildirgesi'
    if 'foto' in ad or 'vesikal' in ad or 'resim' in ad:
        return 'Fotograf'
    return 'Diger_Evraklar'


def _evrak_dosya_ekle(items, klasor, dosya_adi, dosya_yolu):
    """Var olan bir dosyayı items listesine ekler (klasor, dosya_adi, tam_yol)."""
    tam = _evrak_tam_yol(dosya_yolu)
    if tam and os.path.exists(tam) and os.path.isfile(tam):
        ad = dosya_adi or os.path.basename(tam)
        items.append((klasor, ad, tam))


def _calisan_evrak_items(calisan):
    """Çalışanın kendi (sabit alan + CalisanEvrak) dosyaları."""
    items = []
    _evrak_dosya_ekle(items, 'Fotograf', None, calisan.foto)
    _evrak_dosya_ekle(items, 'Sozlesme', None, calisan.sozlesme_pdf)
    _evrak_dosya_ekle(items, 'SGK_Cikis_Bildirgesi', None, calisan.sgk_cikis_bildirgesi)
    for ev in calisan.evraklar:
        _evrak_dosya_ekle(items, _evrak_klasoru(ev.evrak_tipi), ev.dosya_adi, ev.dosya_yolu)
    return items


def _aday_evrak_items(aday):
    """Adayın (sabit alan + AdayEvrak) dosyaları."""
    items = []
    _evrak_dosya_ekle(items, 'CV', None, aday.cv_dosya)
    _evrak_dosya_ekle(items, 'Fotograf', None, aday.foto)
    _evrak_dosya_ekle(items, 'Kimlik', None, aday.kimlik_on)
    _evrak_dosya_ekle(items, 'Kimlik', None, aday.kimlik_arka)
    _evrak_dosya_ekle(items, 'Ehliyet', None, aday.ehliyet_foto)
    _evrak_dosya_ekle(items, 'Diploma', None, aday.diploma_foto)
    _evrak_dosya_ekle(items, 'SRC', None, aday.src_foto)
    _evrak_dosya_ekle(items, 'Diger_Evraklar', None, aday.ikametgah)
    _evrak_dosya_ekle(items, 'Diger_Evraklar', None, aday.adli_sicil)
    _evrak_dosya_ekle(items, 'Diger_Evraklar', None, aday.kargo_barkod_foto)
    _evrak_dosya_ekle(items, 'SGK_Giris_Bildirgesi', None, aday.sgk_bildirgesi)
    for ev in aday.evraklar:
        _evrak_dosya_ekle(items, _evrak_klasoru(ev.evrak_tipi), ev.dosya_adi, ev.dosya_yolu)
    return items


def _calisan_tum_evrak_items(calisan):
    """Çalışan evrakları + bağlı aday(lar)ın evrakları."""
    items = _calisan_evrak_items(calisan)
    for aday in Aday.query.filter_by(calisan_id=calisan.id, is_deleted=False).all():
        items += _aday_evrak_items(aday)
    return items


def _evrak_zip_indir(kisiler, zip_filename):
    """kisiler: [(kok_klasor, items[(klasor, dosya_adi, tam_yol)]), ...]

    Boş sonuç (hiç dosya yoksa) None döner.
    """
    mem = io.BytesIO()
    kullanilan_arc = set()
    kullanilan_root = set()
    dosya_sayisi = 0
    with zipfile.ZipFile(mem, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, items in kisiler:
            # Kök klasör adını benzersizleştir (aynı ad+tc iki kişide olursa)
            uroot, i = root, 1
            while uroot in kullanilan_root:
                i += 1
                uroot = f"{root}_{i}"
            kullanilan_root.add(uroot)

            gorulen_yol = set()
            for klasor, dosya_adi, tam in items:
                if tam in gorulen_yol:
                    continue  # aynı fiziksel dosya iki kez eklenmesin
                gorulen_yol.add(tam)

                guv_ad = (dosya_adi or os.path.basename(tam)).replace('/', '_').replace('\\', '_')
                arc = f"{uroot}/{klasor}/{guv_ad}"
                base, j = arc, 1
                while arc in kullanilan_arc:
                    n, e = os.path.splitext(base)
                    arc = f"{n}_{j}{e}"
                    j += 1
                kullanilan_arc.add(arc)
                try:
                    zf.write(tam, arc)
                    dosya_sayisi += 1
                except OSError:
                    continue

    if dosya_sayisi == 0:
        return None
    mem.seek(0)
    return send_file(mem, mimetype='application/zip', as_attachment=True,
                     download_name=zip_filename)


@ik_bp.route('/calisan/<int:id>/evrak-zip')
@login_required
@permission_required('ik.view')
def calisan_evrak_zip(id):
    """Çalışanın tüm evraklarını (kendi + bağlı aday) ZIP olarak indir."""
    calisan = Calisan.query.get_or_404(id)
    if not calisan_in_scope(calisan):
        flash('Bu çalışanın evraklarını indirme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.liste'))

    items = _calisan_tum_evrak_items(calisan)
    root = _zip_ad_temizle(f"{calisan.full_name}_{calisan.tc_kimlik or calisan.id}")
    resp = _evrak_zip_indir([(root, items)], f"{root}_evraklar.zip")
    if resp is None:
        flash('Bu çalışana ait indirilebilecek evrak dosyası bulunamadı.', 'warning')
        return redirect(url_for('ik.detay', id=id))
    return resp


@ik_bp.route('/aday/<int:id>/evrak-zip')
@login_required
@permission_required('ik.view')
def aday_evrak_zip(id):
    """Adayın tüm evraklarını ZIP olarak indir."""
    aday = Aday.query.get_or_404(id)
    if not aday_in_scope(aday):
        flash('Bu adayın evraklarını indirme yetkiniz yok.', 'danger')
        return redirect(url_for('ik.aday_liste'))

    items = _aday_evrak_items(aday)
    root = _zip_ad_temizle(f"{aday.full_name}_{aday.tc_kimlik or aday.id}")
    resp = _evrak_zip_indir([(root, items)], f"{root}_evraklar.zip")
    if resp is None:
        flash('Bu adaya ait indirilebilecek evrak dosyası bulunamadı.', 'warning')
        return redirect(url_for('ik.aday_detay', id=id))
    return resp


@ik_bp.route('/calisanlar/evrak-zip-toplu', methods=['POST'])
@login_required
@permission_required('ik.view')
def calisan_evrak_zip_toplu():
    """Seçili çalışanların evraklarını tek ZIP'te (kişi başı klasör) indir."""
    ids = [int(x) for x in request.form.getlist('ids') if x.isdigit()]
    if not ids:
        flash('Lütfen en az bir çalışan seçin.', 'warning')
        return redirect(url_for('ik.liste'))

    kisiler = []
    for calisan in Calisan.query.filter(Calisan.id.in_(ids), Calisan.is_deleted == False).all():
        if not calisan_in_scope(calisan):
            continue
        items = _calisan_tum_evrak_items(calisan)
        if items:
            root = _zip_ad_temizle(f"{calisan.full_name}_{calisan.tc_kimlik or calisan.id}")
            kisiler.append((root, items))

    resp = _evrak_zip_indir(kisiler, f"calisan_evraklari_{datetime.now().strftime('%Y%m%d_%H%M')}.zip")
    if resp is None:
        flash('Seçilen çalışanlara ait indirilebilecek evrak bulunamadı.', 'warning')
        return redirect(url_for('ik.liste'))
    return resp


@ik_bp.route('/adaylar/evrak-zip-toplu', methods=['POST'])
@login_required
@permission_required('ik.view')
def aday_evrak_zip_toplu():
    """Seçili adayların evraklarını tek ZIP'te (kişi başı klasör) indir."""
    ids = [int(x) for x in request.form.getlist('ids') if x.isdigit()]
    if not ids:
        flash('Lütfen en az bir aday seçin.', 'warning')
        return redirect(url_for('ik.aday_liste'))

    kisiler = []
    for aday in Aday.query.filter(Aday.id.in_(ids), Aday.is_deleted == False).all():
        if not aday_in_scope(aday):
            continue
        items = _aday_evrak_items(aday)
        if items:
            root = _zip_ad_temizle(f"{aday.full_name}_{aday.tc_kimlik or aday.id}")
            kisiler.append((root, items))

    resp = _evrak_zip_indir(kisiler, f"aday_evraklari_{datetime.now().strftime('%Y%m%d_%H%M')}.zip")
    if resp is None:
        flash('Seçilen adaylara ait indirilebilecek evrak bulunamadı.', 'warning')
        return redirect(url_for('ik.aday_liste'))
    return resp
