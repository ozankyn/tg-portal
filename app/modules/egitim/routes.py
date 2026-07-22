# -*- coding: utf-8 -*-
from app import csrf
"""
TG Portal - Eğitim Routes
Eğitim yönetimi, katılımcı takibi
"""
from datetime import datetime, date, timedelta
from flask import (Blueprint, render_template, redirect, url_for, flash, request,
                   jsonify, current_app, send_file, session)
from flask_login import login_required, current_user
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename
import os
import random
import re

from app import db
from app.models.egitim import (
    EgitimTipi, Egitim, EgitimKatilimci, EgitimKatilimLog, EgitimMateryali,
    CalisanZorunluEgitim, PozisyonZorunluEgitim,
    EgitimOturumu, EgitimKayit, EgitimAnket
)
from app.models.quiz import (
    SoruKategorisi, Soru, SoruSecenegi,
    Test, TestSorusu, TestSonuc, TestCevap
)
from app.models.ik import Calisan, Pozisyon, Aday
from app.models.proje import Proje, HedefKadro
from app.models.base import CalisanDurumu
from app.utils import permission_required, paginate_query
from app.services.jitsi import JitsiService

egitim_bp = Blueprint('egitim', __name__)

# Eğitim yönetim yetkisi olan roller (bunlara ek olarak eğitimi oluşturan da yönetebilir)
EGITIM_YONETICI_ROLLER = {'Sistem Yoneticisi', 'Ajans Baskani', 'Direktor', 'Egitim Uzmani'}


def _egitim_yonetebilir(egitim, user=None):
    """Kullanıcı bu eğitimi yönetebilir mi? (admin roller / Egitim Uzmani / oluşturan)."""
    user = user or current_user
    if not getattr(user, 'is_authenticated', False):
        return False
    if getattr(user, 'is_admin', False):
        return True
    if egitim is not None and egitim.olusturan_id and egitim.olusturan_id == user.id:
        return True
    return bool({r.name for r in user.roles} & EGITIM_YONETICI_ROLLER)


def _yonetim_yetki_hatasi(egitim):
    """Yetki yoksa flash + redirect döndürür, varsa None."""
    if _egitim_yonetebilir(egitim):
        return None
    flash('Bu eğitimi yönetme yetkiniz yok. Yalnızca görüntüleyebilirsiniz.', 'danger')
    return redirect(url_for('egitim.detay', id=egitim.id))


@egitim_bp.app_context_processor
def inject_egitim_yetki():
    """Template'lerde {% if egitim_yonetebilir(egitim) %} kullanımı için."""
    return dict(egitim_yonetebilir=_egitim_yonetebilir)


def _katilim_loglarini_kapat(egitim_id):
    """Eğitim sonlandırılınca ayrılış zamanı boş olan katılım loglarını now() ile kapatır.
    commit çağıranın sorumluluğunda."""
    EgitimKatilimLog.query.filter(
        EgitimKatilimLog.egitim_id == egitim_id,
        EgitimKatilimLog.ayrilma_zamani.is_(None),
    ).update({EgitimKatilimLog.ayrilma_zamani: datetime.now()}, synchronize_session=False)


ALLOWED_EXTENSIONS = {
    'pdf': 'dokuman',
    'ppt': 'sunum',
    'pptx': 'sunum',
    'doc': 'dokuman',
    'docx': 'dokuman',
    'xls': 'dokuman',
    'xlsx': 'dokuman',
    'mp4': 'video',
    'webm': 'video',
    'jpg': 'gorsel',
    'jpeg': 'gorsel',
    'png': 'gorsel',
    'gif': 'gorsel'
}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_materyal_tipi(filename):
    ext = filename.rsplit('.', 1)[1].lower()
    return ALLOWED_EXTENSIONS.get(ext, 'dokuman')



# ============================================================
# DASHBOARD
# ============================================================

@egitim_bp.route('/dashboard')
@login_required
@permission_required('egitim.view')
def dashboard():
    """Eğitim Dashboard"""
    # Yaklaşan eğitimler (7 gün içinde)
    from datetime import timedelta
    bugun = datetime.now()
    bir_hafta_sonra = bugun + timedelta(days=7)
    
    yaklasan_egitimler = Egitim.query.filter(
        Egitim.is_deleted == False,
        Egitim.durum == 'planli',
        Egitim.baslangic_tarihi >= bugun,
        Egitim.baslangic_tarihi <= bir_hafta_sonra
    ).order_by(Egitim.baslangic_tarihi).limit(5).all()
    
    # Devam eden eğitimler
    devam_eden = Egitim.query.filter(
        Egitim.is_deleted == False,
        Egitim.durum == 'devam_ediyor'
    ).count()
    
    # Bu ay tamamlanan
    ay_basi = bugun.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    tamamlanan_bu_ay = Egitim.query.filter(
        Egitim.is_deleted == False,
        Egitim.durum == 'tamamlandi',
        Egitim.bitis_tarihi >= ay_basi
    ).count()
    
    # Toplam eğitim alan çalışan (bu yıl)
    yil_basi = bugun.replace(month=1, day=1)
    egitim_alan_calisan = db.session.query(
        db.func.count(db.distinct(EgitimKatilimci.calisan_id))
    ).join(Egitim).filter(
        Egitim.baslangic_tarihi >= yil_basi,
        EgitimKatilimci.durum.in_(['katildi', 'gecti'])
    ).scalar() or 0
    
    # Zorunlu eğitimi eksik çalışanlar
    eksik_zorunlu = CalisanZorunluEgitim.query.filter(
        CalisanZorunluEgitim.tamamlandi == False
    ).count()
    
    # Eğitim tiplerine göre dağılım
    tip_dagilim = db.session.query(
        EgitimTipi.ad,
        db.func.count(Egitim.id)
    ).join(Egitim).filter(
        Egitim.is_deleted == False,
        Egitim.baslangic_tarihi >= yil_basi
    ).group_by(EgitimTipi.ad).all()
    
    return render_template('egitim/dashboard.html',
                          yaklasan_egitimler=yaklasan_egitimler,
                          devam_eden=devam_eden,
                          tamamlanan_bu_ay=tamamlanan_bu_ay,
                          egitim_alan_calisan=egitim_alan_calisan,
                          eksik_zorunlu=eksik_zorunlu,
                          tip_dagilim=tip_dagilim)


# ============================================================
# EĞİTİM LİSTESİ
# ============================================================

@egitim_bp.route('/')
@egitim_bp.route('/liste')
@login_required
@permission_required('egitim.view')
def liste():
    """Eğitim listesi"""
    page = request.args.get('page', 1, type=int)
    durum = request.args.get('durum')
    tip_id = request.args.get('tip_id', type=int)
    proje_id = request.args.get('proje_id', type=int)
    tarih = request.args.get('tarih')  # gecmis, gelecek, bugun
    
    query = Egitim.query.filter_by(is_deleted=False)
    
    if durum:
        query = query.filter(Egitim.durum == durum)
    if tip_id:
        query = query.filter(Egitim.egitim_tipi_id == tip_id)
    if proje_id:
        query = query.filter(Egitim.proje_id == proje_id)
    
    if tarih == 'gecmis':
        query = query.filter(Egitim.baslangic_tarihi < datetime.now())
    elif tarih == 'gelecek':
        query = query.filter(Egitim.baslangic_tarihi >= datetime.now())
    elif tarih == 'bugun':
        bugun = date.today()
        query = query.filter(db.func.date(Egitim.baslangic_tarihi) == bugun)
    
    query = query.order_by(Egitim.baslangic_tarihi.desc())
    pagination = paginate_query(query, page, 20)
    
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.ad).all()
    projeler = Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all()
    
    return render_template('egitim/liste.html',
                          egitimler=pagination.items,
                          pagination=pagination,
                          egitim_tipleri=egitim_tipleri,
                          projeler=projeler)


# ============================================================
# EĞİTİM EKLE / DÜZENLE
# ============================================================

@egitim_bp.route('/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('egitim.create')
def ekle():
    """Yeni eğitim ekle"""
    if request.method == 'POST':
        egitim = Egitim(
            egitim_tipi_id=int(request.form['egitim_tipi_id']),
            baslik=request.form.get('baslik', '').strip(),
            aciklama=request.form.get('aciklama', '').strip() or None,
            proje_id=int(request.form['proje_id']) if request.form.get('proje_id') else None,
            baslangic_tarihi=datetime.strptime(request.form['baslangic_tarihi'], '%Y-%m-%dT%H:%M'),
            bitis_tarihi=datetime.strptime(request.form['bitis_tarihi'], '%Y-%m-%dT%H:%M') if request.form.get('bitis_tarihi') else None,
            sure_saat=float(request.form['sure_saat']) if request.form.get('sure_saat') else None,
            lokasyon_tipi=request.form.get('lokasyon_tipi', 'yuz_yuze'),
            lokasyon=request.form.get('lokasyon', '').strip() or None,
            egitmen_tipi=request.form.get('egitmen_tipi'),
            egitmen_id=int(request.form['egitmen_id']) if request.form.get('egitmen_id') else None,
            dis_egitmen_ad=request.form.get('dis_egitmen_ad', '').strip() or None,
            dis_egitmen_kurum=request.form.get('dis_egitmen_kurum', '').strip() or None,
            kontenjan=int(request.form['kontenjan']) if request.form.get('kontenjan') else None,
            min_katilimci=int(request.form['min_katilimci']) if request.form.get('min_katilimci') else None,
            maliyet=float(request.form['maliyet']) if request.form.get('maliyet') else None,
            notlar=request.form.get('notlar', '').strip() or None,
            olusturan_id=current_user.id
        )
        
        db.session.add(egitim)
        db.session.commit()
        
        flash(f'"{egitim.baslik}" eğitimi oluşturuldu.', 'success')
        return redirect(url_for('egitim.detay', id=egitim.id))
    
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.sira, EgitimTipi.ad).all()
    projeler = Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all()
    egitmenler = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).order_by(Calisan.ad).all()
    
    return render_template('egitim/form.html',
                          egitim=None,
                          egitim_tipleri=egitim_tipleri,
                          projeler=projeler,
                          egitmenler=egitmenler)


@egitim_bp.route('/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('egitim.edit')
def duzenle(id):
    """Eğitim düzenle"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    if request.method == 'POST':
        egitim.egitim_tipi_id = int(request.form['egitim_tipi_id'])
        egitim.baslik = request.form.get('baslik', '').strip()
        egitim.aciklama = request.form.get('aciklama', '').strip() or None
        egitim.proje_id = int(request.form['proje_id']) if request.form.get('proje_id') else None
        egitim.baslangic_tarihi = datetime.strptime(request.form['baslangic_tarihi'], '%Y-%m-%dT%H:%M')
        egitim.bitis_tarihi = datetime.strptime(request.form['bitis_tarihi'], '%Y-%m-%dT%H:%M') if request.form.get('bitis_tarihi') else None
        egitim.sure_saat = float(request.form['sure_saat']) if request.form.get('sure_saat') else None
        egitim.lokasyon_tipi = request.form.get('lokasyon_tipi', 'yuz_yuze')
        egitim.lokasyon = request.form.get('lokasyon', '').strip() or None
        egitim.egitmen_tipi = request.form.get('egitmen_tipi')
        egitim.egitmen_id = int(request.form['egitmen_id']) if request.form.get('egitmen_id') else None
        egitim.dis_egitmen_ad = request.form.get('dis_egitmen_ad', '').strip() or None
        egitim.dis_egitmen_kurum = request.form.get('dis_egitmen_kurum', '').strip() or None
        egitim.kontenjan = int(request.form['kontenjan']) if request.form.get('kontenjan') else None
        egitim.min_katilimci = int(request.form['min_katilimci']) if request.form.get('min_katilimci') else None
        egitim.maliyet = float(request.form['maliyet']) if request.form.get('maliyet') else None
        egitim.notlar = request.form.get('notlar', '').strip() or None
        
        db.session.commit()
        
        flash('Eğitim güncellendi.', 'success')
        return redirect(url_for('egitim.detay', id=id))
    
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.sira, EgitimTipi.ad).all()
    projeler = Proje.query.filter_by(is_deleted=False, aktif=True).order_by(Proje.ad).all()
    egitmenler = Calisan.query.filter_by(is_deleted=False, durum=CalisanDurumu.AKTIF).order_by(Calisan.ad).all()
    
    return render_template('egitim/form.html',
                          egitim=egitim,
                          egitim_tipleri=egitim_tipleri,
                          projeler=projeler,
                          egitmenler=egitmenler)


# ============================================================
# EĞİTİM DETAY
# ============================================================

@egitim_bp.route('/<int:id>')
@login_required
@permission_required('egitim.view')
def detay(id):
    """Eğitim detay"""
    egitim = Egitim.query.get_or_404(id)
    
    katilimcilar = egitim.katilimcilar.join(Calisan).order_by(Calisan.ad).all()
    materyaller = egitim.materyaller.order_by(EgitimMateryali.sira).all()
    
    # Eklenebilecek çalışanlar (henüz eklenmemiş)
    mevcut_calisan_ids = [k.calisan_id for k in katilimcilar]
    eklenebilir_calisanlar = Calisan.query.filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        ~Calisan.id.in_(mevcut_calisan_ids) if mevcut_calisan_ids else True
    ).order_by(Calisan.ad).all()
    
    katilim_loglari = egitim.katilim_loglari.limit(500).all()

    # Booking: oturumlar, kayıtlar ve anket özeti
    oturumlar = egitim.oturumlar.all()
    kayitlar = (egitim.kayitlar
                .join(EgitimOturumu, EgitimKayit.oturum_id == EgitimOturumu.id)
                .order_by(EgitimOturumu.tarih, EgitimOturumu.baslangic_saati,
                          EgitimKayit.kayit_zamani)
                .all())
    aktif_kayit_sayisi = sum(1 for k in kayitlar if k.durum == 'onaylandi')

    anketler = egitim.anketler.order_by(EgitimAnket.created_at.desc()).all()
    anket_ortalama = round(sum(a.puan for a in anketler) / len(anketler), 1) if anketler else None

    return render_template('egitim/detay.html',
                          egitim=egitim,
                          katilimcilar=katilimcilar,
                          materyaller=materyaller,
                          eklenebilir_calisanlar=eklenebilir_calisanlar,
                          katilim_loglari=katilim_loglari,
                          oturumlar=oturumlar,
                          kayitlar=kayitlar,
                          aktif_kayit_sayisi=aktif_kayit_sayisi,
                          anketler=anketler,
                          anket_ortalama=anket_ortalama)


# ============================================================
# EĞİTİM DURUM DEĞİŞTİR
# ============================================================

@egitim_bp.route('/<int:id>/durum/<durum>', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def durum_degistir(id, durum):
    """Eğitim durumunu değiştir"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    if durum not in ['planli', 'devam_ediyor', 'tamamlandi', 'iptal']:
        flash('Geçersiz durum.', 'danger')
        return redirect(url_for('egitim.detay', id=id))

    egitim.durum = durum

    if durum == 'tamamlandi' and not egitim.bitis_tarihi:
        egitim.bitis_tarihi = datetime.now()

    if durum == 'iptal':
        egitim.iptal_nedeni = request.form.get('iptal_nedeni')

    # Eğitim sonlandırıldıysa: canlı yayını kapat + aktif katılımcıların ayrılış zamanını işaretle
    if durum in ('tamamlandi', 'iptal'):
        egitim.jitsi_aktif = False
        _katilim_loglarini_kapat(egitim.id)

    db.session.commit()
    
    flash(f'Eğitim durumu güncellendi: {durum}', 'success')
    return redirect(url_for('egitim.detay', id=id))


# ============================================================
# KATILIMCI YÖNETİMİ
# ============================================================

@egitim_bp.route('/<int:id>/katilimci/ekle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def katilimci_ekle(id):
    """Eğitime katılımcı ekle"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    calisan_ids = request.form.getlist('calisan_ids')
    
    eklenen = 0
    for calisan_id in calisan_ids:
        # Zaten var mı kontrol
        mevcut = EgitimKatilimci.query.filter_by(
            egitim_id=id,
            calisan_id=int(calisan_id)
        ).first()
        
        if not mevcut:
            katilimci = EgitimKatilimci(
                egitim_id=id,
                calisan_id=int(calisan_id),
                davet_eden_id=current_user.id
            )
            db.session.add(katilimci)
            eklenen += 1
    
    db.session.commit()
    
    flash(f'{eklenen} katılımcı eklendi.', 'success')
    return redirect(url_for('egitim.detay', id=id))


@egitim_bp.route('/katilimci/<int:id>/durum', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def katilimci_durum(id):
    """Katılımcı durumunu güncelle"""
    katilimci = EgitimKatilimci.query.get_or_404(id)

    if not _egitim_yonetebilir(katilimci.egitim):
        return jsonify({'success': False, 'error': 'Yetki yok'}), 403

    durum = request.form.get('durum')
    if durum:
        katilimci.durum = durum
        
        if durum in ['katildi', 'gecti']:
            katilimci.katilim_tarihi = datetime.now()
        
        if durum == 'gecti':
            # Sertifika bilgisi
            katilimci.sertifika_tarihi = date.today()
            if katilimci.egitim.egitim_tipi.gecerlilik_gun:
                from datetime import timedelta
                katilimci.sertifika_gecerlilik = date.today() + timedelta(days=katilimci.egitim.egitim_tipi.gecerlilik_gun)
    
    puan = request.form.get('puan')
    if puan:
        katilimci.puan = int(puan)
    
    katilimci.degerlendirme = request.form.get('degerlendirme')
    katilimci.katilim_notu = request.form.get('katilim_notu')
    
    if durum == 'mazeret':
        katilimci.mazeret_nedeni = request.form.get('mazeret_nedeni')
    
    db.session.commit()
    
    return jsonify({'success': True})


@egitim_bp.route('/katilimci/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def katilimci_sil(id):
    """Katılımcıyı sil"""
    katilimci = EgitimKatilimci.query.get_or_404(id)
    egitim_id = katilimci.egitim_id

    if not _egitim_yonetebilir(katilimci.egitim):
        flash('Bu eğitimi yönetme yetkiniz yok.', 'danger')
        return redirect(url_for('egitim.detay', id=egitim_id))

    db.session.delete(katilimci)
    db.session.commit()
    
    flash('Katılımcı silindi.', 'success')
    return redirect(url_for('egitim.detay', id=egitim_id))


# ============================================================
# TOPLU KATILIMCI EKLEME (Proje/Kadro bazlı)
# ============================================================

@egitim_bp.route('/<int:id>/toplu-katilimci', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def toplu_katilimci(id):
    """Proje veya kadrodan toplu katılımcı ekle"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    kaynak = request.form.get('kaynak')  # proje, kadro
    kaynak_id = request.form.get('kaynak_id', type=int)
    
    if kaynak == 'proje' and kaynak_id:
        # Projedeki tüm aktif çalışanları ekle
        calisanlar = Calisan.query.join(HedefKadro).filter(
            HedefKadro.proje_id == kaynak_id,
            Calisan.is_deleted == False,
            Calisan.durum == CalisanDurumu.AKTIF
        ).all()
    elif kaynak == 'kadro' and kaynak_id:
        # Kadrodaki tüm aktif çalışanları ekle
        calisanlar = Calisan.query.filter(
            Calisan.kadro_id == kaynak_id,
            Calisan.is_deleted == False,
            Calisan.durum == CalisanDurumu.AKTIF
        ).all()
    else:
        flash('Geçersiz kaynak.', 'danger')
        return redirect(url_for('egitim.detay', id=id))
    
    eklenen = 0
    for calisan in calisanlar:
        mevcut = EgitimKatilimci.query.filter_by(
            egitim_id=id,
            calisan_id=calisan.id
        ).first()
        
        if not mevcut:
            katilimci = EgitimKatilimci(
                egitim_id=id,
                calisan_id=calisan.id,
                davet_eden_id=current_user.id
            )
            db.session.add(katilimci)
            eklenen += 1
    
    db.session.commit()
    
    flash(f'{eklenen} katılımcı eklendi.', 'success')
    return redirect(url_for('egitim.detay', id=id))


# ============================================================
# EĞİTİM TİPLERİ YÖNETİMİ
# ============================================================

@egitim_bp.route('/tipler')
@login_required
@permission_required('egitim.view')
def tip_liste():
    """Eğitim tipleri listesi"""
    tipler = EgitimTipi.query.order_by(EgitimTipi.sira, EgitimTipi.ad).all()
    return render_template('egitim/tip_liste.html', egitim_tipleri=tipler)


@egitim_bp.route('/tip/ekle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def tip_ekle():
    """Yeni eğitim tipi ekle"""
    tip = EgitimTipi(
        ad=request.form.get('ad'),
        kod=request.form.get('kod'),
        kategori=request.form.get('kategori'),
        aciklama=request.form.get('aciklama'),
        sure_saat=float(request.form['sure_saat']) if request.form.get('sure_saat') else None,
        gecerlilik_gun=int(request.form['gecerlilik_gun']) if request.form.get('gecerlilik_gun') else None,
        sertifika_gerekli=request.form.get('sertifika_gerekli') == 'on',
        tekrar_periyot_gun=int(request.form['tekrar_periyot_gun']) if request.form.get('tekrar_periyot_gun') else None
    )
    db.session.add(tip)
    db.session.commit()
    
    flash('Eğitim tipi eklendi.', 'success')
    return redirect(url_for('egitim.tip_liste'))


# ============================================================
# ÇALIŞAN EĞİTİM GEÇMİŞİ
# ============================================================

@egitim_bp.route('/calisan/<int:id>')
@login_required
@permission_required('egitim.view')
def calisan_egitimler(id):
    """Çalışanın eğitim geçmişi"""
    calisan = Calisan.query.get_or_404(id)
    
    egitim_kayitlari = calisan.egitim_kayitlari.join(Egitim).order_by(
        Egitim.baslangic_tarihi.desc()
    ).all()
    
    zorunlu_egitimler = calisan.zorunlu_egitimler.all()
    
    return render_template('egitim/calisan_egitimler.html',
                          calisan=calisan,
                          egitim_kayitlari=egitim_kayitlari,
                          zorunlu_egitimler=zorunlu_egitimler)


# ============================================================
# ZORUNLU EĞİTİM TAKİBİ
# ============================================================

@egitim_bp.route('/zorunlu-egitimler')
@login_required
@permission_required('egitim.view')
def zorunlu_egitim_takip():
    """Zorunlu eğitim durumu takibi"""
    # Eksik zorunlu eğitimleri olan çalışanlar
    eksik_egitimler = db.session.query(
        Calisan,
        CalisanZorunluEgitim
    ).join(CalisanZorunluEgitim).filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        CalisanZorunluEgitim.tamamlandi == False
    ).all()
    
    # Yenileme gereken eğitimler (30 gün içinde süresi dolacak)
    from datetime import timedelta
    otuz_gun_sonra = date.today() + timedelta(days=30)
    
    yenileme_gerekli = db.session.query(
        Calisan,
        CalisanZorunluEgitim
    ).join(CalisanZorunluEgitim).filter(
        Calisan.is_deleted == False,
        Calisan.durum == CalisanDurumu.AKTIF,
        CalisanZorunluEgitim.tamamlandi == True,
        CalisanZorunluEgitim.son_gecerlilik <= otuz_gun_sonra
    ).all()
    
    return render_template('egitim/zorunlu_takip.html',
                          eksik_egitimler=eksik_egitimler,
                          yenileme_gerekli=yenileme_gerekli)


# ============================================================
# RAPORLAR
# ============================================================

@egitim_bp.route('/rapor')
@login_required
@permission_required('egitim.view')
def rapor():
    """Eğitim raporları"""
    # Yıllık eğitim istatistikleri
    yil = request.args.get('yil', date.today().year, type=int)
    yil_basi = date(yil, 1, 1)
    yil_sonu = date(yil, 12, 31)
    
    # Aylık eğitim sayısı
    aylik_egitim = db.session.query(
        db.func.extract('month', Egitim.baslangic_tarihi).label('ay'),
        db.func.count(Egitim.id).label('sayi')
    ).filter(
        Egitim.is_deleted == False,
        Egitim.baslangic_tarihi >= yil_basi,
        Egitim.baslangic_tarihi <= yil_sonu
    ).group_by('ay').all()
    
    # Toplam eğitim saati
    toplam_saat = db.session.query(
        db.func.sum(Egitim.sure_saat)
    ).filter(
        Egitim.is_deleted == False,
        Egitim.durum == 'tamamlandi',
        Egitim.baslangic_tarihi >= yil_basi,
        Egitim.baslangic_tarihi <= yil_sonu
    ).scalar() or 0
    
    # Eğitim başarı oranı
    toplam_katilimci = EgitimKatilimci.query.join(Egitim).filter(
        Egitim.baslangic_tarihi >= yil_basi,
        Egitim.baslangic_tarihi <= yil_sonu
    ).count()
    
    basarili_katilimci = EgitimKatilimci.query.join(Egitim).filter(
        Egitim.baslangic_tarihi >= yil_basi,
        Egitim.baslangic_tarihi <= yil_sonu,
        EgitimKatilimci.durum == 'gecti'
    ).count()
    
    basari_orani = round((basarili_katilimci / toplam_katilimci * 100), 1) if toplam_katilimci > 0 else 0
    
    return render_template('egitim/rapor.html',
                          yil=yil,
                          aylik_egitim=aylik_egitim,
                          toplam_saat=toplam_saat,
                          toplam_katilimci=toplam_katilimci,
                          basarili_katilimci=basarili_katilimci,
                          basari_orani=basari_orani)

# ============================================================
# MATERYAL YÜKLEME
# ============================================================

@egitim_bp.route('/<int:id>/materyal/yukle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def materyal_yukle(id):
    """Eğitime materyal yükle"""
    egitim = Egitim.query.get_or_404(id)
    
    # Harici link mi?
    if request.form.get('harici_link'):
        materyal = EgitimMateryali(
            egitim_id=id,
            ad=request.form.get('ad', 'Harici İçerik').strip(),
            aciklama=request.form.get('aciklama', '').strip() or None,
            materyal_tipi=request.form.get('materyal_tipi', 'link'),
            harici_link=request.form.get('harici_link').strip(),
            yukleyen_id=current_user.id
        )
        db.session.add(materyal)
        db.session.commit()
        flash('Harici link eklendi.', 'success')
        return redirect(url_for('egitim.detay', id=id))
    
    # Dosya yükleme
    if 'dosya' not in request.files:
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('egitim.detay', id=id))
    
    dosya = request.files['dosya']
    
    if dosya.filename == '':
        flash('Dosya seçilmedi.', 'danger')
        return redirect(url_for('egitim.detay', id=id))
    
    if dosya and allowed_file(dosya.filename):
        filename = secure_filename(dosya.filename)
        # Benzersiz isim oluştur
        import uuid
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        
        # Upload klasörü
        upload_folder = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'), 'egitim', str(id))
        os.makedirs(upload_folder, exist_ok=True)
        
        filepath = os.path.join(upload_folder, unique_filename)
        dosya.save(filepath)
        
        # Dosya boyutu
        file_size = os.path.getsize(filepath)
        
        materyal = EgitimMateryali(
            egitim_id=id,
            ad=request.form.get('ad', filename).strip(),
            aciklama=request.form.get('aciklama', '').strip() or None,
            materyal_tipi=get_materyal_tipi(filename),
            dosya_adi=filename,
            dosya_yolu=filepath,
            dosya_boyut=file_size,
            mime_type=dosya.content_type,
            yukleyen_id=current_user.id
        )
        db.session.add(materyal)
        db.session.commit()
        
        flash(f'"{filename}" yüklendi.', 'success')
    else:
        flash('Desteklenmeyen dosya formatı.', 'danger')
    
    return redirect(url_for('egitim.detay', id=id))


# ============================================================
# MATERYAL GÖRÜNTÜLEME
# ============================================================

@egitim_bp.route('/materyal/<int:id>/goruntule')
@login_required
@permission_required('egitim.view')
def materyal_goruntule(id):
    """Materyali görüntüle"""
    materyal = EgitimMateryali.query.get_or_404(id)
    
    return render_template('egitim/materyal_goruntule.html', materyal=materyal)


@egitim_bp.route('/materyal/<int:id>/indir')
@login_required
@permission_required('egitim.view')
def materyal_indir(id):
    """Materyali indir"""
    materyal = EgitimMateryali.query.get_or_404(id)
    
    if not materyal.dosya_yolu or not os.path.exists(materyal.dosya_yolu):
        flash('Dosya bulunamadı.', 'danger')
        return redirect(url_for('egitim.detay', id=materyal.egitim_id))
    
    return send_file(
        materyal.dosya_yolu,
        download_name=materyal.dosya_adi,
        as_attachment=True
    )


@egitim_bp.route('/materyal/<int:id>/embed')
@login_required
@permission_required('egitim.view')
def materyal_embed(id):
    """Materyal embed (iframe için)"""
    materyal = EgitimMateryali.query.get_or_404(id)
    
    if not materyal.dosya_yolu or not os.path.exists(materyal.dosya_yolu):
        return "Dosya bulunamadı", 404
    
    return send_file(
        materyal.dosya_yolu,
        mimetype=materyal.mime_type
    )


# ============================================================
# MATERYAL SİL
# ============================================================

@egitim_bp.route('/materyal/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def materyal_sil(id):
    """Materyali sil"""
    materyal = EgitimMateryali.query.get_or_404(id)
    egitim_id = materyal.egitim_id
    
    # Dosyayı da sil
    if materyal.dosya_yolu and os.path.exists(materyal.dosya_yolu):
        os.remove(materyal.dosya_yolu)
    
    db.session.delete(materyal)
    db.session.commit()
    
    flash('Materyal silindi.', 'success')
    return redirect(url_for('egitim.detay', id=egitim_id))


# ============================================================
# MATERYAL SIRALAMA
# ============================================================

@egitim_bp.route('/<int:id>/materyal/sirala', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def materyal_sirala(id):
    """Materyal sırasını güncelle (AJAX)"""
    siralama = request.json.get('siralama', [])
    
    for index, materyal_id in enumerate(siralama):
        materyal = EgitimMateryali.query.get(materyal_id)
        if materyal and materyal.egitim_id == id:
            materyal.sira = index
    
    db.session.commit()
    return jsonify({'success': True})

# ============================================================
# SORU BANKASI YÖNETİMİ
# ============================================================

@egitim_bp.route('/sorular')
@login_required
@permission_required('egitim.view')
def soru_liste():
    """Soru bankası listesi"""
    page = request.args.get('page', 1, type=int)
    kategori_id = request.args.get('kategori_id', type=int)
    egitim_tipi_id = request.args.get('egitim_tipi_id', type=int)
    zorluk = request.args.get('zorluk', type=int)
    soru_tipi = request.args.get('soru_tipi')
    
    query = Soru.query.filter_by(is_deleted=False)
    
    if kategori_id:
        query = query.filter(Soru.kategori_id == kategori_id)
    if egitim_tipi_id:
        query = query.filter(Soru.egitim_tipi_id == egitim_tipi_id)
    if zorluk:
        query = query.filter(Soru.zorluk == zorluk)
    if soru_tipi:
        query = query.filter(Soru.soru_tipi == soru_tipi)
    
    query = query.order_by(Soru.created_at.desc())
    pagination = paginate_query(query, page, 20)
    
    kategoriler = SoruKategorisi.query.filter_by(aktif=True).order_by(SoruKategorisi.ad).all()
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.ad).all()
    
    return render_template('egitim/soru_liste.html',
                          sorular=pagination.items,
                          pagination=pagination,
                          kategoriler=kategoriler,
                          egitim_tipleri=egitim_tipleri)


@egitim_bp.route('/soru/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('egitim.edit')
def soru_ekle():
    """Yeni soru ekle"""
    if request.method == 'POST':
        soru = Soru(
            soru_metni=request.form.get('soru_metni', '').strip(),
            soru_tipi=request.form.get('soru_tipi', 'coktan_secmeli'),
            kategori_id=int(request.form['kategori_id']) if request.form.get('kategori_id') else None,
            egitim_tipi_id=int(request.form['egitim_tipi_id']) if request.form.get('egitim_tipi_id') else None,
            zorluk=int(request.form.get('zorluk', 1)),
            puan=int(request.form.get('puan', 10)),
            aciklama=request.form.get('aciklama', '').strip() or None,
            olusturan_id=current_user.id
        )
        db.session.add(soru)
        db.session.flush()  # ID al
        
        # Seçenekleri ekle
        secenek_metinleri = request.form.getlist('secenek_metni')
        dogru_secenekler = request.form.getlist('dogru_secenek')
        
        for i, metin in enumerate(secenek_metinleri):
            if metin.strip():
                secenek = SoruSecenegi(
                    soru_id=soru.id,
                    secenek_metni=metin.strip(),
                    dogru=str(i) in dogru_secenekler,
                    sira=i
                )
                db.session.add(secenek)
        
        db.session.commit()
        flash('Soru eklendi.', 'success')
        return redirect(url_for('egitim.soru_liste'))
    
    kategoriler = SoruKategorisi.query.filter_by(aktif=True).order_by(SoruKategorisi.ad).all()
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.ad).all()
    
    return render_template('egitim/soru_form.html',
                          soru=None,
                          kategoriler=kategoriler,
                          egitim_tipleri=egitim_tipleri)


@egitim_bp.route('/soru/<int:id>/duzenle', methods=['GET', 'POST'])
@login_required
@permission_required('egitim.edit')
def soru_duzenle(id):
    """Soru düzenle"""
    soru = Soru.query.get_or_404(id)
    
    if request.method == 'POST':
        soru.soru_metni = request.form.get('soru_metni', '').strip()
        soru.soru_tipi = request.form.get('soru_tipi', 'coktan_secmeli')
        soru.kategori_id = int(request.form['kategori_id']) if request.form.get('kategori_id') else None
        soru.egitim_tipi_id = int(request.form['egitim_tipi_id']) if request.form.get('egitim_tipi_id') else None
        soru.zorluk = int(request.form.get('zorluk', 1))
        soru.puan = int(request.form.get('puan', 10))
        soru.aciklama = request.form.get('aciklama', '').strip() or None
        
        # Mevcut seçenekleri sil
        SoruSecenegi.query.filter_by(soru_id=soru.id).delete()
        
        # Yeni seçenekleri ekle
        secenek_metinleri = request.form.getlist('secenek_metni')
        dogru_secenekler = request.form.getlist('dogru_secenek')
        
        for i, metin in enumerate(secenek_metinleri):
            if metin.strip():
                secenek = SoruSecenegi(
                    soru_id=soru.id,
                    secenek_metni=metin.strip(),
                    dogru=str(i) in dogru_secenekler,
                    sira=i
                )
                db.session.add(secenek)
        
        db.session.commit()
        flash('Soru güncellendi.', 'success')
        return redirect(url_for('egitim.soru_liste'))
    
    kategoriler = SoruKategorisi.query.filter_by(aktif=True).order_by(SoruKategorisi.ad).all()
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.ad).all()
    
    return render_template('egitim/soru_form.html',
                          soru=soru,
                          kategoriler=kategoriler,
                          egitim_tipleri=egitim_tipleri)


@egitim_bp.route('/soru/<int:id>/sil', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def soru_sil(id):
    """Soru sil (soft delete)"""
    soru = Soru.query.get_or_404(id)
    soru.is_deleted = True
    soru.deleted_at = datetime.now()
    db.session.commit()
    
    flash('Soru silindi.', 'success')
    return redirect(url_for('egitim.soru_liste'))


# ============================================================
# SORU KATEGORİLERİ
# ============================================================

@egitim_bp.route('/soru-kategorileri')
@login_required
@permission_required('egitim.view')
def soru_kategori_liste():
    """Soru kategorileri"""
    kategoriler = SoruKategorisi.query.order_by(SoruKategorisi.sira, SoruKategorisi.ad).all()
    return render_template('egitim/soru_kategori_liste.html', kategoriler=kategoriler)


@egitim_bp.route('/soru-kategori/ekle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def soru_kategori_ekle():
    """Yeni kategori ekle"""
    kategori = SoruKategorisi(
        ad=request.form.get('ad'),
        aciklama=request.form.get('aciklama'),
        ust_kategori_id=int(request.form['ust_kategori_id']) if request.form.get('ust_kategori_id') else None
    )
    db.session.add(kategori)
    db.session.commit()
    
    flash('Kategori eklendi.', 'success')
    return redirect(url_for('egitim.soru_kategori_liste'))


# ============================================================
# TEST YÖNETİMİ
# ============================================================

@egitim_bp.route('/testler')
@login_required
@permission_required('egitim.view')
def test_liste():
    """Test listesi"""
    page = request.args.get('page', 1, type=int)
    egitim_id = request.args.get('egitim_id', type=int)
    aktif = request.args.get('aktif')
    
    query = Test.query.filter_by(is_deleted=False)
    
    if egitim_id:
        query = query.filter(Test.egitim_id == egitim_id)
    if aktif == '1':
        query = query.filter(Test.aktif == True)
    elif aktif == '0':
        query = query.filter(Test.aktif == False)
    
    query = query.order_by(Test.created_at.desc())
    pagination = paginate_query(query, page, 20)
    
    return render_template('egitim/test_liste.html',
                          testler=pagination.items,
                          pagination=pagination)


@egitim_bp.route('/test/ekle', methods=['GET', 'POST'])
@login_required
@permission_required('egitim.edit')
def test_ekle():
    """Yeni test oluştur"""
    if request.method == 'POST':
        test = Test(
            baslik=request.form.get('baslik', '').strip(),
            aciklama=request.form.get('aciklama', '').strip() or None,
            egitim_id=int(request.form['egitim_id']) if request.form.get('egitim_id') else None,
            egitim_tipi_id=int(request.form['egitim_tipi_id']) if request.form.get('egitim_tipi_id') else None,
            sure_dakika=int(request.form['sure_dakika']) if request.form.get('sure_dakika') else None,
            gecme_puani=int(request.form.get('gecme_puani', 70)),
            soru_karistir=request.form.get('soru_karistir') == 'on',
            secenek_karistir=request.form.get('secenek_karistir') == 'on',
            sonucu_goster=request.form.get('sonucu_goster') == 'on',
            dogru_cevaplari_goster=request.form.get('dogru_cevaplari_goster') == 'on',
            tekrar_hak=int(request.form['tekrar_hak']) if request.form.get('tekrar_hak') else None,
            aktif=request.form.get('aktif') == 'on',
            olusturan_id=current_user.id
        )
        
        # Tarihler
        if request.form.get('baslangic_tarihi'):
            test.baslangic_tarihi = datetime.strptime(request.form['baslangic_tarihi'], '%Y-%m-%dT%H:%M')
        if request.form.get('bitis_tarihi'):
            test.bitis_tarihi = datetime.strptime(request.form['bitis_tarihi'], '%Y-%m-%dT%H:%M')
        
        db.session.add(test)
        db.session.commit()
        
        flash('Test oluşturuldu. Şimdi soru ekleyebilirsiniz.', 'success')
        return redirect(url_for('egitim.test_detay', id=test.id))
    
    egitimler = Egitim.query.filter_by(is_deleted=False).order_by(Egitim.baslangic_tarihi.desc()).all()
    egitim_tipleri = EgitimTipi.query.filter_by(aktif=True).order_by(EgitimTipi.ad).all()
    
    return render_template('egitim/test_form.html',
                          test=None,
                          egitimler=egitimler,
                          egitim_tipleri=egitim_tipleri)


@egitim_bp.route('/test/<int:id>')
@login_required
@permission_required('egitim.view')
def test_detay(id):
    """Test detay - sorular ve sonuçlar"""
    test = Test.query.get_or_404(id)
    
    # Test soruları
    test_sorulari = test.test_sorulari.order_by(TestSorusu.sira).all()
    
    # Eklenebilecek sorular
    mevcut_soru_ids = [ts.soru_id for ts in test_sorulari]
    eklenebilir_sorular = Soru.query.filter(
        Soru.is_deleted == False,
        Soru.aktif == True,
        ~Soru.id.in_(mevcut_soru_ids) if mevcut_soru_ids else True
    )
    
    # Eğitim tipi filtresi
    if test.egitim_tipi_id:
        eklenebilir_sorular = eklenebilir_sorular.filter(
            db.or_(Soru.egitim_tipi_id == test.egitim_tipi_id, Soru.egitim_tipi_id == None)
        )
    
    eklenebilir_sorular = eklenebilir_sorular.order_by(Soru.created_at.desc()).limit(100).all()
    
    # Son sonuçlar
    son_sonuclar = test.sonuclar.filter_by(tamamlandi=True).order_by(TestSonuc.bitis_zamani.desc()).limit(10).all()
    
    return render_template('egitim/test_detay.html',
                          test=test,
                          test_sorulari=test_sorulari,
                          eklenebilir_sorular=eklenebilir_sorular,
                          son_sonuclar=son_sonuclar)


@egitim_bp.route('/test/<int:id>/soru-ekle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def test_soru_ekle(id):
    """Teste soru ekle"""
    test = Test.query.get_or_404(id)
    
    soru_ids = request.form.getlist('soru_ids')
    
    # Mevcut max sıra
    max_sira = db.session.query(db.func.max(TestSorusu.sira)).filter_by(test_id=id).scalar() or 0
    
    eklenen = 0
    for soru_id in soru_ids:
        # Zaten var mı?
        mevcut = TestSorusu.query.filter_by(test_id=id, soru_id=int(soru_id)).first()
        if not mevcut:
            max_sira += 1
            ts = TestSorusu(
                test_id=id,
                soru_id=int(soru_id),
                sira=max_sira
            )
            db.session.add(ts)
            eklenen += 1
    
    db.session.commit()
    flash(f'{eklenen} soru eklendi.', 'success')
    return redirect(url_for('egitim.test_detay', id=id))


@egitim_bp.route('/test/<int:id>/soru-cikar/<int:soru_id>', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def test_soru_cikar(id, soru_id):
    """Testten soru çıkar"""
    ts = TestSorusu.query.filter_by(test_id=id, soru_id=soru_id).first_or_404()
    db.session.delete(ts)
    db.session.commit()
    
    flash('Soru testten çıkarıldı.', 'success')
    return redirect(url_for('egitim.test_detay', id=id))


@egitim_bp.route('/test/<int:id>/soru-sirala', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def test_soru_sirala(id):
    """Test sorularını sırala (AJAX)"""
    siralama = request.json.get('siralama', [])
    
    for index, ts_id in enumerate(siralama):
        ts = TestSorusu.query.get(ts_id)
        if ts and ts.test_id == id:
            ts.sira = index
    
    db.session.commit()
    return jsonify({'success': True})


# ============================================================
# TEST ÇÖZME
# ============================================================

@egitim_bp.route('/test/<int:id>/baslat', methods=['GET', 'POST'])
@login_required
def test_baslat(id):
    """Testi başlat"""
    test = Test.query.get_or_404(id)
    
    # Kullanıcının çalışan kaydını bul
    calisan = current_user.calisan if hasattr(current_user, "calisan") else None
    if not calisan:
        flash('Çalışan kaydınız bulunamadı.', 'danger')
        return redirect(url_for('egitim.test_liste'))
    
    # Çözebilir mi kontrol
    cozebilir, mesaj = test.kullanici_cozebilir_mi(calisan.id)
    if not cozebilir:
        flash(mesaj, 'warning')
        return redirect(url_for('egitim.test_liste'))
    
    # Devam eden sınav var mı?
    devam_eden = TestSonuc.query.filter_by(
        test_id=id,
        calisan_id=calisan.id,
        tamamlandi=False
    ).first()
    
    if devam_eden:
        # Süre kontrolü
        if test.sure_dakika:
            gecen_sure = (datetime.now() - devam_eden.baslangic_zamani).total_seconds()
            if gecen_sure > test.sure_dakika * 60:
                # Süre dolmuş, otomatik bitir
                devam_eden.tamamlandi = True
                devam_eden.bitis_zamani = datetime.now()
                devam_eden.gecen_sure_saniye = int(gecen_sure)
                _hesapla_sonuc(devam_eden)
                db.session.commit()
                flash('Önceki sınavınızın süresi dolmuştu, otomatik değerlendirildi.', 'info')
            else:
                # Devam et
                return redirect(url_for('egitim.test_coz', sonuc_id=devam_eden.id))
        else:
            return redirect(url_for('egitim.test_coz', sonuc_id=devam_eden.id))
    
    if request.method == 'POST':
        # Yeni sınav başlat
        sonuc = TestSonuc(
            test_id=id,
            calisan_id=calisan.id,
            toplam_puan=test.toplam_puan
        )
        db.session.add(sonuc)
        db.session.commit()
        
        return redirect(url_for('egitim.test_coz', sonuc_id=sonuc.id))
    
    return render_template('egitim/test_baslat.html', test=test)


@egitim_bp.route('/test/coz/<int:sonuc_id>', methods=['GET', 'POST'])
@login_required
def test_coz(sonuc_id):
    """Test çözme sayfası"""
    sonuc = TestSonuc.query.get_or_404(sonuc_id)
    test = sonuc.test
    
    # Yetki kontrolü
    calisan = current_user.calisan if hasattr(current_user, "calisan") else None
    if not calisan or sonuc.calisan_id != calisan.id:
        flash('Bu sınava erişim yetkiniz yok.', 'danger')
        return redirect(url_for('egitim.test_liste'))
    
    if sonuc.tamamlandi:
        return redirect(url_for('egitim.test_sonuc', sonuc_id=sonuc_id))
    
    # Süre kontrolü
    kalan_sure = None
    if test.sure_dakika:
        gecen_sure = (datetime.now() - sonuc.baslangic_zamani).total_seconds()
        kalan_sure = max(0, test.sure_dakika * 60 - int(gecen_sure))
        
        if kalan_sure <= 0:
            # Süre doldu
            sonuc.tamamlandi = True
            sonuc.bitis_zamani = datetime.now()
            sonuc.gecen_sure_saniye = test.sure_dakika * 60
            _hesapla_sonuc(sonuc)
            db.session.commit()
            flash('Süre doldu! Sınavınız otomatik değerlendirildi.', 'warning')
            return redirect(url_for('egitim.test_sonuc', sonuc_id=sonuc_id))
    
    # Soruları al
    test_sorulari = test.test_sorulari.order_by(TestSorusu.sira).all()
    
    if test.soru_karistir:
        random.shuffle(test_sorulari)
    
    # Mevcut cevapları al
    mevcut_cevaplar = {c.soru_id: c for c in sonuc.cevaplar.all()}
    
    if request.method == 'POST':
        # Cevapları kaydet
        for ts in test_sorulari:
            soru = ts.soru
            cevap_key = f'soru_{soru.id}'
            
            if soru.soru_tipi == 'coklu_secim':
                secilen = request.form.getlist(cevap_key)
                secilen_ids = [int(s) for s in secilen] if secilen else None
            else:
                secilen = request.form.get(cevap_key)
                secilen_ids = int(secilen) if secilen else None
            
            # Mevcut cevabı güncelle veya yeni oluştur
            cevap = mevcut_cevaplar.get(soru.id)
            if not cevap:
                cevap = TestCevap(sonuc_id=sonuc_id, soru_id=soru.id)
                db.session.add(cevap)
            
            if soru.soru_tipi == 'coklu_secim':
                cevap.secilen_secenekler = secilen_ids
            else:
                cevap.secilen_secenek_id = secilen_ids
        
        db.session.commit()
        
        # Bitir mi?
        if request.form.get('bitir'):
            sonuc.tamamlandi = True
            sonuc.bitis_zamani = datetime.now()
            sonuc.gecen_sure_saniye = int((sonuc.bitis_zamani - sonuc.baslangic_zamani).total_seconds())
            _hesapla_sonuc(sonuc)
            db.session.commit()
            
            flash('Sınavınız tamamlandı!', 'success')
            return redirect(url_for('egitim.test_sonuc', sonuc_id=sonuc_id))
        
        flash('Cevaplarınız kaydedildi.', 'info')
    
    return render_template('egitim/test_coz.html',
                          test=test,
                          sonuc=sonuc,
                          test_sorulari=test_sorulari,
                          mevcut_cevaplar=mevcut_cevaplar,
                          kalan_sure=kalan_sure)


def _hesapla_sonuc(sonuc):
    """Test sonucunu hesapla"""
    test = sonuc.test
    
    dogru = 0
    yanlis = 0
    bos = 0
    alinan_puan = 0
    
    for ts in test.test_sorulari.all():
        soru = ts.soru
        cevap = sonuc.cevaplar.filter_by(soru_id=soru.id).first()
        
        if not cevap or (not cevap.secilen_secenek_id and not cevap.secilen_secenekler):
            bos += 1
            if cevap:
                cevap.dogru = False
                cevap.alinan_puan = 0
            continue
        
        # Doğru mu kontrol et
        if soru.soru_tipi == 'coklu_secim':
            # Çoklu seçim
            dogru_ids = set(s.id for s in soru.dogru_secenekler)
            secilen_ids = set(cevap.secilen_secenekler or [])
            dogru_mu = dogru_ids == secilen_ids
        else:
            # Tekli seçim
            dogru_secenek = soru.dogru_secenek
            dogru_mu = dogru_secenek and cevap.secilen_secenek_id == dogru_secenek.id
        
        cevap.dogru = dogru_mu
        
        if dogru_mu:
            dogru += 1
            cevap.alinan_puan = ts.puan
            alinan_puan += ts.puan
        else:
            yanlis += 1
            cevap.alinan_puan = 0
    
    sonuc.dogru_sayisi = dogru
    sonuc.yanlis_sayisi = yanlis
    sonuc.bos_sayisi = bos
    sonuc.alinan_puan = alinan_puan
    sonuc.yuzde = round((alinan_puan / sonuc.toplam_puan * 100), 1) if sonuc.toplam_puan > 0 else 0
    sonuc.gecti = sonuc.yuzde >= test.gecme_puani


@egitim_bp.route('/test/sonuc/<int:sonuc_id>')
@login_required
def test_sonuc(sonuc_id):
    """Test sonuç sayfası"""
    sonuc = TestSonuc.query.get_or_404(sonuc_id)
    test = sonuc.test
    
    # Yetki kontrolü
    calisan = current_user.calisan if hasattr(current_user, "calisan") else None
    is_owner = calisan and sonuc.calisan_id == calisan.id
    is_admin = current_user.has_permission('egitim.edit')
    
    if not is_owner and not is_admin:
        flash('Bu sonuca erişim yetkiniz yok.', 'danger')
        return redirect(url_for('egitim.test_liste'))
    
    # Cevapları al
    cevaplar = []
    for ts in test.test_sorulari.order_by(TestSorusu.sira).all():
        cevap = sonuc.cevaplar.filter_by(soru_id=ts.soru_id).first()
        cevaplar.append({
            'soru': ts.soru,
            'cevap': cevap
        })
    
    return render_template('egitim/test_sonuc.html',
                          test=test,
                          sonuc=sonuc,
                          cevaplar=cevaplar)


# ============================================================
# EĞİTİME BAĞLI TESTLERİ GÖR
# ============================================================

@egitim_bp.route('/<int:id>/testler')
@login_required
@permission_required('egitim.view')
def egitim_testleri(id):
    """Eğitime bağlı testler"""
    egitim = Egitim.query.get_or_404(id)
    testler = egitim.testler.filter_by(is_deleted=False).all()
    
    return render_template('egitim/egitim_testleri.html',
                          egitim=egitim,
                          testler=testler)

# ============================================================
# JITSI ONLINE EĞİTİM
# ============================================================

@egitim_bp.route('/<int:id>/jitsi/baslat', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def jitsi_baslat(id):
    """Online eğitimi başlat - Jitsi odası oluştur"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    if egitim.lokasyon_tipi not in ['online', 'hibrit']:
        flash('Bu eğitim online değil.', 'warning')
        return redirect(url_for('egitim.detay', id=id))

    # Oda adı oluştur
    if not egitim.jitsi_room_name:
        egitim.jitsi_room_name = JitsiService.create_room_name('egitim', egitim.id)

    egitim.jitsi_aktif = True
    egitim.durum = 'devam_ediyor'
    db.session.commit()

    flash('Online eğitim başlatıldı!', 'success')
    return redirect(url_for('egitim.detay', id=id))


@egitim_bp.route('/<int:id>/jitsi/durdur', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def jitsi_durdur(id):
    """Online eğitimi durdur"""
    egitim = Egitim.query.get_or_404(id)

    hata = _yonetim_yetki_hatasi(egitim)
    if hata:
        return hata

    egitim.jitsi_aktif = False
    # Aktif katılımcıların ayrılış zamanını işaretle (kalma süresi hesabı için)
    _katilim_loglarini_kapat(egitim.id)
    db.session.commit()

    flash('Online eğitim durduruldu.', 'info')
    return redirect(url_for('egitim.detay', id=id))


@egitim_bp.route('/<int:id>/katilim-log-export')
@login_required
@permission_required('egitim.view')
def katilim_log_export(id):
    """Eğitim katılım logunu Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    egitim = Egitim.query.get_or_404(id)
    loglar = egitim.katilim_loglari.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Katilim Logu'

    headers = ['Ad Soyad', 'Telefon', 'Eşleşme', 'Giriş Zamanı',
               'Ayrılış Zamanı', 'Süre (dk)', 'IP']
    ws.append(headers)
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='137FEC')
    for c in ws[1]:
        c.font = hf
        c.fill = hfill

    for log in loglar:
        ws.append([
            log.ad_soyad or '',
            ('0' + log.telefon) if log.telefon else '',
            log.eslesme_tipi,
            log.giris_zamani.strftime('%d.%m.%Y %H:%M') if log.giris_zamani else '',
            log.ayrilma_zamani.strftime('%d.%m.%Y %H:%M') if log.ayrilma_zamani else '',
            log.kalma_suresi_dk if log.kalma_suresi_dk is not None else '',
            log.ip or '',
        ])

    for i, w in enumerate([28, 15, 12, 18, 18, 10, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"katilim_log_{id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


# ============================================================
# DIŞ KATILIM (Public - login gerektirmez, telefon eşleştirmeli)
# ============================================================

class _MisafirKullanici:
    """JitsiService için basit kullanıcı objesi (full_name + email)."""
    def __init__(self, full_name, email=''):
        self.full_name = full_name
        self.email = email or ''


def _istek_ip():
    """Nginx arkasında gerçek IP (X-Forwarded-For ilk değer)."""
    xff = request.headers.get('X-Forwarded-For', '')
    if xff:
        return xff.split(',')[0].strip()
    return request.remote_addr or ''


def _normalize_tel(ham):
    """Telefonu son 10 haneye normalize eder (5XXXXXXXXX).

    Geçersiz Türk cep numarası ise None döner (5 ile başlayan 10 hane şartı).
    """
    if not ham:
        return None
    digits = re.sub(r'\D', '', ham)
    if digits.startswith('90') and len(digits) == 12:
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith('0'):
        digits = digits[1:]
    if len(digits) < 10:
        return None
    digits = digits[-10:]
    if not digits.startswith('5'):
        return None
    return digits


def _telefon_eslesme(norm10):
    """norm10 (10 hane) ile önce Calisan, yoksa onaylı Aday eşleştirir.

    Returns (calisan, aday) - en fazla biri dolu.
    """
    if not norm10:
        return None, None
    # DB tarafında telefonu rakamlaştırıp son 10 haneyi karşılaştır (format bağımsız)
    cal_norm = db.func.right(db.func.regexp_replace(Calisan.telefon, r'[^0-9]', '', 'g'), 10)
    calisan = Calisan.query.filter(
        Calisan.is_deleted == False,
        Calisan.telefon.isnot(None),
        cal_norm == norm10,
    ).first()
    if calisan:
        return calisan, None

    aday_norm = db.func.right(db.func.regexp_replace(Aday.telefon, r'[^0-9]', '', 'g'), 10)
    aday = Aday.query.filter(
        Aday.is_deleted == False,
        Aday.calisan_id.is_(None),  # henüz çalışana dönüşmemiş
        Aday.telefon.isnot(None),
        Aday.durum.in_(['onaylandi', 'sgk_giris_talebi', 'sgk_girisi_yapildi']),
        aday_norm == norm10,
    ).first()
    return None, aday


def _rate_limit_ok(anahtar, limit=5, pencere=60):
    """Redis ile IP başına pencere içindeki istek limiti. Redis yoksa izin verir (fail-open)."""
    try:
        import redis as _redis
        url = os.environ.get('REDIS_URL') or current_app.config.get('REDIS_URL') or 'redis://redis:6379/0'
        r = _redis.from_url(url)
        sayi = r.incr(anahtar)
        if sayi == 1:
            r.expire(anahtar, pencere)
        return sayi <= limit
    except Exception as e:
        current_app.logger.warning(f"Rate limit kontrolü yapılamadı, izin veriliyor: {e}")
        return True


@egitim_bp.route('/katil/<int:id>', methods=['GET', 'POST'])
def katil(id):
    """Online eğitime dış katılım (login gerektirmez).

    Ad Soyad + Telefon alınır; telefon ile çalışan/aday eşleştirilir,
    JWT üretilip Jitsi odasına yönlendirilir. Her giriş loglanır.
    """
    egitim = Egitim.query.get_or_404(id)

    # Sadece aktif/canlı eğitimler için link çalışsın
    kapali = (egitim.is_deleted or egitim.durum == 'iptal'
              or not egitim.jitsi_aktif or not egitim.jitsi_room_name)
    if kapali:
        return render_template('egitim/katil.html', egitim=egitim, kapali=True)

    if request.method == 'POST':
        ip = _istek_ip()

        # Rate limit: aynı IP'den dakikada max 5 istek
        if not _rate_limit_ok(f'egitim_katil:{id}:{ip}', limit=5, pencere=60):
            return render_template(
                'egitim/katil.html', egitim=egitim,
                hata='Çok fazla deneme yaptınız. Lütfen bir dakika sonra tekrar deneyin.',
                ad_soyad=request.form.get('ad_soyad', ''),
                telefon=request.form.get('telefon', '')), 429

        ad_soyad = (request.form.get('ad_soyad') or '').strip()
        telefon = (request.form.get('telefon') or '').strip()
        norm10 = _normalize_tel(telefon)

        if len(ad_soyad) < 3 or not norm10:
            return render_template(
                'egitim/katil.html', egitim=egitim,
                hata='Lütfen adınızı ve geçerli bir cep telefonu numarası girin (5XX XXX XX XX).',
                ad_soyad=ad_soyad, telefon=telefon)

        calisan, aday = _telefon_eslesme(norm10)

        # Eşleşirse resmi ad/email kullan
        goruntu_ad = ad_soyad
        email = ''
        if calisan:
            goruntu_ad = calisan.full_name
            email = calisan.email or ''
        elif aday:
            goruntu_ad = aday.full_name
            email = aday.email or ''

        # Katılım logu
        db.session.add(EgitimKatilimLog(
            egitim_id=egitim.id,
            ad_soyad=goruntu_ad,
            telefon=norm10,
            calisan_id=calisan.id if calisan else None,
            aday_id=aday.id if aday else None,
            giris_zamani=datetime.now(),
            ip=ip,
        ))

        # Çalışansa varsa katılımcı kaydını da 'katildi' yap
        if calisan:
            katilimci = EgitimKatilimci.query.filter_by(
                egitim_id=egitim.id, calisan_id=calisan.id).first()
            if katilimci and katilimci.durum in ('davetli', None):
                katilimci.durum = 'katildi'
                katilimci.katilim_tarihi = datetime.now()

        db.session.commit()

        # Dış katılımcı: JWT verme, guest (participant) olarak yönlendir
        meeting_url = JitsiService.get_guest_url(
            egitim.jitsi_room_name,
            goruntu_ad,
        )
        return redirect(meeting_url)

    return render_template('egitim/katil.html', egitim=egitim)


@egitim_bp.route('/<int:id>/jitsi/katil')
@login_required
def jitsi_katil(id):
    """Online eğitime katıl - JWT ile Jitsi'ye yönlendir"""
    egitim = Egitim.query.get_or_404(id)

    if not egitim.jitsi_aktif:
        flash('Bu eğitimin canlı yayını şu an aktif değil.', 'warning')
        return redirect(url_for('egitim.detay', id=id))

    # Kullanıcı bilgilerini al
    calisan = current_user.calisan if hasattr(current_user, "calisan") else None

    if calisan:
        user_obj = calisan
    else:
        user_obj = current_user

    # Eğitmen mi kontrol et (moderatör yetkisi için)
    is_moderator = False
    if egitim.egitmen_id and calisan and egitim.egitmen_id == calisan.id:
        is_moderator = True
    if current_user.has_permission('egitim.edit'):
        is_moderator = True

    # Moderatör → JWT ile owner olur; katılımcı → JWT'siz guest (participant) kalır
    if is_moderator:
        meeting_url = JitsiService.get_meeting_url(
            egitim.jitsi_room_name,
            user_obj,
            is_moderator=True
        )
    else:
        goruntu_ad = getattr(user_obj, 'full_name', None) or getattr(user_obj, 'username', '')
        meeting_url = JitsiService.get_guest_url(
            egitim.jitsi_room_name,
            goruntu_ad
        )

    # Katılım kaydı oluştur/güncelle
    if calisan:
        katilimci = EgitimKatilimci.query.filter_by(
            egitim_id=id,
            calisan_id=calisan.id
        ).first()

        if katilimci:
            katilimci.katilim_tarihi = datetime.now()
            katilimci.durum = 'katildi'
            db.session.commit()

    return redirect(meeting_url)


@egitim_bp.route('/<int:id>/jitsi/embed')
@login_required
def jitsi_embed(id):
    """Jitsi'yi iframe içinde göster (opsiyonel)"""
    egitim = Egitim.query.get_or_404(id)

    if not egitim.jitsi_aktif:
        flash('Bu eğitimin canlı yayını şu an aktif değil.', 'warning')
        return redirect(url_for('egitim.detay', id=id))

    calisan = current_user.calisan if hasattr(current_user, "calisan") else None
    user_obj = calisan if calisan else current_user

    is_moderator = False
    if egitim.egitmen_id and calisan and egitim.egitmen_id == calisan.id:
        is_moderator = True
    if current_user.has_permission('egitim.edit'):
        is_moderator = True

    # Jitsi URL oluştur (JWT dahil) + yönlendir
    meeting_url = JitsiService.get_meeting_url(
        egitim.jitsi_room_name,
        user_obj,
        is_moderator
    )
    return redirect(meeting_url)


# ============================================================
# JITSI WEBHOOK - Katılım Takibi
# ============================================================

@csrf.exempt
@egitim_bp.route('/jitsi/webhook', methods=['POST'])
def jitsi_webhook():
    """
    Jitsi webhook endpoint - katılım takibi için
    Jitsi'den gelen eventler:
    - participant_joined: Kullanıcı odaya girdi
    - participant_left: Kullanıcı odadan çıktı
    """
    from flask import request, jsonify
    import json

    # Webhook secret kontrolü (güvenlik için)
    webhook_secret = request.headers.get('X-Webhook-Secret', '')
    if webhook_secret != 'TgPortalJitsiWebhook2025!':
        return jsonify({'error': 'Unauthorized'}), 401

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data'}), 400

        event_type = data.get('event')
        room_name = data.get('room_name', '')
        participant = data.get('participant', {})
        user_email = participant.get('email', '')
        user_name = participant.get('name', '')
        timestamp = datetime.now()

        # Room name'den egitim_id çıkar (format: egitim_123_xxx)
        egitim_id = None
        if room_name.startswith('egitim_'):
            parts = room_name.split('_')
            if len(parts) >= 2:
                try:
                    egitim_id = int(parts[1])
                except ValueError:
                    pass

        if not egitim_id:
            return jsonify({'status': 'ignored', 'reason': 'not an education room'}), 200

        # Eğitimi bul
        egitim = Egitim.query.get(egitim_id)
        if not egitim:
            return jsonify({'status': 'ignored', 'reason': 'education not found'}), 200

        # Kullanıcıyı email ile bul
        calisan = None
        if user_email:
            calisan = Calisan.query.filter_by(email=user_email, is_deleted=False).first()

        if not calisan:
            # Email ile bulunamadıysa isimle dene
            if user_name:
                parts = user_name.strip().split(' ', 1)
                if len(parts) == 2:
                    calisan = Calisan.query.filter_by(
                        ad=parts[0],
                        soyad=parts[1],
                        is_deleted=False
                    ).first()

        if not calisan:
            return jsonify({'status': 'ignored', 'reason': 'participant not found'}), 200

        # Katılımcı kaydını bul
        katilimci = EgitimKatilimci.query.filter_by(
            egitim_id=egitim_id,
            calisan_id=calisan.id
        ).first()

        if not katilimci:
            return jsonify({'status': 'ignored', 'reason': 'not a registered participant'}), 200

        if event_type == 'participant_joined':
            # Katılım başlangıcını kaydet
            katilimci.jitsi_katilim_baslangic = timestamp
            katilimci.jitsi_katilim_sayisi = (katilimci.jitsi_katilim_sayisi or 0) + 1
            katilimci.durum = 'katildi'
            if not katilimci.katilim_tarihi:
                katilimci.katilim_tarihi = timestamp
            db.session.commit()

        elif event_type == 'participant_left':
            # Süreyi hesapla ve kaydet
            if katilimci.jitsi_katilim_baslangic:
                sure_saniye = int((timestamp - katilimci.jitsi_katilim_baslangic).total_seconds())
                katilimci.jitsi_toplam_sure = (katilimci.jitsi_toplam_sure or 0) + sure_saniye
                katilimci.jitsi_katilim_bitis = timestamp
            db.session.commit()

        return jsonify({'status': 'ok'}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@egitim_bp.route('/<int:id>/katilim-raporu')
@login_required
@permission_required('egitim.view')
def katilim_raporu(id):
    """Online eğitim katılım raporu"""
    egitim = Egitim.query.get_or_404(id)

    katilimcilar = EgitimKatilimci.query.filter_by(egitim_id=id).join(Calisan).order_by(Calisan.ad).all()

    # İstatistikler
    toplam_katilimci = len(katilimcilar)
    online_katilan = sum(1 for k in katilimcilar if k.jitsi_toplam_sure and k.jitsi_toplam_sure > 0)

    # Ortalama katılım süresi
    toplam_sure = sum(k.jitsi_toplam_sure or 0 for k in katilimcilar)
    ortalama_sure = toplam_sure // online_katilan if online_katilan > 0 else 0

    return render_template('egitim/katilim_raporu.html',
                          egitim=egitim,
                          katilimcilar=katilimcilar,
                          toplam_katilimci=toplam_katilimci,
                          online_katilan=online_katilan,
                          ortalama_sure=ortalama_sure)


# ============================================================
# EĞİTİM BOOKING - OTURUM YÖNETİMİ (İK tarafı)
# ============================================================

def _saat_parse(ham):
    """'HH:MM' -> time. Geçersizse None."""
    if not ham:
        return None
    try:
        return datetime.strptime(ham.strip()[:5], '%H:%M').time()
    except ValueError:
        return None


def _oturum_formu_oku(form):
    """Oturum formunu okur. (veri, hata) döner."""
    tarih_ham = (form.get('tarih') or '').strip()
    try:
        tarih = datetime.strptime(tarih_ham, '%Y-%m-%d').date()
    except ValueError:
        return None, 'Geçerli bir tarih seçin.'

    baslangic = _saat_parse(form.get('baslangic_saati'))
    if not baslangic:
        return None, 'Geçerli bir başlangıç saati girin.'

    bitis = _saat_parse(form.get('bitis_saati'))
    if bitis and bitis <= baslangic:
        return None, 'Bitiş saati başlangıç saatinden sonra olmalı.'

    try:
        kontenjan = int(form.get('kontenjan') or 0)
    except ValueError:
        kontenjan = 0
    if kontenjan < 1:
        return None, 'Kontenjan en az 1 olmalı.'

    return {
        'tarih': tarih,
        'baslangic_saati': baslangic,
        'bitis_saati': bitis,
        'kontenjan': kontenjan,
        'aciklama': (form.get('aciklama') or '').strip() or None,
        'toplanti_linki': (form.get('toplanti_linki') or '').strip() or None,
        'jitsi_otomatik': form.get('jitsi_otomatik') == 'on',
    }, None


def _jitsi_oturum_linki(egitim, oturum):
    """Oturuma özel Jitsi oda linki üretir."""
    oda = JitsiService.create_room_name('egitim', f'{egitim.id}_{oturum.id}')
    return f'https://{JitsiService.JITSI_DOMAIN}/{oda}'


@egitim_bp.route('/<int:id>/oturum/ekle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def oturum_ekle(id):
    """Eğitime yeni oturum ekle."""
    egitim = Egitim.query.get_or_404(id)
    hata_redirect = _yonetim_yetki_hatasi(egitim)
    if hata_redirect:
        return hata_redirect

    veri, hata = _oturum_formu_oku(request.form)
    if hata:
        flash(hata, 'danger')
        return redirect(url_for('egitim.detay', id=id))

    jitsi_otomatik = veri.pop('jitsi_otomatik')
    oturum = EgitimOturumu(egitim_id=egitim.id, **veri)
    db.session.add(oturum)
    db.session.flush()  # oturum.id gerekiyor (Jitsi oda adı için)

    if jitsi_otomatik and not oturum.toplanti_linki:
        oturum.toplanti_linki = _jitsi_oturum_linki(egitim, oturum)

    db.session.commit()
    flash(f'Oturum eklendi: {oturum.zaman_text}', 'success')
    return redirect(url_for('egitim.detay', id=id) + '#oturumlar')


@egitim_bp.route('/oturum/<int:oturum_id>/duzenle', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def oturum_duzenle(oturum_id):
    """Oturum bilgilerini güncelle (kontenjan artır/azalt dahil)."""
    oturum = EgitimOturumu.query.get_or_404(oturum_id)
    egitim = oturum.egitim
    hata_redirect = _yonetim_yetki_hatasi(egitim)
    if hata_redirect:
        return hata_redirect

    veri, hata = _oturum_formu_oku(request.form)
    if hata:
        flash(hata, 'danger')
        return redirect(url_for('egitim.detay', id=egitim.id))

    # Kontenjan mevcut kayıt sayısının altına indirilemez
    mevcut_kayit = oturum.kayit_sayisi
    if veri['kontenjan'] < mevcut_kayit:
        flash(f'Kontenjan mevcut kayıt sayısının ({mevcut_kayit}) altına indirilemez.', 'danger')
        return redirect(url_for('egitim.detay', id=egitim.id))

    jitsi_otomatik = veri.pop('jitsi_otomatik')
    for alan, deger in veri.items():
        setattr(oturum, alan, deger)
    oturum.aktif = request.form.get('aktif') == 'on'

    if jitsi_otomatik and not oturum.toplanti_linki:
        oturum.toplanti_linki = _jitsi_oturum_linki(egitim, oturum)

    db.session.commit()
    flash('Oturum güncellendi.', 'success')
    return redirect(url_for('egitim.detay', id=egitim.id) + '#oturumlar')


@egitim_bp.route('/oturum/<int:oturum_id>/sil', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def oturum_sil(oturum_id):
    """Oturumu sil. Aktif kaydı varsa silinmez (önce kayıtlar iptal edilmeli)."""
    oturum = EgitimOturumu.query.get_or_404(oturum_id)
    egitim = oturum.egitim
    hata_redirect = _yonetim_yetki_hatasi(egitim)
    if hata_redirect:
        return hata_redirect

    if oturum.kayit_sayisi > 0:
        flash('Bu oturumda aktif kayıt var. Önce kayıtları iptal edin veya oturumu pasife alın.', 'danger')
        return redirect(url_for('egitim.detay', id=egitim.id))

    db.session.delete(oturum)
    db.session.commit()
    flash('Oturum silindi.', 'success')
    return redirect(url_for('egitim.detay', id=egitim.id) + '#oturumlar')


@egitim_bp.route('/kayit/<int:kayit_id>/iptal-yonetim', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def kayit_iptal_yonetim(kayit_id):
    """İK tarafından kayıt iptali."""
    kayit = EgitimKayit.query.get_or_404(kayit_id)
    egitim = kayit.egitim
    hata_redirect = _yonetim_yetki_hatasi(egitim)
    if hata_redirect:
        return hata_redirect

    if kayit.durum != 'iptal':
        kayit.durum = 'iptal'
        kayit.iptal_zamani = datetime.now()
        kayit.iptal_eden = 'ik'
        db.session.commit()
        flash(f'{kayit.ad_soyad} kaydı iptal edildi.', 'success')
    return redirect(url_for('egitim.detay', id=egitim.id) + '#kayitlar')


@egitim_bp.route('/<int:id>/kayit-export')
@login_required
@permission_required('egitim.view')
def kayit_export(id):
    """Eğitim booking kayıtlarını Excel olarak indir."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from flask import Response

    egitim = Egitim.query.get_or_404(id)
    kayitlar = (EgitimKayit.query
                .filter_by(egitim_id=id)
                .join(EgitimOturumu, EgitimKayit.oturum_id == EgitimOturumu.id)
                .order_by(EgitimOturumu.tarih, EgitimOturumu.baslangic_saati,
                          EgitimKayit.kayit_zamani)
                .all())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kayitlar'

    headers = ['Oturum', 'Ad Soyad', 'Telefon', 'E-posta', 'Eşleşme',
               'Durum', 'Kayıt Zamanı', 'Anket Puanı']
    ws.append(headers)
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='137FEC')
    for c in ws[1]:
        c.font = hf
        c.fill = hfill

    for k in kayitlar:
        ws.append([
            k.oturum.zaman_text if k.oturum else '',
            k.ad_soyad or '',
            ('0' + k.telefon) if k.telefon else '',
            k.email or '',
            k.eslesme_tipi,
            'İptal' if k.durum == 'iptal' else 'Onaylandı',
            k.kayit_zamani.strftime('%d.%m.%Y %H:%M') if k.kayit_zamani else '',
            k.anket.puan if k.anket else '',
        ])

    for i, w in enumerate([24, 28, 15, 26, 12, 12, 18, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"egitim_kayitlari_{id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@egitim_bp.route('/oturum/<int:oturum_id>/anket-sms', methods=['POST'])
@login_required
@permission_required('egitim.edit')
def oturum_anket_sms(oturum_id):
    """Oturumdaki katılımcılara anket linkini SMS ile gönder."""
    oturum = EgitimOturumu.query.get_or_404(oturum_id)
    egitim = oturum.egitim
    hata_redirect = _yonetim_yetki_hatasi(egitim)
    if hata_redirect:
        return hata_redirect

    from app.modules.basvuru.routes import send_netgsm_sms

    gonderilen, basarisiz = 0, 0
    for kayit in oturum.kayitlar.filter_by(durum='onaylandi').all():
        if kayit.anket:
            continue  # zaten yanıtlamış
        link = url_for('egitim.anket', token=kayit.token, _external=True)
        mesaj = (f"Team Guerilla - {egitim.baslik} egitimine katiliminiz icin tesekkurler. "
                 f"Kisa anketimizi doldurur musunuz? {link}")
        sonuc = send_netgsm_sms('0' + kayit.telefon, mesaj)
        if sonuc.get('success'):
            gonderilen += 1
        else:
            basarisiz += 1
            current_app.logger.warning(
                f"Anket SMS gonderilemedi (kayit={kayit.id}): {sonuc.get('error')}")

    if gonderilen:
        flash(f'{gonderilen} katılımcıya anket linki gönderildi.', 'success')
    if basarisiz:
        flash(f'{basarisiz} SMS gönderilemedi. Detay için log kayıtlarına bakın.', 'warning')
    if not gonderilen and not basarisiz:
        flash('Anket gönderilecek katılımcı bulunamadı.', 'info')
    return redirect(url_for('egitim.detay', id=egitim.id) + '#oturumlar')


# ============================================================
# PUBLIC BOOKING (login gerektirmez - telefon + OTP doğrulamalı)
# ============================================================

KAYIT_SESSION_KEY = 'egitim_kayit'


def _kayit_oturumu(egitim_id):
    """Session'daki booking durumunu döner (farklı eğitime aitse sıfırlar)."""
    veri = session.get(KAYIT_SESSION_KEY)
    if not isinstance(veri, dict) or veri.get('egitim_id') != egitim_id:
        return None
    return veri


def _kayit_oturumu_yaz(veri):
    session[KAYIT_SESSION_KEY] = veri
    session.modified = True


def _otp_uret(veri):
    """Yeni OTP üretip session'a yazar, kodu döner."""
    kod = str(random.randint(100000, 999999))
    veri['otp_kod'] = kod
    veri['otp_expires'] = (datetime.now() + timedelta(minutes=5)).isoformat()
    veri['otp_deneme'] = 0
    _kayit_oturumu_yaz(veri)
    return kod


def _otp_dogrula(veri, girilen):
    """(basarili, mesaj) döner. Başarılıysa session'da dogrulandi=True olur."""
    kod = (girilen or '').strip()
    if not veri.get('otp_kod') or not veri.get('otp_expires'):
        return False, 'Doğrulama kodu bulunamadı. Lütfen yeni kod isteyin.'
    if datetime.now() > datetime.fromisoformat(veri['otp_expires']):
        return False, 'Kodun süresi doldu. Lütfen yeni kod isteyin.'
    if veri.get('otp_deneme', 0) >= 3:
        return False, 'Çok fazla yanlış deneme. Lütfen yeni kod isteyin.'
    if veri['otp_kod'] != kod:
        veri['otp_deneme'] = veri.get('otp_deneme', 0) + 1
        _kayit_oturumu_yaz(veri)
        return False, f"Yanlış kod. {3 - veri['otp_deneme']} deneme hakkınız kaldı."

    veri['dogrulandi'] = True
    veri['otp_kod'] = None
    _kayit_oturumu_yaz(veri)
    return True, None


def _booking_acik_mi(egitim):
    """Booking sayfası bu eğitim için açık mı? (kapalıysa sebep metni döner)"""
    if egitim.is_deleted or egitim.durum == 'iptal':
        return False, 'Bu eğitim iptal edilmiş.'
    if egitim.durum == 'tamamlandi':
        return False, 'Bu eğitim tamamlanmış.'
    if not egitim.oturumlar.filter_by(aktif=True).count():
        return False, 'Bu eğitim için henüz kayıt açılmamış.'
    return True, None


def _booking_render(egitim, adim, **kwargs):
    """Booking sayfasını uygun adımla render eder."""
    oturumlar = []
    if adim == 'oturum':
        oturumlar = egitim.oturumlar.filter_by(aktif=True).all()
        oturumlar = [o for o in oturumlar if not o.gecmis_mi]
    return render_template('egitim/kayit.html', egitim=egitim, adim=adim,
                           oturumlar=oturumlar, **kwargs)


@egitim_bp.route('/kayit/<int:id>')
def kayit(id):
    """Public booking sayfası - eğitime kayıt (login gerektirmez)."""
    egitim = Egitim.query.get_or_404(id)

    acik, kapali_mesaj = _booking_acik_mi(egitim)
    if not acik:
        return render_template('egitim/kayit.html', egitim=egitim,
                               adim='kapali', kapali_mesaj=kapali_mesaj)

    veri = _kayit_oturumu(id)
    if veri and veri.get('dogrulandi'):
        return _booking_render(egitim, 'oturum', telefon=veri.get('telefon'))
    if veri and veri.get('otp_kod'):
        return _booking_render(egitim, 'kod', telefon=veri.get('telefon'))
    return _booking_render(egitim, 'telefon')


@egitim_bp.route('/kayit/<int:id>/kod-gonder', methods=['POST'])
def kayit_kod_gonder(id):
    """Telefona OTP kodu gönder."""
    egitim = Egitim.query.get_or_404(id)
    acik, kapali_mesaj = _booking_acik_mi(egitim)
    if not acik:
        return render_template('egitim/kayit.html', egitim=egitim,
                               adim='kapali', kapali_mesaj=kapali_mesaj)

    ad_soyad = (request.form.get('ad_soyad') or '').strip()
    telefon_ham = (request.form.get('telefon') or '').strip()
    email = (request.form.get('email') or '').strip() or None
    norm10 = _normalize_tel(telefon_ham)

    if len(ad_soyad) < 3 or not norm10:
        return _booking_render(
            egitim, 'telefon', ad_soyad=ad_soyad, telefon=telefon_ham, email=email,
            hata='Lütfen adınızı ve geçerli bir cep telefonu numarası girin (5XX XXX XX XX).')

    ip = _istek_ip()
    # SMS maliyeti ve numara tacizini önlemek için IP + numara bazlı limit
    if not _rate_limit_ok(f'egitim_kayit_otp_ip:{ip}', limit=5, pencere=300) or \
       not _rate_limit_ok(f'egitim_kayit_otp_tel:{norm10}', limit=3, pencere=600):
        return _booking_render(
            egitim, 'telefon', ad_soyad=ad_soyad, telefon=telefon_ham, email=email,
            hata='Çok fazla kod talebi gönderdiniz. Lütfen bir süre sonra tekrar deneyin.'), 429

    # Bu eğitime bu telefondan aktif kayıt var mı?
    mevcut = EgitimKayit.query.filter_by(
        egitim_id=egitim.id, telefon=norm10, durum='onaylandi').first()
    if mevcut:
        return render_template('egitim/kayit_tamam.html', kayit=mevcut,
                               egitim=egitim, zaten_kayitli=True)

    veri = {
        'egitim_id': egitim.id,
        'ad_soyad': ad_soyad,
        'telefon': norm10,
        'email': email,
        'dogrulandi': False,
        'ip': ip,
    }
    kod = _otp_uret(veri)

    from app.modules.basvuru.routes import send_netgsm_sms
    mesaj = f"Team Guerilla egitim kayit dogrulama kodunuz: {kod} - Bu kod 5 dakika gecerlidir."
    sonuc = send_netgsm_sms('0' + norm10, mesaj)
    if not sonuc.get('success'):
        current_app.logger.error(f"Egitim kayit OTP gonderilemedi ({norm10}): {sonuc.get('error')}")
        return _booking_render(
            egitim, 'telefon', ad_soyad=ad_soyad, telefon=telefon_ham, email=email,
            hata='Doğrulama kodu gönderilemedi. Lütfen numaranızı kontrol edip tekrar deneyin.')

    return _booking_render(egitim, 'kod', telefon=norm10,
                           bilgi='Doğrulama kodu telefonunuza gönderildi.')


@egitim_bp.route('/kayit/<int:id>/kod-tekrar', methods=['POST'])
def kayit_kod_tekrar(id):
    """OTP kodunu yeniden gönder."""
    egitim = Egitim.query.get_or_404(id)
    veri = _kayit_oturumu(id)
    if not veri or not veri.get('telefon'):
        return redirect(url_for('egitim.kayit', id=id))

    ip = _istek_ip()
    if not _rate_limit_ok(f'egitim_kayit_otp_ip:{ip}', limit=5, pencere=300) or \
       not _rate_limit_ok(f'egitim_kayit_otp_tel:{veri["telefon"]}', limit=3, pencere=600):
        return _booking_render(
            egitim, 'kod', telefon=veri['telefon'],
            hata='Çok fazla kod talebi gönderdiniz. Lütfen bir süre sonra tekrar deneyin.'), 429

    kod = _otp_uret(veri)
    from app.modules.basvuru.routes import send_netgsm_sms
    mesaj = f"Team Guerilla egitim kayit dogrulama kodunuz: {kod} - Bu kod 5 dakika gecerlidir."
    sonuc = send_netgsm_sms('0' + veri['telefon'], mesaj)
    if not sonuc.get('success'):
        return _booking_render(egitim, 'kod', telefon=veri['telefon'],
                               hata='Kod gönderilemedi. Lütfen tekrar deneyin.')

    return _booking_render(egitim, 'kod', telefon=veri['telefon'],
                           bilgi='Yeni doğrulama kodu gönderildi.')


@egitim_bp.route('/kayit/<int:id>/dogrula', methods=['POST'])
def kayit_dogrula(id):
    """OTP kodunu doğrula, oturum seçim adımına geç."""
    egitim = Egitim.query.get_or_404(id)
    veri = _kayit_oturumu(id)
    if not veri:
        return redirect(url_for('egitim.kayit', id=id))

    ip = _istek_ip()
    if not _rate_limit_ok(f'egitim_kayit_dogrula:{ip}', limit=15, pencere=300):
        return _booking_render(egitim, 'kod', telefon=veri.get('telefon'),
                               hata='Çok fazla deneme yaptınız. Lütfen biraz bekleyin.'), 429

    basarili, mesaj = _otp_dogrula(veri, request.form.get('kod'))
    if not basarili:
        return _booking_render(egitim, 'kod', telefon=veri.get('telefon'), hata=mesaj)

    return _booking_render(egitim, 'oturum', telefon=veri.get('telefon'))


@egitim_bp.route('/kayit/<int:id>/kaydol', methods=['POST'])
def kayit_kaydol(id):
    """Seçilen oturuma kaydı oluştur (kontenjan kilidi ile)."""
    egitim = Egitim.query.get_or_404(id)
    veri = _kayit_oturumu(id)
    if not veri or not veri.get('dogrulandi'):
        flash('Lütfen önce telefon numaranızı doğrulayın.', 'warning')
        return redirect(url_for('egitim.kayit', id=id))

    try:
        oturum_id = int(request.form.get('oturum_id') or 0)
    except ValueError:
        oturum_id = 0

    # Kontenjan kontrolünü serialize etmek için oturum satırını kilitle
    oturum = (db.session.query(EgitimOturumu)
              .filter_by(id=oturum_id, egitim_id=egitim.id)
              .with_for_update()
              .first())
    if not oturum:
        return _booking_render(egitim, 'oturum', telefon=veri.get('telefon'),
                               hata='Lütfen bir oturum seçin.')

    if not oturum.aktif or oturum.gecmis_mi:
        db.session.rollback()
        return _booking_render(egitim, 'oturum', telefon=veri.get('telefon'),
                               hata='Seçtiğiniz oturum artık kayda kapalı.')

    if oturum.dolu_mu:
        db.session.rollback()
        return _booking_render(egitim, 'oturum', telefon=veri.get('telefon'),
                               hata='Seçtiğiniz oturumun kontenjanı doldu. Lütfen başka bir oturum seçin.')

    telefon = veri['telefon']
    mevcut = EgitimKayit.query.filter_by(
        egitim_id=egitim.id, telefon=telefon, durum='onaylandi').first()
    if mevcut:
        db.session.rollback()
        return render_template('egitim/kayit_tamam.html', kayit=mevcut,
                               egitim=egitim, zaten_kayitli=True)

    calisan, aday = _telefon_eslesme(telefon)
    ad_soyad = veri.get('ad_soyad') or ''
    email = veri.get('email')
    if calisan:
        ad_soyad = calisan.full_name
        email = email or calisan.email
    elif aday:
        ad_soyad = aday.full_name
        email = email or aday.email

    kayit_obj = EgitimKayit(
        oturum_id=oturum.id,
        egitim_id=egitim.id,
        ad_soyad=ad_soyad,
        telefon=telefon,
        email=email,
        calisan_id=calisan.id if calisan else None,
        aday_id=aday.id if aday else None,
        kayit_zamani=datetime.now(),
        durum='onaylandi',
        ip=_istek_ip(),
    )
    db.session.add(kayit_obj)

    # Çalışansa eğitim katılımcı listesine de davetli olarak ekle
    if calisan:
        katilimci = EgitimKatilimci.query.filter_by(
            egitim_id=egitim.id, calisan_id=calisan.id).first()
        if not katilimci:
            db.session.add(EgitimKatilimci(
                egitim_id=egitim.id, calisan_id=calisan.id,
                durum='davetli', davet_tarihi=datetime.now()))

    try:
        db.session.commit()
    except IntegrityError:
        # Partial unique index: aynı telefondan eşzamanlı ikinci kayıt
        db.session.rollback()
        mevcut = EgitimKayit.query.filter_by(
            egitim_id=egitim.id, telefon=telefon, durum='onaylandi').first()
        if mevcut:
            return render_template('egitim/kayit_tamam.html', kayit=mevcut,
                                   egitim=egitim, zaten_kayitli=True)
        return _booking_render(egitim, 'oturum', telefon=telefon,
                               hata='Kayıt oluşturulamadı. Lütfen tekrar deneyin.')

    _kayit_bilgi_sms(kayit_obj)
    session.pop(KAYIT_SESSION_KEY, None)
    return redirect(url_for('egitim.kayit_tamam', token=kayit_obj.token))


def _kayit_bilgi_sms(kayit_obj):
    """Kayıt sonrası bilgilendirme SMS'i (hata durumunda kaydı bozmaz)."""
    try:
        from app.modules.basvuru.routes import send_netgsm_sms
        link = url_for('egitim.kayit_tamam', token=kayit_obj.token, _external=True)
        oturum = kayit_obj.oturum
        mesaj = (f"Egitime kaydiniz alinmistir. {kayit_obj.egitim.baslik} - "
                 f"Tarih: {oturum.tarih.strftime('%d.%m.%Y')} "
                 f"Saat: {oturum.baslangic_saati.strftime('%H:%M')}. "
                 f"Katilim linki ve iptal icin: {link}")
        sonuc = send_netgsm_sms('0' + kayit_obj.telefon, mesaj)
        if sonuc.get('success'):
            kayit_obj.sms_gonderildi = True
            db.session.commit()
        else:
            current_app.logger.warning(
                f"Kayit SMS gonderilemedi (kayit={kayit_obj.id}): {sonuc.get('error')}")
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Kayit SMS hatasi (kayit={kayit_obj.id}): {e}")


@egitim_bp.route('/kayit/tamam/<token>')
def kayit_tamam(token):
    """Kayıt onay sayfası - toplantı linki, iptal ve anket erişimi."""
    kayit_obj = EgitimKayit.query.filter_by(token=token).first_or_404()
    return render_template('egitim/kayit_tamam.html',
                           kayit=kayit_obj, egitim=kayit_obj.egitim)


@egitim_bp.route('/kayit/iptal/<token>', methods=['GET', 'POST'])
def kayit_iptal(token):
    """Katılımcının kendi kaydını iptal etmesi."""
    kayit_obj = EgitimKayit.query.filter_by(token=token).first_or_404()

    if request.method == 'POST':
        if not kayit_obj.iptal_edilebilir_mi:
            return render_template('egitim/kayit_iptal.html', kayit=kayit_obj,
                                   hata='Bu kayıt artık iptal edilemez.')
        kayit_obj.durum = 'iptal'
        kayit_obj.iptal_zamani = datetime.now()
        kayit_obj.iptal_eden = 'katilimci'
        db.session.commit()
        return render_template('egitim/kayit_iptal.html', kayit=kayit_obj, iptal_edildi=True)

    return render_template('egitim/kayit_iptal.html', kayit=kayit_obj)


@egitim_bp.route('/anket/<token>', methods=['GET', 'POST'])
def anket(token):
    """Eğitim sonrası memnuniyet anketi (kayıt token'ı ile)."""
    kayit_obj = EgitimKayit.query.filter_by(token=token).first_or_404()

    if kayit_obj.anket:
        return render_template('egitim/anket.html', kayit=kayit_obj,
                               egitim=kayit_obj.egitim, tamamlandi=True)

    if request.method == 'POST':
        def _puan(alan):
            try:
                deger = int(request.form.get(alan) or 0)
            except ValueError:
                return None
            if 1 <= deger <= 5:
                return deger
            return None

        puan = _puan('puan')
        if not puan:
            return render_template('egitim/anket.html', kayit=kayit_obj,
                                   egitim=kayit_obj.egitim,
                                   hata='Lütfen genel memnuniyet puanı verin.')

        db.session.add(EgitimAnket(
            egitim_id=kayit_obj.egitim_id,
            oturum_id=kayit_obj.oturum_id,
            kayit_id=kayit_obj.id,
            puan=puan,
            egitmen_puan=_puan('egitmen_puan'),
            icerik_puan=_puan('icerik_puan'),
            yorum=(request.form.get('yorum') or '').strip() or None,
            ip=_istek_ip(),
        ))
        try:
            db.session.commit()
        except IntegrityError:
            db.session.rollback()  # kayit_id unique - çift gönderim
        return render_template('egitim/anket.html', kayit=kayit_obj,
                               egitim=kayit_obj.egitim, tesekkur=True)

    return render_template('egitim/anket.html', kayit=kayit_obj, egitim=kayit_obj.egitim)
