"""
TG Portal - HedefKadro Excel Export
====================================
Tüm kadro kayıtlarını Excel'e aktarır.
Aktif kadrolar ve silinmiş kadrolar ayrı sheet'lerde.

KULLANIM:
  cd /app && PYTHONPATH=/app python3 scripts/kadro_export.py
"""
import sys
from datetime import datetime


def run():
    from app import create_app, db
    from app.models.proje import Proje, HedefKadro
    from app.models.ik import Calisan
    from app.models.base import CalisanDurumu
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    app = create_app()
    with app.app_context():
        kadrolar = HedefKadro.query.order_by(HedefKadro.proje_id, HedefKadro.pozisyon_adi).all()
        print(f"Toplam kadro: {len(kadrolar)}")

        aktif = [k for k in kadrolar if not k.is_deleted]
        deleted = [k for k in kadrolar if k.is_deleted]
        print(f"Aktif: {len(aktif)}, Deleted: {len(deleted)}")

        wb = openpyxl.Workbook()

        # Stil
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        headers = ['Kadro ID', 'Proje ID', 'Proje Adı', 'Pozisyon Adı',
                   'İl', 'İlçe', 'Bölge', 'Hedef Sayı', 'Mevcut Çalışan',
                   'Doluluk %', 'Aktif']
        widths = [10, 10, 25, 40, 15, 15, 15, 12, 14, 12, 8]

        def write_sheet(ws, rows, title):
            ws.title = title

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row_num = 1
            for k in rows:
                row_num += 1
                proje = Proje.query.get(k.proje_id) if k.proje_id else None
                proje_adi = proje.ad if proje else '-'

                mevcut = Calisan.query.filter(
                    Calisan.kadro_id == k.id,
                    Calisan.is_deleted == False,
                    Calisan.durum == CalisanDurumu.AKTIF
                ).count()

                hedef = k.hedef_sayi or 0
                doluluk = round((mevcut / hedef * 100), 1) if hedef > 0 else 0

                il_adi = ''
                if hasattr(k, 'il_obj') and k.il_obj:
                    il_adi = k.il_obj.ad
                elif k.il:
                    il_adi = k.il

                ilce_adi = ''
                if hasattr(k, 'ilce_obj') and k.ilce_obj:
                    ilce_adi = k.ilce_obj.ad
                elif k.ilce:
                    ilce_adi = k.ilce

                values = [
                    k.id, k.proje_id, proje_adi, k.pozisyon_adi,
                    il_adi, ilce_adi, k.bolge or '',
                    hedef, mevcut, doluluk,
                    'Evet' if not k.is_deleted else 'Hayır'
                ]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.border = thin_border
                    if col == 10:
                        cell.number_format = '0.0'
                        cell.alignment = Alignment(horizontal='center')
                    elif col in (1, 2, 8, 9):
                        cell.alignment = Alignment(horizontal='center')

            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{row_num}'
            print(f"  {title}: {row_num - 1} satır")

        # Sheet 1: Aktif
        ws_aktif = wb.active
        write_sheet(ws_aktif, aktif, 'Aktif Kadrolar')

        # Sheet 2: Silinmiş
        if deleted:
            ws_del = wb.create_sheet()
            write_sheet(ws_del, deleted, 'Silinmiş Kadrolar')

        output = '/app/data/mevcut_kadrolar.xlsx'
        wb.save(output)
        print(f"\nKaydedildi: {output}")


if __name__ == '__main__':
    run()
