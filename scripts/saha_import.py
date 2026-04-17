"""
TG Portal - Saha Calisani Import Script'i
==========================================
Excel'den (data/PERSONEL BILGILERI.xlsx) saha calisanlarini import eder.
- DB temizlik (duplicate departman, musteri)
- Yeni proje olusturma (Marka Elcisi, Modern Kanal Sniper, Paylasimli Merch)
- Calisan kayitlarini guncelleme/olusturma
- Proje/kadro/koordinator eslestirme
- Ayrilanlari AYRILDI durumuna cekme

KULLANIM:
  cd /app && PYTHONPATH=/app python3 scripts/saha_import.py --dry-run
  cd /app && PYTHONPATH=/app python3 scripts/saha_import.py --apply
  cd /app && PYTHONPATH=/app python3 scripts/saha_import.py --force-apply
"""
import sys
import os
from datetime import datetime, date

# Excel dosya yolu (Docker container icinde /app altinda)
EXCEL_PATH = '/app/data/PERSONEL_BILGILERI.xlsx'

# Atlanacak departmanlar (brief'e gore)
ATLA_DEPT = {'Vena', 'Triodor', 'Event', 'Ofis'}

# ============================================================
# PROJE-KOORDINATOR ESLESTIRME TABLOSU (docs/CLAUDE_CODE_BRIEF.md'den)
# ============================================================
# Excel dept adi -> (proje_adi, musteri_kisa_ad, koordinator_ad_soyad)
DEPT_PROJE_MAP = {
    'Blues':                       ('Efes Blues', 'Efes', 'Yakup Ateş'),
    'Blues Spv':                   ('Efes Blues', 'Efes', 'Yakup Ateş'),
    'KK Merch':                   ('Efes KK Merch', 'Efes', 'Kazım Sifoğlu'),
    'Sse':                        ('PMI SSE', 'PMI', 'Erdi Aslantaş'),
    'Sse Spv':                    ('PMI SSE', 'PMI', 'Erdi Aslantaş'),
    'Part Time Sse':              ('PMI SSE', 'PMI', 'Erdi Aslantaş'),
    'Brown Forman':               ('BF Merchandiser', 'Brown Forman', 'Muhammet Aslan'),
    'Beylerbeyi Merch':           ('Beylerbeyi Merch', 'Beylerbeyi', 'Havvanur Mahmutoğlu'),
    'Beylerbeyi Merch Yaya':      ('Beylerbeyi Merch', 'Beylerbeyi', 'Havvanur Mahmutoğlu'),
    'Adco Merch':                 ('Adco Merchandiser', 'Kemer Gıda', 'Ufuk Dinç'),
    'Marka Elçisi':               ('Efes Marka Elçisi', 'Efes', 'Halil Karaoğlan'),
    'Modern Kanal Part Time':     ('Efes Modern Kanal', 'Efes', 'Ufuk Dinç'),
    'Efes Spv':                   ('Efes Sniper', 'Efes', 'Hakan Alpan'),
    'Part Time Sniper':           ('Efes Sniper', 'Efes', 'Ufuk Dinç'),
    'Paylaşımlı Merch (PMI-BF)': ('Paylaşımlı Merch', 'PMI', 'Erdi Aslantaş'),
}

# Bilinen musteri ID'leri (DB'deki gercek degerler)
MUSTERI_ID_MAP = {
    'Efes': 1,
    'PMI': 4,
    'Brown Forman': 5,
}

# Yeni olusturulacak projeler (DB'de yoksa)
YENI_PROJELER = [
    # (proje_adi, musteri_id, kod)
    ('Efes Marka Elçisi', 1, 'ME'),   # Efes
    ('Efes Modern Kanal', 1, 'MKS'),  # Efes
    ('Paylaşımlı Merch', 4, 'PM'),    # PMI (PMI+BF ortak pilot, DB'de PMI)
]

# Pozisyon eslestirme: Excel dept adi -> HedefKadro pozisyon_adi
DEPT_POZISYON_MAP = {
    'Blues':                       'Merchandiser',
    'Blues Spv':                   'Süpervizör',
    'KK Merch':                   'Merchandiser',
    'Sse':                        'SSE',
    'Sse Spv':                    'Süpervizör',
    'Part Time Sse':              'Part Time SSE',
    'Brown Forman':               'Merchandiser',
    'Beylerbeyi Merch':           'Merchandiser',
    'Beylerbeyi Merch Yaya':      'Yaya Merchandiser',
    'Adco Merch':                 'Merchandiser',
    'Marka Elçisi':               'Marka Elçisi',
    'Modern Kanal Part Time':     'Part Time Sniper',
    'Efes Spv':                   'Süpervizör',
    'Part Time Sniper':           'Part Time Sniper',
    'Paylaşımlı Merch (PMI-BF)': 'Merchandiser',
}


def parse_date(val):
    """Excel datetime veya string'den date parse et."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), '%Y-%m-%d').date()
    except Exception:
        return None


def parse_ad_soyad(full_name):
    """'OZAN EREN KAYAN' -> ('OZAN EREN', 'KAYAN')"""
    if not full_name:
        return None, None
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], ''
    return ' '.join(parts[:-1]), parts[-1]


# Ad ile bulunamayan koordinatorler icin email fallback
KOORDINATOR_EMAIL_MAP = {
    'Erdi Aslantaş':      'erdiaslantas@teamguerilla.com',
    'Halil Karaoğlan':    'halilkaraoglan@teamguerilla.com',
    'Kazım Sifoğlu':      'abidinkazimsifoglu@teamguerilla.com',
    'Ufuk Dinç':          'ufukdinc@teamguerilla.com',
}


def find_koordinator(calisan_map, ad_soyad, Calisan=None):
    """Koordinator calisanini ad soyad ile bul, bulunamazsa email ile dene."""
    if not ad_soyad:
        return None
    # 1. Ad-soyad ile ara
    parts = ad_soyad.upper().split()
    for c in calisan_map.values():
        if c and c.ad and c.soyad:
            full = f"{c.ad} {c.soyad}".upper()
            if all(p in full for p in parts):
                return c
    # 2. Email fallback
    email = KOORDINATOR_EMAIL_MAP.get(ad_soyad)
    if email and Calisan:
        c = Calisan.query.filter_by(email=email, is_deleted=False).first()
        if c:
            return c
    return None


def read_excel():
    """Excel'den saha calisanlarini oku."""
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = []

    for row_idx in range(2, ws.max_row + 1):
        dept_adi = ws.cell(row=row_idx, column=9).value
        ad_soyad = ws.cell(row=row_idx, column=4).value

        if not ad_soyad or not dept_adi:
            continue
        if dept_adi in ATLA_DEPT:
            continue

        rows.append({
            'sicil_no':       ws.cell(row=row_idx, column=2).value,
            'tc_kimlik':      ws.cell(row=row_idx, column=3).value,
            'ad_soyad':       ad_soyad.strip(),
            'dept_adi':       dept_adi.strip(),
            'sgk_sube':       ws.cell(row=row_idx, column=10).value or '',
            'sgk_no':         ws.cell(row=row_idx, column=11).value or '',
            'kidem_tarihi':   ws.cell(row=row_idx, column=12).value,
            'giris_tarihi':   ws.cell(row=row_idx, column=13).value,
            'cikis_tarihi':   ws.cell(row=row_idx, column=14).value,
            'telefon':        ws.cell(row=row_idx, column=16).value or '',
            'dogum_tarihi':   ws.cell(row=row_idx, column=17).value,
            'cinsiyet':       ws.cell(row=row_idx, column=18).value or '',
            'egitim_durumu':  ws.cell(row=row_idx, column=19).value or '',
            'email':          ws.cell(row=row_idx, column=20).value or '',
            'yemek_karti':    ws.cell(row=row_idx, column=21).value or '',
            'beden':          ws.cell(row=row_idx, column=22).value or '',
            'kargo_subesi':   ws.cell(row=row_idx, column=23).value or '',
            'adres':          ws.cell(row=row_idx, column=24).value or '',
            'cikis_sebebi':   ws.cell(row=row_idx, column=25).value or '',
            'cikis_nedeni':   ws.cell(row=row_idx, column=26).value or '',
        })

    return rows


def run(mode='dry-run'):
    from app import create_app, db
    from app.models.core import User
    from app.models.ik import Calisan, Departman, Pozisyon
    from app.models.proje import Musteri, Proje, HedefKadro
    from app.models.base import CalisanDurumu

    app = create_app()
    with app.app_context():
        print("=" * 80)
        print(f"  SAHA CALISANI IMPORT [{mode.upper()}]")
        print(f"  Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print("=" * 80)

        # ============================================================
        # ADIM 0: DB TEMIZLIK
        # ============================================================
        print("\n[0] DB TEMIZLIK")
        print("-" * 40)

        # 0a. Duplicate departmanlar (id 1-16 bos, 17-24 aktif)
        dup_depts = Departman.query.filter(Departman.id <= 16).all()
        print(f"  Duplicate departmanlar (id<=16): {len(dup_depts)} kayit")
        for d in dup_depts:
            calisan_count = Calisan.query.filter_by(departman_id=d.id).count()
            print(f"    id={d.id} '{d.ad}' -> {calisan_count} calisan bagli")
            if calisan_count > 0:
                print(f"    !!! DIKKAT: {calisan_count} calisan bagli, silinemez!")
            elif mode == 'apply':
                # Pozisyonlari da temizle
                poz_count = Pozisyon.query.filter_by(departman_id=d.id).count()
                if poz_count > 0:
                    Pozisyon.query.filter_by(departman_id=d.id).delete()
                    print(f"    -> {poz_count} pozisyon silindi")
                db.session.delete(d)
                print(f"    -> Departman silindi")

        # 0b. Duplicate musteri: Brown Forman id=6 (id=5 kalacak)
        dup_musteri = Musteri.query.filter_by(id=6).first()
        if dup_musteri:
            proje_count = Proje.query.filter_by(musteri_id=6).count()
            print(f"  Duplicate musteri: id=6 '{dup_musteri.ad}' -> {proje_count} proje bagli")
            if proje_count > 0:
                print(f"    !!! {proje_count} proje bagli, once projeleri tasimali")
                if mode == 'apply':
                    Proje.query.filter_by(musteri_id=6).update({'musteri_id': 5})
                    print(f"    -> Projeler musteri_id=5'e tasindi")
            if mode == 'apply':
                db.session.delete(dup_musteri)
                print(f"    -> Musteri id=6 silindi")
        else:
            print("  Duplicate musteri id=6 zaten yok")

        if mode == 'apply':
            db.session.flush()

        # ============================================================
        # ADIM 1: MUSTERI KONTROL
        # ============================================================
        print(f"\n[1] MUSTERILER")
        print("-" * 40)
        for label, mid in MUSTERI_ID_MAP.items():
            m = Musteri.query.get(mid)
            if m:
                print(f"  [OK] {label} -> id={mid} '{m.ad}'")
            else:
                print(f"  [!!!] {label} -> id={mid} BULUNAMADI")

        # ============================================================
        # ADIM 2: PROJELER
        # ============================================================
        print(f"\n[2] PROJELER")
        print("-" * 40)
        projeler = Proje.query.filter_by(is_deleted=False).all()
        proje_map = {}  # proje_adi -> Proje
        for p in projeler:
            proje_map[p.ad] = p
            print(f"  [VAR] id={p.id} '{p.ad}' (musteri_id={p.musteri_id})")

        for proje_adi, musteri_id, kod in YENI_PROJELER:
            if proje_adi not in proje_map:
                print(f"  [YENI] '{proje_adi}' -> musteri_id={musteri_id}, kod={kod}")
                if mode == 'apply':
                    p = Proje(ad=proje_adi, kod=kod, musteri_id=musteri_id,
                              aktif=True, durum='aktif', is_deleted=False,
                              created_at=datetime.now(), updated_at=datetime.now())
                    db.session.add(p)
                    db.session.flush()
                    proje_map[proje_adi] = p

        # ============================================================
        # ADIM 3: HEDEF KADROLAR (proje basina pozisyon_adi)
        # ============================================================
        print(f"\n[3] HEDEF KADROLAR")
        print("-" * 40)
        kadro_map = {}  # (proje_adi, pozisyon_adi) -> HedefKadro

        # Mevcut kadrolari yukle
        for k in HedefKadro.query.filter_by(is_deleted=False).all():
            proje = Proje.query.get(k.proje_id)
            if proje:
                key = (proje.ad, k.pozisyon_adi)
                kadro_map[key] = k

        # Her dept icin kadro olustur
        needed_kadros = set()
        for dept_adi, (proje_adi, _, _) in DEPT_PROJE_MAP.items():
            poz_adi = DEPT_POZISYON_MAP.get(dept_adi, 'Merchandiser')
            needed_kadros.add((proje_adi, poz_adi))

        for proje_adi, poz_adi in sorted(needed_kadros):
            key = (proje_adi, poz_adi)
            if key in kadro_map:
                k = kadro_map[key]
                print(f"  [VAR] '{proje_adi}' / '{poz_adi}' (id={k.id})")
            else:
                proje = proje_map.get(proje_adi)
                proje_id = proje.id if proje else None
                print(f"  [YENI] '{proje_adi}' / '{poz_adi}' -> proje_id={proje_id}")
                if mode == 'apply' and proje_id:
                    k = HedefKadro(proje_id=proje_id, pozisyon_adi=poz_adi,
                                   hedef_sayi=50, aktif=True, is_deleted=False,
                                   created_at=datetime.now(), updated_at=datetime.now())
                    db.session.add(k)
                    db.session.flush()
                    kadro_map[key] = k

        # ============================================================
        # ADIM 4: KOORDINATOR ESLESTIRME HAZIRLIGI
        # ============================================================
        print(f"\n[4] KOORDINATOR HAZIRLIGI")
        print("-" * 40)
        # Ofis calisanlarini (koordinatorler) yukle
        koordinator_map = {}  # ad_soyad -> Calisan
        ofis_calisanlar = Calisan.query.filter(
            Calisan.is_deleted == False
        ).all()
        for c in ofis_calisanlar:
            full = f"{c.ad} {c.soyad}"
            koordinator_map[full] = c

        # Brief'teki koordinatorleri kontrol et
        koordinator_ids = {}  # koordinator_ad -> calisan_id
        all_koord_names = set(v[2] for v in DEPT_PROJE_MAP.values())
        for name in sorted(all_koord_names):
            c = find_koordinator(koordinator_map, name, Calisan=Calisan)
            if c:
                koordinator_ids[name] = c.id
                print(f"  [OK] {name} -> calisan_id={c.id}")
            else:
                koordinator_ids[name] = None
                print(f"  [!!!] {name} -> BULUNAMADI")

        # ============================================================
        # ADIM 5: EXCEL'DEN SAHA CALISANLARI
        # ============================================================
        print(f"\n[5] SAHA CALISANLARI")
        print("-" * 40)

        rows = read_excel()
        print(f"  Excel'den {len(rows)} saha kaydi okundu")
        print()

        stats = {
            'yeni': 0, 'guncellenen': 0, 'ayrilan': 0,
            'kadro_atanan': 0, 'koordinator_atanan': 0,
            'bilinmeyen_dept': 0,
        }

        for r in rows:
            sicil = r['sicil_no']
            tc = r['tc_kimlik']
            ad, soyad = parse_ad_soyad(r['ad_soyad'])
            dept_adi = r['dept_adi']
            cikis = parse_date(r['cikis_tarihi'])

            # Proje/kadro/koordinator bilgisi
            proje_info = DEPT_PROJE_MAP.get(dept_adi)
            if not proje_info:
                print(f"  [?] {sicil} {ad} {soyad} -> bilinmeyen dept: '{dept_adi}'")
                stats['bilinmeyen_dept'] += 1
                continue

            proje_adi, _, koord_adi = proje_info
            poz_adi = DEPT_POZISYON_MAP.get(dept_adi, 'Merchandiser')
            kadro_key = (proje_adi, poz_adi)
            kadro = kadro_map.get(kadro_key)
            koord_id = koordinator_ids.get(koord_adi)

            # Mevcut calisan ara (sicil_no veya tc_kimlik ile)
            calisan = None
            if sicil:
                calisan = Calisan.query.filter_by(sicil_no=sicil).first()
            if not calisan and tc:
                calisan = Calisan.query.filter_by(tc_kimlik=tc).first()

            durum = CalisanDurumu.AYRILDI if cikis else CalisanDurumu.AKTIF

            if calisan:
                # GUNCELLE
                changes = []
                kadro_id = kadro.id if kadro else None
                if calisan.kadro_id != kadro_id and kadro_id:
                    changes.append(f"kadro={kadro_id}")
                if calisan.yonetici_id != koord_id and koord_id:
                    changes.append(f"yon={koord_id}")
                if str(calisan.durum) != str(durum):
                    changes.append(f"durum={durum.value}")
                if cikis and calisan.isten_ayrilma != cikis:
                    changes.append(f"cikis={cikis}")

                if changes:
                    print(f"  [UPD] {sicil} {ad} {soyad} -> {', '.join(changes)}")
                    stats['guncellenen'] += 1
                    if mode == 'apply':
                        if kadro_id:
                            calisan.kadro_id = kadro_id
                            stats['kadro_atanan'] += 1
                        if koord_id:
                            calisan.yonetici_id = koord_id
                            stats['koordinator_atanan'] += 1
                        calisan.durum = durum
                        if cikis:
                            calisan.isten_ayrilma = cikis
                            ayrilma = r['cikis_nedeni'] or r['cikis_sebebi']
                            if ayrilma and ayrilma != '0':
                                calisan.ayrilma_nedeni = str(ayrilma).strip()
                            stats['ayrilan'] += 1
                        calisan.updated_at = datetime.now()
            else:
                # YENI CALISAN
                email = r['email'].strip().lower() if r['email'] else ''
                print(f"  [YENI] {sicil} {ad} {soyad} | {dept_adi} | {proje_adi} | durum={durum.value}")
                stats['yeni'] += 1

                if mode == 'apply':
                    kadro_id = kadro.id if kadro else None
                    calisan = Calisan(
                        sicil_no=sicil or None,
                        ad=ad,
                        soyad=soyad,
                        tc_kimlik=tc or None,
                        dogum_tarihi=parse_date(r['dogum_tarihi']),
                        cinsiyet=r['cinsiyet'] or None,
                        email=email or None,
                        telefon=str(r['telefon']).strip() if r['telefon'] else None,
                        adres=r['adres'].strip() if r['adres'] else None,
                        egitim_durumu=r['egitim_durumu'] or None,
                        yemek_karti=r['yemek_karti'] or None,
                        beden=r['beden'] or None,
                        kargo_subesi=r['kargo_subesi'] or None,
                        kidem_tarihi=parse_date(r['kidem_tarihi']),
                        ise_baslama=parse_date(r['giris_tarihi']),
                        isten_ayrilma=cikis,
                        kadro_id=kadro_id,
                        yonetici_id=koord_id,
                        durum=durum,
                        is_deleted=False,
                        created_at=datetime.now(),
                        updated_at=datetime.now(),
                    )
                    if cikis:
                        ayrilma = r['cikis_nedeni'] or r['cikis_sebebi']
                        if ayrilma and ayrilma != '0':
                            calisan.ayrilma_nedeni = str(ayrilma).strip()
                        stats['ayrilan'] += 1
                    if kadro_id:
                        stats['kadro_atanan'] += 1
                    if koord_id:
                        stats['koordinator_atanan'] += 1
                    db.session.add(calisan)

        # ============================================================
        # COMMIT
        # ============================================================
        if mode == 'apply':
            db.session.commit()
            print(f"\n>>> TUM DEGISIKLIKLER KAYDEDILDI <<<")
        else:
            print(f"\n>>> DRY-RUN: Hicbir degisiklik yapilmadi <<<")

        # ============================================================
        # OZET
        # ============================================================
        print(f"\n{'=' * 80}")
        print(f"  OZET")
        print(f"  Excel satir: {len(rows)}")
        print(f"  Yeni calisan: {stats['yeni']}")
        print(f"  Guncellenen: {stats['guncellenen']}")
        print(f"  Ayrilan (AYRILDI): {stats['ayrilan']}")
        print(f"  Kadro atanan: {stats['kadro_atanan']}")
        print(f"  Koordinator atanan: {stats['koordinator_atanan']}")
        print(f"  Bilinmeyen dept: {stats['bilinmeyen_dept']}")
        print(f"{'=' * 80}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Kullanim:")
        print("  python3 scripts/saha_import.py --dry-run")
        print("  python3 scripts/saha_import.py --apply")
        print("  python3 scripts/saha_import.py --force-apply")
        sys.exit(1)

    arg = sys.argv[1]
    if arg == '--dry-run':
        run(mode='dry-run')
    elif arg == '--apply':
        print("DIKKAT: Asagidaki islemler yapilacak:")
        print("  - Duplicate departmanlar (id<=16) silinecek")
        print("  - Duplicate musteri (id=6) silinecek")
        print("  - Yeni projeler olusturulacak")
        print("  - ~309 saha calisani import/guncellenecek")
        print("  - Ayrilanlar AYRILDI durumuna cekilecek")
        print()
        print("Devam etmek icin 'EVET' yazin: ", end='')
        onay = input().strip()
        if onay == 'EVET':
            run(mode='apply')
        else:
            print("Iptal edildi.")
    elif arg == '--force-apply':
        run(mode='apply')
    else:
        print(f"Bilinmeyen arguman: {arg}")
        sys.exit(1)
