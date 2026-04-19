"""
TG Portal - Kadro Eslestirme Script'i
======================================
Kadro_Eslestirme_Formu.xlsx dosyasindan kadro_id'leri okuyup
calisanlarin kadro_id alanini gunceller.

Her sheet bir koordinatorun ekibi. Row 3 header:
  A: Personel Kodu (sicil_no)
  F: KADRO SEÇİMİ (DOLDURUN) -> kadro_id

KULLANIM:
  cd /app && PYTHONPATH=/app python3 scripts/kadro_eslestir.py --dry-run
  cd /app && PYTHONPATH=/app python3 scripts/kadro_eslestir.py --apply
"""
import sys
import os
import unicodedata

EXCEL_PATH = '/app/data/Kadro_Eslestirme_Formu.xlsx'
SKIP_SHEETS = {'KADRO LİSTESİ'}


def normalize_tr(text):
    """Turkce karakter normalize: kucuk harf + unicode NFKC"""
    if not text:
        return ''
    # Turkce buyuk harfleri kucult (İ->i, I->ı ozel durumlari)
    text = text.replace('İ', 'i').replace('I', 'ı')
    text = text.lower()
    text = unicodedata.normalize('NFKC', text)
    return text.strip()


def main():
    apply = '--apply' in sys.argv
    mode = 'APPLY' if apply else 'DRY-RUN'
    print(f'=== Kadro Eslestirme ({mode}) ===\n')

    # Flask app context
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from app import create_app, db
    from app.models.ik import Calisan
    from app.models.proje import HedefKadro

    app = create_app()

    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    toplam = 0
    guncellenen = 0
    atlanan = 0
    bulunamayan = []
    bos_kadro = []
    zaten_ayni = 0

    with app.app_context():
        # Mevcut kadro_id'leri dogrula
        mevcut_kadro_ids = set(k.id for k in HedefKadro.query.all())
        print(f'Veritabaninda {len(mevcut_kadro_ids)} hedef kadro mevcut.')

        # Ad+soyad fallback icin tum calisanlari once yukle
        tum_calisanlar = Calisan.query.all()
        print(f'Veritabaninda {len(tum_calisanlar)} calisan mevcut.\n')

        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=4, values_only=True))  # skip header (row 1-3)
            print(f'--- {sheet_name} ({len(rows)} satir) ---')

            for row in rows:
                if not row or not row[0]:
                    continue

                sicil_no = str(row[0]).strip()
                ad_soyad = row[1] if len(row) > 1 else '?'
                kadro_val = row[5] if len(row) > 5 else None

                toplam += 1

                # Bos kadro secimi
                if kadro_val is None or str(kadro_val).strip() == '':
                    bos_kadro.append(f'  [BOS] {sicil_no} - {ad_soyad}')
                    atlanan += 1
                    continue

                # kadro_id integer olmali
                try:
                    kadro_id = int(kadro_val)
                except (ValueError, TypeError):
                    print(f'  [HATA] {sicil_no} - Gecersiz kadro degeri: {kadro_val}')
                    atlanan += 1
                    continue

                # Kadro var mi kontrol
                if kadro_id not in mevcut_kadro_ids:
                    print(f'  [HATA] {sicil_no} - Kadro ID {kadro_id} veritabaninda yok!')
                    atlanan += 1
                    continue

                # Calisani bul: once sicil_no, sonra ad+soyad fallback
                calisan = Calisan.query.filter_by(sicil_no=sicil_no).first()
                if not calisan and ad_soyad and ad_soyad != '?':
                    # Ad soyad ile fallback ara
                    parcalar = str(ad_soyad).strip().split(None, 1)
                    if len(parcalar) == 2:
                        aranan_ad, aranan_soyad = parcalar
                    else:
                        aranan_ad, aranan_soyad = parcalar[0], ''

                    # Case-insensitive + Turkce normalize arama
                    eslesenler = [
                        c for c in tum_calisanlar
                        if normalize_tr(c.ad) == normalize_tr(aranan_ad)
                        and normalize_tr(c.soyad) == normalize_tr(aranan_soyad)
                    ]

                    if len(eslesenler) == 1:
                        calisan = eslesenler[0]
                        print(f'  [FALLBACK] {sicil_no} -> ad/soyad ile bulundu: {calisan.ad} {calisan.soyad} (id={calisan.id})')
                    elif len(eslesenler) > 1:
                        ids = [str(c.id) for c in eslesenler]
                        print(f'  [COKLU] {sicil_no} - {ad_soyad}: {len(eslesenler)} eslesme (id: {", ".join(ids)}) - ATLANDI')
                        atlanan += 1
                        continue

                if not calisan:
                    bulunamayan.append(f'  [YOK] {sicil_no} - {ad_soyad}')
                    atlanan += 1
                    continue

                # Zaten ayni kadro mu?
                if calisan.kadro_id == kadro_id:
                    zaten_ayni += 1
                    continue

                eski = calisan.kadro_id
                print(f'  {sicil_no} ({calisan.ad} {calisan.soyad}): kadro {eski} -> {kadro_id}')

                if apply:
                    calisan.kadro_id = kadro_id

                guncellenen += 1

        if apply:
            db.session.commit()
            print('\nDegisiklikler veritabanina kaydedildi.')

    wb.close()

    # Ozet
    print(f'\n=== OZET ({mode}) ===')
    print(f'Toplam satir:       {toplam}')
    print(f'Guncellenecek:      {guncellenen}')
    print(f'Zaten dogru:        {zaten_ayni}')
    print(f'Atlanan:            {atlanan}')

    if bulunamayan:
        print(f'\nBulunamayan calisanlar ({len(bulunamayan)}):')
        for b in bulunamayan:
            print(b)

    if bos_kadro:
        print(f'\nBos kadro secimi ({len(bos_kadro)}):')
        for b in bos_kadro:
            print(b)

    if not apply and guncellenen > 0:
        print(f'\nGercek guncelleme icin: python3 scripts/kadro_eslestir.py --apply')


if __name__ == '__main__':
    main()
